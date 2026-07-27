"""_build_promptbakeoff_xlsx.py -- reviewable spreadsheet for the prompt-bakeoff
raw captions (see PLAN: prompt-bakeoff-harness, 2026-07-27). Reuses the styling
conventions of _build_caption_xlsx.py (same header fill, borders, green/red
verdict coloring) so this reads as the same family of artifact, and reuses
semsup_caption_qa.py's banned-word/token-limit logic directly rather than
duplicating the rules in a second place.

Does NOT overwrite Caption_Train_All_Clips.xlsx (the 267-row incumbent record) -
writes a new file, Caption_PromptBakeoff_300.xlsx (or whatever --out is given;
the row count follows whatever --input actually contains).
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "student_training" / "scripts"))
from semsup_caption_qa import BANNED_RE, MAX_TOKENS  # noqa: E402

DEFAULT_OUT = HERE / "Caption_PromptBakeoff_300.xlsx"

# Styling constants copied verbatim from _build_caption_xlsx.py, so this file
# reads as the same family of artifact as the incumbent caption spreadsheet.
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
STRONG_GREEN = PatternFill("solid", fgColor="00B050")
RED = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# New QA-specific colors (deliverable 3b) - distinct from the verdict red/green
# above so a reviewer never confuses "wrong verdict" with "failed a QA gate".
BANNED_WORD_FILL = PatternFill("solid", fgColor="FFC000")   # amber
OVER_TOKEN_FILL = PatternFill("solid", fgColor="B4A7D6")    # lavender

HEADERS = ["video_id", "requested_time_to_event", "gt_verdict", "verdict", "confidence",
           "caption_neutral", "risk_clause", "arm_a", "arm_b", "arm_c",
           "n_tokens_a", "n_tokens_b"]
WIDTHS = [10, 20, 11, 9, 11, 45, 30, 45, 55, 30, 12, 12]
WRAP_COLS = {"caption_neutral", "risk_clause", "arm_a", "arm_b", "arm_c"}


def _norm(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    if s in ("1", "YES", "TRUE"):
        return "YES"
    if s in ("0", "NO", "FALSE"):
        return "NO"
    return s or None


def build_records(raw_path: Path, manifest_path: Path, tok) -> list:
    raw_rows = [json.loads(l) for l in open(raw_path, encoding="utf-8") if l.strip()]
    manifest_rows = [json.loads(l) for l in open(manifest_path, encoding="utf-8") if l.strip()]
    manifest_by_vid = {r["video_id"]: r for r in manifest_rows}

    records = []
    for r in raw_rows:
        m = manifest_by_vid.get(r["video_id"])
        if m is None:
            continue
        neutral = r["caption_neutral"]
        risk = r["risk_clause"]
        arm_a = neutral
        arm_b = neutral + ", " + risk
        target = m["target"]
        gt_verdict = "YES" if target == 1 else "NO"
        secs = {"TTE_0.5": 0.5, "TTE_1.0": 1.0, "TTE_1.5": 1.5}.get(m.get("horizon_label"))
        arm_c = (f"collision imminent in {secs} seconds" if target == 1 and secs is not None
                 else "collision imminent" if target == 1 else "normal driving, no collision")
        records.append({
            "video_id": r["video_id"],
            "requested_time_to_event": m.get("requested_time_to_event"),
            "gt_verdict": gt_verdict,
            "verdict": _norm(r.get("verdict")),
            "confidence": r.get("confidence"),
            "caption_neutral": neutral,
            "risk_clause": risk,
            "arm_a": arm_a, "arm_b": arm_b, "arm_c": arm_c,
            "n_tokens_a": len(tok(arm_a)["input_ids"]),
            "n_tokens_b": len(tok(arm_b)["input_ids"]),
            "_banned_word_hit": bool(BANNED_RE.search(neutral)),
        })
    records.sort(key=lambda r: (r["video_id"], str(r["requested_time_to_event"])))
    return records


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", required=True, help="raw teacher-output JSONL (new schema)")
    ap.add_argument("--manifest", required=True, help="sampler manifest from semsup_sample_clips.py")
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--siglip-model", default="google/siglip-base-patch16-224")
    args = ap.parse_args()

    from transformers import AutoTokenizer
    tok = AutoTokenizer.from_pretrained(args.siglip_model)

    records = build_records(Path(args.input), Path(args.manifest), tok)

    wb = Workbook()
    ws = wb.active
    ws.title = "prompt_bakeoff"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    n_match = n_mismatch = n_banned = n_overtoken = 0
    for i, rec in enumerate(records, start=2):
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                        horizontal="left" if wrap else "center")
            cell.border = border

        gt, vd = rec["gt_verdict"], rec["verdict"]
        vcell = ws.cell(row=i, column=HEADERS.index("verdict") + 1)
        if gt and vd:
            if vd == gt:
                vcell.fill = STRONG_GREEN
                n_match += 1
            else:
                vcell.fill = RED
                n_mismatch += 1

        if rec["_banned_word_hit"]:
            ws.cell(row=i, column=HEADERS.index("caption_neutral") + 1).fill = BANNED_WORD_FILL
            n_banned += 1
        for col in ("n_tokens_a", "n_tokens_b"):
            if rec[col] > MAX_TOKENS:
                ws.cell(row=i, column=HEADERS.index(col) + 1).fill = OVER_TOKEN_FILL
                n_overtoken += 1

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"wrote {len(records)} rows -> {out_path}")
    print(f"verdict vs gt_verdict: match(green)={n_match}  mismatch(red)={n_mismatch}")
    print(f"QA flags: banned-word(amber)={n_banned}  over-{MAX_TOKENS}-token(lavender)={n_overtoken}")


if __name__ == "__main__":
    main()
