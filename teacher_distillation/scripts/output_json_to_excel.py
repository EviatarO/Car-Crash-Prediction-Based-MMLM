import argparse
import json
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re


def _normalize_verdict(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    if text in {"YES", "NO"}:
        return text
    return None


def _is_mismatch(target: object, gt_verdict: object) -> bool:
    verdict = _normalize_verdict(gt_verdict)
    try:
        target_int = int(target)
    except Exception:
        return False
    if verdict is None:
        return False
    # Requested rule:
    # (gt_verdict == YES and target == 0) or (gt_verdict == NO and target == 1)
    return (verdict == "YES" and target_int == 0) or (verdict == "NO" and target_int == 1)


def _is_pred_mismatch(target: object, pred_verdict: object) -> bool:
    verdict = _normalize_verdict(pred_verdict)
    try:
        target_int = int(target)
    except Exception:
        return False
    if verdict is None:
        return False
    return (verdict == "YES" and target_int == 0) or (verdict == "NO" and target_int == 1)


def _extract_json_object_from_text(text: str) -> Optional[Dict]:
    if not text:
        return None

    cleaned = text.strip()
    # Handle fenced blocks like: ```json ... ```
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)
        cleaned = cleaned.strip()

    # Try direct parse first.
    try:
        parsed = json.loads(cleaned)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        pass

    # Fallback: extract first {...} object from text.
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    try:
        parsed = json.loads(cleaned[start : end + 1])
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        return None
    return None


def _extract_causal_reasoning(obj: Dict) -> Optional[str]:
    top_level = obj.get("causal_reasoning")
    if top_level:
        return str(top_level).strip()

    teacher_text = obj.get("teacher_text")
    if not teacher_text:
        return None

    parsed = _extract_json_object_from_text(str(teacher_text))
    if not parsed:
        return None
    value = parsed.get("causal_reasoning")
    return str(value).strip() if value else None


def _extract_gt_verdict(obj: Dict) -> Optional[str]:
    top_level = _normalize_verdict(obj.get("gt_verdict"))
    if top_level:
        return top_level

    teacher_text = obj.get("teacher_text")
    if not teacher_text:
        return None
    parsed = _extract_json_object_from_text(str(teacher_text))
    if not parsed:
        return None
    return _normalize_verdict(parsed.get("collision_verdict"))


def _extract_collision_verdict(obj: Dict) -> Optional[str]:
    top_level = _normalize_verdict(obj.get("collision_verdict"))
    if top_level:
        return top_level
    teacher_text = obj.get("teacher_text")
    if not teacher_text:
        return None
    parsed = _extract_json_object_from_text(str(teacher_text))
    if not parsed:
        return None
    return _normalize_verdict(parsed.get("collision_verdict"))


def _extract_field(obj: Dict, key: str) -> Optional[object]:
    top_level = obj.get(key)
    if top_level is not None:
        return top_level
    teacher_text = obj.get("teacher_text")
    if not teacher_text:
        return None
    parsed = _extract_json_object_from_text(str(teacher_text))
    if not parsed:
        return None
    return parsed.get(key)


def output_json_to_excel(input_jsonl: Path, output_xlsx: Path) -> int:
    rows: List[Dict] = []
    with open(input_jsonl, "r", encoding="utf-8") as f:
        for line in f:
            raw = line.strip()
            if not raw:
                continue
            obj = json.loads(raw)
            gt_verdict = _extract_gt_verdict(obj)
            collision_verdict = _extract_collision_verdict(obj)
            target = obj.get("target")
            requested_tte = obj.get("requested_time_to_event")
            time_to_event = obj.get("time_to_event")
            if target == 0:
                if requested_tte is None:
                    requested_tte = "TN_MIDPOINT"
                if time_to_event is None:
                    time_to_event = "N/A_NO_EVENT"
            dynamic_objects = _extract_field(obj, "dynamic_objects")
            if isinstance(dynamic_objects, list):
                parts = []
                for item in dynamic_objects:
                    if isinstance(item, dict):
                        t = item.get("type", "?")
                        p = item.get("position", "?")
                        f = item.get("feature", "")
                        parts.append(f"{t} ({p}{', ' + f if f else ''})")
                    else:
                        parts.append(str(item))
                dynamic_objects = " | ".join(parts) if parts else None
            elif isinstance(dynamic_objects, dict):
                dynamic_objects = json.dumps(dynamic_objects, ensure_ascii=True)
            rows.append(
                {
                    "video_id": obj.get("video_id"),
                    "target": target,
                    "gt_verdict": gt_verdict,
                    "collision_verdict": collision_verdict,
                    "confidence": _extract_field(obj, "confidence"),
                    "t_seconds": obj.get("t_seconds"),
                    "time_to_event": time_to_event,
                    "requested_time_to_event": requested_tte,
                    "scene_context": _extract_field(obj, "scene_context"),
                    "dynamic_objects": dynamic_objects,
                    "temporal_analysis": _extract_field(obj, "temporal_analysis"),
                    "occlusion_check": _extract_field(obj, "occlusion_check"),
                    "time_to_contact": _extract_field(obj, "time_to_contact"),
                    "verdict_reasoning": _extract_field(obj, "verdict_reasoning"),
                    "causal_reasoning": _extract_causal_reasoning(obj),
                    "error": obj.get("error"),
                    "_mismatch": _is_pred_mismatch(obj.get("target"), collision_verdict)
                    or _is_mismatch(obj.get("target"), gt_verdict),
                }
            )

    if not rows:
        # Create an empty Excel with headers for consistency.
        empty_df = pd.DataFrame(
            columns=[
                "video_id",
                "target",
                "gt_verdict",
                "collision_verdict",
                "confidence",
                "t_seconds",
                "time_to_event",
                "requested_time_to_event",
                "scene_context",
                "dynamic_objects",
                "temporal_analysis",
                "occlusion_check",
                "time_to_contact",
                "verdict_reasoning",
                "causal_reasoning",
                "error",
            ]
        )
        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        empty_df.to_excel(output_xlsx, index=False)
        return 0

    df = pd.DataFrame(rows)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.drop(columns=["_mismatch"]).to_excel(output_xlsx, index=False)

    # Highlight entire row in red when mismatch rule is met.
    wb = load_workbook(output_xlsx)
    ws = wb.active
    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")

    for idx, mismatch in enumerate(df["_mismatch"].tolist(), start=2):  # row 1 is header
        if mismatch:
            for col in range(1, len(df.columns)):  # all output columns except _mismatch
                ws.cell(row=idx, column=col).fill = red_fill

    wb.save(output_xlsx)
    return len(df)


def output_jason_to_excel(input_jsonl: Path, output_xlsx: Path) -> int:
    # Backward-compat alias for requested name with original typo.
    return output_json_to_excel(input_jsonl, output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert teacher_dataset.jsonl to Excel with mismatch highlighting."
    )
    parser.add_argument(
        "--input_jsonl",
        type=Path,
        default=Path("outputs/teacher_dataset.jsonl"),
        help="Path to teacher dataset JSONL file.",
    )
    parser.add_argument(
        "--output_xlsx",
        type=Path,
        default=Path("outputs/teacher_dataset.xlsx"),
        help="Path to output Excel file.",
    )
    args = parser.parse_args()

    if not args.input_jsonl.exists():
        raise RuntimeError(f"Input file not found: {args.input_jsonl}")

    count = output_json_to_excel(args.input_jsonl, args.output_xlsx)
    print(f"Wrote {count} rows to {args.output_xlsx}")


if __name__ == "__main__":
    main()
