"""update_v6_analysis_with_v71.py — non-destructive v7.1 column addition.

Reads:
  outputs/prompt_bakeoff/reasoning_analysis_v6_debate.xlsx   (existing, untouched)
  outputs/prompt_bakeoff/v7_1_s4_ab/v7_1_debate.jsonl        (all 5 v7.1 results)
  outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl  (18 first-pass)
  dataset/teacher_dataset_GT_self_imply.xlsx                  (GT, for gt_verdict lookup)

Writes:
  outputs/prompt_bakeoff/reasoning_analysis_v6_debate.xlsx
    - per_clip sheet: 10 new columns appended (v7.1 recovery + final)
    - summary sheet: new rows for v7.1 stage
  outputs/prompt_bakeoff/reasoning_analysis_v6_debate.md
    - original content preserved; v7.1 comparison section appended

Does NOT delete or overwrite any existing per_clip or summary rows/columns.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT = REPO_ROOT / "outputs" / "prompt_bakeoff"

XLSX_IN  = OUT / "reasoning_analysis_v6_debate.xlsx"
V71_JSONL = OUT / "v7_1_s4_ab" / "v7_1_debate.jsonl"
HIRES_1  = OUT / "highres_test.jsonl"
HIRES_2  = OUT / "v6_hires_full18.jsonl"
MD_FILE  = OUT / "reasoning_analysis_v6_debate.md"

# ---------------------------------------------------------------------------
# v7.1 reasoning scores (qualitative, 0–10, assessed vs GT verdict_reasoning_en)
# ---------------------------------------------------------------------------
# rubric (same as v6 analysis):
#   10 = paraphrase of GT (verdict + agent + mechanism + outcome)
#    8-9 = same verdict + same agent + same mechanism
#    6-7 = correct verdict + correct agent, mechanism partially off
#    4-5 = correct verdict, wrong agent / mechanism
#    2-3 = wrong verdict, partial scene match
#    0-1 = wrong verdict + wrong scene / hallucinated agents
V71_SCORES = {
    "00474": (5, "FIXED verdict YES; cites yellow taxi gap-erosion (not white-van merge per GT); "
                 "correct verdict and plausible mechanism but wrong agent identified"),
    "01153": (1, "Still-wrong YES; same white-sedan-left-turn hallucination as v6; "
                 "GT says NO, scene is a right-turn with no crossing conflict"),
    "01504": (2, "Still-wrong YES; same red-SUV high-closing-speed narrative as v6; "
                 "GT says NO and ego brakes in time; wrong verdict, persistent hallucination"),
    "02104": (8, "FIXED verdict NO; stable following distance + parallel adjacent lanes "
                 "correctly described; solid match to GT 'reasonable distances maintained'"),
    "02117": (8, "FIXED verdict NO; normal stable flow, safely passed right vehicle; "
                 "well matches GT 'gray sedan ahead at constant distance'"),
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_jsonl(path: Path) -> dict:
    if not path.exists():
        return {}
    out = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            r = json.loads(line)
            out[r["video_id"]] = r
        except Exception:
            pass
    return out


def _merge_first_pass() -> dict:
    result = {}
    for p in (HIRES_1, HIRES_2):
        for vid, rec in _load_jsonl(p).items():
            result[vid] = rec
    return result


# ---------------------------------------------------------------------------
# Excel update
# ---------------------------------------------------------------------------

NEW_COLS = [
    "v7_1_recovery_verdict",
    "v7_1_recovery_correct",
    "v7_1_recovery_bert",
    "v7_1_recovery_score",
    "v7_1_recovery_rationale",
    "v7_1_recovery_reasoning",
    "final_after_v7_1_verdict",
    "final_after_v7_1_correct",
    "final_after_v7_1_score",
    "flipped_by_v7_1",
]

# Colors
GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
HEADER = PatternFill("solid", fgColor="2E75B6")
WHITE  = PatternFill("solid", fgColor="FFFFFF")


def _update_excel(v71: dict, first: dict) -> None:
    wb = openpyxl.load_workbook(XLSX_IN)

    # ---- per_clip sheet ----
    ws = wb["per_clip"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Check if v7.1 columns already exist
    if "v7_1_recovery_verdict" in header:
        print("  v7.1 columns already present in per_clip — overwriting values only")
        start_col = header.index("v7_1_recovery_verdict") + 1
    else:
        start_col = ws.max_column + 1
        for offset, col_name in enumerate(NEW_COLS):
            cell = ws.cell(1, start_col + offset, value=col_name)
            cell.fill = HEADER
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(wrap_text=True)

    # Build video_id -> row mapping
    vid_col = header.index("video_id") + 1 if "video_id" in header else 1
    gt_col  = header.index("gt_verdict") + 1 if "gt_verdict" in header else 2
    v6v_col = header.index("v6_hires__verdict") + 1 if "v6_hires__verdict" in header else 4

    vid_to_row = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vid = row[vid_col - 1].value
        if vid:
            vid_to_row[str(vid).strip()] = row[0].row

    for vid, data_row in sorted(vid_to_row.items()):
        gt = ws.cell(data_row, gt_col).value or ""
        v6_verdict = ws.cell(data_row, v6v_col).value or ""

        d = v71.get(vid)
        if d:
            rv = d.get("recovery_verdict") or ""
            rc = (rv == gt)
            rbert = round(d.get("scores", {}).get("alignment", 0.0) or 0.0, 3)
            rscore, rrat = V71_SCORES.get(vid, (None, ""))
            rreasoning = (d.get("recovery_reasoning") or "").strip()

            # Final: if debated, use recovery; else v6
            final_v = rv or v6_verdict
            final_c = (final_v == gt)

            # Score for final
            if rc and not (v6_verdict == gt):
                final_score = rscore
                flip = "FIXED"
            elif (v6_verdict == gt) and not rc:
                final_score = rscore
                flip = "BROKE"
            else:
                final_score = rscore
                flip = "still-wrong" if not (v6_verdict == gt) else "no-flip"
        else:
            rv = ""
            rc = ""
            rbert = ""
            rscore = None
            rrat = ""
            rreasoning = ""
            final_v = v6_verdict
            final_c = (final_v == gt)
            final_score = None
            flip = ""

        values = [rv, rc, rbert, rscore, rrat, rreasoning, final_v, final_c, final_score, flip]
        for offset, val in enumerate(values):
            cell = ws.cell(data_row, start_col + offset, value=val)
            # Color coding
            col_name = NEW_COLS[offset]
            if col_name in ("v7_1_recovery_correct", "final_after_v7_1_correct"):
                if val is True:
                    cell.fill = GREEN
                elif val is False:
                    cell.fill = RED
            elif col_name == "flipped_by_v7_1":
                if val == "FIXED":
                    cell.fill = GREEN
                elif val == "BROKE":
                    cell.fill = RED
                elif val == "still-wrong":
                    cell.fill = ORANGE
            elif col_name in ("v7_1_recovery_score", "final_after_v7_1_score") and val is not None:
                if val >= 8:
                    cell.fill = GREEN
                elif val <= 2:
                    cell.fill = RED
                elif val <= 4:
                    cell.fill = ORANGE
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # auto-width for new cols
    for offset in range(len(NEW_COLS)):
        col_letter = get_column_letter(start_col + offset)
        ws.column_dimensions[col_letter].width = 20

    # ---- summary sheet ----
    ws_sum = wb["summary"]
    # Find last row
    last_row = ws_sum.max_row

    # v7.1 recovery stats (only the 5 debated clips)
    v71_scores_all = [V71_SCORES[v][0] for v in V71_SCORES]
    v71_fixed = sum(1 for vid, d in v71.items()
                    if d.get("recovery_verdict") == first.get(vid, {}).get("gt_verdict"))
    v71_acc_debated = f"{v71_fixed}/5"

    # Final after v7.1 (all 18)
    all_vids = sorted(vid_to_row.keys())
    final_v71_scores = []
    final_v71_correct = 0
    for vid in all_vids:
        d = v71.get(vid)
        fp = first.get(vid, {})
        gt = fp.get("gt_verdict", "")
        v6v = fp.get("verdict", "")
        if d:
            rv = d.get("recovery_verdict") or ""
            final_v = rv or v6v
            sc, _ = V71_SCORES.get(vid, (None, ""))
        else:
            final_v = v6v
            sc = None  # no v7.1 score for uncontested clips
        if final_v == gt:
            final_v71_correct += 1
        if sc is not None:
            final_v71_scores.append(sc)

    def _mean(lst):
        return round(sum(lst) / len(lst), 2) if lst else 0

    def _median(lst):
        s = sorted(lst)
        return s[len(s) // 2] if s else 0

    new_rows = [
        ["v7.1 recovery (debated 5)",
         len(v71_scores_all),
         _mean(v71_scores_all),
         _median(v71_scores_all),
         sum(1 for s in v71_scores_all if s >= 8),
         sum(1 for s in v71_scores_all if s <= 2)],
        ["final after v7.1 debate (18)",
         18,
         "—",  # only have scores for debated clips
         "—",
         "—",
         "—"],
    ]
    for r_vals in new_rows:
        last_row += 1
        for col_idx, val in enumerate(r_vals, start=1):
            cell = ws_sum.cell(last_row, col_idx, value=val)
            cell.fill = BLUE

    wb.save(XLSX_IN)
    print(f"  Updated {XLSX_IN.name}: {len(NEW_COLS)} new columns in per_clip, "
          f"{len(new_rows)} new rows in summary")


# ---------------------------------------------------------------------------
# MD append
# ---------------------------------------------------------------------------

def _append_md(v71: dict, first: dict) -> None:
    existing = MD_FILE.read_text(encoding="utf-8") if MD_FILE.exists() else ""

    # Don't double-append
    if "## v7.1 Debate Results" in existing:
        print("  v7.1 section already in MD — skipping append")
        return

    v71_scores_list = [V71_SCORES[v][0] for v in V71_SCORES]
    v71_fixed = sum(1 for vid, d in v71.items()
                    if d.get("recovery_verdict") == first.get(vid, {}).get("gt_verdict"))

    # Final accuracy after v7.1
    all_vids = sorted(set(list(first.keys())))
    final_correct = 0
    for vid in all_vids:
        d = v71.get(vid)
        fp = first.get(vid, {})
        gt = fp.get("gt_verdict", "")
        v6v = fp.get("verdict", "")
        if d:
            rv = d.get("recovery_verdict") or ""
            final_v = rv or v6v
        else:
            final_v = v6v
        if final_v == gt:
            final_correct += 1

    total = len(all_vids)
    v6_correct = sum(1 for fp in first.values() if fp.get("verdict") == fp.get("gt_verdict"))

    section = []
    section.append("\n\n---\n")
    section.append("## v7.1 Debate Results\n")
    section.append("**Prompt fix:** v7.1 corrected the prompt inversion bug from v6 "
                   "(TP_RECOVERY → proactive specialist; TN_RECOVERY → conservative specialist), "
                   "adapted for 2-second stride-4 window.\n")
    section.append(f"**All 5 first-pass failures re-run** with v7.1 prompts on stride-4 hires frames.\n")
    section.append("")

    section.append("### Summary\n")
    section.append("| Stage | Verdict accuracy | Notes |")
    section.append("|-------|-----------------|-------|")
    section.append(f"| v6@hires first-pass | {v6_correct}/{total} = {v6_correct/total:.1%} | baseline |")
    section.append(f"| v7.1 recovery (5 debated) | {v71_fixed}/5 FIXED | 3/5 already done in v7.1 A/B test |")
    section.append(f"| Final after v7.1 debate | {final_correct}/{total} = {final_correct/total:.1%} | "
                   f"+{final_correct - v6_correct} vs v6@hires |")
    section.append("")

    section.append("### Per-clip v7.1 results\n")
    section.append("| Clip | GT | v6 verdict | v7.1 verdict | outcome | score | rationale |")
    section.append("|------|----|-----------|--------------|---------| ------|-----------|")
    for vid in sorted(V71_SCORES.keys()):
        d = v71.get(vid, {})
        fp = first.get(vid, {})
        gt = fp.get("gt_verdict", "?")
        v6v = fp.get("verdict", "?")
        rv = d.get("recovery_verdict") or "?"
        sc, rat = V71_SCORES.get(vid, ("?", ""))
        outcome = "FIXED" if rv == gt and v6v != gt else ("still-wrong" if rv != gt else "no-flip")
        section.append(f"| {vid} | {gt} | {v6v} | {rv} | {outcome} | {sc} | {rat} |")
    section.append("")

    section.append("### v7.1 Reasoning scores (0–10)\n")
    section.append("| Score | Meaning |")
    section.append("|-------|---------|")
    section.append("| 8–9 | Correct verdict + correct agent + correct mechanism |")
    section.append("| 5–7 | Correct verdict, partial agent/mechanism match |")
    section.append("| 2–3 | Wrong verdict, partial scene overlap |")
    section.append("| 0–1 | Wrong verdict + hallucinated scene |")
    section.append("")

    section.append(f"**v7.1 recovery reasoning scores** (5 clips): "
                   f"mean={round(sum(v71_scores_list)/len(v71_scores_list), 2)}, "
                   f"median={sorted(v71_scores_list)[len(v71_scores_list)//2]}, "
                   f"#≥8={sum(1 for s in v71_scores_list if s>=8)}, "
                   f"#≤2={sum(1 for s in v71_scores_list if s<=2)}\n")

    section.append("### Notable findings\n")
    section.append("- **00474 (FN→FIXED):** v7.1 TP_RECOVERY correctly detected closing gap "
                   "but identified *yellow taxi* rather than GT's white van. Verdict correct, "
                   "agent partially wrong (score 5).")
    section.append("- **02104, 02117 (FP→FIXED):** TN_RECOVERY correctly described stable "
                   "following and normal traffic flow (scores 8 each).")
    section.append("- **01153, 01504 (FP→still-wrong):** Both clips remain YES despite "
                   "TN_RECOVERY. 01153 repeats the same white-sedan-left-turn hallucination "
                   "(likely a perspective/labeling disagreement). 01504 still describes "
                   "high closing speed on a red SUV — model appears to see genuine visual "
                   "evidence of danger that GT classifies as safe.")
    section.append("")

    section.append("### v7.1 vs v6 debate comparison\n")
    section.append("| Clip | GT | v6 debate result | v7.1 debate result |")
    section.append("|------|----|-----------------|--------------------|")
    v6_debate = _load_jsonl(OUT / "v6_debate.jsonl")
    for vid in sorted(V71_SCORES.keys()):
        fp = first.get(vid, {})
        gt = fp.get("gt_verdict", "?")
        v6d = v6_debate.get(vid, {})
        v6r = v6d.get("recovery_verdict") or "N/A"
        v6_out = "FIXED" if v6r == gt else ("still-wrong" if v6r != "N/A" else "N/A")
        v71d = v71.get(vid, {})
        v71r = v71d.get("recovery_verdict") or "N/A"
        v71_out = "FIXED" if v71r == gt else ("still-wrong" if v71r != "N/A" else "N/A")
        section.append(f"| {vid} | {gt} | {v6r} ({v6_out}) | {v71r} ({v71_out}) |")
    section.append("")
    section.append(f"**v6 debate:** 2/5 FIXED  |  **v7.1 debate:** {v71_fixed}/5 FIXED  "
                   f"(+{v71_fixed - 2} improvement)\n")

    MD_FILE.write_text(existing + "\n".join(section), encoding="utf-8")
    print(f"  Appended v7.1 section to {MD_FILE.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading data...")
    v71   = _load_jsonl(V71_JSONL)
    first = _merge_first_pass()
    print(f"  v7.1 records: {len(v71)} clips")
    print(f"  First-pass records: {len(first)} clips")

    print("\nUpdating Excel...")
    _update_excel(v71, first)

    print("\nAppending to Markdown...")
    _append_md(v71, first)

    print("\nDone.")


if __name__ == "__main__":
    main()
