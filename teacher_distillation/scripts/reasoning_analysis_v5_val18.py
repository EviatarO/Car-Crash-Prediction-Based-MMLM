"""reasoning_analysis_v5_val18.py -- scores PROMPT_SEMSUP_V5_BALANCED captions
from qwen/qwen3-vl-235b-a22b-thinking against the 18-clip GT validation set,
formatted like the other semsup_val18 workbooks (whole-row colour fill,
wrapped text, sized rows). Adds risk_score and counter_evidence columns not
present in the V4 workbook, since V5's headline change is exactly those two
fields (continuous score instead of binary-only verdict, mandatory pre-mortem).

Reads:
- dataset/manifests/val_e3a.jsonl (video_id, gt_verdict, requested_time_to_event,
  gt_reasoning_en)
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds -
  NOT present in val_e3a.jsonl; these two files together cover exactly the 18
  val clips and are the runs that used these same _hires frames, so they are
  authoritative. Other files disagree on t_seconds for some clips because
  different experiment generations placed windows differently - do not use them)
- outputs/prompt_bakeoff/semsup_val18/raw_v5_qwen3vl235b.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v5_val18.xlsx

Colour rule (whole row, same convention as reasoning_analysis_qwen3vl_val18.py):
  GREEN  = verdict correct AND caption matches GT reasoning
  ORANGE = verdict correct but caption is middling/generic or contains a minor
           fabricated detail, OR verdict wrong but the caption still describes
           the scene well (caption is the training target - a good description
           with a miscalibrated verdict retains real value)
  RED    = verdict wrong AND caption wrong/hallucinated, OR verdict happens to
           be right but the caption names a different, non-existent mechanism
           (a hallucinated caption is not usable as a training target even if
           the mechanical threshold on risk_score landed on the right side)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
V5_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v5_qwen3vl235b.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v5_val18.xlsx"

# (score 0-10, rationale, colour) per clip, hand-scored against gt_reasoning_en.
SCORES = {
    "00077": (8, "'Black sedan ahead in ego lane braking with closing distance' captures GT's rear-end mechanism (sedan merges in, brakes, ego fails to react) closely; doesn't name the tow truck that triggered the braking. Correct verdict.", "green"),
    "00147": (1, "GT: ego itself deviates into the right-lane vehicle's path. Caption says the opposite - 'maintaining consistent following distance' - and misses the ego-deviation mechanism entirely. Wrong verdict.", "red"),
    "00283": (7, "'White SUV merging from left lane... while vehicle ahead braking' matches GT's cut-in-and-block mechanism directionally; object misidentified (SUV vs pickup+trailer) and merge direction differs from GT's right-to-left turn. Correct verdict.", "green"),
    "00319": (2, "Verdict correct only by coincidence: caption invents a stationary 'large truck ahead' and completely misses GT's actual mechanism (a car entering the intersection from the right without slowing). Not a usable caption despite the right-side threshold.", "orange"),
    "00372": (1, "Describes stable following distance; misses GT's stopping-sedan/crosswalk-pedestrian mechanism entirely. Wrong verdict.", "red"),
    "00474": (1, "Describes 'all vehicles maintaining position' - directly contradicts GT's van-cuts-into-ego-lane event, the same miss recorded for every model tested on this clip so far. Wrong verdict.", "red"),
    "00493": (2, "Correctly places the silver sedan ahead and pickup in the adjacent lane, but frames it as stable following rather than GT's ego-fails-to-react-to-braking-sedan mechanism. Wrong verdict.", "red"),
    "00529": (1, "States the silver SUV 'maintains parallel position' - the opposite of GT's claim that it drifts into the ego lane after an obstruction. Wrong verdict.", "red"),
    "00687": (2, "Regression vs the V4 run on this same clip: V4 correctly perceived the gray SUV as merging/drifting into ego's lane (just miscalibrated the risk); V5 describes it as 'parked on right side' - now a perception miss, not just calibration. Wrong verdict.", "red"),
    "01153": (1, "Hallucinates a 'white sedan turning left across ego path' - GT has no crossing conflict, ego performs its own smooth right turn with all other vehicles stationary/parallel. The one false positive this round.", "red"),
    "01281": (8, "Matches GT's controlled-following framing closely (jeep/SUV in right lane, pickup ahead at steady distance); does not fabricate a merge event. Correct verdict.", "green"),
    "01504": (6, "Matches GT's controlled-following framing and verdict; 'red SUV' vs GT's 'dark SUV' is the same colour-naming mismatch seen across every model tested on this clip.", "green"),
    "01550": (7, "Matches GT's controlled-closing, brake-lights-ahead description reasonably closely. Correct verdict.", "green"),
    "01552": (6, "Captures the gas-station scene (minivan exiting, truck moving away) but invents a 'school bus' detail not in GT - the same fabrication flagged on this clip in the V4 run. Correct verdict.", "orange"),
    "01643": (5, "Correct empty-road verdict, but introduces 'parked cars' not present in GT's 'no vehicles around it' - same fabrication flagged on this clip in the V4 run.", "orange"),
    "01737": (9, "Clean match to GT's empty, curving, night-time interchange-under-a-bridge description. Correct verdict.", "green"),
    "02104": (6, "Generic 'multiple vehicles maintaining constant distance' avoids fabricating a hazard but misses GT's specific merging-sedan detail - same vagueness flagged on this clip in the V4 run. Correct verdict.", "orange"),
    "02117": (8, "Matches GT's core mechanism (car ahead at constant distance, green light, no crossing pedestrians) without hallucinating the black-SUV-merge event that broke earlier Gemini runs on this clip. Correct verdict.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "v5_verdict", "risk_score", "v5_caption",
           "counter_evidence", "score", "score_explanation"]
WIDTHS = [11, 11, 11, 18, 55, 10, 10, 55, 45, 8, 55]
WRAP_COLS = {"gt_reasoning_en", "v5_caption", "counter_evidence", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    v5 = {r["video_id"]: r for r in _load_jsonl(V5_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t} - refusing to substitute a "
                            f"value from a disagreeing source (see module docstring).")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], v5[vid]
        verdict_str = "YES" if q["verdict"] == 1 else "NO"
        caption = f"{q['caption_neutral']}, {q['risk_clause']}"
        score, rationale, colour = SCORES.get(vid, (None, "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"], "gt_reasoning_en": v["gt_reasoning_en"],
            "v5_verdict": verdict_str, "risk_score": q["risk_score"], "v5_caption": caption,
            "counter_evidence": q.get("counter_evidence", ""),
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    # summary sheet
    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["v5_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["v5_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["v5_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["v5_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["v5_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    colour_counts = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}

    try:
        from sklearn.metrics import average_precision_score, roc_auc_score
        y_true = [1 if r["gt_verdict"] == "YES" else 0 for r in rows]
        y_score = [r["risk_score"] for r in rows]
        auc = round(roc_auc_score(y_true, y_score), 3)
        ap = round(average_precision_score(y_true, y_score), 3)
    except ImportError:
        auc = ap = "sklearn not available"

    metrics = [
        ("n", n), ("verdict_accuracy (score>=50 cut)", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP", tp), ("FP", fp), ("TN", tn), ("FN", fn),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("risk_score AUC", auc), ("risk_score AP", ap),
        ("mean_hand_score", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green", colour_counts["green"]), ("n_orange", colour_counts["orange"]),
        ("n_red", colour_counts["red"]),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
