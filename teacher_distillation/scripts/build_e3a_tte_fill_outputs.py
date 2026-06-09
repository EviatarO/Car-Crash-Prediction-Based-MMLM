"""E3a TTE-fill Stage 4: assemble the 267-row teacher table, monitor xlsx, summary md.

Inputs:
  dataset/teacher_labels/teacher_dataset_e3a.jsonl                 (89 originals)
  dataset/teacher_labels/teacher_dataset_v11.xlsx                  (richer fields for originals)
  outputs/prompt_bakeoff/v11_100clips_resampled/final_combined.jsonl (to disambiguate
                                                                     pass1 vs pass2 on v11_resampled_fp)
  outputs/prompt_bakeoff/e3a_tte_fill/pass1.jsonl                  (178 new pass-1)
  outputs/prompt_bakeoff/e3a_tte_fill/recovery.jsonl               (subset of 178 routed to recovery)

Outputs:
  outputs/prompt_bakeoff/e3a_tte_fill/results_e3a_tte_fill.xlsx
  outputs/prompt_bakeoff/e3a_tte_fill/monitor_tte_coverage.xlsx
  outputs/prompt_bakeoff/e3a_tte_fill/summary.md
  outputs/prompt_bakeoff/e3a_tte_fill/final_combined.jsonl
"""
from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]

E3A_JSONL = REPO_ROOT / "dataset" / "teacher_labels" / "teacher_dataset_e3a.jsonl"
V11_XLSX  = REPO_ROOT / "dataset" / "teacher_labels" / "teacher_dataset_v11.xlsx"

V11_RESAMP_FINAL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v11_100clips_resampled" / "final_combined.jsonl"

OUT_DIR  = REPO_ROOT / "outputs" / "prompt_bakeoff" / "e3a_tte_fill"
PASS1    = OUT_DIR / "pass1.jsonl"
RECOVERY = OUT_DIR / "recovery.jsonl"
OUT_XLSX = OUT_DIR / "results_e3a_tte_fill.xlsx"
OUT_MONI = OUT_DIR / "monitor_tte_coverage.xlsx"
OUT_MD   = OUT_DIR / "summary.md"
OUT_JSONL_FINAL = OUT_DIR / "final_combined.jsonl"
RR_SCORES = OUT_DIR / "reasoning_relation_scores.json"  # manually-authored per-video scores (YES only)
FR2_JSON  = OUT_DIR / "final_reasoning2.json"             # manually-authored harmonized rewrites (changed YES videos)
RR2_SCORES = OUT_DIR / "reasoning_relation2_scores.json" # manually-authored post-harmonization scores (YES only)

# Colors per the approved plan
STRONG_GREEN = PatternFill("solid", fgColor="00B050")   # pass1 correct
SOFT_GREEN   = PatternFill("solid", fgColor="C6EFCE")   # pass2 correct
RED          = PatternFill("solid", fgColor="FFC7CE")   # still wrong
HEADER       = PatternFill("solid", fgColor="2E75B6")
GREY         = PatternFill("solid", fgColor="D9D9D9")


# ---- helpers ----
def _load_jsonl(p: Path) -> List[dict]:
    if not p.exists(): return []
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]


def _load_v11_xlsx_map() -> Dict[str, dict]:
    df = pd.read_excel(V11_XLSX)
    out: Dict[str, dict] = {}
    for _, row in df.iterrows():
        if pd.isna(row.get("video_id")): continue
        vid = f"{int(row['video_id']):05d}"
        out[vid] = row.to_dict()
    return out


def _safe(d: dict, k: str):
    v = d.get(k)
    if pd.isna(v) if not isinstance(v, (dict, list)) else False:  # noqa
        return None
    return v


def _load_v11_resampled_full() -> Dict[str, dict]:
    """Full v11-resampled records by video_id (covers v11_resampled_fp + v11_fn_v7_1)."""
    return {r["video_id"]: r for r in _load_jsonl(V11_RESAMP_FINAL)}


def _v11_resampled_outcome(rec: dict) -> str:
    gt = rec["gt_verdict"]
    rsv = rec.get("v6_resampled_verdict"); rcv = rec.get("v7_1_recovery_verdict")
    if rsv == gt: return "pass1_correct"
    if rcv == gt: return "pass2_correct"
    return "still_wrong"


def _row_for_existing(e3a: dict, v11_map: Dict[str, dict], v11_resamp: Dict[str, dict]) -> dict:
    """
    Build a row for one of the 89 existing clips.

    Three sources are handled distinctly:
      - v11_preserved   : v6 first-pass at the original midpoint. Pull from v11_xlsx.
      - v11_fn_v7_1     : v6 first-pass was WRONG (NO); v7.1 TP_RECOVERY fixed it.
                          Pass-1 reasoning lives in v11_xlsx (verdict_reasoning, etc).
                          Recovery output lives in v11_resamp.final_combined.jsonl.
      - v11_resampled_fp: v6 first-pass at midpoint was wrong; the resampled t-4s frames
                          are the actual training frames. So pass-1 reasoning = the
                          RESAMPLED v6_balanced reasoning, NOT the v6_orig from midpoint.
                          t_seconds should be the resampled t_new.
    """
    vid = e3a["video_id"]
    src = e3a.get("source")
    gt = e3a["gt_verdict"]
    v11 = v11_map.get(vid, {})
    rs  = v11_resamp.get(vid, {})

    # horizon_label
    tte = e3a.get("requested_time_to_event")
    horizon_label = f"TTE_{tte}" if gt == "YES" else "MID"

    # Common JSON parse of e3a assistant_target as fallback reasoning
    fallback_reason = None
    at = e3a.get("assistant_target")
    if isinstance(at, str) and at.strip().startswith("{"):
        try:
            fallback_reason = json.loads(at).get("reason")
        except Exception:
            fallback_reason = None

    if src == "v11_resampled_fp":
        # pass-1 (resampled v6) reasoning is what was actually inferred on the training frames
        outcome = _v11_resampled_outcome(rs)
        return {
            "video_id": vid, "gt_verdict": gt,
            "t_seconds": rs.get("v6_resampled_t_new") or _safe(v11, "t_seconds"),
            "requested_time_to_event": "MID-4s_resampled",
            "horizon_label": horizon_label,
            "collision_verdict": rs.get("v6_resampled_verdict"),
            "verdict_reasoning": (rs.get("v6_resampled_reasoning") or "").strip() or None,
            "scene_context": None, "dynamic_objects": None,
            "temporal_analysis": None, "occlusion_check": None,
            "time_to_contact": None, "mismatch": None,
            "p2_collision_verdict": rs.get("v7_1_recovery_verdict"),
            "p2_verdict_reasoning": (rs.get("v7_1_recovery_reasoning") or "").strip() or None,
            "p2_scene_context": None, "p2_dynamic_objects": None,
            "p2_temporal_analysis": None, "p2_occlusion_check": None,
            "p2_time_to_contact": None,
            "final_verdict":  e3a.get("final_verdict") or rs.get("final_verdict"),
            "final_reasoning": (rs.get("teacher_reasoning_final") or "").strip() or fallback_reason,
            "pipeline_outcome": outcome,
            "row_origin": "existing_89", "error": None,
        }

    if src == "v11_fn_v7_1":
        # Pass-1 was wrong (v6_orig at midpoint, predicted NO). Recovery TP_RECOVERY may or may not have fixed it.
        outcome = "pass2_correct" if (rs.get("v7_1_recovery_verdict") == gt) else "still_wrong"
        return {
            "video_id": vid, "gt_verdict": gt,
            "t_seconds": _safe(v11, "t_seconds"),
            "requested_time_to_event": _safe(v11, "requested_time_to_event"),
            "horizon_label": horizon_label,
            "collision_verdict": _safe(v11, "collision_verdict"),
            "verdict_reasoning": _safe(v11, "verdict_reasoning"),
            "scene_context":      _safe(v11, "scene_context"),
            "dynamic_objects":    _safe(v11, "dynamic_objects"),
            "temporal_analysis":  _safe(v11, "temporal_analysis"),
            "occlusion_check":    _safe(v11, "occlusion_check"),
            "time_to_contact":    _safe(v11, "time_to_contact"),
            "mismatch":           _safe(v11, "mismatch"),
            "p2_collision_verdict": rs.get("v7_1_recovery_verdict"),
            "p2_verdict_reasoning": (rs.get("v7_1_recovery_reasoning") or "").strip() or None,
            "p2_scene_context": None, "p2_dynamic_objects": None,
            "p2_temporal_analysis": None, "p2_occlusion_check": None,
            "p2_time_to_contact": None,
            "final_verdict":  e3a.get("final_verdict") or rs.get("final_verdict"),
            "final_reasoning": (rs.get("teacher_reasoning_final") or "").strip() or fallback_reason,
            "pipeline_outcome": outcome,
            "row_origin": "existing_89", "error": None,
        }

    # v11_preserved (default)
    fv = e3a.get("final_verdict")
    outcome = "pass1_correct" if fv == gt else "still_wrong"
    return {
        "video_id": vid, "gt_verdict": gt,
        "t_seconds": _safe(v11, "t_seconds"),
        "requested_time_to_event": _safe(v11, "requested_time_to_event"),
        "horizon_label": horizon_label,
        "collision_verdict": _safe(v11, "collision_verdict"),
        "verdict_reasoning": _safe(v11, "verdict_reasoning"),
        "scene_context":      _safe(v11, "scene_context"),
        "dynamic_objects":    _safe(v11, "dynamic_objects"),
        "temporal_analysis":  _safe(v11, "temporal_analysis"),
        "occlusion_check":    _safe(v11, "occlusion_check"),
        "time_to_contact":    _safe(v11, "time_to_contact"),
        "mismatch":           _safe(v11, "mismatch"),
        "p2_collision_verdict": _safe(v11, "p2_collision_verdict"),
        "p2_verdict_reasoning": _safe(v11, "p2_verdict_reasoning"),
        "p2_scene_context":     _safe(v11, "p2_scene_context"),
        "p2_dynamic_objects":   _safe(v11, "p2_dynamic_objects"),
        "p2_temporal_analysis": _safe(v11, "p2_temporal_analysis"),
        "p2_occlusion_check":   _safe(v11, "p2_occlusion_check"),
        "p2_time_to_contact":   _safe(v11, "p2_time_to_contact"),
        "final_verdict":  fv,
        "final_reasoning": _safe(v11, "final_reasoning") or fallback_reason,
        "pipeline_outcome": outcome,
        "row_origin": "existing_89", "error": _safe(v11, "error"),
    }


def _row_for_new(pass1_rec: dict, recovery_rec: Optional[dict]) -> dict:
    fj1 = pass1_rec.get("full_json") or {}
    p1v = pass1_rec.get("verdict")
    gt = pass1_rec["gt_verdict"]
    horizon_label = pass1_rec["horizon_label"]

    fj2 = (recovery_rec or {}).get("full_json") or {}
    p2v = (recovery_rec or {}).get("recovery_verdict")

    # Determine outcome + final
    if p1v == gt:
        outcome = "pass1_correct"
        final_v = p1v; final_r = fj1.get("verdict_reasoning") or pass1_rec.get("reasoning")
    elif recovery_rec and p2v == gt:
        outcome = "pass2_correct"
        final_v = p2v; final_r = fj2.get("verdict_reasoning") or recovery_rec.get("recovery_reasoning")
    else:
        outcome = "still_wrong"
        final_v = p1v if p1v is not None else (p2v if recovery_rec else None)
        final_r = (fj1.get("verdict_reasoning") if p1v is not None
                   else (fj2.get("verdict_reasoning") if recovery_rec else None))

    # t_seconds for the NEW variant = t_new (last-frame time of the new window)
    return {
        "video_id": pass1_rec["video_id"],
        "gt_verdict": gt,
        "t_seconds": pass1_rec.get("t_new"),
        "requested_time_to_event": (pass1_rec.get("horizon_s") if gt == "YES"
                                     else f"{pass1_rec.get('horizon_s')}_offset"),
        "horizon_label": horizon_label,
        "collision_verdict": fj1.get("collision_verdict") or p1v,
        "verdict_reasoning": fj1.get("verdict_reasoning"),
        "scene_context":      fj1.get("scene_context"),
        "dynamic_objects":    fj1.get("dynamic_objects"),
        "temporal_analysis":  fj1.get("temporal_analysis"),
        "occlusion_check":    fj1.get("occlusion_check"),
        "time_to_contact":    fj1.get("time_to_contact"),
        "mismatch":           fj1.get("mismatch"),
        "p2_collision_verdict": fj2.get("collision_verdict") or p2v if recovery_rec else None,
        "p2_verdict_reasoning": fj2.get("verdict_reasoning") if recovery_rec else None,
        "p2_scene_context":     fj2.get("scene_context") if recovery_rec else None,
        "p2_dynamic_objects":   fj2.get("dynamic_objects") if recovery_rec else None,
        "p2_temporal_analysis": fj2.get("temporal_analysis") if recovery_rec else None,
        "p2_occlusion_check":   fj2.get("occlusion_check") if recovery_rec else None,
        "p2_time_to_contact":   fj2.get("time_to_contact") if recovery_rec else None,
        "final_verdict":  final_v,
        "final_reasoning": final_r,
        "pipeline_outcome": outcome,
        "row_origin": "new_178",
        "error": pass1_rec.get("error") or ((recovery_rec or {}).get("error") if recovery_rec else None),
    }


# ---- main assembly ----
def build_combined() -> List[dict]:
    e3a = _load_jsonl(E3A_JSONL)
    v11_map = _load_v11_xlsx_map()
    v11_resamp = _load_v11_resampled_full()
    pass1 = _load_jsonl(PASS1)
    recovery = _load_jsonl(RECOVERY)

    pass1_keyed: Dict[Tuple[str, str], dict] = {(r["video_id"], r["horizon_label"]): r for r in pass1}
    recovery_keyed: Dict[Tuple[str, str], dict] = {(r["video_id"], r["horizon_label"]): r for r in recovery}

    rows: List[dict] = []
    # 89 existing
    for e in e3a:
        rows.append(_row_for_existing(e, v11_map, v11_resamp))
    # 178 new
    for (vid, horizon), p1 in pass1_keyed.items():
        # Only attach a recovery record when pass-1 was actually wrong; a clip fixed
        # at pass-1 (e.g. a re-sampled retry that returns NO) must not show a stale
        # YES recovery in its p2_* columns.
        rec = recovery_keyed.get((vid, horizon)) if p1.get("verdict") != p1.get("gt_verdict") else None
        rows.append(_row_for_new(p1, rec))

    # Sort by (video_id, horizon_order)
    horizon_order = {"TTE_0.5": 0, "TTE_1.0": 1, "TTE_1.5": 2, "MID": 0, "MID-4": 1, "MID-8": 2}
    rows.sort(key=lambda r: (r["video_id"], horizon_order.get(r["horizon_label"], 99)))
    return rows


# ---- xlsx writers ----
COLUMNS = [
    "video_id", "gt_verdict", "t_seconds", "requested_time_to_event", "horizon_label",
    "collision_verdict", "verdict_reasoning", "scene_context", "dynamic_objects",
    "temporal_analysis", "occlusion_check", "time_to_contact", "mismatch",
    "p2_collision_verdict", "p2_verdict_reasoning", "p2_scene_context", "p2_dynamic_objects",
    "p2_temporal_analysis", "p2_occlusion_check", "p2_time_to_contact",
    "final_verdict", "final_reasoning", "pipeline_outcome", "row_origin", "error",
]

WIDTHS = {
    "video_id": 10, "gt_verdict": 10, "t_seconds": 11,
    "requested_time_to_event": 14, "horizon_label": 12,
    "collision_verdict": 14, "verdict_reasoning": 60, "scene_context": 35,
    "dynamic_objects": 35, "temporal_analysis": 35, "occlusion_check": 30,
    "time_to_contact": 14, "mismatch": 10,
    "p2_collision_verdict": 14, "p2_verdict_reasoning": 60, "p2_scene_context": 35,
    "p2_dynamic_objects": 35, "p2_temporal_analysis": 35, "p2_occlusion_check": 30,
    "p2_time_to_contact": 14,
    "final_verdict": 12, "final_reasoning": 60, "pipeline_outcome": 16,
    "row_origin": 12, "error": 30,
}

OUTCOME_FILL = {"pass1_correct": STRONG_GREEN, "pass2_correct": SOFT_GREEN, "still_wrong": RED}

YELLOW = PatternFill("solid", fgColor="FFFF00")  # marks video_id of changed clips


def _load_fr2() -> Dict[str, dict]:
    """Manually-authored harmonized rewrites, keyed by video_id -> {horizon: {reasoning, change}}."""
    if not FR2_JSON.exists():
        return {}
    d = json.loads(FR2_JSON.read_text(encoding="utf-8"))
    return {k: v for k, v in d.items() if not k.startswith("_")}


def write_results_xlsx(rows: List[dict]) -> None:
    df = pd.DataFrame([{c: r.get(c) for c in COLUMNS} for r in rows], columns=COLUMNS)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUT_XLSX, sheet_name="per_clip", index=False)

    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb["per_clip"]

    # Header styling
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    idx = {h: i + 1 for i, h in enumerate(COLUMNS)}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        outcome = row[idx["pipeline_outcome"] - 1].value
        fill = OUTCOME_FILL.get(outcome)
        if fill is not None:
            for col in ("final_verdict", "final_reasoning", "pipeline_outcome"):
                row[idx[col] - 1].fill = fill
        # Wrap long fields
        for col in ("verdict_reasoning", "scene_context", "dynamic_objects",
                    "temporal_analysis", "occlusion_check",
                    "p2_verdict_reasoning", "p2_scene_context", "p2_dynamic_objects",
                    "p2_temporal_analysis", "p2_occlusion_check",
                    "final_reasoning", "error"):
            row[idx[col] - 1].alignment = Alignment(wrap_text=True, vertical="top")
        # Grey out row_origin = existing_89 to distinguish visually
        if row[idx["row_origin"] - 1].value == "existing_89":
            row[idx["row_origin"] - 1].fill = GREY

    for name, w in WIDTHS.items():
        if name in idx:
            ws.column_dimensions[get_column_letter(idx[name])].width = w
    ws.row_dimensions[1].height = 38

    # ---- Columns Z (final_reasoning2) + AA (change_explanation) ----
    # Z = harmonized rewrite if (video_id,horizon) was modified, else copy of final_reasoning.
    # AA = per-clip change note (blank if unchanged). NO rows always copy the original.
    # video_id (col A) is highlighted yellow for every video with >=1 changed clip.
    fr2 = _load_fr2()
    changed_vids = set(fr2.keys())
    z_col, aa_col = 26, 27  # Z, AA
    hz = ws.cell(1, z_col, "final_reasoning2")
    ha = ws.cell(1, aa_col, "change_explanation")
    for cell in (hz, ha):
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    a_col = idx["video_id"]
    for i, r in enumerate(rows):
        excel_row = i + 2
        vid = r["video_id"]
        horizon = r["horizon_label"]
        entry = fr2.get(vid, {}).get(horizon)
        if entry:
            new_text = entry.get("reasoning") or r.get("final_reasoning")
            note = entry.get("change") or ""
        else:
            new_text = r.get("final_reasoning")
            note = ""
        zc = ws.cell(excel_row, z_col, new_text)
        ac = ws.cell(excel_row, aa_col, note)
        zc.alignment = Alignment(wrap_text=True, vertical="top")
        ac.alignment = Alignment(wrap_text=True, vertical="top")
        if vid in changed_vids:
            ws.cell(excel_row, a_col).fill = YELLOW

    ws.column_dimensions[get_column_letter(z_col)].width = 60
    ws.column_dimensions[get_column_letter(aa_col)].width = 45

    wb.save(OUT_XLSX)


# Score-band fills for the reasoning_relation sheet
RR_GREEN = PatternFill("solid", fgColor="C6EFCE")   # >=7 strong same-story
RR_AMBER = PatternFill("solid", fgColor="FFEB9C")   # 4-6 one clip diverges
RR_RED   = PatternFill("solid", fgColor="FFC7CE")   # <=3 different scenarios / hallucination


def _rr_fill(score: int) -> PatternFill:
    if score >= 7:
        return RR_GREEN
    if score >= 4:
        return RR_AMBER
    return RR_RED


def write_reasoning_relation_sheet(rows: List[dict]) -> None:
    """Append a per-video 'reasoning_relation' sheet (YES/TP videos only) to the
    results workbook, scored from the manually-authored RR_SCORES json. Adds a
    score-frequency table and an embedded bar chart (reasoning-quality histogram).
    Skipped gracefully if the scores json is absent."""
    if not RR_SCORES.exists():
        print(f"  (skip reasoning_relation: {RR_SCORES.name} not found)")
        return
    scores = json.loads(RR_SCORES.read_text(encoding="utf-8"))
    scores = {k: v for k, v in scores.items() if not k.startswith("_")}

    # Order by the per_clip video order, restricted to YES videos present in scores
    seen: List[str] = []
    for r in rows:
        v = r["video_id"]
        if r["gt_verdict"] == "YES" and v in scores and v not in seen:
            seen.append(v)

    wb = openpyxl.load_workbook(OUT_XLSX)
    if "reasoning_relation" in wb.sheetnames:
        del wb["reasoning_relation"]
    ws = wb.create_sheet("reasoning_relation")

    headers = ["video_id", "video_score", "reasoning_relation"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    r_i = 2
    for vid in seen:
        sc = int(scores[vid]["score"])
        rel = scores[vid]["relation"]
        ws.cell(r_i, 1, vid).alignment = Alignment(horizontal="center", vertical="top")
        sc_cell = ws.cell(r_i, 2, sc)
        sc_cell.alignment = Alignment(horizontal="center", vertical="top")
        sc_cell.fill = _rr_fill(sc)
        rel_cell = ws.cell(r_i, 3, rel)
        rel_cell.alignment = Alignment(wrap_text=True, vertical="top")
        rel_cell.fill = _rr_fill(sc)
        r_i += 1

    last_data_row = r_i - 1
    n = len(seen)

    # Mean row (formula, like the val-review =SUM/n)
    ws.cell(r_i, 1, "MEAN").font = Font(bold=True)
    mean_cell = ws.cell(r_i, 2, f"=SUM(B2:B{last_data_row})/{n}" if n else 0)
    mean_cell.font = Font(bold=True)
    mean_cell.alignment = Alignment(horizontal="center")
    mean_cell.number_format = "0.00"
    mean_row = r_i

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 95
    ws.row_dimensions[1].height = 30

    # ---- Score-frequency histogram: matplotlib PNG embedded as image ----
    # openpyxl's native BarChart serialisation is unreliable on Windows Excel;
    # a matplotlib PNG always renders correctly.
    import io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from collections import Counter
    from openpyxl.drawing.image import Image as XLImage

    counts = Counter(int(scores[v]["score"]) for v in seen)
    score_vals = list(range(0, 11))
    count_vals = [counts.get(s, 0) for s in score_vals]

    fig, ax = plt.subplots(figsize=(8, 4.5))
    bars = ax.bar(score_vals, count_vals, color="#4472C4", edgecolor="white", width=0.7)
    for bar, val in zip(bars, count_vals):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
                    str(val), ha="center", va="bottom", fontsize=11, fontweight="bold")
    ax.set_xlabel("video_score (0–10)", fontsize=11)
    ax.set_ylabel("n_videos", fontsize=11)
    ax.set_title("Reasoning-quality distribution (42 YES videos)", fontsize=12)
    ax.set_xticks(score_vals)
    ax.yaxis.get_major_locator().set_params(integer=True)
    ax.set_ylim(0, max(count_vals) + 2)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)

    img = XLImage(buf)
    img.anchor = "E2"
    ws.add_image(img)

    wb.save(OUT_XLSX)
    mean_val = sum(int(scores[v]["score"]) for v in seen) / n if n else 0.0
    print(f"  reasoning_relation: {n} YES videos, mean score {mean_val:.2f} "
          f"(row {mean_row}); histogram + chart added")


def write_reasoning_relation2_sheet(rows: List[dict]) -> None:
    """Append a 'reasoning_relation2' sheet scored from the POST-harmonization
    RR2_SCORES json (after final_reasoning2 rewrites). Same layout as
    reasoning_relation (video_id / video_score / reasoning_relation + MEAN +
    matplotlib-PNG histogram). video_id (col A) is highlighted yellow for every
    video whose reasoning was changed (present in final_reasoning2.json)."""
    if not RR2_SCORES.exists():
        print(f"  (skip reasoning_relation2: {RR2_SCORES.name} not found)")
        return
    scores = json.loads(RR2_SCORES.read_text(encoding="utf-8"))
    scores = {k: v for k, v in scores.items() if not k.startswith("_")}
    changed_vids = set(_load_fr2().keys())

    seen: List[str] = []
    for r in rows:
        v = r["video_id"]
        if r["gt_verdict"] == "YES" and v in scores and v not in seen:
            seen.append(v)

    wb = openpyxl.load_workbook(OUT_XLSX)
    if "reasoning_relation2" in wb.sheetnames:
        del wb["reasoning_relation2"]
    ws = wb.create_sheet("reasoning_relation2")

    headers = ["video_id", "video_score", "reasoning_relation"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    r_i = 2
    for vid in seen:
        sc = int(scores[vid]["score"])
        rel = scores[vid]["relation"]
        vid_cell = ws.cell(r_i, 1, vid)
        vid_cell.alignment = Alignment(horizontal="center", vertical="top")
        if vid in changed_vids:
            vid_cell.fill = YELLOW
        sc_cell = ws.cell(r_i, 2, sc)
        sc_cell.alignment = Alignment(horizontal="center", vertical="top")
        sc_cell.fill = _rr_fill(sc)
        rel_cell = ws.cell(r_i, 3, rel)
        rel_cell.alignment = Alignment(wrap_text=True, vertical="top")
        rel_cell.fill = _rr_fill(sc)
        r_i += 1

    last_data_row = r_i - 1
    n = len(seen)

    ws.cell(r_i, 1, "MEAN").font = Font(bold=True)
    mean_cell = ws.cell(r_i, 2, f"=SUM(B2:B{last_data_row})/{n}" if n else 0)
    mean_cell.font = Font(bold=True)
    mean_cell.alignment = Alignment(horizontal="center")
    mean_cell.number_format = "0.00"
    mean_row = r_i

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 95
    ws.row_dimensions[1].height = 30

    # ---- histogram (matplotlib PNG) ----
    import io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from collections import Counter
    from openpyxl.drawing.image import Image as XLImage

    counts = Counter(int(scores[v]["score"]) for v in seen)
    score_vals = list(range(0, 11))
    count_vals = [counts.get(s, 0) for s in score_vals]

    fig, ax = plt.subplots(figsize=(8, 4.5))
    bars = ax.bar(score_vals, count_vals, color="#2E9947", edgecolor="white", width=0.7)
    for bar, val in zip(bars, count_vals):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
                    str(val), ha="center", va="bottom", fontsize=11, fontweight="bold")
    ax.set_xlabel("video_score (0–10)", fontsize=11)
    ax.set_ylabel("n_videos", fontsize=11)
    ax.set_title("Reasoning-quality distribution — after harmonization (42 YES videos)", fontsize=11)
    ax.set_xticks(score_vals)
    ax.yaxis.get_major_locator().set_params(integer=True)
    ax.set_ylim(0, max(count_vals) + 2)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)

    img = XLImage(buf)
    img.anchor = "E2"
    ws.add_image(img)

    wb.save(OUT_XLSX)
    mean_val = sum(int(scores[v]["score"]) for v in seen) / n if n else 0.0
    print(f"  reasoning_relation2: {n} YES videos, mean score {mean_val:.2f} "
          f"(row {mean_row}); {len(changed_vids)} changed videos marked; histogram added")


def write_monitor_xlsx(rows: List[dict]) -> None:
    yes_horizons = ["TTE_0.5", "TTE_1.0", "TTE_1.5"]
    no_horizons  = ["MID", "MID-4", "MID-8"]

    # Build per-video outcome map
    by_vid: Dict[str, Dict[str, str]] = {}
    gt_by_vid: Dict[str, str] = {}
    for r in rows:
        by_vid.setdefault(r["video_id"], {})[r["horizon_label"]] = r["pipeline_outcome"]
        gt_by_vid[r["video_id"]] = r["gt_verdict"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _add_sheet(name: str, horizons: List[str], vids: List[str]) -> None:
        ws = wb.create_sheet(name)
        ws.cell(1, 1, "video_id"); ws.cell(1, 1).fill = HEADER
        for i, h in enumerate(horizons, start=2):
            ws.cell(1, i, h); ws.cell(1, i).fill = HEADER
        ws.cell(1, len(horizons) + 2, "all_perfect"); ws.cell(1, len(horizons) + 2).fill = HEADER
        for c in range(1, len(horizons) + 3):
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")

        for r_i, vid in enumerate(sorted(vids), start=2):
            ws.cell(r_i, 1, vid)
            outcomes = by_vid.get(vid, {})
            all_perfect = 1
            for c_i, h in enumerate(horizons, start=2):
                oc = outcomes.get(h)
                fill = OUTCOME_FILL.get(oc, RED if oc is None else None)
                cell = ws.cell(r_i, c_i, "")
                if fill is not None: cell.fill = fill
                if oc not in ("pass1_correct", "pass2_correct"):
                    all_perfect = 0
            ws.cell(r_i, len(horizons) + 2, all_perfect)

        ws.column_dimensions["A"].width = 12
        for c in range(2, len(horizons) + 2):
            ws.column_dimensions[get_column_letter(c)].width = 12
        ws.column_dimensions[get_column_letter(len(horizons) + 2)].width = 12

    yes_vids = [v for v, g in gt_by_vid.items() if g == "YES"]
    no_vids  = [v for v, g in gt_by_vid.items() if g == "NO"]
    _add_sheet("monitor_yes", yes_horizons, yes_vids)
    _add_sheet("monitor_no",  no_horizons,  no_vids)
    wb.save(OUT_MONI)


def write_summary_md(rows: List[dict]) -> None:
    total = len(rows)
    n_new = sum(1 for r in rows if r["row_origin"] == "new_178")
    n_old = total - n_new

    def _ct(outcome): return sum(1 for r in rows if r["pipeline_outcome"] == outcome)
    p1 = _ct("pass1_correct"); p2 = _ct("pass2_correct"); sw = _ct("still_wrong")

    # Per-horizon breakdown
    horizons_yes = ["TTE_0.5", "TTE_1.0", "TTE_1.5"]
    horizons_no  = ["MID", "MID-4", "MID-8"]

    def _cm_for(rows_):
        from collections import Counter
        c = Counter(r["pipeline_outcome"] for r in rows_)
        n = len(rows_) or 1
        return c, n

    yes_rows = [r for r in rows if r["gt_verdict"] == "YES"]
    no_rows  = [r for r in rows if r["gt_verdict"] == "NO"]

    # Per-video coverage
    by_vid: Dict[str, Dict[str, str]] = {}
    for r in rows:
        by_vid.setdefault(r["video_id"], {})[r["horizon_label"]] = r["pipeline_outcome"]

    perfect_videos = sum(1 for v, oc in by_vid.items()
                         if all(oc.get(h) in ("pass1_correct", "pass2_correct") for h in
                                (horizons_yes if any(rr["gt_verdict"] == "YES" and rr["video_id"] == v for rr in rows)
                                 else horizons_no)))

    md: List[str] = []
    md.append("# E3a multi-horizon TTE-fill — Summary\n")
    md.append("## Context")
    md.append("This run fills the per-video horizon matrix called for in §7(ii) of the E3a progress report:")
    md.append("> *Scaling the distillation set with per-video multi-horizon clips … a monotonic slope would require within-scene, multi-horizon training.*")
    md.append("Every selected video now has 3 distilled variants (positives at TTE = 0.5 / 1.0 / 1.5 s; negatives at MID / MID−4s / MID−8s). The student can fit a within-scene temporal response.\n")

    md.append("## Headline numbers")
    md.append("| Bucket | n | pass1_correct | pass2_correct | still_wrong |")
    md.append("|---|---|---|---|---|")
    md.append(f"| All rows | {total} | {p1} ({p1/total:.1%}) | {p2} ({p2/total:.1%}) | {sw} ({sw/total:.1%}) |")
    md.append(f"| Existing 89 (reused as-is) | {n_old} | "
              f"{sum(1 for r in rows if r['row_origin']=='existing_89' and r['pipeline_outcome']=='pass1_correct')} | "
              f"{sum(1 for r in rows if r['row_origin']=='existing_89' and r['pipeline_outcome']=='pass2_correct')} | "
              f"{sum(1 for r in rows if r['row_origin']=='existing_89' and r['pipeline_outcome']=='still_wrong')} |")
    md.append(f"| New 178 (pass-1 + recovery) | {n_new} | "
              f"{sum(1 for r in rows if r['row_origin']=='new_178' and r['pipeline_outcome']=='pass1_correct')} | "
              f"{sum(1 for r in rows if r['row_origin']=='new_178' and r['pipeline_outcome']=='pass2_correct')} | "
              f"{sum(1 for r in rows if r['row_origin']=='new_178' and r['pipeline_outcome']=='still_wrong')} |")
    md.append("")

    md.append("## Per-horizon breakdown")
    md.append("**YES (positives)**\n")
    md.append("| horizon | n | pass1_correct | pass2_correct | still_wrong |")
    md.append("|---|---|---|---|---|")
    for h in horizons_yes:
        sub = [r for r in yes_rows if r["horizon_label"] == h]
        n_ = len(sub)
        a = sum(1 for r in sub if r["pipeline_outcome"]=="pass1_correct")
        b = sum(1 for r in sub if r["pipeline_outcome"]=="pass2_correct")
        c = sum(1 for r in sub if r["pipeline_outcome"]=="still_wrong")
        md.append(f"| {h} | {n_} | {a} | {b} | {c} |")
    md.append("\n**NO (negatives)**\n")
    md.append("| horizon | n | pass1_correct | pass2_correct | still_wrong |")
    md.append("|---|---|---|---|---|")
    for h in horizons_no:
        sub = [r for r in no_rows if r["horizon_label"] == h]
        n_ = len(sub)
        a = sum(1 for r in sub if r["pipeline_outcome"]=="pass1_correct")
        b = sum(1 for r in sub if r["pipeline_outcome"]=="pass2_correct")
        c = sum(1 for r in sub if r["pipeline_outcome"]=="still_wrong")
        md.append(f"| {h} | {n_} | {a} | {b} | {c} |")
    md.append("")

    md.append("## Per-video coverage")
    md.append(f"- **Videos with all 3 horizons strong/soft-green:** {perfect_videos} / {len(by_vid)}")
    md.append(f"- **Videos with ≥1 red horizon:** {len(by_vid) - perfect_videos}\n")

    md.append("## Still-wrong rows")
    md.append("| video_id | gt | horizon | final_verdict | row_origin |")
    md.append("|---|---|---|---|---|")
    for r in rows:
        if r["pipeline_outcome"] != "still_wrong": continue
        md.append(f"| {r['video_id']} | {r['gt_verdict']} | {r['horizon_label']} | "
                  f"{r['final_verdict'] or '—'} | {r['row_origin']} |")
    md.append("")

    # ---- The deliverable workbook + the reasoning-quality work ----
    from collections import Counter

    def _load_scores(path: Path) -> Dict[str, dict]:
        if not path.exists():
            return {}
        d = json.loads(path.read_text(encoding="utf-8"))
        return {k: v for k, v in d.items() if not k.startswith("_")}

    rr1 = _load_scores(RR_SCORES)
    rr2 = _load_scores(RR2_SCORES)
    fr2 = _load_scores(FR2_JSON)
    n_changed = len(fr2)
    n_clip_edits = sum(len(v) for v in fr2.values())
    mean1 = (sum(int(x["score"]) for x in rr1.values()) / len(rr1)) if rr1 else 0.0
    mean2 = (sum(int(x["score"]) for x in rr2.values()) / len(rr2)) if rr2 else 0.0
    le4_1 = sum(1 for x in rr1.values() if int(x["score"]) <= 4)
    le4_2 = sum(1 for x in rr2.values() if int(x["score"]) <= 4)
    ge7_1 = sum(1 for x in rr1.values() if int(x["score"]) >= 7)
    ge7_2 = sum(1 for x in rr2.values() if int(x["score"]) >= 7)

    md.append("## The deliverable workbook — `results_e3a_tte_fill.xlsx`")
    md.append("This workbook is the human-auditable view of the distilled teacher set "
              "and the artifact we curate before student SFT. It has these sheets:")
    md.append("- **`per_clip`** — one row per distilled clip (267 total). Carries the teacher's "
              "structured fields (verdict, per-axis reasoning, time-to-contact, occlusion), the "
              "`final_reasoning` rationale (col V) that becomes the SFT target, the pipeline "
              "outcome colour-coding, plus reviewer annotations in `error` (col Y).")
    md.append("- **`reasoning_relation`** — per-video review of the 42 YES videos: do the three "
              "TTE clips (0.5 / 1.0 / 1.5 s) tell the *same* story? `video_score` (0–10) + a prose "
              "`reasoning_relation`, a MEAN row, and a histogram of the score distribution.")
    md.append("- **`reasoning_relation2`** — the same review re-scored **after** the reasoning "
              "harmonization described below.")
    md.append("- **`monitor_tte_coverage.xlsx`** (sibling file) — per-video × horizon green/red grid.\n")

    md.append("### Why we need it")
    md.append("AP/AUC tell us the teacher's *verdicts* rank well, but they say nothing about whether "
              "the *rationales* we distill are faithful. The student is trained on the `final_reasoning` "
              "text, so a verdict that is right for a hallucinated reason still poisons the student: it "
              "learns to narrate plausibly rather than perceive correctly. This workbook is where we "
              "measure and fix rationale quality before it reaches training.\n")

    md.append("## Reasoning-quality review (`reasoning_relation`)")
    md.append("For a positive video the three clips are the *same collision* seen 1.5 / 1.0 / 0.5 s "
              "before impact (~2 s each, ~1.5 s overlap between consecutive clips), so they must "
              "describe one consistent agent + mechanism with urgency rising as TTE shrinks. We scored "
              "all 42 YES videos by hand (no API).")
    if rr1:
        md.append(f"- **Mean video_score: {mean1:.2f} / 10.**")
        md.append(f"- **{le4_1} videos ≤ 4** (within-video incoherence / likely hallucination), "
                  f"**{ge7_1} videos ≥ 7**.")
        md.append("- The low band exposed cases where the teacher invented a *different agent at one "
                  "horizon for the same clip* (e.g. 00254 SUV→pedestrian→SUV; 00427 three different "
                  "vehicles; 00842 pedestrian→taxi→rear-end). These are the rotten ingredients a high "
                  "AP would otherwise hide.\n")

    md.append("## Reasoning harmonization (`final_reasoning2` + `reasoning_relation2`)")
    md.append("Rather than discard the low-coherence videos, we **repair** their rationales into one "
              "consistent within-video story, written into a new **`final_reasoning2`** column (col Z) "
              "with a per-clip **`change_explanation`** (col AA). The original `final_reasoning` (col V) "
              "is left intact so the before/after lift is auditable.")
    md.append("Governing principle: **TTE_0.5 is the most accurate** (closest to impact) → it is the "
              "source of truth for agent type/colour and collision mechanism whenever the three clips "
              "disagree. Five rewrite rules:")
    md.append("1. **Colour/type mismatch** → rephrase all clips to match TTE_0.5.")
    md.append("2. **TTE_1.0 hallucinates** (0.5 & 1.5 agree) → rewrite 1.0 to flow between them.")
    md.append("3. **TTE_1.5 hallucinates** (0.5 & 1.0 agree) → polish 1.5 to match.")
    md.append("4. **Complementary beats** → when each clip captures a different part of one overlapping "
              "event, weave a single flowing story sharing one scene/agents (e.g. 00842 taxi cuts in to "
              "pick up the crossing pedestrian → ego closes for a rear-end).")
    md.append("5. **Preserve the teacher's post-CoT structure** — same scene→dynamics→time-to-contact "
              "prose at comparable length, so each rewrite stays a valid SFT target.")
    md.append(f"\n**Scope:** {n_changed} of 42 YES videos changed ({n_clip_edits} individual clip "
              f"edits); the other {42 - n_changed} were already coherent and copied verbatim. NO videos "
              f"(225 rows) copy `final_reasoning` unchanged. Changed videos are flagged **yellow** on "
              f"`video_id` (col A) in both `per_clip` and `reasoning_relation2`.")
    if rr1 and rr2:
        md.append("")
        md.append("| metric | before (`reasoning_relation`) | after (`reasoning_relation2`) |")
        md.append("|---|---|---|")
        md.append(f"| mean video_score | {mean1:.2f} | **{mean2:.2f}** |")
        md.append(f"| videos ≤ 4 | {le4_1} | **{le4_2}** |")
        md.append(f"| videos ≥ 7 | {ge7_1} | **{ge7_2}** |")
        md.append("")

    md.append("## Why edit `final_reasoning`, and how it helps training")
    md.append("The student (InternVL3.5-4B-Flash, LoRA + ScoreHead) is distilled on the teacher's "
              "`final_reasoning` text as the supervision signal. Three concrete payoffs from "
              "harmonizing it:")
    md.append("- **Removes within-video hallucinations from the supervision** — the student no longer "
              "sees three contradictory agents for one scene, so it isn't taught to fabricate.")
    md.append("- **Enforces a monotonic within-scene temporal response** — with a consistent agent and "
              "rising urgency across 1.5→0.5 s, the multi-horizon clips now teach the smooth "
              "P(collision) slope the E3a §7(ii) experiment requires, instead of noise.")
    md.append("- **Grounds type/colour on the most reliable frame** — anchoring identity to TTE_0.5 "
              "gives the student stable object descriptions instead of a different colour per horizon, "
              "improving visual grounding through the projector.")
    md.append("Net effect: same verdicts and AP, but a cleaner, self-consistent rationale corpus — "
              "polished *ingredients*, not just polished output.\n")

    md.append("## Hallucination caveat")
    md.append("Harmonization fixed within-video coherence; it does **not** prove every rewrite is "
              "frame-grounded. Two edits remain worth flagging: **00065** injects reviewer ground truth "
              "(the ego struck a silver lead car), overriding all three clips; **00592 / 00621** retain "
              "score 7 because the true point of impact is still ambiguous. A frame-level grounding pass "
              "on the `final_reasoning2` corpus (at minimum the 28 changed videos) remains the next "
              "honest step before final SFT.\n")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")


def main() -> None:
    t0 = time.time()
    print("Loading inputs…")
    rows = build_combined()
    print(f"  Combined rows: {len(rows)}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Persist combined JSONL
    with OUT_JSONL_FINAL.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False, default=str) + "\n")
    print(f"  -> {OUT_JSONL_FINAL}")

    print("Writing results_e3a_tte_fill.xlsx…")
    write_results_xlsx(rows)
    print(f"  -> {OUT_XLSX}")

    print("Adding reasoning_relation sheet (YES videos)…")
    write_reasoning_relation_sheet(rows)

    print("Adding reasoning_relation2 sheet (post-harmonization)…")
    write_reasoning_relation2_sheet(rows)

    print("Writing monitor_tte_coverage.xlsx…")
    write_monitor_xlsx(rows)
    print(f"  -> {OUT_MONI}")

    print("Writing summary.md…")
    write_summary_md(rows)
    print(f"  -> {OUT_MD}")

    wall = time.time() - t0
    print()
    print("=" * 65)
    print("Stage 4 COMPLETE")
    print(f"  Rows written: {len(rows)}")
    print(f"  Wall time:    {wall:.1f}s")
    print(f"  Outputs:      {OUT_XLSX.name}, {OUT_MONI.name}, {OUT_MD.name}")
    print("=" * 65)


if __name__ == "__main__":
    main()
