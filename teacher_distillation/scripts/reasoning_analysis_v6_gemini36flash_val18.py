"""reasoning_analysis_v6_gemini36flash_val18.py -- scores PROMPT_G_OPT_v6_balanced
(unmodified, native output fields) from google/gemini-3.6-flash against the
18-clip GT validation set.

Direct within-model comparison point for reasoning_analysis_v9_gemini36flash_val18.py
-- same model, same 18 clips, only the prompt differs (V9's 866-token minimal
caption-first design vs v6_balanced's original ~965-token 7-step CoT-gated
verdict-first design, no caption fields at all).

Reads:
- dataset/manifests/val_e3a.jsonl
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds)
- outputs/prompt_bakeoff/semsup_val18/raw_v6_gemini36flash.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v6_gemini36flash_val18.xlsx

Colour rule (whole row, same convention as every prior round this session):
  GREEN  = verdict correct AND reasoning substantively matches GT's mechanism
  ORANGE = verdict correct but reasoning generic/partial, OR verdict wrong while
           the reasoning still substantively matches GT
  RED    = reasoning misses or CONTRADICTS GT's stated mechanism (regardless of
           verdict)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
V6G_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v6_gemini36flash.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v6_gemini36flash_val18.xlsx"

# (score 0-10, reasoning_match, rationale, colour) per clip, hand-scored against
# gt_reasoning_en, using verdict_reasoning + temporal_analysis as the caption
# equivalent (v6_balanced has no caption_neutral field).
SCORES = {
    "00077": (8, "MATCH", "Correctly identifies the closing gap and ego's lack of deceleration -- 'no evidence of effective deceleration or lateral evasive action' -- matching GT's 'EGO fails to brake in time' directly. Misses the initiating merge. Verdict correct.", "green"),
    "00147": (5, "PARTIAL", "Correct verdict, but causality is inverted: reports the sedan 'cuts sharply across the ego vehicle's path' where GT has EGO deviating into that sedan's lane. Right conflict, wrong attribution.", "orange"),
    "00283": (8, "MATCH", "Correctly identifies a pickup truck (not an SUV, unlike every Qwen3-VL round on this clip) performing a turning/crossing maneuver across ego's lane, matching GT's object and mechanism closely. Verdict correct.", "green"),
    "00319": (1, "CONTRADICT", "Asserts surrounding vehicles 'remain outside the ego vehicle's direct path with no evidence of converging' -- the opposite of GT's entering-car hazard. The 12th consecutive miss on this clip across every round and model tested to date.", "red"),
    "00372": (1, "MISS", "Frames the scene as stable following with 'risk remains low and stable throughout' -- misses GT's actual mechanism (lead sedan stops for crosswalk pedestrians during a right turn) entirely.", "red"),
    "00474": (1, "CONTRADICT", "States ego is 'smoothly passing the white van on the right... without any sudden braking or lateral intrusion' -- directly contradicts GT's van-performs-a-sharp-left-turn-into-ego-lane mechanism.", "red"),
    "00493": (8, "MATCH", "Correctly identifies the braking sedan and that 'the ego vehicle maintains a high rate of approach, severely reducing headway' -- matches GT's 'ego does not slow down' directly. Misses the initiating merge. Verdict correct.", "green"),
    "00529": (1, "CONTRADICT", "States the silver SUV 'maintain[s] parallel, stable trajectories' with 'no lateral lane cut-ins' -- a direct contradiction of GT's SUV-drifts-into-ego-lane mechanism.", "red"),
    "00687": (1, "CONTRADICT", "Confidently asserts 'no trajectory conflict exists. The grey SUV on the right is being safely passed by ego, receding' -- the opposite of GT's SUV-drifts-in-and-closes-rapidly mechanism. An apparent-motion misread stated as fact with no hedge.", "red"),
    "01153": (8, "MATCH", "Correctly resolves this clip without the fabricated crossing-sedan hallucination that broke every Qwen3-VL round tested (V5, V6, V8, V9) -- reports the sedan passing 'without encroaching,' matching GT's 'all vehicles remain in their respective lanes' directly. Verdict correct.", "green"),
    "01281": (5, "PARTIAL", "Correct verdict and no fabricated danger, but misses the specific mechanism GT describes (a braking lead pickup, controlled closing distance), framing it only as generically stable.", "orange"),
    "01504": (2, "PARTIAL", "States the lead vehicle maintains 'consistent speed' -- misses GT's central point that vehicles ahead are braking and ego notices and brakes in time. Verdict correct via a generic safe read, not the actual mechanism.", "orange"),
    "01550": (8, "MATCH", "Closely matches GT's 'controlled manner' framing: 'closing speed is moderate... ego gradually closes distance at a controlled rate.' Verdict correct.", "green"),
    "01552": (7, "MATCH", "Reasonable match (turn into a driveway, minivan clearing, box truck ahead), no fabricated 'school bus' -- the persistent artifact seen on this clip in V4 through V7. Verdict correct.", "green"),
    "01643": (8, "MATCH", "Clean, accurate match with no fabricated road-work sign (unlike several Qwen3-VL rounds on this clip). Verdict correct.", "green"),
    "01737": (8, "MATCH", "Accurately captures the empty, curving night road with no other agents. Verdict correct.", "green"),
    "02104": (6, "PARTIAL", "Mentions both the sedan and the truck ahead, matching GT's actors reasonably, but does not narrate the merge event GT describes -- frames it as steady throughout rather than a discrete event. No fabricated danger. Verdict correct.", "green"),
    "02117": (8, "MATCH", "Correctly reports a constant, safe distance to the lead vehicle (not 'closing,' unlike Qwen3-VL's V9 caption on this clip) and notes the right-side vehicles without encroachment. Verdict correct.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "gem_verdict", "reasoning_match",
           "verdict_reasoning", "temporal_analysis", "score", "score_explanation"]
WIDTHS = [11, 11, 11, 15, 55, 11, 14, 55, 55, 8, 60]
WRAP_COLS = {"gt_reasoning_en", "verdict_reasoning", "temporal_analysis", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    v6g = {r["video_id"]: r for r in _load_jsonl(V6G_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t}")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], v6g[vid]
        score, match, rationale, colour = SCORES.get(vid, (None, "MISS", "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"],
            "gt_reasoning_en": v["gt_reasoning_en"],
            "gem_verdict": q["verdict"], "reasoning_match": match,
            "verdict_reasoning": q["verdict_reasoning"], "temporal_analysis": q["temporal_analysis"],
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["gem_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["gem_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["gem_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["gem_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["gem_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    cc = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}
    match_counts = {m: sum(1 for r in rows if r["reasoning_match"] == m)
                     for m in ("MATCH", "PARTIAL", "MISS", "CONTRADICT")}

    metrics = [
        ("n", n),
        ("--- REASONING ALIGNMENT ---", ""),
        ("reasoning MATCH (substantively correct)", match_counts["MATCH"]),
        ("reasoning PARTIAL (right shape, wrong detail)", match_counts["PARTIAL"]),
        ("reasoning MISS (misses GT's mechanism)", match_counts["MISS"]),
        ("reasoning CONTRADICT (states the opposite of / fabricates beyond GT)", match_counts["CONTRADICT"]),
        ("mean_hand_score (0-10)", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green / n_orange / n_red", f"{cc['green']} / {cc['orange']} / {cc['red']}"),
        ("--- verdict metrics ---", ""),
        ("verdict_accuracy", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP / FP / TN / FN", f"{tp} / {fp} / {tn} / {fn}"),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("model", "google/gemini-3.6-flash"),
        ("prompt", "PROMPT_G_OPT_v6_balanced (unmodified, native fields, no caption)"),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 30

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
