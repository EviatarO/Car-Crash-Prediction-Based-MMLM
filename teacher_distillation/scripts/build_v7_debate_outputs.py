"""Build results_v7_debate.xlsx and leaderboard_v7_debate.md from
   outputs/prompt_bakeoff/v7_stride8/{v6_s8_hires,v7_debate}.jsonl

xlsx is color-coded:
  - green   (C6EFCE)  pass_v6     : correct on first pass
  - orange  (FFEB9C)  pass_debate : wrong v6, fixed by v7 recovery
  - red     (FFC7CE)  fail        : wrong after debate (or first-pass error)

GT reasoning column comes from dataset/teacher_dataset_GT_self_imply.xlsx (col G).
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
V7_DIR    = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v7_stride8"
FIRST_JSONL  = V7_DIR / "v6_s8_hires.jsonl"
DEBATE_JSONL = V7_DIR / "v7_debate.jsonl"
GT_XLSX      = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"

OUT_XLSX = V7_DIR / "results_v7_debate.xlsx"
OUT_MD   = V7_DIR / "leaderboard_v7_debate.md"

# ── colours ─────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
HEADER = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
BUCKET_FILL = {"pass_v6": GREEN, "pass_debate": ORANGE, "fail": RED}

# v6@stride-4 18-clip baseline numbers (from prior misaligned run, for reference)
V6_BASELINE = {
    "first_pass_acc":     13/18,    # 72.2%
    "post_debate_acc":    15/18,    # 83.3%
    "v6_TP_FP_TN_FN":     (8, 4, 5, 1),
    "final_TP_FP_TN_FN":  (8, 2, 7, 1),
}


def _load_jsonl(path: Path) -> List[Dict]:
    if not path.exists():
        return []
    return [json.loads(l) for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


def _norm_vid(v) -> str:
    s = str(v).strip().split(".")[0]
    return f"{int(s):05d}" if s.isdigit() else s


def _confusion(df: pd.DataFrame, pred_col: str) -> Dict[str, int]:
    tp = int(((df[pred_col] == "YES") & (df["gt_verdict"] == "YES")).sum())
    fp = int(((df[pred_col] == "YES") & (df["gt_verdict"] == "NO")).sum())
    tn = int(((df[pred_col] == "NO")  & (df["gt_verdict"] == "NO")).sum())
    fn = int(((df[pred_col] == "NO")  & (df["gt_verdict"] == "YES")).sum())
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn}


def _style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def _color_rows(ws, bucket_col_idx: int) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        bucket_cell = row[bucket_col_idx - 1]
        fill = BUCKET_FILL.get(str(bucket_cell.value or ""))
        if fill:
            for cell in row:
                cell.fill = fill


def main() -> None:
    first  = {r["video_id"]: r for r in _load_jsonl(FIRST_JSONL)}
    debate = {r["video_id"]: r for r in _load_jsonl(DEBATE_JSONL)}

    gt_df = pd.read_excel(GT_XLSX)
    gt_map = {
        _norm_vid(row[gt_df.columns[0]]): str(row[gt_df.columns[6]]).strip()
        for _, row in gt_df.iterrows()
    }

    rows: List[Dict] = []
    for vid in sorted(first):
        h = first[vid]
        gt        = h["gt_verdict"]
        gt_reason = gt_map.get(vid, "")
        s8_verdict = h.get("verdict") or ""
        s8_correct = (s8_verdict == gt)
        s8_bert    = round(h.get("scores", {}).get("alignment", 0.0) or 0.0, 3)
        s8_reason  = (h.get("reasoning") or "").strip()

        d = debate.get(vid)
        if d is None:
            recovery_prompt  = ""
            recovery_verdict = ""
            recovery_correct = ""
            recovery_bert    = ""
            recovery_reason  = ""
            final_verdict    = s8_verdict
            flipped          = ""
        else:
            recovery_prompt  = d.get("recovery_prompt", "")
            recovery_verdict = d.get("recovery_verdict") or ""
            recovery_correct = (recovery_verdict == gt) if recovery_verdict else ""
            recovery_bert    = round(d.get("scores", {}).get("alignment", 0.0) or 0.0, 3)
            recovery_reason  = (d.get("recovery_reasoning") or "").strip()
            final_verdict    = recovery_verdict if recovery_verdict else s8_verdict
            if not s8_correct and recovery_verdict == gt:
                flipped = "FIXED"
            elif s8_correct and recovery_verdict and recovery_verdict != gt:
                flipped = "BROKE"
            else:
                flipped = "still-wrong"
        final_correct = (final_verdict == gt)

        if s8_correct:
            bucket = "pass_v6"
        elif flipped == "FIXED":
            bucket = "pass_debate"
        else:
            bucket = "fail"

        rows.append({
            "video_id": vid,
            "gt_verdict": gt,
            "gt_reasoning_en": gt_reason,
            "v6_s8__verdict": s8_verdict,
            "v6_s8__correct": s8_correct,
            "v6_s8__bert": s8_bert,
            "v6_s8__reasoning": s8_reason,
            "recovery_prompt": recovery_prompt,
            "recovery__verdict": recovery_verdict,
            "recovery__correct": recovery_correct,
            "recovery__bert": recovery_bert,
            "recovery__reasoning": recovery_reason,
            "final_after_debate": final_verdict,
            "final_correct": final_correct,
            "flipped_by_debate": flipped,
            "passes_for_student_training": bucket,
        })

    df = pd.DataFrame(rows)
    n = len(df)

    # Accuracy
    s8_valid = df["v6_s8__verdict"].isin(["YES", "NO"])
    s8_correct_n = int(df.loc[s8_valid, "v6_s8__correct"].sum())
    s8_total     = int(s8_valid.sum())
    s8_acc       = s8_correct_n / s8_total if s8_total else 0.0

    final_correct_n = int(df["final_correct"].sum())
    final_acc       = final_correct_n / n if n else 0.0

    df["_final_cm"] = df["final_after_debate"].apply(lambda x: x if x in ("YES", "NO") else None)
    cm_s8    = _confusion(df[s8_valid], "v6_s8__verdict")
    cm_final = _confusion(df[df["_final_cm"].notna()], "_final_cm")
    df.drop(columns=["_final_cm"], inplace=True)

    n_debated  = int((df["recovery_prompt"] != "").sum())
    n_fixed    = int((df["flipped_by_debate"] == "FIXED").sum())
    n_broke    = int((df["flipped_by_debate"] == "BROKE").sum())
    n_still    = int((df["flipped_by_debate"] == "still-wrong").sum())
    fn_rescue  = int(((df["flipped_by_debate"] == "FIXED") &
                      (df["recovery_prompt"] == "PROMPT_G_OPT_v7_TP_RECOVERY")).sum())
    fp_rescue  = int(((df["flipped_by_debate"] == "FIXED") &
                      (df["recovery_prompt"] == "PROMPT_G_OPT_v7_TN_RECOVERY")).sum())
    n_pass_v6     = int((df["passes_for_student_training"] == "pass_v6").sum())
    n_pass_debate = int((df["passes_for_student_training"] == "pass_debate").sum())
    n_fail        = int((df["passes_for_student_training"] == "fail").sum())

    summary = pd.DataFrame([
        {"metric": "n_clips",                            "value": n},
        {"metric": "v6_balanced_s8 accuracy",            "value": f"{s8_acc:.1%}  ({s8_correct_n}/{s8_total})"},
        {"metric": "after-debate accuracy (v7)",         "value": f"{final_acc:.1%}  ({final_correct_n}/{n})"},
        {"metric": "v6_s8 confusion (TP/FP/TN/FN)",     "value": f"{cm_s8['TP']}/{cm_s8['FP']}/{cm_s8['TN']}/{cm_s8['FN']}"},
        {"metric": "after-debate confusion (TP/FP/TN/FN)","value": f"{cm_final['TP']}/{cm_final['FP']}/{cm_final['TN']}/{cm_final['FN']}"},
        {"metric": "clips debated",                      "value": n_debated},
        {"metric": "FIXED (wrong->right)",               "value": n_fixed},
        {"metric": "BROKE (right->wrong)",               "value": n_broke},
        {"metric": "still-wrong",                        "value": n_still},
        {"metric": "FN rescued by v7 TP_RECOVERY",       "value": fn_rescue},
        {"metric": "FP rescued by v7 TN_RECOVERY",       "value": fp_rescue},
        {"metric": "--- Student-training yield ---",     "value": ""},
        {"metric": "pass_v6 (green)",                    "value": n_pass_v6},
        {"metric": "pass_debate (orange)",               "value": n_pass_debate},
        {"metric": "fail (red)",                         "value": n_fail},
        {"metric": "TOTAL valid for training",           "value": n_pass_v6 + n_pass_debate},
        {"metric": "--- vs. v6@stride-4 baseline (misaligned recovery, 18 clips) ---", "value": ""},
        {"metric": "v6@s4 first-pass acc",               "value": f"{V6_BASELINE['first_pass_acc']:.1%}  (13/18)"},
        {"metric": "v6@s4 post-debate acc",              "value": f"{V6_BASELINE['post_debate_acc']:.1%}  (15/18)"},
    ])

    # Write xlsx
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="per_clip", index=False)
        summary.to_excel(w, sheet_name="summary", index=False)

    wb = load_workbook(OUT_XLSX)
    _style_sheet(wb["per_clip"])
    bucket_col_idx = df.columns.get_loc("passes_for_student_training") + 1
    _color_rows(wb["per_clip"], bucket_col_idx)
    _style_sheet(wb["summary"])
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print()
    print(summary.to_string(index=False))

    # ── leaderboard.md ───────────────────────────────────────────────────────
    md = []
    md.append("# Leaderboard - v6_balanced_s8 (stride=8, 4s window) + v7 Debate Recovery (18 GT clips)\n")
    md.append("**Setup:** 18 GT-reasoning clips, frames @ NATIVE 1280x720, **stride=8 (16 frames over ~4 seconds, same final-frame anchor as the stride-4 test)**, Gemini `detail=\"high\"`, `temperature=0.1`.")
    md.append("")
    md.append("**Debate:** Each clip where the first pass disagreed with GT was retested with a v7 recovery prompt:")
    md.append("- GT=YES, first=NO (FN) -> `PROMPT_G_OPT_v7_TP_RECOVERY` (senior collision-prevention specialist; YES-biased, late-frame focus)")
    md.append("- GT=NO, first=YES (FP) -> `PROMPT_G_OPT_v7_TN_RECOVERY` (senior false-alarm-reduction specialist; NO-biased, sustained-conflict requirement)")
    md.append("")
    md.append("> *Note:* The v6@stride-4 baseline numbers below are from a run where the v6 recovery file/variable names were inverted vs their content. Direct A/B comparability is therefore partial. See `quiet-orbiting-clock.md` plan, 'Critical finding'.")
    md.append("")

    md.append("## Headline metrics\n")
    md.append("| Stage | Accuracy | TP | FP | TN | FN |")
    md.append("|-------|----------|----|----|----|----|")
    md.append(f"| v6_balanced_s8 (first pass) | **{s8_acc:.1%}** ({s8_correct_n}/{s8_total}) "
              f"| {cm_s8['TP']} | {cm_s8['FP']} | {cm_s8['TN']} | {cm_s8['FN']} |")
    md.append(f"| after v7 debate | **{final_acc:.1%}** ({final_correct_n}/{n}) "
              f"| {cm_final['TP']} | {cm_final['FP']} | {cm_final['TN']} | {cm_final['FN']} |")
    md.append(f"| _ref: v6@stride-4 first pass_ | _72.2%_ (13/18) | _8_ | _4_ | _5_ | _1_ |")
    md.append(f"| _ref: v6@stride-4 + v6 recovery (misaligned)_ | _83.3%_ (15/18) | _8_ | _2_ | _7_ | _1_ |")
    md.append("")

    md.append("## Debate yield\n")
    md.append(f"- Clips debated: **{n_debated}**")
    md.append(f"- FIXED (wrong->right): **{n_fixed}**")
    md.append(f"  - FN rescued by v7 TP_RECOVERY: {fn_rescue} / {cm_s8['FN']} FN clips")
    md.append(f"  - FP rescued by v7 TN_RECOVERY: {fp_rescue} / {cm_s8['FP']} FP clips")
    md.append(f"- BROKE (right->wrong): **{n_broke}**")
    md.append(f"- still-wrong: **{n_still}**")
    md.append("")

    md.append("## Student-training yield\n")
    md.append("| Bucket | Count | Meaning |")
    md.append("|--------|-------|---------|")
    md.append(f"| GREEN  pass_v6      | **{n_pass_v6}** | correct on first pass |")
    md.append(f"| ORANGE pass_debate  | **{n_pass_debate}** | fixed by v7 recovery |")
    md.append(f"| RED    fail         | **{n_fail}** | wrong after debate |")
    md.append(f"| **Valid total**     | **{n_pass_v6 + n_pass_debate} / {n}** | available for distillation |")
    md.append("")

    md.append("## Per-clip table\n")
    md.append("| Clip | GT | v6_s8 | recovery | final | flip | bucket |")
    md.append("|------|----|-------|----------|-------|------|--------|")
    for _, row in df.iterrows():
        rv = row["recovery__verdict"] or "-"
        rp = row["recovery_prompt"] or ""
        rp_short = rp.replace("PROMPT_G_OPT_v7_", "").replace("_RECOVERY", "") if rp else ""
        rec_cell  = f"{rv} ({rp_short})" if rp else "-"
        flip      = row["flipped_by_debate"] or "-"
        s8_mark   = "OK" if row["v6_s8__correct"] else "XX"
        fin_mark  = "OK" if row["final_correct"] else "XX"
        bkt_label = {"pass_v6": "GREEN", "pass_debate": "ORANGE", "fail": "RED"}.get(
            row["passes_for_student_training"], "")
        md.append(
            f"| {row['video_id']} | {row['gt_verdict']} "
            f"| {row['v6_s8__verdict'] or '??'} {s8_mark} "
            f"| {rec_cell} "
            f"| {row['final_after_debate']} {fin_mark} "
            f"| {flip} "
            f"| {bkt_label} {row['passes_for_student_training']} |"
        )
    md.append("")

    md.append("## Files\n")
    md.append(f"- Per-clip xlsx: `outputs/prompt_bakeoff/v7_stride8/{OUT_XLSX.name}`")
    md.append(f"- Leaderboard:   `outputs/prompt_bakeoff/v7_stride8/{OUT_MD.name}`")
    md.append("- First-pass jsonl: `outputs/prompt_bakeoff/v7_stride8/v6_s8_hires.jsonl`")
    md.append("- Debate jsonl:     `outputs/prompt_bakeoff/v7_stride8/v7_debate.jsonl`")
    md.append("")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")
    print(f"\nWrote {OUT_MD}")


if __name__ == "__main__":
    main()
