"""reasoning_analysis_qwen3vl_val18.py -- scores PROMPT_SEMSUP_V4_QWEN captions
from qwen/qwen3-vl-235b-a22b-thinking against the 18-clip GT validation set,
formatted like outputs/prompt_bakeoff/v11_100clips/results_v6_debate_v11
(version 1).xlsb.xlsx (whole-row colour fill, wrapped text, sized rows).

Reads:
- dataset/manifests/val_e3a.jsonl (video_id, gt_verdict, requested_time_to_event,
  gt_reasoning_en)
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds -
  NOT present in val_e3a.jsonl; these two files together cover exactly the 18
  val clips and are the runs that used these same _hires frames, so they are
  authoritative. Other files disagree on t_seconds for some clips because
  different experiment generations placed windows differently - do not use them)
- outputs/prompt_bakeoff/semsup_val18/raw_v4_qwen3vl235b.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_qwen3vl_val18.xlsx

Colour rule (whole row, agreed with the user):
  GREEN  = verdict correct AND caption matches GT reasoning
  ORANGE = verdict correct but caption is middling/generic,
           OR verdict wrong but the caption still describes the scene well
           (this is the caption-is-the-training-target case - a good
           description with a miscalibrated verdict retains real value)
  RED    = verdict wrong AND caption wrong/hallucinated
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
QWEN3VL_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v4_qwen3vl235b.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_qwen3vl_val18.xlsx"

# (score 0-10, rationale, colour) per clip. Colour decided per the rule above,
# using GT reasoning as the reference. Caption = caption_neutral + risk_clause.
SCORES = {
    "00077": (3, "Identifies the black sedan ahead and notes decreasing distance, but calls it 'normal highway following' - misses the braking/rapid-closure severity GT describes; verdict wrong.", "orange"),
    "00147": (6, "Captures the two-sedan converging scenario and correct verdict, but inverts causality (says the OTHER sedan merges into ego's lane; GT says EGO deviates into the other vehicle's lane).", "orange"),
    "00283": (7, "'White SUV cutting left... directly ahead at close proximity with brake lights' matches GT's blocking-turn mechanism and verdict well; object misidentified as SUV rather than pickup+trailer.", "green"),
    "00319": (1, "Describes calm following-traffic through an intersection; misses GT's crossing vehicle from the right entirely. Wrong verdict.", "red"),
    "00372": (2, "Describes a stable following scenario; misses the stopping-sedan/pedestrian-crosswalk hazard GT describes. Wrong verdict.", "red"),
    "00474": (1, "States the van maintains 'consistent distance' - directly contradicts GT's left-turn-into-lane event, the same miss every model has made on this clip. Wrong verdict.", "red"),
    "00493": (2, "Describes stable following distance; misses EGO's own merge context and the sedan's braking that GT says ego fails to react to. Wrong verdict.", "red"),
    "00529": (1, "States the silver SUV stays 'parallel' - the opposite of GT's claim that it drifts into the ego lane due to an obstruction. Wrong verdict.", "red"),
    "00687": (6, "'Gray SUV merging from right lane into ego lane' correctly captures GT's core mechanism, but the risk_clause calls it 'normal merging traffic' and the verdict is NO - a calibration miss on a correctly-perceived hazard.", "orange"),
    "01153": (7, "Correctly identifies a crossing sedan while ego proceeds through a green light, without escalating to a hallucinated collision - matches GT's calm resolution and verdict.", "green"),
    "01281": (8, "Matches GT's controlled, reasonable-distance framing closely; does not fabricate a merge event as several other models did. Correct verdict.", "green"),
    "01504": (6, "Matches GT's controlled-following framing and verdict; 'red SUV' vs GT's 'dark SUV' is a colour-naming mismatch seen consistently across every model tested on this clip.", "green"),
    "01550": (8, "Closely matches GT's controlled-closing, maintained-distance description. Correct verdict.", "green"),
    "01552": (6, "Captures the gas-station-adjacent scene (crossing minivan, truck ahead) reasonably; invents a 'school bus' detail not in GT. Correct verdict.", "green"),
    "01643": (5, "Correct empty-road verdict, but introduces parked cars and oncoming traffic not present in GT's 'no vehicles around it'.", "orange"),
    "01737": (9, "Clean match to GT's empty, curving, night-time interchange description. Correct verdict.", "green"),
    "02104": (6, "Generic 'multiple vehicles maintaining consistent distance' avoids fabricating a merge/hazard object (unlike other models on this clip) and is directionally consistent with GT, but is vague relative to GT's specific merging-sedan detail.", "orange"),
    "02117": (8, "Correctly identifies the sedan at constant distance without hallucinating the black-SUV-merge event that broke every Gemini run and is the third of three Qwen/GPT-family model tested to resolve this clip correctly.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "qwen_verdict", "qwen_caption", "score", "score_explanation"]
WIDTHS = [11, 11, 11, 18, 60, 11, 60, 8, 55]
WRAP_COLS = {"gt_reasoning_en", "qwen_caption", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    qwen = {r["video_id"]: r for r in _load_jsonl(QWEN3VL_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t} - refusing to substitute a "
                            f"value from a disagreeing source (see module docstring).")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], qwen[vid]
        qwen_verdict = "YES" if q["verdict"] == 1 else "NO"
        caption = f"{q['caption_neutral']}, {q['risk_clause']}"
        score, rationale, colour = SCORES.get(vid, (None, "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"], "gt_reasoning_en": v["gt_reasoning_en"],
            "qwen_verdict": qwen_verdict, "qwen_caption": caption,
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        # size the row to the longest wrapped cell (~90 chars/line at these widths)
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    # summary sheet
    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["qwen_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["qwen_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["qwen_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["qwen_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["qwen_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    colour_counts = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}
    metrics = [
        ("n", n), ("verdict_accuracy", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP", tp), ("FP", fp), ("TN", tn), ("FN", fn),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("mean_score", round(sum(scores) / len(scores), 2)),
        ("median_score", sorted(scores)[len(scores) // 2]),
        ("n_green", colour_counts["green"]), ("n_orange", colour_counts["orange"]),
        ("n_red", colour_counts["red"]),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
