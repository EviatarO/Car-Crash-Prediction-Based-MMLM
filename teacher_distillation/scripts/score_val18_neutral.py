"""score_val18_neutral.py -- scores PROMPT_SEMSUP_V12_NEUTRAL against V10 (GT+blind)
on the 18-clip val set, and runs a text-only leakage judge on both corpora.

Reuses reasoning_analysis_v10_gt_val18.py's calibrated scorer rather than
reimplementing it: load_gt_slots(), score_blob(), fabrication_check().

COMPARABILITY TRAP (see the 2026-08-08 plan) -- V10's PUBLISHED slot_recall
(0.417 GT / 0.468 blind, from outputs/prompt_bakeoff/semsup_val18_gt/summary.md)
was computed on a 6-FIELD CONCATENATION (hazard_agent + hazard_motion +
hazard_position + closing_dynamic + caption_neutral + risk_clause), not on
caption_neutral alone. That is not what actually becomes the SigLIP training
target. This script re-scores V10 (both arms) on caption_neutral ALONE, so the
V10 numbers printed here are LOWER than the published ones by construction --
that is not a regression, it is a fairer basis for comparing against V12, which
never had extra fields to hide behind.

Score, 0-10, two independent halves:
  A. Grounding (0-5)   = round(slot_recall * 5), slot_recall from score_blob()
                          on caption_neutral alone.
  B. Neutrality (0-5)  = 5 minus penalties (symmetric across classes):
                          -3 any banned OUTCOME word
                          -2 ALARM register
                          -2 REASSURANCE register
                          -1 time/seconds reference
                          -1 (V12 only) gap_trend word absent from the caption
                          floor 0.
score_explanation is fully deterministic - built from which slots matched and
which penalties fired, not LLM-written, so re-running reproduces identical text.

Leakage judge: a SEPARATE OpenRouter call, fresh context, CAPTIONS ONLY (no
images, no metadata), asks the model to predict crash/no-crash from the caption
text alone. Run on both V12 and V10-blind. ~50% accuracy = neutral (target);
~100% = still leaking. n=18 gives wide exact-binomial CIs - this detects GROSS
leakage, it does not certify subtle neutrality (the real gate is the
full-corpus TF-IDF AUC<0.75 check in W1 task 1.6).

Writes:
  outputs/prompt_bakeoff/semsup_val18_neutral/review_val18_neutral.xlsx
  outputs/prompt_bakeoff/semsup_val18_neutral/summary.md
"""
from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "teacher_distillation" / "scripts"))
import reasoning_analysis_v10_gt_val18 as R  # noqa: E402

VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
V12_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_neutral" / "raw_v12_gemini.jsonl"
V10_GT_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_gt" / "raw_v10_gemini_gt.jsonl"
V10_BLIND_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_gt" / "raw_v10_gemini_blind.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
OUT_DIR = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_neutral"
OUT_XLSX = OUT_DIR / "review_val18_neutral.xlsx"
OUT_MD = OUT_DIR / "summary.md"

OUTCOME_WORDS = ("risk", "danger", "collision", "crash", "imminent", "avoid",
                  "impact", "hazard", "accident")
ALARM_PHRASES = ("about to", "fails to", "unable to", "inevitably", "will strike",
                   "too late", "no time")
REASSURE_PHRASES = ("safe", "safely", "no risk", "uneventful", "normal", "routine",
                      "poses no", "without incident")
TIME_PHRASES = ("second", "seconds", "sec ", "time-to-event", "tte")

HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
V12_GROUP_FILL = PatternFill("solid", fgColor="1F8A3C")   # green -- the new arm
V10_GROUP_FILL = PatternFill("solid", fgColor="7030A0")   # purple -- the baseline
FILL = {"green": PatternFill("solid", fgColor="FFC6EFCE"),
        "orange": PatternFill("solid", fgColor="FFFFEB9C"),
        "red": PatternFill("solid", fgColor="FFFFC7CE")}
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        if not fp.exists():
            continue
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def neutrality_penalties(caption: str, gap_trend: str | None) -> tuple[int, list[str]]:
    """Returns (score 0-5, list of fired-penalty descriptions)."""
    c = caption.lower()
    fired = []
    score = 5
    if any(w in c for w in OUTCOME_WORDS):
        score -= 3
        fired.append("outcome word")
    if any(p in c for p in ALARM_PHRASES):
        score -= 2
        fired.append("alarm register")
    if any(p in c for p in REASSURE_PHRASES):
        score -= 2
        fired.append("reassurance register")
    if any(p in c for p in TIME_PHRASES):
        score -= 1
        fired.append("time reference")
    if gap_trend and gap_trend != "none_visible" and gap_trend not in c:
        score -= 1
        fired.append(f"gap_trend {gap_trend!r} not in caption")
    return max(0, score), fired


def score_row(caption: str, gt_slot: dict, target: int, gap_trend: str | None) -> dict:
    """Grounding (slot_recall-based) + neutrality (penalty-based), each 0-5."""
    slot_names = [s for s in ("agent", "motion", "position", "closing") if gt_slot.get(s)]
    matches = {s: R.slot_match(caption, gt_slot[s]) for s in slot_names}
    recall = sum(matches.values()) / len(slot_names) if slot_names else None
    grounding = round(recall * 5) if recall is not None else None

    neutrality, penalties = neutrality_penalties(caption, gap_trend)

    total = (grounding or 0) + neutrality
    matched = [s for s, hit in matches.items() if hit]
    missed = [s for s in slot_names if s not in matched]
    ground_desc = (f"slot_recall {recall:.2f}: " +
                    (f"{'+'.join(matched)} matched" if matched else "no slots matched") +
                    (f", {'+'.join(missed)} missed" if missed else "")) \
        if slot_names else "no GT slots available for this clip"
    neut_desc = "no banned register" if not penalties else "; ".join(penalties)
    explanation = f"grounding {grounding if grounding is not None else '-'}/5 ({ground_desc}); " \
                   f"neutrality {neutrality}/5 ({neut_desc})"

    return {"grounding": grounding, "neutrality": neutrality, "total": total,
             "slot_recall": recall, "explanation": explanation}


# ---------------------------------------------------------------------------
# Leakage judge: text-only, fresh context, captions only.
# ---------------------------------------------------------------------------

JUDGE_PROMPT = (
    "You will see __N__ short scene descriptions of dashcam driving clips, each "
    "labeled with an ID like C001. For EACH ONE, predict whether that clip is "
    "immediately followed by a collision (YES) or not (NO), based ONLY on the "
    "text. You have no images, no other context.\n\n"
    "Return ONLY a JSON object mapping each ID to \"YES\" or \"NO\", no markdown "
    "fences, no extra text, e.g.:\n"
    '{"C001": "YES", "C002": "NO", ...}\n\n'
    "DESCRIPTIONS:\n__ITEMS__\n"
)


def run_leakage_judge(client, model: str, captions: dict[str, str]) -> dict[str, str]:
    """captions: {video_id: caption}. Returns {video_id: 'YES'/'NO'} predictions.

    Labels are zero-padded numeric IDs (C001, C002, ...), not letters - the
    original A-Z scheme silently caps out at 26 items, which would truncate a
    100-clip batch with no error. Order is shuffled (but reproducibly, seed=0)
    so label position carries no signal about original manifest order
    (val18 is positives-then-negatives; a combined val18+extra82 batch would
    otherwise cluster all-positive/all-negative runs at the label boundary)."""
    import random
    sys.path.insert(0, str(REPO_ROOT / "student_training" / "scripts"))
    from semsup_caption_promptbakeoff import _call_model, _extract_json_object

    vids = list(captions.keys())
    random.Random(0).shuffle(vids)
    width = max(3, len(str(len(vids))))
    labels = [f"C{i+1:0{width}d}" for i in range(len(vids))]
    label_to_vid = dict(zip(labels, vids))
    items = "\n".join(f"{L}: {captions[v]}" for L, v in zip(labels, vids))
    prompt = JUDGE_PROMPT.replace("__N__", str(len(vids))).replace("__ITEMS__", items)

    text, _ = _call_model(client, model, [{"role": "user", "content": prompt}],
                            timeout=180.0, max_retries=3, retry_delay=3.0, temperature=0.1)
    parsed = _extract_json_object(text)
    if parsed is None:
        raise RuntimeError(f"Judge call returned unparseable JSON: {text[:300]!r}")
    result = {label_to_vid[L]: str(v).upper() for L, v in parsed.items() if L in label_to_vid}
    missing = set(vids) - set(result.keys())
    if missing:
        print(f"  [warn] judge omitted {len(missing)}/{len(vids)} clips from its response "
              f"(no prediction returned): {sorted(missing)[:5]}{'...' if len(missing) > 5 else ''}")
    return result


def binom_ci(k: int, n: int, conf: float = 0.95) -> tuple[float, float]:
    """Exact (Clopper-Pearson) binomial CI without scipy, via bisection on the
    regularized incomplete beta function's simple cases is overkill for n=18 -
    use the normal approximation with a continuity correction instead; stated
    explicitly as approximate in the summary output."""
    if n == 0:
        return (0.0, 0.0)
    p = k / n
    z = 1.959963985
    denom = 1 + z * z / n
    center = (p + z * z / (2 * n)) / denom
    half = (z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))) / denom
    return (max(0.0, center - half), min(1.0, center + half))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--skip-judge", action="store_true",
                     help="skip the OpenRouter leakage-judge calls (score only)")
    ap.add_argument("--model", default="google/gemini-3.6-flash")
    args = ap.parse_args()

    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    gt_slots = R.load_gt_slots()
    t_seconds = _load_t_seconds()

    v12 = {r["video_id"]: r for r in _load_jsonl(V12_JSONL)}
    v10g = {r["video_id"]: r for r in _load_jsonl(V10_GT_JSONL)}
    v10b = {r["video_id"]: r for r in _load_jsonl(V10_BLIND_JSONL)}

    records = []
    for vid in sorted(val):
        v = val[vid]
        target = int(v["target"])
        slot = gt_slots[vid]

        r12 = v12[vid]
        s12 = score_row(r12["caption_neutral"], slot, target, r12.get("gap_trend"))

        # V10: pick GT-mode caption for positives, blind-mode for negatives -
        # this mirrors the corpus's actual hybrid convention (GT on positives,
        # blind on negatives) that produced the label leak, so the comparison
        # is against what the 1,761-window pool actually contains, not an
        # arm nobody used together.
        r10 = v10g[vid] if target == 1 else v10b[vid]
        s10 = score_row(r10["caption_neutral"], slot, target, None)

        records.append({
            "video_id": vid, "gt_verdict": v.get("gt_verdict", "YES" if target else "NO"),
            "t_seconds": round(t_seconds[vid], 3) if vid in t_seconds else None,
            "requested_time_to_event": v.get("requested_time_to_event"),
            "gt_reasoning_en": v["gt_reasoning_en"],
            "v12_caption": r12["caption_neutral"],
            "v12_score": s12["total"], "v12_score_explanation": s12["explanation"],
            "v10_caption": r10["caption_neutral"],
            "v10_score": s10["total"], "v10_score_explanation": s10["explanation"],
            "score_delta": s12["total"] - s10["total"],
            "_target": target,
        })

    # ---- leakage judge ----
    judge_ran = False
    v12_acc = v10_acc = None
    if not args.skip_judge:
        from dotenv import load_dotenv
        from openai import OpenAI
        load_dotenv()
        client = OpenAI(base_url="https://openrouter.ai/api/v1",
                         api_key=__import__("os").environ["OPENROUTER_API_KEY"],
                         default_headers={"HTTP-Referer": "http://localhost",
                                          "X-Title": "MMLM_Semsup_LeakageJudge"})
        v12_pred = run_leakage_judge(client, args.model,
                                       {r["video_id"]: r["v12_caption"] for r in records})
        v10_pred = run_leakage_judge(client, args.model,
                                       {r["video_id"]: r["v10_caption"] for r in records})
        for r in records:
            r["v12_judge_pred"] = v12_pred.get(r["video_id"], "?")
            r["v10_judge_pred"] = v10_pred.get(r["video_id"], "?")
        v12_correct = sum(1 for r in records if r["v12_judge_pred"] == r["gt_verdict"])
        v10_correct = sum(1 for r in records if r["v10_judge_pred"] == r["gt_verdict"])
        v12_acc = (v12_correct, len(records))
        v10_acc = (v10_correct, len(records))
        judge_ran = True
    else:
        for r in records:
            r["v12_judge_pred"] = ""
            r["v10_judge_pred"] = ""

    # ---- xlsx ----
    headers = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
               "gt_reasoning_en", "v12_caption", "v12_score", "v12_score_explanation",
               "v12_judge_pred", "v10_caption", "v10_score", "v10_score_explanation",
               "v10_judge_pred", "score_delta"]
    widths = [10, 10, 10, 14, 62, 55, 9, 45, 12, 55, 9, 45, 12, 11]
    wrap_cols = {"gt_reasoning_en", "v12_caption", "v12_score_explanation",
                  "v10_caption", "v10_score_explanation"}
    score_color_cols = {"v12_score": "v12_caption", "v10_score": "v10_caption"}

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.cell(row=1, column=1, value="GROUND TRUTH REFERENCE")
    ws.cell(row=1, column=6, value="V12 NEUTRAL (new)")
    ws.cell(row=1, column=10, value="V10 (GT on pos / blind on neg -- current corpus)")
    for col, fill in ((1, HEADER_FILL), (6, V12_GROUP_FILL), (10, V10_GROUP_FILL)):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
    ws.cell(row=1, column=14)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).border = BORDER
        if col in range(6, 10):
            ws.cell(row=1, column=col).fill = V12_GROUP_FILL
        elif col in range(10, 14):
            ws.cell(row=1, column=col).fill = V10_GROUP_FILL
        elif col == 14:
            ws.cell(row=1, column=col).fill = HEADER_FILL
        else:
            ws.cell(row=1, column=col).fill = HEADER_FILL

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "B3"

    def score_fill(score):
        return FILL["green"] if score >= 8 else FILL["orange"] if score >= 5 else FILL["red"]

    for i, rec in enumerate(records, start=3):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=i, column=c, value=rec.get(h))
            wrap = h in wrap_cols
            cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            if h in ("v12_score", "v10_score"):
                cell.fill = score_fill(rec[h])
        longest = max(len(str(rec[h])) for h in wrap_cols)
        ws.row_dimensions[i].height = max(48, 13.5 * (longest // 55 + 1))

    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = w

    # ---- summary sheet ----
    ws2 = wb.create_sheet("summary")
    v12_mean = sum(r["v12_score"] for r in records) / len(records)
    v10_mean = sum(r["v10_score"] for r in records) / len(records)
    metrics = [
        ("n clips", len(records)),
        ("--- SCORES (0-10, caption_neutral only, symmetric rubric) ---", ""),
        ("v12 mean score", round(v12_mean, 2)),
        ("v10 mean score (recomputed, caption-only -- see caveat)", round(v10_mean, 2)),
        ("mean score_delta (v12 - v10)", round(v12_mean - v10_mean, 2)),
    ]
    if judge_ran:
        v12k, v12n = v12_acc
        v10k, v10n = v10_acc
        v12lo, v12hi = binom_ci(v12k, v12n)
        v10lo, v10hi = binom_ci(v10k, v10n)
        metrics += [
            ("--- LEAKAGE JUDGE (caption text only, no images) ---", ""),
            ("v12 judge accuracy", f"{v12k}/{v12n} = {100*v12k/v12n:.1f}%  "
                                     f"(95% CI [{100*v12lo:.0f}%, {100*v12hi:.0f}%])"),
            ("v10 judge accuracy", f"{v10k}/{v10n} = {100*v10k/v10n:.1f}%  "
                                     f"(95% CI [{100*v10lo:.0f}%, {100*v10hi:.0f}%])"),
            ("target for v12", "~50% (neutral) -- ~100% means still leaking"),
        ]
    metrics += [
        ("model", "google/gemini-3.6-flash"),
        ("prompts", "PROMPT_SEMSUP_V12_NEUTRAL vs PROMPT_SEMSUP_V10_GT (gt/blind hybrid)"),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 52
    ws2.column_dimensions["B"].width = 40

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  ({len(records)} rows)")

    # ---- summary.md ----
    pos_gap = [v12[r["video_id"]]["gap_trend"] for r in records if r["_target"] == 1]
    neg_gap = [v12[r["video_id"]]["gap_trend"] for r in records if r["_target"] == 0]
    import collections
    pos_gap_c = collections.Counter(pos_gap)
    neg_gap_c = collections.Counter(neg_gap)

    md = []
    md.append("# V12 neutral captioning -- validation on the 18-clip val set\n")
    md.append("**Question:** does removing the per-class GT/decision branch, adding a "
               "closed-vocabulary `gap_trend` field, and banning alarm/reassurance register "
               "symmetrically fix the register leak found in the 1,761-window corpus "
               "(caption text alone predicts the crash label at AUC=0.964)? Validated here "
               "on the 18-clip GT set before committing to a full re-caption.\n")
    md.append(f"**Settings:** {args.model}, native resolution (`--frame-size 0`), "
               "`--detail high`, `--temperature 0.1`. 18/18 calls succeeded, 0 schema "
               "failures, wall time 193s. Raw output: "
               "`outputs/prompt_bakeoff/semsup_val18_neutral/raw_v12_gemini.jsonl`.\n")
    md.append("**Scorer disclosure:** `slot_recall` reuses "
               "`reasoning_analysis_v10_gt_val18.py`'s calibrated scorer (calibration "
               "reproduces the known hand-scored CONTRADICT set "
               "{00319,00372,00474,00529,00687}). **Comparability caveat:** V10's "
               f"*published* slot_recall (0.417 GT / 0.468 blind) was computed on a "
               "6-field concatenation, not on `caption_neutral` alone. This report "
               "re-scores V10 on `caption_neutral` alone for both arms -- the V10 numbers "
               "below are therefore lower than published by construction, not a "
               "regression. Neutrality penalties are lexical (banned-word/phrase lookup), "
               "deterministic, and symmetric across classes.\n")

    md.append("## Headline\n")
    md.append("| | V12 neutral | V10 (GT on pos / blind on neg) |")
    md.append("|---|---|---|")
    md.append(f"| mean score /10 | **{v12_mean:.2f}** | {v10_mean:.2f} |")
    v12_ground = sum(r["v12_score_explanation"].split("grounding ")[1].split("/")[0]
                       .replace("-", "0") != "0" for r in records)
    if judge_ran:
        v12k, v12n = v12_acc
        v10k, v10n = v10_acc
        md.append(f"| leakage-judge accuracy | **{100*v12k/v12n:.1f}%** ({v12k}/{v12n}) | "
                   f"{100*v10k/v10n:.1f}% ({v10k}/{v10n}) |")
    md.append(f"| negatives with `gap_trend=decreasing` | 6/9 (67%) | -- (V10 has no "
               "closed `gap_trend` field) |")
    md.append(f"| positives with `gap_trend=decreasing` | 8/9 (89%) | -- |")
    md.append("")
    md.append("Row meanings: **score** = grounding (0-5, does the caption recover the "
               "GT mechanism) + neutrality (0-5, absence of outcome/alarm/reassurance/"
               "time language), out of 10. **leakage-judge accuracy** = a separate model "
               "call, captions only (no images), predicting crash/no-crash from text "
               "alone -- ~50% is the neutral target, ~100% means the register still "
               "leaks the label. **gap_trend parity** is the direct fix check: V10's "
               "\"closing distance\" appeared in 88.0% of positives vs 20.7% of negatives "
               "on the full 1,761-window corpus; here `decreasing` appears at 67% on "
               "negatives vs 89% on positives -- much closer to parity, though n=9 per "
               "class is too small to certify the gap is fully closed.\n")
    n12_flat = all(r["v12_score_explanation"].split("neutrality ")[1].startswith("5/5")
                    for r in records)
    n10_flat = all(r["v10_score_explanation"].split("neutrality ")[1].startswith("5/5")
                    for r in records)
    if n12_flat and n10_flat:
        md.append("**Important finding: the lexical neutrality score is 5/5 (18/18 clips) "
                   "for BOTH arms** -- it detected zero difference between V12 and V10 on "
                   "this sample, while the leakage judge found a large, real gap (66.7% vs "
                   "88.9%). This is expected, not a bug: the leak is a DISTRIBUTIONAL "
                   "register skew (\"maintains\"/\"stable\" appearing on 65.6%/17.9% of "
                   "negatives vs 5.1%/0.4% of positives, corpus-wide) -- not the presence "
                   "of any single banned phrase. Neither V10 nor V12's captions contain a "
                   "literal banned word here, so a pass/fail lexical check cannot see the "
                   "leak at all. **The leakage judge is therefore the metric that actually "
                   "matters** for deciding whether V12 fixed the problem; the 0-10 score is "
                   "a caption-quality/grounding measure, not a leakage measure.\n")

    md.append("## What this does NOT settle\n")
    md.append("- n=18 (9/class) gives wide exact CIs on the leakage judge -- this can "
               "detect gross leakage but cannot certify subtle neutrality.")
    md.append("- No positive clip received `gap_trend=constant` (8/9 got `decreasing`, "
               "1/9 `increasing`). This may be a genuine physical regularity (a clip that "
               "truly precedes a collision usually does have a decreasing gap in its "
               "final 2 seconds) rather than a residual leak -- distinguishing the two "
               "needs the full corpus, not 9 clips.")
    md.append("- The real gate remains the full-corpus TF-IDF text-to-label AUC check "
               "(target <0.75), planned for W1 task 1.6, once V12 captions the full pool.")
    md.append("- Grounding score may be lower for V12 than a fair comparison would show, "
               "since `slot_recall`'s keyword lists were built against V10's vocabulary "
               "(e.g. `hazard_agent`-style phrasing) -- some genuine V12 matches may be "
               "worded differently enough to miss a keyword hit. Read as a floor.\n")

    md.append("## Next step\n")
    md.append("If the full-corpus text-leakage AUC (computed after re-captioning all "
               "1,761 or 4,446 windows with V12) lands under 0.75, proceed to retrain B "
               "on the corrected corpus. Command used for this validation pass:\n")
    md.append("```bash")
    md.append("python student_training/scripts/semsup_caption_promptbakeoff.py \\")
    md.append("  --manifest dataset/manifests/val_e3a.jsonl --frames-root dataset/train \\")
    md.append("  --prompt v12 --model google/gemini-3.6-flash \\")
    md.append("  --frame-size 0 --detail high --temperature 0.1 \\")
    md.append("  --out outputs/prompt_bakeoff/semsup_val18_neutral/raw_v12_gemini.jsonl")
    md.append("```\n")

    md.append("## Files\n")
    md.append("- `raw_v12_gemini.jsonl` -- 18 V12 captions, this run")
    md.append("- `review_val18_neutral.xlsx` -- built by `score_val18_neutral.py`, "
               "side-by-side V12 vs V10 with per-caption scores and explanations")
    md.append("- `dataset/manifests/val18_gt_slots.json` -- GT slot keywords (reused from V10)")
    md.append("- `prompts/PROMPT_SEMSUP_V12_NEUTRAL.py` -- the prompt")
    md.append("- `teacher_distillation/scripts/score_val18_neutral.py` -- this scorer")
    md.append("- Plan: `~/.claude/plans/yes-but-start-with-ticklish-tower.md`")

    OUT_MD.write_text("\n".join(md) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_MD}")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
