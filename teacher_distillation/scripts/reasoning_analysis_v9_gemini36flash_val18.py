"""reasoning_analysis_v9_gemini36flash_val18.py -- scores PROMPT_SEMSUP_V9_MINIMAL
captions from google/gemini-3.6-flash against the 18-clip GT validation set.

Cross-model check: same prompt (V9), same 18 clips, different model family
(Gemini 3.6 Flash vs qwen3-vl-235b-a22b-thinking). Isolates whether V9's strong
verdict metrics on Qwen3-VL transfer to a different, cheaper/faster model.

Reads:
- dataset/manifests/val_e3a.jsonl
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds)
- outputs/prompt_bakeoff/semsup_val18/raw_v9_gemini36flash.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v9_gemini36flash_val18.xlsx

Colour rule (whole row, same convention as V4-V9 Qwen):
  GREEN  = verdict correct AND caption substantively matches GT's mechanism
  ORANGE = verdict correct but reasoning generic/partial, OR verdict wrong while
           the reasoning still substantively matches GT
  RED    = caption misses or CONTRADICTS GT's stated mechanism (regardless of
           verdict)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
GEM_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v9_gemini36flash.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v9_gemini36flash_val18.xlsx"

# (score 0-10, reasoning_match, rationale, colour) per clip, hand-scored against
# gt_reasoning_en. ego_inverted flags the specific failure this run is dominated
# by: the caption states ego DID/DID NOT react when GT says the opposite.
SCORES = {
    "00077": (8, "PARTIAL", False, "Matches GT's decisive mechanism directly: brake lights, gap closing rapidly, ego not slowing. Misses the initiating merge event. Verdict correct at high confidence (93).", "green"),
    "00147": (1, "CONTRADICT", True, "States ego is 'remaining stopped' -- GT has ego performing a left turn and moving. Also inverts causality (the other vehicle 'enters across ego path' vs GT's ego deviating into its lane). Wrong verdict.", "red"),
    "00283": (8, "MATCH", False, "Best-aligned reading of this clip across every round and model tested: correctly identifies a pickup truck (not the SUV every Qwen round reported) entering from the right, closing steadily, ego not slowing -- matches GT's mechanism and direction closely. Verdict correct.", "green"),
    "00319": (1, "CONTRADICT", False, "Asserts 'clear path ahead' -- the opposite of GT's entering-car hazard. Misses the crossing vehicle entirely, the 11th consecutive miss on this clip across every round and model tested to date.", "red"),
    "00372": (2, "CONTRADICT", True, "States 'ego braking in response' -- directly contradicts GT's explicit 'the EGO vehicle does not slow down.' Misses the pedestrian-crosswalk mechanism. Wrong verdict.", "red"),
    "00474": (0, "CONTRADICT", True, "Three separate contradictions of GT in one caption: 'gap opening' (GT: closing), 'van stays alongside' (GT: van turns into ego's lane), 'ego slowing' (GT: ego continues at the same speed). Wrong verdict.", "red"),
    "00493": (2, "CONTRADICT", True, "Correctly identifies the braking sedan but states 'ego braking in response' -- directly contradicting GT's 'the EGO vehicle does not slow down.' The exact clip where Qwen3-VL's V9 run got this same detail right. Wrong verdict.", "red"),
    "00529": (1, "CONTRADICT", True, "Claims the gap is 'remaining steady' and 'ego slowing in response' -- both contradict GT (SUV drifts into ego's lane; ego does not maintain safe distance). Wrong verdict.", "red"),
    "00687": (2, "CONTRADICT", False, "Describes the SUV as merely 'shifting rearward in frame as ego passes' -- dismissing a genuine lateral drift GT confirms is real (not apparent-only) as ego's own motion. Wrong verdict.", "red"),
    "01153": (8, "MATCH", False, "Does not reproduce the fabricated crossing-sedan hallucination seen on this identical clip in the Qwen3-VL V5/V6/V8/V9 runs -- correctly reports the sedan passing in its own lane and the SUV stopped. Matches GT's 'all vehicles remain in their respective lanes' directly. Verdict correct.", "green"),
    "01281": (6, "PARTIAL", False, "Reasonably captures the lead pickup and lane structure; 'ego not slowing' is a mild mismatch against GT's 'controlled closing distance' framing but not a direct contradiction. Verdict correct.", "orange"),
    "01504": (1, "CONTRADICT", True, "States 'ego not slowing' -- directly contradicting GT's central claim that 'the EGO vehicle noticed this and also braked in time.' Caused this false positive directly.", "red"),
    "01550": (1, "CONTRADICT", True, "States 'ego not slowing' -- contradicts GT's 'closes the gap in a controlled manner while maintaining distance.' The same detail Qwen3-VL's V9 run got correct on this identical clip. Caused this false positive directly.", "red"),
    "01552": (6, "PARTIAL", False, "Reasonably captures the minivan exiting and ego turning into the gas station area, no fabricated hazard. Verdict correct.", "green"),
    "01643": (6, "PARTIAL", False, "Correctly reports a clear road and no danger, but invents a 'road work sign' -- the same minor fabrication seen on this clip across multiple prior rounds. Verdict correct.", "orange"),
    "01737": (9, "MATCH", False, "Detailed, accurate match to GT (curved road, overpass, empty, night) -- more descriptive than Qwen3-VL's terse V9 caption on this same clip. Verdict correct.", "green"),
    "02104": (6, "PARTIAL", False, "Captures the flatbed truck and silver sedan both present in ego's lane (partially reflecting GT's merge outcome) without fabricating danger or falsely claiming 'no changes,' unlike Qwen3-VL's V9 caption on this clip. Verdict correct.", "green"),
    "02117": (8, "MATCH", False, "Correctly reports the following distance as steady (matching GT's 'constant distance,' unlike Qwen3-VL's V9 caption which wrongly said 'closing') and notes a stopped vehicle on the right. Verdict correct.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "gem_verdict", "risk_score", "reasoning_match",
           "ego_reaction_inverted", "gem_caption", "score", "score_explanation"]
WIDTHS = [11, 11, 11, 15, 55, 11, 9, 14, 16, 55, 8, 60]
WRAP_COLS = {"gt_reasoning_en", "gem_caption", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    gem = {r["video_id"]: r for r in _load_jsonl(GEM_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t}")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], gem[vid]
        score, match, ego_inv, rationale, colour = SCORES.get(vid, (None, "MISS", False, "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"],
            "gt_reasoning_en": v["gt_reasoning_en"],
            "gem_verdict": "YES" if q["verdict"] == 1 else "NO", "risk_score": q["risk_score"],
            "reasoning_match": match, "ego_reaction_inverted": "YES" if ego_inv else "",
            "gem_caption": f"{q['caption_neutral']}, {q['risk_clause']}",
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["gem_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["gem_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["gem_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["gem_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["gem_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    cc = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}
    match_counts = {m: sum(1 for r in rows if r["reasoning_match"] == m)
                     for m in ("MATCH", "PARTIAL", "MISS", "CONTRADICT")}
    ego_inv_n = sum(1 for r in rows if r["ego_reaction_inverted"] == "YES")

    from sklearn.metrics import average_precision_score, roc_auc_score
    y_true = [1 if r["gt_verdict"] == "YES" else 0 for r in rows]
    y_score = [r["risk_score"] for r in rows]

    metrics = [
        ("n", n),
        ("--- REASONING ALIGNMENT ---", ""),
        ("reasoning MATCH (substantively correct)", match_counts["MATCH"]),
        ("reasoning PARTIAL (right shape, wrong detail)", match_counts["PARTIAL"]),
        ("reasoning MISS (misses GT's mechanism)", match_counts["MISS"]),
        ("reasoning CONTRADICT (states the opposite of / fabricates beyond GT)", match_counts["CONTRADICT"]),
        ("  of which: ego-reaction direction inverted specifically", ego_inv_n),
        ("mean_hand_score (0-10)", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green / n_orange / n_red", f"{cc['green']} / {cc['orange']} / {cc['red']}"),
        ("--- verdict metrics ---", ""),
        ("verdict_accuracy (score>=50 cut)", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP / FP / TN / FN", f"{tp} / {fp} / {tn} / {fn}"),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("risk_score AUC", round(roc_auc_score(y_true, y_score), 3)),
        ("risk_score AP", round(average_precision_score(y_true, y_score), 3)),
        ("n_distinct_risk_scores", len(set(y_score))),
        ("model", "google/gemini-3.6-flash"),
        ("prompt", "PROMPT_SEMSUP_V9_MINIMAL (unchanged from the Qwen3-VL run)"),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 30

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
