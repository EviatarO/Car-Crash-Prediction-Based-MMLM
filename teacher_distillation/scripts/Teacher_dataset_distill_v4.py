"""Teacher Dataset Distill v4 -- Replay-only with enhanced PROMPT_G (PROMPT_I).

Replays the exact 18 clips from teacher_dataset_v3.jsonl (including the
replacement clip 00613) using PROMPT_I.

No video replacement is performed; the clip set is identical to v3.
"""

import argparse
import base64
import json
import math
import os
import re
import time
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from prompts.templates import PROMPT_I

try:
    import httpx
except Exception:
    httpx = None


DEFAULT_MODEL = "google/gemini-3-pro-preview"
PROMPT_ID = "I"


def _normalize_video_id(value: object) -> str:
    return f"{int(float(value)):05d}"


def _normalize_verdict(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


def _normalize_confidence(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"HIGH", "MEDIUM", "LOW"} else None


def _encode_image(path: Path, frame_size: int) -> str:
    if path.exists():
        img = Image.open(path).convert("RGB")
    else:
        img = Image.new("RGB", (frame_size, frame_size), color=(0, 0, 0))
    if frame_size and img.size != (frame_size, frame_size):
        img = img.resize((frame_size, frame_size))
    tmp = BytesIO()
    img.save(tmp, format="JPEG")
    b64 = base64.b64encode(tmp.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"


def _build_messages(prompt: str, image_b64s: Sequence[str], detail: Optional[str]) -> List[Dict]:
    content = [{"type": "text", "text": prompt}]
    for b64 in image_b64s:
        image_url = {"url": b64}
        if detail:
            image_url["detail"] = detail
        content.append({"type": "image_url", "image_url": image_url})
    return [{"role": "user", "content": content}]


def _extract_json_object(raw_response: str) -> Optional[Dict]:
    if not raw_response:
        return None
    try:
        obj = json.loads(raw_response)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass
    fenced = re.search(r"```(?:json)?\s*(\{[\s\S]*\})\s*```", raw_response, flags=re.IGNORECASE)
    if fenced:
        try:
            obj = json.loads(fenced.group(1))
            return obj if isinstance(obj, dict) else None
        except Exception:
            return None
    start = raw_response.find("{")
    end = raw_response.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    try:
        obj = json.loads(raw_response[start : end + 1])
        return obj if isinstance(obj, dict) else None
    except Exception:
        return None


def _extract_structured_fields(raw_response: str) -> Tuple[Optional[Dict], Optional[str]]:
    parsed_obj = _extract_json_object(raw_response)
    if not parsed_obj:
        return None, None
    parsed_obj["collision_verdict"] = _normalize_verdict(parsed_obj.get("collision_verdict"))
    parsed_obj["confidence"] = _normalize_confidence(parsed_obj.get("confidence"))
    return parsed_obj, parsed_obj["collision_verdict"]


def _call_model(
    client: OpenAI,
    model: str,
    messages: List[Dict],
    timeout: float,
    max_retries: int,
    retry_delay: float,
    temperature: float = 0.2,
) -> Tuple[str, Dict]:
    last_exc = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                timeout=timeout,
            )
            text = response.choices[0].message.content if response.choices else ""
            usage = response.usage.model_dump() if hasattr(response, "usage") and response.usage else {}
            return text or "", usage
        except Exception as exc:
            is_timeout = isinstance(exc, TimeoutError)
            if httpx is not None and isinstance(exc, httpx.TimeoutException):
                is_timeout = True
            if "timeout" in str(exc).lower():
                is_timeout = True
            if is_timeout:
                raise RuntimeError(f"OpenRouter timeout after {timeout}s: {exc}") from exc
            last_exc = exc
            time.sleep(retry_delay * attempt)
    raise RuntimeError(f"OpenRouter call failed after {max_retries} attempts: {last_exc}")


def _load_replay_clips(path: Path) -> List[Dict]:
    clips: List[Dict] = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            raw = line.strip()
            if not raw:
                continue
            clips.append(json.loads(raw))
    if not clips:
        raise RuntimeError(f"No clips found in replay_jsonl: {path}")
    return clips


def _materialize_clips(
    replay_clips: List[Dict],
    frames_root: Path,
    frame_size: int,
    fps: float,
) -> List[Dict]:
    out: List[Dict] = []
    for rec in replay_clips:
        video_id = _normalize_video_id(rec.get("video_id"))
        frame_indices = rec.get("frame_indices")
        if not isinstance(frame_indices, list) or not frame_indices:
            raise RuntimeError(f"Replay clip missing frame_indices for video {video_id}")

        frames_dir = frames_root / video_id
        image_b64s = []
        for idx in frame_indices:
            frame_path = frames_dir / f"frame_{int(idx):05d}.jpg"
            image_b64s.append(_encode_image(frame_path, frame_size=frame_size))

        end_frame_idx = int(rec.get("end_frame_idx", frame_indices[-1]))
        t_seconds = rec.get("t_seconds")
        if t_seconds is None:
            t_seconds = float(end_frame_idx) / max(1e-6, fps)

        out.append(
            {
                "video_id": video_id,
                "end_frame_idx": end_frame_idx,
                "frame_indices": frame_indices,
                "window_size": int(rec.get("window_size", 16)),
                "stride": int(rec.get("stride", 4)),
                "target": int(rec.get("target", 0)),
                "gt_verdict": rec.get("gt_verdict", "YES" if int(rec.get("target", 0)) == 1 else "NO"),
                "time_to_event": rec.get("time_to_event"),
                "time_of_alert": rec.get("time_of_alert"),
                "time_of_event": rec.get("time_of_event"),
                "t_seconds": t_seconds,
                "target_risk": rec.get("target_risk"),
                "requested_time_to_event": rec.get("requested_time_to_event"),
                "image_b64s": image_b64s,
            }
        )
    return out


def _dynamic_objects_to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                t = item.get("type", "?")
                lane = item.get("lane", "?")
                pos = item.get("position", "?")
                feat = item.get("feature", "")
                suffix = f", {feat}" if feat else ""
                parts.append(f"{t} ({lane}, {pos}{suffix})")
            else:
                parts.append(str(item))
        return " | ".join(parts) if parts else None
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=True)
    return str(value)


def _write_excel(records: List[Dict], output_xlsx: Path) -> None:
    rows = []
    for rec in records:
        target = rec.get("target")
        requested_tte = rec.get("requested_time_to_event")
        time_to_event = rec.get("time_to_event")
        if target == 0:
            if requested_tte is None:
                requested_tte = "TN_MIDPOINT"
            if time_to_event is None:
                time_to_event = "N/A_NO_EVENT"
        rows.append(
            {
                "video_id": rec.get("video_id"),
                "target": target,
                "gt_verdict": rec.get("gt_verdict"),
                "collision_verdict": rec.get("collision_verdict"),
                "confidence": rec.get("confidence"),
                "t_seconds": rec.get("t_seconds"),
                "time_to_event": time_to_event,
                "requested_time_to_event": requested_tte,
                "scene_context": rec.get("scene_context"),
                "dynamic_objects": _dynamic_objects_to_text(rec.get("dynamic_objects")),
                "temporal_analysis": rec.get("temporal_analysis"),
                "occlusion_check": rec.get("occlusion_check"),
                "time_to_contact": rec.get("time_to_contact"),
                "collision_target": rec.get("collision_target"),
                "verdict_reasoning": rec.get("verdict_reasoning"),
                "error": rec.get("error"),
                "review": "",
            }
        )

    df = pd.DataFrame(rows)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_xlsx, index=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active
    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    cols = len(df.columns)
    for row_idx, rec in enumerate(records, start=2):
        target = rec.get("target")
        pred = _normalize_verdict(rec.get("collision_verdict"))
        mismatch = (target == 1 and pred == "NO") or (target == 0 and pred == "YES")
        if mismatch:
            for col_idx in range(1, cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

    # Auto-fit all columns for readability, cap width for very long text cells.
    max_width_cap = 60
    for col_idx in range(1, cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, len(df) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                continue
            max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width_cap)

    wb.save(output_xlsx)


def _confusion_counts(records: List[Dict]) -> Dict[str, int]:
    tp = fp = tn = fn = 0
    for rec in records:
        target = rec.get("target")
        pred = _normalize_verdict(rec.get("collision_verdict"))
        if target == 1 and pred == "YES":
            tp += 1
        elif target == 1 and pred == "NO":
            fn += 1
        elif target == 0 and pred == "NO":
            tn += 1
        elif target == 0 and pred == "YES":
            fp += 1
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--replay_jsonl", required=True, help="Path to teacher_dataset_v3.jsonl")
    parser.add_argument("--frames_root", required=True)
    parser.add_argument("--output_jsonl", required=True)
    parser.add_argument("--output_xlsx", default="")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--fps", type=float, default=30.0)
    parser.add_argument("--frame_size", type=int, default=256)
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument("--max_retries", type=int, default=3)
    parser.add_argument("--retry_delay", type=float, default=2.0)
    parser.add_argument("--detail", default="low")
    parser.add_argument("--temperature", type=float, default=0.2)
    args = parser.parse_args()

    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY not set")

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "MMLM_Teacher_Distill_V4"),
        },
    )

    replay_path = Path(args.replay_jsonl)
    frames_root = Path(args.frames_root)
    output_jsonl = Path(args.output_jsonl)
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx = Path(args.output_xlsx) if args.output_xlsx else output_jsonl.with_suffix(".xlsx")

    replay_clips = _load_replay_clips(replay_path)
    clips = _materialize_clips(replay_clips, frames_root=frames_root, frame_size=args.frame_size, fps=args.fps)

    print(
        f"Loaded {len(clips)} clips from {replay_path} (no replacement)\n"
        f"  model={args.model} temp={args.temperature} frame_size={args.frame_size} detail={args.detail}",
        flush=True,
    )

    prompt = PROMPT_I
    run_records: List[Dict] = []
    total = len(clips)
    with open(output_jsonl, "w", encoding="utf-8") as fout:
        for i, clip in enumerate(clips, start=1):
            messages = _build_messages(prompt, clip["image_b64s"], detail=args.detail)
            raw_response = ""
            usage: Dict = {}
            err = None
            t0 = time.time()
            try:
                raw_response, usage = _call_model(
                    client=client,
                    model=args.model,
                    messages=messages,
                    timeout=args.timeout,
                    max_retries=args.max_retries,
                    retry_delay=args.retry_delay,
                    temperature=args.temperature,
                )
                latency_s = time.time() - t0
            except Exception as exc:
                err = str(exc)
                latency_s = time.time() - t0

            parsed_obj, verdict = _extract_structured_fields(raw_response)
            record = {
                "video_id": clip["video_id"],
                "end_frame_idx": clip["end_frame_idx"],
                "frame_indices": clip["frame_indices"],
                "window_size": clip["window_size"],
                "stride": clip["stride"],
                "target": clip["target"],
                "gt_verdict": clip["gt_verdict"],
                "time_to_event": clip["time_to_event"],
                "time_of_alert": clip["time_of_alert"],
                "time_of_event": clip["time_of_event"],
                "t_seconds": clip["t_seconds"],
                "target_risk": clip["target_risk"],
                "model_id": args.model,
                "prompt_id": PROMPT_ID,
                "requested_time_to_event": clip["requested_time_to_event"],
                "latency_s": latency_s,
                "input_tokens": usage.get("prompt_tokens"),
                "output_tokens": usage.get("completion_tokens"),
                "scene_context": parsed_obj.get("scene_context") if parsed_obj else None,
                "dynamic_objects": parsed_obj.get("dynamic_objects") if parsed_obj else None,
                "temporal_analysis": parsed_obj.get("temporal_analysis") if parsed_obj else None,
                "occlusion_check": parsed_obj.get("occlusion_check") if parsed_obj else None,
                "time_to_contact": parsed_obj.get("time_to_contact") if parsed_obj else None,
                "collision_verdict": verdict,
                "collision_target": parsed_obj.get("collision_target") if parsed_obj else None,
                "confidence": parsed_obj.get("confidence") if parsed_obj else None,
                "verdict_reasoning": parsed_obj.get("verdict_reasoning") if parsed_obj else None,
                "raw_response": raw_response,
                "error": err,
            }
            run_records.append(record)
            fout.write(json.dumps(record, ensure_ascii=True) + "\n")
            status = "ok" if err is None else "error"
            print(
                f"[{i}/{total}] video={clip['video_id']} target={clip['target']} "
                f"verdict={verdict} latency={latency_s:.2f}s status={status}",
                flush=True,
            )

    _write_excel(run_records, output_xlsx=output_xlsx)
    cm = _confusion_counts(run_records)
    print("\nConfusion Matrix (v4):", flush=True)
    print(f"TP={cm['TP']} FP={cm['FP']} TN={cm['TN']} FN={cm['FN']}", flush=True)
    total_correct = cm["TP"] + cm["TN"]
    print(f"Accuracy: {total_correct}/{total} = {total_correct/max(1,total):.1%}", flush=True)
    print(f"Wrote JSONL: {output_jsonl}", flush=True)
    print(f"Wrote XLSX:  {output_xlsx}", flush=True)


if __name__ == "__main__":
    main()
