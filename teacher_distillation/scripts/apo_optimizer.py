"""APO Optimizer: ProTeGi-style failure-driven prompt optimization loop.

Runs phases A-D from the plan:
  A. Setup & smoke check
  B. PROMPT_G2 baseline lock on holdout
  C. ProTeGi iterations (up to 5) with cost cap
  D. Holdout validation of winning prompt vs baseline

Outputs:
  prompts/PROMPT_G_OPTIMIZED.py
  outputs/apo/protegi_trial_log.jsonl
  outputs/apo/convergence.png

Usage:
    py teacher_distillation/scripts/apo_optimizer.py
    py teacher_distillation/scripts/apo_optimizer.py --max_iterations 3 --cost_cap 10
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import dspy
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI
from PIL import Image

# Repo path setup
REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT.parent))  # so we can import the script as module
SCRIPTS_DIR = REPO_ROOT / "teacher_distillation" / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

# Local imports
from apo_metric import score_one, mean_composite, warmup_bertscore, ScoreBreakdown
from apo_signature import CollisionAnalysis, SEED_INSTRUCTION, make_program
from apo_proposer import (
    build_failure_brief, propose_candidates,
    PROPOSER_MODEL, PROPOSER_PRICE_IN, PROPOSER_PRICE_OUT,
)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

TASK_MODEL = "openrouter/google/gemini-3.1-pro-preview"
TASK_MODEL_SLUG_NO_PREFIX = "google/gemini-3.1-pro-preview"
TASK_PRICE_IN = 2.00     # USD per 1M tokens
TASK_PRICE_OUT = 12.00

GT_XLSX = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
FRAMES_ROOT = REPO_ROOT / "dataset" / "train"
OUT_DIR = REPO_ROOT / "outputs" / "apo"
TRIAL_LOG = OUT_DIR / "protegi_trial_log.jsonl"
CONVERGENCE_PNG = OUT_DIR / "convergence.png"
OPTIMIZED_PROMPT_PY = REPO_ROOT / "prompts" / "PROMPT_G_OPTIMIZED.py"

# Stratified split (deterministic). Per the plan.
TRAIN_VIDEO_IDS = ["00319", "00077", "00687", "00283", "00147", "00529", "00493",
                   "01153", "01504", "01643", "01281", "01550", "01737"]
HOLDOUT_VIDEO_IDS = ["00474", "00372", "02104", "02117", "01552"]

# Frame loading
WINDOW_SIZE = 16
STRIDE = 4
FPS = 30


# ---------------------------------------------------------------------------
# Clip loading
# ---------------------------------------------------------------------------

def _normalize_video_id(value) -> str:
    return f"{int(float(str(value).strip())):05d}"


def load_clips(xlsx_path: Path) -> List[Dict]:
    """Read GT excel; return list of clip dicts. Requires column G = verdict_reasoning_en."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["video_id", "target", "gt_verdict", "t_seconds", "verdict_reasoning_en"]
    idx = {}
    for col in required:
        if col not in header:
            raise RuntimeError(f"Missing column '{col}' in {xlsx_path}. Headers: {header}")
        idx[col] = header.index(col)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["video_id"]] is None:
            continue
        gt_en = row[idx["verdict_reasoning_en"]]
        rows.append({
            "video_id": _normalize_video_id(row[idx["video_id"]]),
            "target": int(row[idx["target"]]) if row[idx["target"]] is not None else None,
            "gt_verdict": str(row[idx["gt_verdict"]]).strip().upper() if row[idx["gt_verdict"]] else None,
            "t_seconds": float(row[idx["t_seconds"]]) if row[idx["t_seconds"]] is not None else None,
            "gt_reasoning_en": (gt_en or "").strip(),
        })
    return rows


def split_clips(all_clips: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    by_id = {c["video_id"]: c for c in all_clips}
    train = [by_id[v] for v in TRAIN_VIDEO_IDS if v in by_id]
    holdout = [by_id[v] for v in HOLDOUT_VIDEO_IDS if v in by_id]
    if len(train) != len(TRAIN_VIDEO_IDS) or len(holdout) != len(HOLDOUT_VIDEO_IDS):
        missing_train = [v for v in TRAIN_VIDEO_IDS if v not in by_id]
        missing_holdout = [v for v in HOLDOUT_VIDEO_IDS if v not in by_id]
        raise RuntimeError(f"Missing clips. train_missing={missing_train} holdout_missing={missing_holdout}")
    return train, holdout


def _frame_indices(t_seconds: float, fps: float, window: int, stride: int) -> List[int]:
    end = round(t_seconds * fps)
    return [end - (window - 1 - i) * stride for i in range(window)]


def load_clip_frames(clip: Dict) -> List[dspy.Image]:
    folder = FRAMES_ROOT / clip["video_id"]
    indices = _frame_indices(clip["t_seconds"], FPS, WINDOW_SIZE, STRIDE)
    frames = []
    for idx in indices:
        path = folder / f"frame_{idx:05d}.jpg"
        if path.exists():
            img = Image.open(path).convert("RGB")
        else:
            img = Image.new("RGB", (256, 256), color=(0, 0, 0))
        frames.append(dspy.Image(img))
    return frames


# ---------------------------------------------------------------------------
# Evaluation (one instruction across N clips)
# ---------------------------------------------------------------------------

@dataclass
class ClipEvalResult:
    video_id: str
    gt_verdict: str
    gt_reasoning_en: str
    pred_verdict: Optional[str]
    pred_reasoning: Optional[str]
    pred_temporal: Optional[str]
    pred_spatial: Optional[str]
    pred_full: Dict
    score: ScoreBreakdown
    error: Optional[str] = None


def _failure_type(r: ClipEvalResult) -> str:
    if r.error or not r.pred_verdict:
        return "parse_failure"
    if r.score.verdict == 0.0:
        # Wrong verdict
        if r.gt_verdict == "YES" and r.pred_verdict == "NO":
            return "FN"
        if r.gt_verdict == "NO" and r.pred_verdict == "YES":
            return "FP"
        return "verdict_other"
    # Verdict correct
    if r.score.length < 1.0:
        return "length_overflow"
    if r.score.alignment < 0.85:
        return "verdict_correct_low_alignment"
    return "ok"


def _load_frames_from_dir(frames_dir: Path, frame_indices: List[int], frame_size: int = 256) -> List[dspy.Image]:
    """Load 16 frames given an explicit folder + indices list (for APOClip flow)."""
    frames = []
    for idx in frame_indices:
        path = frames_dir / f"frame_{idx:05d}.jpg"
        if path.exists():
            img = Image.open(path).convert("RGB")
        else:
            img = Image.new("RGB", (frame_size, frame_size), color=(0, 0, 0))
        frames.append(dspy.Image(img))
    return frames


def evaluate_apo_clips(
    instruction: str,
    apo_clips,                     # List[APOClip] from apo_v11_loader
    mode: str,                     # "train" / "regression" / "val"
    label: str = "",
    cost_monitor=None,             # optional CallCostMonitor for Layer 3
) -> List["ClipEvalResult"]:
    """Run an instruction across APOClip objects, return per-clip eval results.

    Uses split-specific scoring:
      - "train": score_train_only (verdict + length, no BERTScore)
      - "regression": verdict-only (binary), alignment + length set to 0/None
      - "val": score_one (full composite with BERTScore)
    """
    from apo_metric import score_one, score_train_only

    program = make_program(instruction)
    results = []
    for i, c in enumerate(apo_clips, start=1):
        frames = _load_frames_from_dir(c.frames_dir, c.frame_indices)
        pred_verdict = None
        pred_reasoning = None
        pred_temporal = None
        pred_spatial = None
        pred_full = {}
        err = None
        try:
            pred = program(frames=frames)
            pred_verdict = (pred.collision_verdict or "").strip().upper()
            pred_verdict = pred_verdict if pred_verdict in {"YES", "NO"} else None
            pred_reasoning = pred.verdict_reasoning or ""
            pred_temporal = pred.temporal_analysis or ""
            pred_spatial = pred.spatiotemporal_attention or ""
            pred_full = {
                "scene_context": pred.scene_context,
                "ego_state": pred.ego_state,
                "dynamic_objects": pred.dynamic_objects,
                "temporal_analysis": pred.temporal_analysis,
                "spatiotemporal_attention": pred.spatiotemporal_attention,
                "time_to_contact": pred.time_to_contact,
                "collision_verdict": pred.collision_verdict,
                "verdict_reasoning": pred.verdict_reasoning,
            }
        except Exception as exc:
            err = str(exc)

        # Score by mode
        if mode == "train":
            score = score_train_only(pred_verdict, pred_reasoning, c.gt_verdict)
        elif mode == "regression":
            # Binary verdict; we still build a ScoreBreakdown for uniformity
            v = 1.0 if (pred_verdict and pred_verdict == c.gt_verdict) else 0.0
            score = ScoreBreakdown(composite=v, verdict=v, alignment=0.0, length=0.0, word_count=len((pred_reasoning or "").split()))
        elif mode == "val":
            gt_en = c.gt_reasoning_en or ""
            score = score_one(pred_verdict, pred_reasoning, c.gt_verdict, gt_en)
        else:
            raise ValueError(f"Unknown mode: {mode}")

        result = ClipEvalResult(
            video_id=c.video_id,
            gt_verdict=c.gt_verdict,
            gt_reasoning_en=c.gt_reasoning_en or "",
            pred_verdict=pred_verdict,
            pred_reasoning=pred_reasoning,
            pred_temporal=pred_temporal,
            pred_spatial=pred_spatial,
            pred_full=pred_full,
            score=score,
            error=err,
        )
        results.append(result)
        flag = (f"err: {err[:60]}" if err else
                f"composite={score.composite:.3f}  v={score.verdict:.0f}  a={score.alignment:.2f}  l={score.length:.2f}")
        print(f"      [{label} {i}/{len(apo_clips)}] {c.video_id} gt={c.gt_verdict} pred={pred_verdict}  {flag}",
              flush=True)

        # Layer 3 cost anomaly check (optional)
        if cost_monitor is not None:
            # Pull the most recent task_lm history entry for cost
            try:
                hist = dspy.settings.lm.history
                if hist:
                    usage = hist[-1].get("usage", {}) or {}
                    in_tok = usage.get("prompt_tokens", 0) or 0
                    out_tok = usage.get("completion_tokens", 0) or 0
                    call_cost = in_tok * TASK_PRICE_IN / 1_000_000 + out_tok * TASK_PRICE_OUT / 1_000_000
                    if not cost_monitor.record_and_check(call_cost):
                        print("  ABORT: cost anomaly threshold exceeded", flush=True)
                        sys.exit(3)
            except Exception:
                pass

    return results


def evaluate_instruction(
    instruction: str,
    clips: List[Dict],
    label: str = "",
) -> List[ClipEvalResult]:
    """LEGACY (18-clip mode): Run an instruction across a list of clips, return per-clip eval results."""
    program = make_program(instruction)
    results = []
    for i, clip in enumerate(clips, start=1):
        frames = load_clip_frames(clip)
        try:
            pred = program(frames=frames)
            pred_verdict = (pred.collision_verdict or "").strip().upper()
            pred_verdict = pred_verdict if pred_verdict in {"YES", "NO"} else None
            pred_reasoning = pred.verdict_reasoning or ""
            pred_temporal = pred.temporal_analysis or ""
            pred_spatial = pred.spatiotemporal_attention or ""
            pred_full = {
                "scene_context": pred.scene_context,
                "ego_state": pred.ego_state,
                "dynamic_objects": pred.dynamic_objects,
                "temporal_analysis": pred.temporal_analysis,
                "spatiotemporal_attention": pred.spatiotemporal_attention,
                "time_to_contact": pred.time_to_contact,
                "collision_verdict": pred.collision_verdict,
                "verdict_reasoning": pred.verdict_reasoning,
            }
            score = score_one(pred_verdict, pred_reasoning, clip["gt_verdict"], clip["gt_reasoning_en"])
            err = None
        except Exception as exc:
            pred_verdict = None
            pred_reasoning = None
            pred_temporal = None
            pred_spatial = None
            pred_full = {}
            score = ScoreBreakdown(0.0, 0.0, 0.0, 0.0, 0)
            err = str(exc)

        result = ClipEvalResult(
            video_id=clip["video_id"],
            gt_verdict=clip["gt_verdict"],
            gt_reasoning_en=clip["gt_reasoning_en"],
            pred_verdict=pred_verdict,
            pred_reasoning=pred_reasoning,
            pred_temporal=pred_temporal,
            pred_spatial=pred_spatial,
            pred_full=pred_full,
            score=score,
            error=err,
        )
        results.append(result)
        flag = f"err: {err[:60]}" if err else f"composite={score.composite:.3f}  v={score.verdict:.0f}  a={score.alignment:.2f}  l={score.length:.2f}"
        print(f"      [{label} {i}/{len(clips)}] {clip['video_id']} gt={clip['gt_verdict']} pred={pred_verdict}  {flag}",
              flush=True)

    return results


def _to_failure_dict(r: ClipEvalResult) -> Dict:
    """Convert ClipEvalResult to the dict format expected by build_failure_brief."""
    return {
        "video_id": r.video_id,
        "gt_verdict": r.gt_verdict,
        "gt_reasoning_en": r.gt_reasoning_en,
        "pred_verdict": r.pred_verdict,
        "pred_reasoning": r.pred_reasoning,
        "pred_temporal": r.pred_temporal,
        "pred_spatial": r.pred_spatial,
        "composite": r.score.composite,
        "verdict_score": r.score.verdict,
        "alignment_score": r.score.alignment,
        "length_score": r.score.length,
        "word_count": r.score.word_count,
        "failure_type": _failure_type(r),
    }


# ---------------------------------------------------------------------------
# Cost tracking (DSPy LM history)
# ---------------------------------------------------------------------------

def get_dspy_cost(lm: dspy.LM) -> float:
    """Sum cost across all DSPy LM calls so far (USD)."""
    total = 0.0
    for entry in lm.history:
        usage = entry.get("usage", {}) or {}
        in_tok = usage.get("prompt_tokens", 0) or 0
        out_tok = usage.get("completion_tokens", 0) or 0
        total += in_tok * TASK_PRICE_IN / 1_000_000 + out_tok * TASK_PRICE_OUT / 1_000_000
    return total


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def append_trial_log(record: Dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def load_resume_instruction(path: Path) -> str:
    """Extract a winning instruction from either a .py file or a trial log .jsonl.

    For .py files: imports and reads PROMPT_G_OPTIMIZED (or PROMPT_G).
    For .jsonl files: scans all entries, finds the highest mean_composite from
                      candidates or beam entries.
    """
    suffix = path.suffix.lower()
    if suffix == ".py":
        import importlib.util
        spec = importlib.util.spec_from_file_location("resume_module", path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        if hasattr(mod, "PROMPT_G_OPTIMIZED"):
            return mod.PROMPT_G_OPTIMIZED.strip()
        if hasattr(mod, "PROMPT_G"):
            return mod.PROMPT_G.strip()
        raise RuntimeError(f"{path} does not define PROMPT_G_OPTIMIZED or PROMPT_G")

    if suffix == ".jsonl":
        best_score = -1.0
        best_instruction = None
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                # Look in candidates list
                for c in rec.get("candidates", []) or []:
                    score = c.get("mean_composite")
                    if score is not None and score > best_score:
                        best_score = score
                        best_instruction = c.get("instruction")
                # Look in beam_after
                for c in rec.get("beam_after", []) or []:
                    score = c.get("mean_composite")
                    if score is not None and score > best_score:
                        best_score = score
                        best_instruction = c.get("instruction")
                # Look in seed eval
                if rec.get("phase") == "C_seed":
                    score = rec.get("train_mean_composite")
                    if score is not None and score > best_score:
                        best_score = score
                        best_instruction = rec.get("instruction")
        if best_instruction is None:
            raise RuntimeError(f"No instructions found in trial log {path}")
        print(f"        Resume: best instruction from {path.name} (mean_composite={best_score:.4f})", flush=True)
        return best_instruction.strip()

    raise RuntimeError(f"Unsupported resume file extension: {path}")


def _clip_result_from_failure_dict(d: Dict) -> "ClipEvalResult":
    """Reconstruct a ClipEvalResult from a persisted failure-dict (for resume)."""
    score = ScoreBreakdown(
        composite=d.get("composite", 0.0),
        verdict=d.get("verdict_score", 0.0),
        alignment=d.get("alignment_score", 0.0),
        length=d.get("length_score", 0.0),
        word_count=d.get("word_count", 0),
    )
    return ClipEvalResult(
        video_id=d["video_id"],
        gt_verdict=d.get("gt_verdict", ""),
        gt_reasoning_en=d.get("gt_reasoning_en", ""),
        pred_verdict=d.get("pred_verdict"),
        pred_reasoning=d.get("pred_reasoning", ""),
        pred_temporal=d.get("pred_temporal", ""),
        pred_spatial=d.get("pred_spatial", ""),
        pred_full={},
        score=score,
        error=None,
    )


def _load_v11scale_seed_state(log_path: Path) -> Optional[Dict]:
    """Read Phase B and C_seed results from an existing v11scale trial log.

    Returns a dict with keys:
        baseline_mean, seed_instruction, seed_train_mean,
        seed_regression_acc, seed_results (List[ClipEvalResult])
    or None if the required phases are not both present.
    """
    if not log_path.exists():
        return None

    baseline_mean: Optional[float] = None
    seed_instruction: Optional[str] = None
    seed_train_mean: Optional[float] = None
    seed_regression_acc: Optional[float] = None
    seed_results: List["ClipEvalResult"] = []

    with open(log_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except Exception:
                continue
            phase = rec.get("phase", "")
            if phase == "B_baseline":
                baseline_mean = rec.get("val_mean_composite")
            elif phase == "C_seed":
                seed_instruction = rec.get("instruction")
                seed_train_mean = rec.get("train_mean_composite")
                seed_regression_acc = rec.get("regression_acc")
                seed_results = [
                    _clip_result_from_failure_dict(d)
                    for d in rec.get("per_clip_train", [])
                ]

    if None in (baseline_mean, seed_instruction, seed_train_mean, seed_regression_acc):
        return None

    return {
        "baseline_mean": baseline_mean,
        "seed_instruction": seed_instruction,
        "seed_train_mean": seed_train_mean,
        "seed_regression_acc": seed_regression_acc,
        "seed_results": seed_results,
    }


def save_optimized_prompt(instruction: str, path: Path, history: List[Dict]) -> None:
    """Write PROMPT_G_OPTIMIZED.py with the winning instruction + provenance comment."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write('"""PROMPT_G_OPTIMIZED -- discovered via ProTeGi APO with beam search.\n\n')
        f.write("Generated by teacher_distillation/scripts/apo_optimizer.py.\n")
        f.write(f"Iterations completed: {len(history)}\n")
        f.write(f"Final best train composite: {history[-1]['mean_composite']:.4f}\n")
        f.write('"""\n\n')
        # Use a Python triple-quoted string for the instruction, escape any triple quotes
        safe = instruction.replace('"""', '\\"\\"\\"')
        f.write(f'PROMPT_G_OPTIMIZED = """{safe}"""\n')


def save_convergence_plot(history: List[Dict], path: Path) -> None:
    """Plot best-in-beam composite per iteration."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print(f"  matplotlib not available; skipping plot", flush=True)
        return
    iterations = list(range(len(history)))
    composites = [h["mean_composite"] for h in history]
    plt.figure(figsize=(8, 5))
    plt.plot(iterations, composites, marker="o", linewidth=2, label="Best in beam")
    plt.xlabel("Iteration (0 = seed)")
    plt.ylabel("Mean train composite score (best in beam)")
    plt.title("APO ProTeGi Convergence (beam search)")
    plt.grid(True, alpha=0.3)
    plt.ylim(0, 1)
    plt.legend()
    path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(path, dpi=120, bbox_inches="tight")
    plt.close()


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="APO ProTeGi-style optimizer with beam search")
    parser.add_argument("--mode", choices=["18clip", "v11scale"], default="18clip",
                        help="18clip: original 13/5 GT split. v11scale: 31 train / 67 regression / 18 val.")
    parser.add_argument("--max_iterations", type=int, default=5)
    parser.add_argument("--beam_size", type=int, default=3, help="Top-K instructions kept across iterations")
    parser.add_argument("--candidates_per_iter", type=int, default=5, help="New candidates proposed per iteration")
    parser.add_argument("--worst_k", type=int, default=5, help="# worst (clip, beam_member) pairs shown to proposer")
    parser.add_argument("--cost_cap", type=float, default=20.0)
    parser.add_argument("--early_stop_patience", type=int, default=2)
    parser.add_argument("--regression_tolerance", type=float, default=0.05,
                        help="(v11scale) Acceptance gate: regression_acc must be >= 1.0 - tolerance")
    parser.add_argument("--projected_cost", type=float, default=21.0,
                        help="Estimated cost in USD for the run (used by Layer 1 credit check)")
    parser.add_argument(
        "--phase_d_only", action="store_true",
        help="Skip Phase C entirely. Read winning instruction from --resume_instruction_file (or trial log) "
             "and only run Phase B (baseline) + Phase D (holdout validation)."
    )
    parser.add_argument(
        "--resume_instruction_file", default="",
        help="Path to a Python file with PROMPT_G_OPTIMIZED variable, OR a trial log JSONL. "
             "Used by --phase_d_only and --seed_instruction_from_file."
    )
    parser.add_argument(
        "--seed_instruction_from_file", action="store_true",
        help="Use the instruction in --resume_instruction_file as the seed (instead of SEED_INSTRUCTION). "
             "Used to bootstrap iterative APO from a previous winner."
    )
    parser.add_argument(
        "--keep_trial_log", action="store_true",
        help="Append to existing trial log instead of overwriting it (useful for chained iterations)."
    )
    parser.add_argument(
        "--resume_from_seed", action="store_true",
        help="(v11scale) Skip Phase B and Phase C seed eval by reloading them from the existing "
             "trial log. Requires --keep_trial_log. Saves ~$7 when re-running after a credit stop."
    )
    args = parser.parse_args()

    # Branch to v11scale main if requested
    if args.mode == "v11scale":
        return main_v11scale(args)

    # ---- Phase A: Setup ----
    print("=" * 60)
    print("PHASE A: Setup")
    print("=" * 60)
    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY not set")

    print("[A.1] Configuring DSPy with Gemini via OpenRouter...", flush=True)
    task_lm = dspy.LM(
        model=TASK_MODEL,
        api_key=api_key,
        api_base="https://openrouter.ai/api/v1",
        temperature=0.1,
        max_tokens=4000,   # PROMPT_G2's verbose 8-step protocol can produce up to ~3000 output tokens
    )
    dspy.configure(lm=task_lm)

    print("[A.2] Loading BERTScore (downloads ~1.4GB if not cached)...", flush=True)
    warmup_bertscore()
    print("        BERTScore ready.", flush=True)

    print("[A.3] OpenRouter client for proposer (Claude Opus)...", flush=True)
    proposer_client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": "MMLM_APO_ProTeGi",
        },
    )

    print("[A.4] Loading clips from GT excel...", flush=True)
    all_clips = load_clips(GT_XLSX)
    train, holdout = split_clips(all_clips)
    print(f"        Loaded {len(all_clips)} clips. Train: {len(train)}, Holdout: {len(holdout)}", flush=True)

    # Resolve resume instruction up front (used by both --phase_d_only and --seed_instruction_from_file)
    # IMPORTANT: must happen BEFORE the trial log deletion below, in case the resume file IS the trial log
    resume_instruction = None
    if args.phase_d_only or args.seed_instruction_from_file:
        if not args.resume_instruction_file:
            raise RuntimeError("--phase_d_only and --seed_instruction_from_file require --resume_instruction_file")
        resume_path = Path(args.resume_instruction_file)
        if not resume_path.exists():
            raise RuntimeError(f"Resume file not found: {resume_path}")
        resume_instruction = load_resume_instruction(resume_path)
        print(f"[A.5] Loaded resume instruction ({len(resume_instruction.split())} words):", flush=True)
        print(f"        \"{resume_instruction[:200]}...\"" if len(resume_instruction) > 200 else f"        \"{resume_instruction}\"", flush=True)

    # Auto-keep trial log in resume modes (so we don't lose prior optimization history)
    keep_trial_log = args.keep_trial_log or args.phase_d_only or args.seed_instruction_from_file

    # Reset trial log file (unless caller wants to preserve it, or we're resuming)
    if TRIAL_LOG.exists() and not keep_trial_log:
        TRIAL_LOG.unlink()

    # ---- Phase B: PROMPT_G2 baseline on holdout ----
    print()
    print("=" * 60)
    print("PHASE B: PROMPT_G2 baseline lock on holdout")
    print("=" * 60)
    # Load PROMPT_G2 from the Python file
    import importlib.util
    spec = importlib.util.spec_from_file_location("prompt_g2_module", REPO_ROOT / "prompts" / "PROMPT_G2.py")
    pg2 = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(pg2)
    PROMPT_G2 = pg2.PROMPT_G   # the file uses var name PROMPT_G

    print("[B.1] Evaluating PROMPT_G2 on 5 holdout clips...", flush=True)
    baseline_results = evaluate_instruction(PROMPT_G2, holdout, label="baseline")
    baseline_mean = mean_composite([r.score for r in baseline_results])
    print(f"        PROMPT_G2 holdout mean composite: {baseline_mean:.4f}", flush=True)

    append_trial_log({
        "phase": "B_baseline",
        "instruction_label": "PROMPT_G2",
        "holdout_mean_composite": baseline_mean,
        "per_clip": [_to_failure_dict(r) for r in baseline_results],
    }, TRIAL_LOG)

    # ---- Phase C: ProTeGi iterations (skipped if --phase_d_only) ----
    if args.phase_d_only:
        print()
        print("=" * 60)
        print("PHASE C: SKIPPED (--phase_d_only mode)")
        print("=" * 60)
        # Build a synthetic beam with just the resume instruction; we won't iterate
        beam: List[Dict] = [{
            "instruction": resume_instruction,
            "mean_composite": 0.0,   # unknown; will be measured in Phase D
            "results": [],
            "diagnosis": "[resumed]",
        }]
        history = [{
            "iteration": 0,
            "best_instruction": resume_instruction,
            "mean_composite": 0.0,
            "beam_size": 1,
            "diagnosis": "[resumed]",
        }]
        # Skip the iteration loop entirely
        # (the for-loop below will not execute since we set max_iterations=0 effectively via early return-ish)
        # We just jump to Phase D below.
        skip_phase_c_iterations = True
    else:
        skip_phase_c_iterations = False
        print()
        print("=" * 60)
        print("PHASE C: ProTeGi optimization")
        print("=" * 60)

        # Choose seed: either default SEED_INSTRUCTION or the resumed winner
        if args.seed_instruction_from_file and resume_instruction:
            seed_to_use = resume_instruction
            print(f"[C.0] Using RESUMED instruction as seed (from {args.resume_instruction_file}):", flush=True)
        else:
            seed_to_use = SEED_INSTRUCTION
            print(f"[C.0] Using DEFAULT seed instruction:", flush=True)
        print(f"      Seed: {seed_to_use[:200]}{'...' if len(seed_to_use) > 200 else ''}", flush=True)

        # Initial seed evaluation on TRAIN — seed becomes the only initial beam member
        print(f"      Evaluating seed on {len(train)} train clips...", flush=True)
        seed_results = evaluate_instruction(seed_to_use, train, label="C0")
        seed_mean = mean_composite([r.score for r in seed_results])
        print(f"        SEED train mean composite: {seed_mean:.4f}", flush=True)

        # Beam: list of dicts {instruction, mean_composite, results, diagnosis}
        # Initially holds just the seed; expands to up to beam_size after first iteration.
        beam: List[Dict] = [{
            "instruction": seed_to_use,
            "mean_composite": seed_mean,
            "results": seed_results,
            "diagnosis": "[seed]",
        }]

        history = [{
            "iteration": 0,
            "best_instruction": seed_to_use,
            "mean_composite": seed_mean,            # = best in beam
            "beam_size": 1,
            "diagnosis": "[seed]",
        }]

        append_trial_log({
            "phase": "C_seed",
            "iteration": 0,
            "instruction": seed_to_use,
            "train_mean_composite": seed_mean,
            "per_clip": [_to_failure_dict(r) for r in seed_results],
        }, TRIAL_LOG)

    non_improvement_count = 0
    proposer_total_cost = 0.0
    # If we skipped Phase C (resume mode), seed_mean isn't defined; use placeholder
    best_seen_score = locals().get("seed_mean", 0.0)

    # Phase C iteration loop is skipped in --phase_d_only mode
    iterations_to_run = 0 if skip_phase_c_iterations else args.max_iterations
    for iter_idx in range(1, iterations_to_run + 1):
        # Cost cap check
        cur_dspy_cost = get_dspy_cost(task_lm)
        cur_total = cur_dspy_cost + proposer_total_cost
        print()
        print(f"--- Iteration {iter_idx}/{args.max_iterations} ---")
        print(f"  Cumulative cost: ${cur_total:.4f}  (Gemini=${cur_dspy_cost:.4f}, Claude=${proposer_total_cost:.4f})",
              flush=True)
        if cur_total > args.cost_cap:
            print(f"  COST CAP HIT (${cur_total:.2f} > ${args.cost_cap:.2f}). Stopping.", flush=True)
            break

        # Print current beam state
        print(f"  Current beam (size={len(beam)}):", flush=True)
        for bi, m in enumerate(beam, start=1):
            print(f"    [#{bi}] composite={m['mean_composite']:.4f}  instr={m['instruction'][:90]}...", flush=True)

        # 1. Pool failures across the beam: collect (clip, member_idx) pairs, sort by composite, take worst K
        all_failures: List[Tuple[int, ClipEvalResult]] = []
        for bi, m in enumerate(beam, start=1):
            for r in m["results"]:
                all_failures.append((bi, r))
        # Sort by composite ascending (lowest first); take worst K
        all_failures.sort(key=lambda pair: pair[1].score.composite)
        worst_pairs = all_failures[: args.worst_k]
        worst_dicts = []
        for bi, r in worst_pairs:
            d = _to_failure_dict(r)
            d["source_label"] = f"from beam #{bi}"
            worst_dicts.append(d)

        # 2. Build failure brief (multi-source aware)
        brief = build_failure_brief(
            beam=[{"instruction": m["instruction"], "mean_composite": m["mean_composite"]} for m in beam],
            worst_clips=worst_dicts,
            n_candidates=args.candidates_per_iter,
            score_history=[
                {"iteration": h["iteration"], "instruction": h["best_instruction"],
                 "score": h["mean_composite"]} for h in history
            ],
        )

        # 3. Call Claude proposer
        print(f"  [{iter_idx}.1] Calling Claude proposer with {len(worst_dicts)} failures (pooled from beam)...", flush=True)
        candidates_raw, prop_usage = propose_candidates(
            proposer_client, brief, n=args.candidates_per_iter,
        )
        proposer_total_cost += prop_usage["cost_usd"]
        print(f"        Proposer returned {len(candidates_raw)} candidates "
              f"(cost=${prop_usage['cost_usd']:.4f})", flush=True)
        for ci, c in enumerate(candidates_raw, start=1):
            print(f"          C{ci} diagnosis: {c['diagnosis'][:140]}", flush=True)
            print(f"          C{ci} instruction: {c['instruction'][:140]}", flush=True)

        # 4. Evaluate each candidate on full trainset
        candidate_results = []
        for ci, c in enumerate(candidates_raw, start=1):
            # Skip if exact instruction already in beam (avoid wasted eval)
            if any(c["instruction"] == m["instruction"] for m in beam):
                print(f"  [{iter_idx}.2] Candidate {ci} duplicates an existing beam member — skipping eval", flush=True)
                continue
            print(f"  [{iter_idx}.2] Evaluating candidate {ci}/{len(candidates_raw)} on {len(train)} clips...", flush=True)
            res = evaluate_instruction(c["instruction"], train, label=f"C{iter_idx}.{ci}")
            cm = mean_composite([r.score for r in res])
            print(f"        candidate {ci} mean composite: {cm:.4f}", flush=True)
            candidate_results.append({
                "diagnosis": c["diagnosis"],
                "instruction": c["instruction"],
                "results": res,
                "mean_composite": cm,
            })

        # 5. Update beam: top-K from (current beam ∪ new candidates)
        merged = beam + candidate_results
        merged.sort(key=lambda x: x["mean_composite"], reverse=True)
        new_beam = merged[: args.beam_size]
        new_best_score = new_beam[0]["mean_composite"]

        # Detect improvement on the BEST score in beam
        improved = new_best_score > best_seen_score + 1e-6

        # Append trial log entry
        append_trial_log({
            "phase": "C_iter",
            "iteration": iter_idx,
            "beam_before": [
                {"instruction": m["instruction"], "mean_composite": m["mean_composite"]}
                for m in beam
            ],
            "candidates": [
                {"diagnosis": cr["diagnosis"], "instruction": cr["instruction"],
                 "mean_composite": cr["mean_composite"],
                 "per_clip": [_to_failure_dict(r) for r in cr["results"]]}
                for cr in candidate_results
            ],
            "beam_after": [
                {"instruction": m["instruction"], "mean_composite": m["mean_composite"],
                 "diagnosis": m.get("diagnosis", "")}
                for m in new_beam
            ],
            "best_before": best_seen_score,
            "best_after": new_best_score,
            "improved": improved,
            "proposer_cost_usd": prop_usage["cost_usd"],
        }, TRIAL_LOG)

        history.append({
            "iteration": iter_idx,
            "best_instruction": new_beam[0]["instruction"],
            "mean_composite": new_best_score,
            "beam_size": len(new_beam),
            "diagnosis": new_beam[0].get("diagnosis", ""),
        })

        if improved:
            print(f"  [OK] Best score improved: {best_seen_score:.4f} -> {new_best_score:.4f}", flush=True)
            best_seen_score = new_best_score
            non_improvement_count = 0
        else:
            non_improvement_count += 1
            print(f"  [--] No new best ({non_improvement_count}/{args.early_stop_patience} consecutive)  "
                  f"(best stays {best_seen_score:.4f})", flush=True)

        beam = new_beam

        if non_improvement_count >= args.early_stop_patience:
            print(f"  EARLY STOP triggered after {non_improvement_count} non-improving iterations.", flush=True)
            break

    # Final winner = top-1 of final beam
    best = beam[0]

    # ---- Phase D: Holdout validation ----
    print()
    print("=" * 60)
    print("PHASE D: Holdout validation")
    print("=" * 60)
    print(f"[D.1] Evaluating WINNING instruction on {len(holdout)} holdout clips...", flush=True)
    print(f"      Winning: {best['instruction']}", flush=True)
    winning_holdout = evaluate_instruction(best["instruction"], holdout, label="D")
    winning_mean = mean_composite([r.score for r in winning_holdout])
    delta = winning_mean - baseline_mean
    accept = delta >= 0.05

    # Per-component summaries
    def _summary(results: List[ClipEvalResult]) -> Dict:
        n = len(results)
        verdict_acc = sum(r.score.verdict for r in results) / n
        align_mean = sum(r.score.alignment for r in results) / n
        length_compliance = sum(1 for r in results if r.score.length == 1.0)
        return {
            "n": n,
            "verdict_acc": round(verdict_acc, 4),
            "alignment_mean": round(align_mean, 4),
            "length_compliance": f"{length_compliance}/{n}",
            "composite_mean": round(mean_composite([r.score for r in results]), 4),
        }

    baseline_summary = _summary(baseline_results)
    winning_summary = _summary(winning_holdout)

    print()
    print("=" * 60)
    print("FINAL COMPARISON (5-clip holdout)")
    print("=" * 60)
    print(f"  Metric                  | PROMPT_G2  | PROMPT_G_OPTIMIZED | delta")
    print(f"  Verdict accuracy        | {baseline_summary['verdict_acc']:.3f}     | {winning_summary['verdict_acc']:.3f}             | {winning_summary['verdict_acc']-baseline_summary['verdict_acc']:+.3f}")
    print(f"  BERTScore F1 (mean)     | {baseline_summary['alignment_mean']:.3f}     | {winning_summary['alignment_mean']:.3f}             | {winning_summary['alignment_mean']-baseline_summary['alignment_mean']:+.3f}")
    print(f"  Length compliance       | {baseline_summary['length_compliance']}        | {winning_summary['length_compliance']}                 | -")
    print(f"  COMPOSITE (mean)        | {baseline_summary['composite_mean']:.3f}     | {winning_summary['composite_mean']:.3f}             | {delta:+.3f}")
    print()
    if accept:
        print(f"  [ACCEPT] delta={delta:+.3f} >= 0.05 threshold")
    else:
        print(f"  [REJECT] delta={delta:+.3f} < 0.05 threshold (no improvement). Keep PROMPT_G2.")

    # Save artifacts
    append_trial_log({
        "phase": "D_holdout",
        "winning_instruction": best["instruction"],
        "baseline_summary": baseline_summary,
        "winning_summary": winning_summary,
        "delta_composite": delta,
        "accepted": accept,
        "per_clip": [_to_failure_dict(r) for r in winning_holdout],
    }, TRIAL_LOG)

    if accept:
        save_optimized_prompt(best["instruction"], OPTIMIZED_PROMPT_PY, history)
        print(f"  Saved: {OPTIMIZED_PROMPT_PY}")
    save_convergence_plot(history, CONVERGENCE_PNG)
    print(f"  Saved: {CONVERGENCE_PNG}")
    print(f"  Saved: {TRIAL_LOG}")

    final_total = get_dspy_cost(task_lm) + proposer_total_cost
    print(f"\n  Total cost: ${final_total:.4f}  (Gemini=${get_dspy_cost(task_lm):.4f}, Claude=${proposer_total_cost:.4f})")


# ===========================================================================
# v11scale main flow (Option 2: failure-focused train + regression constraint)
# ===========================================================================

def main_v11scale(args):
    """Main flow for --mode v11scale.

    - Train: 31 v11 Pass-1 failures
    - Regression: 67 v11 Pass-1 successes (hard constraint per iteration)
    - Val: 18 GT clips (final acceptance gate)

    Includes safety Layers 1-4 (credit checks + cost anomaly + heartbeat).
    """
    from apo_v11_loader import build_v11scale_splits, verify_frame_paths
    from apo_safety import (
        assert_sufficient_credit, soft_credit_check,
        CallCostMonitor, heartbeat_log,
    )
    from apo_metric import score_one, score_train_only, mean_composite, warmup_bertscore

    # Output paths (separate from 18clip mode to avoid pollution)
    OUT_DIR_V11 = REPO_ROOT / "outputs" / "apo"
    TRIAL_LOG_V11 = OUT_DIR_V11 / "protegi_v11scale_log.jsonl"
    CONVERGENCE_PNG_V11 = OUT_DIR_V11 / "convergence_v11scale.png"
    OPTIMIZED_PROMPT_PY = REPO_ROOT / "prompts" / "PROMPT_G_OPTIMIZED.py"

    print("=" * 60)
    print("PHASE A: Setup (v11scale mode)")
    print("=" * 60)

    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY not set")

    # ----- LAYER 1: pre-run credit check -----
    print("[A.0] Pre-run credit check (Layer 1)...", flush=True)
    pre_status = assert_sufficient_credit(
        api_key=api_key,
        projected_cost=args.projected_cost,
        safety_margin=1.5,
        label="pre-run",
    )

    print("[A.1] Configuring DSPy with Gemini via OpenRouter...", flush=True)
    task_lm = dspy.LM(
        model=TASK_MODEL,
        api_key=api_key,
        api_base="https://openrouter.ai/api/v1",
        temperature=0.1,
        max_tokens=6000,   # bumped from 4000 — DSPy adapter scaffolding adds tokens
    )
    dspy.configure(lm=task_lm)

    print("[A.2] Loading BERTScore (downloads ~1.4GB if not cached)...", flush=True)
    warmup_bertscore()
    print("        BERTScore ready.", flush=True)

    print("[A.3] OpenRouter client for Claude proposer...", flush=True)
    proposer_client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": "MMLM_APO_v11scale",
        },
    )

    print("[A.4] Building v11scale splits...", flush=True)
    splits = build_v11scale_splits()
    train_clips = splits["train"]
    regression_clips = splits["regression"]
    val_clips = splits["val"]

    print("[A.5] Verifying frame paths...", flush=True)
    errors = verify_frame_paths(splits)
    if errors:
        print(f"\nABORT: {len(errors)} missing frame paths:")
        for e in errors[:10]:
            print(e)
        sys.exit(2)

    # Initialize cost monitor (Layer 3)
    cost_monitor = CallCostMonitor(expected_call_cost=0.06, anomaly_factor=4.0, max_consecutive=3)

    # Reset trial log unless told to keep
    if TRIAL_LOG_V11.exists() and not args.keep_trial_log:
        TRIAL_LOG_V11.unlink()

    # ===== RESUME PATH: reload Phase B + C_seed from existing log =====
    # Use --resume_from_seed --keep_trial_log when picking up after a credit stop.
    # Saves ~$7 by skipping re-evaluation of 18 val clips + 31+67 seed clips.
    _resumed = None
    if getattr(args, "resume_from_seed", False):
        if not args.keep_trial_log:
            print("[RESUME] WARNING: --resume_from_seed has no effect without --keep_trial_log; ignoring.", flush=True)
        else:
            _resumed = _load_v11scale_seed_state(TRIAL_LOG_V11)
            if _resumed is None:
                print("[RESUME] Could not load Phase B / C_seed from existing log — running from scratch.", flush=True)
            else:
                print(f"[RESUME] Loaded Phase B (val_mean={_resumed['baseline_mean']:.4f}) "
                      f"and C_seed (train_mean={_resumed['seed_train_mean']:.4f}, "
                      f"regression_acc={_resumed['seed_regression_acc']:.4f}) from log.", flush=True)

    if _resumed is not None:
        # ===== PHASE B: SKIPPED (loaded from log) =====
        print()
        print("=" * 60)
        print("PHASE B: SKIPPED (loaded from existing trial log)")
        print("=" * 60)
        baseline_mean = _resumed["baseline_mean"]

        # ===== PHASE C seed: SKIPPED (loaded from log) =====
        print()
        print("=" * 60)
        print("PHASE C: ProTeGi optimization (v11scale) — seed loaded from log")
        print("=" * 60)
        seed_instruction = _resumed["seed_instruction"]
        seed_train_mean = _resumed["seed_train_mean"]
        seed_regression_acc = _resumed["seed_regression_acc"]
        seed_results = _resumed["seed_results"]
    else:
        # ===== PHASE B: PROMPT_G2 baseline on val =====
        print()
        print("=" * 60)
        print("PHASE B: PROMPT_G2 baseline on 18 val clips")
        print("=" * 60)
        import importlib.util
        spec = importlib.util.spec_from_file_location("prompt_g2_module", REPO_ROOT / "prompts" / "PROMPT_G2.py")
        pg2 = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(pg2)
        PROMPT_G2 = pg2.PROMPT_G

        heartbeat_log("Phase B: evaluating PROMPT_G2 on 18 val clips")
        baseline_results = evaluate_apo_clips(PROMPT_G2, val_clips, mode="val", label="B", cost_monitor=cost_monitor)
        baseline_mean = mean_composite([r.score for r in baseline_results])
        print(f"        PROMPT_G2 val mean composite: {baseline_mean:.4f}", flush=True)

        append_trial_log({
            "phase": "B_baseline",
            "instruction_label": "PROMPT_G2",
            "val_mean_composite": baseline_mean,
            "per_clip": [_to_failure_dict(r) for r in baseline_results],
        }, TRIAL_LOG_V11)

        # ===== PHASE C: ProTeGi optimization =====
        print()
        print("=" * 60)
        print("PHASE C: ProTeGi optimization (v11scale)")
        print("=" * 60)

        # Seed: PROMPT_G (the prompt that generated v11)
        from prompts.templates import PROMPT_G
        seed_instruction = PROMPT_G

        # Optionally override seed if --seed_instruction_from_file
        if args.seed_instruction_from_file:
            if not args.resume_instruction_file:
                raise RuntimeError("--seed_instruction_from_file requires --resume_instruction_file")
            from pathlib import Path as _P
            seed_instruction = load_resume_instruction(_P(args.resume_instruction_file))
            print(f"[C.0] Using RESUMED instruction as seed (from {args.resume_instruction_file})", flush=True)
        else:
            print(f"[C.0] Using PROMPT_G as seed instruction", flush=True)
        print(f"      Seed (first 200 chars): {seed_instruction[:200]}{'...' if len(seed_instruction) > 200 else ''}", flush=True)

        heartbeat_log(f"Phase C.0: evaluating SEED on {len(train_clips)} train clips")
        seed_results = evaluate_apo_clips(seed_instruction, train_clips, mode="train", label="C0-T", cost_monitor=cost_monitor)
        seed_train_mean = mean_composite([r.score for r in seed_results])
        print(f"        SEED train mean composite: {seed_train_mean:.4f}", flush=True)

        # Initial regression check on seed (sanity — should be ~baseline)
        heartbeat_log(f"Phase C.0: regression check of SEED on {len(regression_clips)} clips")
        seed_regression_results = evaluate_apo_clips(seed_instruction, regression_clips, mode="regression", label="C0-R", cost_monitor=cost_monitor)
        seed_regression_acc = sum(r.score.verdict for r in seed_regression_results) / len(seed_regression_results)
        print(f"        SEED regression verdict_acc: {seed_regression_acc:.4f} (baseline = 1.0 by construction since these were correct in v11)",
              flush=True)

        append_trial_log({
            "phase": "C_seed",
            "iteration": 0,
            "instruction": seed_instruction,
            "train_mean_composite": seed_train_mean,
            "regression_acc": seed_regression_acc,
            "per_clip_train": [_to_failure_dict(r) for r in seed_results],
        }, TRIAL_LOG_V11)

    # Beam initialization (shared by both resume and fresh-start paths)
    beam: List[Dict] = [{
        "instruction": seed_instruction,
        "train_mean": seed_train_mean,
        "regression_acc": seed_regression_acc,
        "results": seed_results,
        "diagnosis": "[seed=PROMPT_G]",
    }]
    history = [{
        "iteration": 0,
        "best_instruction": seed_instruction,
        "train_mean": seed_train_mean,
        "regression_acc": seed_regression_acc,
        "diagnosis": "[seed]",
    }]

    non_improvement_count = 0
    proposer_total_cost = 0.0
    best_seen_score = seed_train_mean
    # IMPORTANT: regression threshold is RELATIVE to the seed's actual measured regression_acc
    # (not absolute 0.95) because DSPy's prompt scaffolding produces different outputs than
    # v11's direct API calls — the seed's reproduction of v11 successes is imperfect.
    # Threshold says: "no candidate may worsen regression by more than `tolerance` vs seed"
    regression_threshold = max(0.0, seed_regression_acc - args.regression_tolerance)
    print(f"  Regression threshold: {regression_threshold:.4f}  "
          f"(seed_regression_acc={seed_regression_acc:.4f} - tolerance={args.regression_tolerance})",
          flush=True)

    for iter_idx in range(1, args.max_iterations + 1):
        # Layer 2: per-iteration credit check
        ok = soft_credit_check(api_key=api_key, minimum_remaining=3.0, label=f"iter{iter_idx}")
        if not ok:
            print(f"  [credit-check] LOW BALANCE — stopping iterations and finalizing.", flush=True)
            break

        cur_dspy_cost = get_dspy_cost(task_lm)
        cur_total = cur_dspy_cost + proposer_total_cost
        print()
        print(f"--- Iteration {iter_idx}/{args.max_iterations} ---")
        print(f"  Cumulative cost: ${cur_total:.4f}  (Gemini=${cur_dspy_cost:.4f}, Claude=${proposer_total_cost:.4f})",
              flush=True)
        if cur_total > args.cost_cap:
            print(f"  COST CAP HIT (${cur_total:.2f} > ${args.cost_cap:.2f}). Stopping.", flush=True)
            break

        print(f"  Current beam (size={len(beam)}):", flush=True)
        for bi, m in enumerate(beam, start=1):
            print(f"    [#{bi}] train={m['train_mean']:.4f}  regr_acc={m['regression_acc']:.3f}  "
                  f"instr={m['instruction'][:80]}...", flush=True)

        # 1. Pool failures across the beam
        all_failures = []
        for bi, m in enumerate(beam, start=1):
            for r in m["results"]:
                all_failures.append((bi, r))
        all_failures.sort(key=lambda pair: pair[1].score.composite)
        worst_pairs = all_failures[: args.worst_k]
        worst_dicts = []
        for bi, r in worst_pairs:
            d = _to_failure_dict(r)
            d["source_label"] = f"from beam #{bi}"
            worst_dicts.append(d)

        # 2. Build failure brief and call Claude proposer
        brief = build_failure_brief(
            beam=[{"instruction": m["instruction"], "mean_composite": m["train_mean"]} for m in beam],
            worst_clips=worst_dicts,
            n_candidates=args.candidates_per_iter,
            score_history=[
                {"iteration": h["iteration"], "instruction": h["best_instruction"], "score": h["train_mean"]}
                for h in history
            ],
        )
        print(f"  [{iter_idx}.1] Calling Claude proposer with {len(worst_dicts)} failures from beam...", flush=True)
        candidates_raw, prop_usage = propose_candidates(proposer_client, brief, n=args.candidates_per_iter)
        proposer_total_cost += prop_usage["cost_usd"]
        print(f"        Proposer returned {len(candidates_raw)} candidates (cost=${prop_usage['cost_usd']:.4f})",
              flush=True)
        for ci, c in enumerate(candidates_raw, start=1):
            print(f"          C{ci} diagnosis: {c['diagnosis'][:140]}", flush=True)
            print(f"          C{ci} instruction: {c['instruction'][:140]}", flush=True)

        # 3. Evaluate each candidate on TRAIN
        candidate_results = []
        for ci, c in enumerate(candidates_raw, start=1):
            if any(c["instruction"] == m["instruction"] for m in beam):
                print(f"  [{iter_idx}.2] Candidate {ci} duplicates a beam member -- skipping", flush=True)
                continue
            heartbeat_log(f"Iter {iter_idx} cand {ci}/{len(candidates_raw)}: train eval on {len(train_clips)} clips")
            train_res = evaluate_apo_clips(c["instruction"], train_clips, mode="train", label=f"C{iter_idx}.{ci}-T", cost_monitor=cost_monitor)
            train_mean = mean_composite([r.score for r in train_res])
            print(f"        candidate {ci} train_mean: {train_mean:.4f}", flush=True)
            candidate_results.append({
                "diagnosis": c["diagnosis"],
                "instruction": c["instruction"],
                "train_results": train_res,
                "train_mean": train_mean,
                "regression_results": None,    # will fill if accepted
                "regression_acc": None,
            })

        # 4. Regression check on each candidate (per-iter constraint)
        # We check candidates that beat current beam minimum on train
        beam_min_train = min(m["train_mean"] for m in beam)
        promising = [c for c in candidate_results if c["train_mean"] >= beam_min_train - 1e-6]

        for cr in promising:
            heartbeat_log(f"Iter {iter_idx}: regression check on {len(regression_clips)} successes")
            reg_res = evaluate_apo_clips(cr["instruction"], regression_clips, mode="regression", label=f"C{iter_idx}-R", cost_monitor=cost_monitor)
            reg_acc = sum(r.score.verdict for r in reg_res) / len(reg_res)
            cr["regression_results"] = reg_res
            cr["regression_acc"] = reg_acc
            passed = reg_acc >= regression_threshold
            print(f"        regression_acc={reg_acc:.4f} threshold={regression_threshold:.4f} -> "
                  f"{'PASS' if passed else 'FAIL'}", flush=True)

        # 5. Update beam: only candidates that PASS the regression check are eligible
        eligible = [cr for cr in promising if cr.get("regression_acc") is not None and cr["regression_acc"] >= regression_threshold]
        # Convert eligible candidates to beam member format
        new_members = [{
            "instruction": cr["instruction"],
            "train_mean": cr["train_mean"],
            "regression_acc": cr["regression_acc"],
            "results": cr["train_results"],
            "diagnosis": cr["diagnosis"],
        } for cr in eligible]

        merged = beam + new_members
        merged.sort(key=lambda x: x["train_mean"], reverse=True)
        new_beam = merged[: args.beam_size]
        new_best_score = new_beam[0]["train_mean"]

        improved = new_best_score > best_seen_score + 1e-6

        # Append trial log entry
        append_trial_log({
            "phase": "C_iter",
            "iteration": iter_idx,
            "beam_before": [
                {"instruction": m["instruction"], "train_mean": m["train_mean"], "regression_acc": m["regression_acc"]}
                for m in beam
            ],
            "candidates": [
                {"diagnosis": cr["diagnosis"], "instruction": cr["instruction"],
                 "train_mean": cr["train_mean"], "regression_acc": cr["regression_acc"],
                 "passed_regression": (cr["regression_acc"] is not None and cr["regression_acc"] >= regression_threshold),
                 "per_clip_train": [_to_failure_dict(r) for r in cr["train_results"]]}
                for cr in candidate_results
            ],
            "beam_after": [
                {"instruction": m["instruction"], "train_mean": m["train_mean"],
                 "regression_acc": m["regression_acc"], "diagnosis": m.get("diagnosis", "")}
                for m in new_beam
            ],
            "improved": improved,
            "proposer_cost_usd": prop_usage["cost_usd"],
        }, TRIAL_LOG_V11)

        history.append({
            "iteration": iter_idx,
            "best_instruction": new_beam[0]["instruction"],
            "train_mean": new_best_score,
            "regression_acc": new_beam[0]["regression_acc"],
            "diagnosis": new_beam[0].get("diagnosis", ""),
        })

        if improved:
            print(f"  [OK] Best train improved: {best_seen_score:.4f} -> {new_best_score:.4f}", flush=True)
            best_seen_score = new_best_score
            non_improvement_count = 0
        else:
            non_improvement_count += 1
            print(f"  [--] No new best ({non_improvement_count}/{args.early_stop_patience} consecutive)  "
                  f"(best stays {best_seen_score:.4f})", flush=True)

        beam = new_beam

        if non_improvement_count >= args.early_stop_patience:
            print(f"  EARLY STOP after {non_improvement_count} non-improving iterations.", flush=True)
            break

    best = beam[0]

    # ===== PHASE D: Final holdout (val) validation =====
    print()
    print("=" * 60)
    print("PHASE D: Holdout val (18 GT clips)")
    print("=" * 60)
    print(f"[D.1] Evaluating WINNING instruction on {len(val_clips)} val clips...", flush=True)
    print(f"      Winning (first 200 chars): {best['instruction'][:200]}{'...' if len(best['instruction']) > 200 else ''}", flush=True)

    heartbeat_log("Phase D: val evaluation")
    winning_val = evaluate_apo_clips(best["instruction"], val_clips, mode="val", label="D", cost_monitor=cost_monitor)
    winning_mean = mean_composite([r.score for r in winning_val])
    delta = winning_mean - baseline_mean
    accept = delta >= 0.05

    def _summary(results):
        n = len(results)
        verdict_acc = sum(r.score.verdict for r in results) / n
        align_mean = sum(r.score.alignment for r in results) / n
        length_compliance = sum(1 for r in results if r.score.length == 1.0)
        return {
            "n": n, "verdict_acc": round(verdict_acc, 4), "alignment_mean": round(align_mean, 4),
            "length_compliance": f"{length_compliance}/{n}",
            "composite_mean": round(mean_composite([r.score for r in results]), 4),
        }

    bs = _summary(baseline_results)
    ws = _summary(winning_val)

    print()
    print("=" * 60)
    print("FINAL COMPARISON (18-clip val, v11scale APO)")
    print("=" * 60)
    print(f"  Metric                  | PROMPT_G2  | PROMPT_G_OPTIMIZED | delta")
    print(f"  Verdict accuracy        | {bs['verdict_acc']:.3f}     | {ws['verdict_acc']:.3f}             | {ws['verdict_acc']-bs['verdict_acc']:+.3f}")
    print(f"  BERTScore F1 (mean)     | {bs['alignment_mean']:.3f}     | {ws['alignment_mean']:.3f}             | {ws['alignment_mean']-bs['alignment_mean']:+.3f}")
    print(f"  Length compliance       | {bs['length_compliance']}        | {ws['length_compliance']}                 | -")
    print(f"  COMPOSITE (mean)        | {bs['composite_mean']:.3f}     | {ws['composite_mean']:.3f}             | {delta:+.3f}")
    print()
    if accept:
        print(f"  [ACCEPT] delta={delta:+.3f} >= 0.05")
    else:
        print(f"  [REJECT] delta={delta:+.3f} < 0.05 (no improvement). Keep PROMPT_G2.")

    # Save artifacts
    append_trial_log({
        "phase": "D_holdout",
        "winning_instruction": best["instruction"],
        "baseline_summary": bs,
        "winning_summary": ws,
        "delta_composite": delta,
        "accepted": accept,
        "per_clip_val": [_to_failure_dict(r) for r in winning_val],
    }, TRIAL_LOG_V11)

    if accept:
        save_optimized_prompt(best["instruction"], OPTIMIZED_PROMPT_PY, [
            {"iteration": h["iteration"], "mean_composite": h["train_mean"]} for h in history
        ])
        print(f"  Saved: {OPTIMIZED_PROMPT_PY}")

    save_convergence_plot(
        [{"iteration": h["iteration"], "mean_composite": h["train_mean"]} for h in history],
        CONVERGENCE_PNG_V11
    )
    print(f"  Saved: {CONVERGENCE_PNG_V11}")
    print(f"  Saved: {TRIAL_LOG_V11}")

    final_total = get_dspy_cost(task_lm) + proposer_total_cost
    print(f"\n  Total cost: ${final_total:.4f}  (Gemini=${get_dspy_cost(task_lm):.4f}, Claude=${proposer_total_cost:.4f})")


if __name__ == "__main__":
    main()
