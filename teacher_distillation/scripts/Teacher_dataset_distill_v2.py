import argparse
import base64
import json
import math
import os
import random
import re
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image

from data.dataset import compute_target_risk
from prompts.templates import PROMPT_G

try:
    import httpx
except Exception:  # pragma: no cover
    httpx = None


DEFAULT_MODEL = "google/gemini-3-pro-preview"
TP_BUCKETS: Tuple[Tuple[str, float], ...] = (("0.5s", 0.5), ("1s", 1.0), ("1.5s", 1.5))
TRACKING_COLUMNS = ["0.5s", "1s", "1.5s"]


def _normalize_video_id(value: object) -> str:
    return f"{int(float(value)):05d}"


def _is_truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y"}


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def _encode_image(path: Path, frame_size: int) -> str:
    if path.exists():
        img = Image.open(path).convert("RGB")
    else:
        img = Image.new("RGB", (frame_size, frame_size), color=(0, 0, 0))
    if frame_size and img.size != (frame_size, frame_size):
        img = img.resize((frame_size, frame_size))
    tmp = BytesIO()
    img.save(tmp, format="JPEG")
    b64 = base64.b64encode(tmp.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"


def _build_messages(prompt: str, image_b64s: Sequence[str], detail: Optional[str]) -> List[Dict]:
    content = [{"type": "text", "text": prompt}]
    for b64 in image_b64s:
        image_url = {"url": b64}
        if detail:
            image_url["detail"] = detail
        content.append({"type": "image_url", "image_url": image_url})
    return [{"role": "user", "content": content}]


def _extract_json_object(raw_response: str) -> Optional[Dict]:
    if not raw_response:
        return None
    try:
        obj = json.loads(raw_response)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass

    fenced = re.search(r"```(?:json)?\s*(\{[\s\S]*\})\s*```", raw_response, flags=re.IGNORECASE)
    if fenced:
        try:
            obj = json.loads(fenced.group(1))
            return obj if isinstance(obj, dict) else None
        except Exception:
            return None

    start = raw_response.find("{")
    end = raw_response.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    try:
        obj = json.loads(raw_response[start : end + 1])
        return obj if isinstance(obj, dict) else None
    except Exception:
        return None


def _normalize_verdict(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


def _normalize_confidence(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"HIGH", "MEDIUM", "LOW"} else None


def _extract_structured_fields(raw_response: str) -> Tuple[Optional[Dict], Optional[str]]:
    parsed_obj = _extract_json_object(raw_response)
    if not parsed_obj:
        return None, None
    verdict = _normalize_verdict(parsed_obj.get("collision_verdict"))
    parsed_obj["collision_verdict"] = verdict
    parsed_obj["confidence"] = _normalize_confidence(parsed_obj.get("confidence"))
    return parsed_obj, verdict


def build_teacher_prompt() -> str:
    return PROMPT_G


def _call_model(
    client: OpenAI,
    model: str,
    messages: List[Dict],
    timeout: float,
    max_retries: int,
    retry_delay: float,
) -> Tuple[str, Dict]:
    last_exc = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0.2,
                timeout=timeout,
            )
            text = response.choices[0].message.content if response.choices else ""
            usage = response.usage.model_dump() if hasattr(response, "usage") and response.usage else {}
            return text or "", usage
        except Exception as exc:  # pragma: no cover - network path
            is_timeout = isinstance(exc, TimeoutError)
            if httpx is not None and isinstance(exc, httpx.TimeoutException):
                is_timeout = True
            if "timeout" in str(exc).lower():
                is_timeout = True
            if is_timeout:
                raise RuntimeError(f"OpenRouter timeout after {timeout}s: {exc}") from exc
            last_exc = exc
            time.sleep(retry_delay * attempt)
    raise RuntimeError(f"OpenRouter call failed after {max_retries} attempts: {last_exc}")


def _build_frame_indices(end_frame_idx: int, window_size: int, stride: int, max_frame: int) -> List[int]:
    start_idx = end_frame_idx - (window_size - 1) * stride
    out: List[int] = []
    for idx in range(start_idx, end_frame_idx + 1, stride):
        out.append(max(0, min(max_frame, idx)))
    return out


def _parse_time(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _bucket_counts(n_tp: int) -> Dict[str, int]:
    base = n_tp // 3
    rem = n_tp % 3
    return {
        "0.5s": base + (1 if rem > 0 else 0),
        "1s": base + (1 if rem > 1 else 0),
        "1.5s": base,
    }


def _select_tp_rows(
    df: pd.DataFrame,
    n_tp: int,
    rng: random.Random,
) -> List[Tuple[int, str, float]]:
    counts = _bucket_counts(n_tp)
    selected: List[Tuple[int, str, float]] = []
    selected_videos = set()

    for col_name, bucket_s in TP_BUCKETS:
        needed = counts[col_name]
        if needed <= 0:
            continue
        candidates = []
        for idx, row in df.iterrows():
            if not _is_truthy(row.get("Has_Event")):
                continue
            if not _is_empty(row.get(col_name)):
                continue
            if _parse_time(row.get("time_of_event")) is None:
                continue
            video_id = _normalize_video_id(row.get("video_index"))
            if video_id in selected_videos:
                continue
            candidates.append((idx, video_id))

        if len(candidates) < needed:
            raise RuntimeError(
                f"Not enough TP candidates for bucket {col_name}. Needed={needed}, available={len(candidates)}."
            )

        picked = rng.sample(candidates, k=needed)
        for idx, video_id in picked:
            selected.append((idx, col_name, bucket_s))
            selected_videos.add(video_id)
    return selected


def _select_tn_rows(
    df: pd.DataFrame,
    n_tn: int,
    rng: random.Random,
) -> List[Tuple[int, str, Optional[float]]]:
    candidates = []
    for idx, row in df.iterrows():
        if _is_truthy(row.get("Has_Event")):
            continue
        if not _is_empty(row.get("0.5s")):
            continue
        candidates.append(idx)
    if len(candidates) < n_tn:
        raise RuntimeError(f"Not enough TN candidates. Needed={n_tn}, available={len(candidates)}.")
    picked = rng.sample(candidates, k=n_tn)
    return [(idx, "0.5s", None) for idx in picked]


def _build_clip(
    row: pd.Series,
    frames_root: Path,
    fps: float,
    window_size: int,
    stride: int,
    bucket_seconds: Optional[float],
    frame_size: int,
) -> Dict:
    video_id = _normalize_video_id(row.get("video_index"))
    frames_dir = frames_root / video_id
    frame_paths = sorted(frames_dir.glob("frame_*.jpg"))
    if not frame_paths:
        raise RuntimeError(f"No frames found for video {video_id} in {frames_dir}")
    max_frame = len(frame_paths) - 1

    has_event = _is_truthy(row.get("Has_Event"))
    t_event = _parse_time(row.get("time_of_event"))
    t_alert = _parse_time(row.get("time_of_alert"))
    if has_event and t_event is None:
        raise RuntimeError(f"Missing time_of_event for TP video {video_id}")

    if has_event:
        end_frame_idx = int(round((t_event - float(bucket_seconds)) * fps))
    else:
        end_frame_idx = max_frame // 2
    end_frame_idx = max(0, min(max_frame, end_frame_idx))
    frame_indices = _build_frame_indices(end_frame_idx, window_size, stride, max_frame)
    t_seconds = end_frame_idx / max(1e-6, fps)
    target = 1 if has_event else 0
    target_risk = compute_target_risk(
        t_seconds=t_seconds,
        time_of_alert=t_alert,
        time_of_event=t_event,
        target=target,
    )

    image_b64s = []
    for idx in frame_indices:
        frame_path = frames_dir / f"frame_{int(idx):05d}.jpg"
        image_b64s.append(_encode_image(frame_path, frame_size=frame_size))

    return {
        "video_id": video_id,
        "frame_indices": frame_indices,
        "end_frame_idx": end_frame_idx,
        "t_seconds": t_seconds,
        "time_of_event": t_event,
        "time_of_alert": t_alert,
        "time_to_event": (t_event - t_seconds) if (t_event is not None) else None,
        "target": target,
        "target_risk": target_risk,
        "gt_verdict": "YES" if has_event else "NO",
        "bucket_seconds": bucket_seconds,
        "image_b64s": image_b64s,
    }


def _write_excel(records: List[Dict], output_xlsx: Path) -> None:
    rows = []
    for rec in records:
        target = rec.get("target")
        requested_tte = rec.get("requested_time_to_event")
        time_to_event = rec.get("time_to_event")
        if target == 0:
            # TN clips have no event anchor; make this explicit in the Excel output.
            if requested_tte is None:
                requested_tte = "TN_MIDPOINT"
            if time_to_event is None:
                time_to_event = "N/A_NO_EVENT"

        rows.append(
            {
                "video_id": rec.get("video_id"),
                "target": target,
                "gt_verdict": rec.get("gt_verdict"),
                "requested_time_to_event": requested_tte,
                "t_seconds": rec.get("t_seconds"),
                "time_to_event": time_to_event,
                "collision_verdict": rec.get("collision_verdict"),
                "confidence": rec.get("confidence"),
                "scene_context": rec.get("scene_context"),
                "temporal_analysis": rec.get("temporal_analysis"),
                "time_to_contact": rec.get("time_to_contact"),
                "verdict_reasoning": rec.get("verdict_reasoning"),
                "error": rec.get("error"),
            }
        )

    df = pd.DataFrame(rows)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_xlsx, index=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active
    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    cols = len(df.columns)
    for row_idx, rec in enumerate(records, start=2):
        target = rec.get("target")
        pred = _normalize_verdict(rec.get("collision_verdict"))
        mismatch = (target == 1 and pred == "NO") or (target == 0 and pred == "YES")
        if mismatch:
            for col_idx in range(1, cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill
    wb.save(output_xlsx)


def _confusion_counts(records: List[Dict]) -> Dict[str, int]:
    tp = fp = tn = fn = 0
    for rec in records:
        target = rec.get("target")
        pred = _normalize_verdict(rec.get("collision_verdict"))
        if target == 1 and pred == "YES":
            tp += 1
        elif target == 1 and pred == "NO":
            fn += 1
        elif target == 0 and pred == "NO":
            tn += 1
        elif target == 0 and pred == "YES":
            fp += 1
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tracking_excel", required=True)
    parser.add_argument("--frames_root", required=True)
    parser.add_argument("--output_jsonl", required=True)
    parser.add_argument("--output_xlsx", default="")
    parser.add_argument("--n_tp", type=int, default=9)
    parser.add_argument("--n_tn", type=int, default=9)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--fps", type=float, default=30.0)
    parser.add_argument("--window_size", type=int, default=16)
    parser.add_argument("--stride", type=int, default=4)
    parser.add_argument("--frame_size", type=int, default=256)
    parser.add_argument("--timeout", type=float, default=30.0)
    parser.add_argument("--max_retries", type=int, default=3)
    parser.add_argument("--retry_delay", type=float, default=2.0)
    parser.add_argument("--detail", default="low")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--overwrite_output", action="store_true")
    args = parser.parse_args()

    if args.n_tp <= 0 or args.n_tn <= 0:
        raise RuntimeError("--n_tp and --n_tn must be > 0")

    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY not set")

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "MMLM_Teacher_Distill_V2"),
        },
    )

    tracking_path = Path(args.tracking_excel)
    if not tracking_path.exists():
        raise RuntimeError(f"Tracking Excel not found: {tracking_path}")
    frames_root = Path(args.frames_root)
    output_path = Path(args.output_jsonl)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx = Path(args.output_xlsx) if args.output_xlsx else output_path.with_suffix(".xlsx")

    df = pd.read_excel(tracking_path)
    needed = {"video_index", "Has_Event", "time_of_event", "time_of_alert"}
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise RuntimeError(f"Missing required columns in tracking excel: {missing}")

    # Initialize tracking columns if the source sheet does not already contain them.
    for col in TRACKING_COLUMNS:
        if col not in df.columns:
            df[col] = None

    for col in TRACKING_COLUMNS:
        df[col] = df[col].astype(object)

    rng = random.Random(args.seed)

    tp_selected = _select_tp_rows(df, n_tp=args.n_tp, rng=rng)
    tn_selected = _select_tn_rows(df, n_tn=args.n_tn, rng=rng)
    selected_specs = tp_selected + tn_selected
    rng.shuffle(selected_specs)

    run_tag = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    selected_clips: List[Dict] = []
    for row_idx, col_name, bucket_seconds in selected_specs:
        row = df.iloc[row_idx]
        clip = _build_clip(
            row=row,
            frames_root=frames_root,
            fps=args.fps,
            window_size=args.window_size,
            stride=args.stride,
            bucket_seconds=bucket_seconds,
            frame_size=args.frame_size,
        )
        selected_clips.append(clip)
        df.at[row_idx, col_name] = run_tag

    # Save tracking immediately after deterministic selection so future runs avoid duplicates.
    df.to_excel(tracking_path, index=False)

    mode = "w" if args.overwrite_output else "a"
    prompt = build_teacher_prompt()
    total = len(selected_clips)
    run_records: List[Dict] = []
    with open(output_path, mode, encoding="utf-8") as fout:
        for i, clip in enumerate(selected_clips, start=1):
            messages = _build_messages(prompt, clip["image_b64s"], detail=args.detail)
            raw_response = ""
            usage: Dict = {}
            err = None
            latency_s = None
            t0 = time.time()
            try:
                raw_response, usage = _call_model(
                    client=client,
                    model=args.model,
                    messages=messages,
                    timeout=args.timeout,
                    max_retries=args.max_retries,
                    retry_delay=args.retry_delay,
                )
                latency_s = time.time() - t0
            except Exception as exc:  # pragma: no cover - network path
                err = str(exc)
                latency_s = time.time() - t0

            parsed_obj, verdict = _extract_structured_fields(raw_response)

            record = {
                "video_id": clip["video_id"],
                "end_frame_idx": clip["end_frame_idx"],
                "frame_indices": clip["frame_indices"],
                "window_size": args.window_size,
                "stride": args.stride,
                "target": clip["target"],
                "gt_verdict": clip["gt_verdict"],
                "time_to_event": clip["time_to_event"],
                "time_of_alert": clip["time_of_alert"],
                "time_of_event": clip["time_of_event"],
                "t_seconds": clip["t_seconds"],
                "target_risk": clip["target_risk"],
                "model_id": args.model,
                "prompt_id": "G",
                "requested_time_to_event": clip["bucket_seconds"],
                "latency_s": latency_s,
                "input_tokens": usage.get("prompt_tokens"),
                "output_tokens": usage.get("completion_tokens"),
                "scene_context": parsed_obj.get("scene_context") if parsed_obj else None,
                "dynamic_objects": parsed_obj.get("dynamic_objects") if parsed_obj else None,
                "temporal_analysis": parsed_obj.get("temporal_analysis") if parsed_obj else None,
                "occlusion_check": parsed_obj.get("occlusion_check") if parsed_obj else None,
                "time_to_contact": parsed_obj.get("time_to_contact") if parsed_obj else None,
                "collision_verdict": verdict,
                "confidence": parsed_obj.get("confidence") if parsed_obj else None,
                "verdict_reasoning": parsed_obj.get("verdict_reasoning") if parsed_obj else None,
                "raw_response": raw_response,
                "error": err,
            }
            run_records.append(record)
            fout.write(json.dumps(record, ensure_ascii=True) + "\n")
            status = "ok" if err is None else "error"
            print(
                f"[{i}/{total}] video={clip['video_id']} bucket={clip['bucket_seconds']} "
                f"target={clip['target']} latency={latency_s:.2f}s status={status}",
                flush=True,
            )

    _write_excel(run_records, output_xlsx=output_xlsx)
    cm = _confusion_counts(run_records)
    print("\nConfusion Matrix (blind prediction):", flush=True)
    print(f"TP={cm['TP']} FP={cm['FP']} TN={cm['TN']} FN={cm['FN']}", flush=True)
    print(f"Updated tracking Excel: {tracking_path}", flush=True)
    print(f"Wrote JSONL records: {output_path}", flush=True)
    print(f"Wrote XLSX report: {output_xlsx}", flush=True)


if __name__ == "__main__":
    main()
