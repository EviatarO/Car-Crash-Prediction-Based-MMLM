"""reasoning_analysis_v7_val18.py -- scores PROMPT_SEMSUP_V7_EGOFRAME captions from
qwen/qwen3-vl-235b-a22b-thinking against the 18-clip GT validation set, in the same
workbook format as the V4/V5/V6 analyses.

Adds an ego_path_correct column: V7's central claim is that estimating ego's own
rotation from static scene structure fixes the dominant error, so whether ego_path
actually matches GT's stated ego manoeuvre is the load-bearing measurement of this
round -- more informative than the verdict, which can be right for wrong reasons.

Reads:
- dataset/manifests/val_e3a.jsonl
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds)
- outputs/prompt_bakeoff/semsup_val18/raw_v7_qwen3vl235b.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v7_val18.xlsx

Colour rule (whole row, same convention as the V4/V5/V6 workbooks):
  GREEN  = verdict correct AND caption matches GT reasoning without fabrication
  ORANGE = verdict correct but caption middling/generic/minor fabrication, OR
           verdict wrong while the caption still describes the scene well
  RED    = verdict wrong AND caption misses or contradicts GT's mechanism
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
V7_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v7_qwen3vl235b.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v7_val18.xlsx"

# GT's stated ego manoeuvre, read from gt_reasoning_en, and whether V7's ego_path
# matches it. This is the round's key measurement (see module docstring).
GT_EGO = {
    "00077": ("straight (highway)", False),
    "00147": ("LEFT turn at intersection", True),
    "00283": ("straight (middle lane, high speed)", False),
    "00319": ("straight through intersection", False),
    "00372": ("straight (right lane)", False),
    "00474": ("straight (crossing intersection)", True),
    "00493": ("LEFT turn then merge", True),
    "00529": ("straight (middle lane)", True),
    "00687": ("LEFT turn at intersection", False),
    "01153": ("RIGHT turn (smooth)", False),
    "01281": ("straight (middle lane)", False),
    "01504": ("straight (rightmost lane)", True),
    "01550": ("straight (right lane)", False),
    "01552": ("perpendicular crossing into gas station", True),
    "01643": ("straight (right lane)", True),
    "01737": ("turn on interchange (right curve)", False),
    "02104": ("straight (middle lane)", True),
    "02117": ("straight (left lane)", True),
}

# (score 0-10, rationale, colour) per clip, hand-scored against gt_reasoning_en.
SCORES = {
    "00077": (6, "Captures GT's actual mechanism well -- black car ahead closing distance with no braking visible -- but invents an ego rightward drift GT does not describe, and the score (49) fell one point below the cut. Verdict wrong.", "orange"),
    "00147": (5, "ego_path correctly identifies the left turn (V6 called this lane-holding), but attribution is still inverted: GT says EGO deviates into the right-hand car's lane; caption says the other car moved leftward into ego's path. Verdict correct.", "orange"),
    "00283": (6, "Detects a genuine cut-in and gets the verdict right, but mirrors it (GT: stationary right-lane pickup turns LEFT into ego lane; caption: SUV moving RIGHTWARD) and misreads ego as turning left when GT has ego straight at speed.", "orange"),
    "00319": (1, "Misses GT's crossing vehicle from the right entirely, describes a vehicle moving AWAY from the intersection, and misreads ego as turning left when GT has it going straight. Wrong verdict.", "red"),
    "00372": (2, "Verdict correct but for an unrelated reason: GT's mechanism is a lead sedan stopping for crosswalk pedestrians; the caption describes a red truck moving leftward across ego's path and misreads ego as turning left. Not a usable caption.", "orange"),
    "00474": (2, "Labels every agent APPARENT ONLY and reports stable traffic, missing GT's van-turns-sharply-into-ego-lane event -- the same miss on this clip in every round to date. Does correctly note the taxi's brake lights. Wrong verdict.", "red"),
    "00493": (8, "Best caption-vs-GT match in the set: ego_path correctly reports the left turn, and the caption names the silver sedan braking with the gap closing -- GT's exact mechanism. Only the score (43) fell below the 50 cut, so the verdict reads NO.", "orange"),
    "00529": (1, "Labels the silver SUV APPARENT ONLY -- the apparent-vs-true test suppressing a genuine lateral event -- so GT's SUV-drifts-into-ego-lane mechanism is missed. Wrong verdict.", "red"),
    "00687": (1, "Regression versus V6, which read this clip correctly. V7 reports ego straight (GT: ego turns left) and the gray SUV as parked (GT: it drifts into the ego lane). Both halves of the mechanism lost. Wrong verdict.", "red"),
    "01153": (6, "Major improvement: the fabricated 'sedan crossing into ego's path' that broke this clip in both V5 and V6 is gone. The caption correctly notes the sedan turns into a DIFFERENT lane, giving the right verdict, though ego's turn direction is inverted (GT right, reported left).", "green"),
    "01281": (5, "Correct verdict with no fabrication -- V6 invented a drifting SUV here. Explicitly reports no truly moving agents. But ego is misread as turning right when GT has it straight in the middle lane.", "orange"),
    "01504": (6, "Correct verdict and no fabrication -- V6 hallucinated a side street and a turning SUV on this clip. Reports all agents APPARENT ONLY, which is right. Misses that ego itself brakes, which is GT's reason for the safe outcome.", "green"),
    "01550": (5, "Correct verdict and correctly reads the closing gap with brake lights ahead, but misreads ego as turning left when GT has it straight in the right lane, and rates the risk moderate where GT calls the closing controlled.", "orange"),
    "01552": (6, "Good match on the real content (minivan crossing, truck ahead, correct verdict) but reproduces the 'school bus' fabrication seen on this clip in V4, V5 and V6 -- now 4-for-4, a stable artifact rather than noise.", "orange"),
    "01643": (5, "Correct verdict, correctly reports no moving agents, but introduces parked cars and a 'Road Work Ahead' sign absent from GT's 'no vehicles around it' -- the same fabrication seen on this clip in every prior round.", "orange"),
    "01737": (8, "Near-exact match to GT's empty night interchange under an overpass, including the road straightening out. Correct verdict. Only flaw: turn direction inverted (GT right curve, reported left).", "green"),
    "02104": (6, "Correct verdict with no fabricated hazard, and explicitly reports no truly moving agents. Generic relative to GT's specific merging-sedan detail.", "orange"),
    "02117": (8, "Clean match to GT (vehicle ahead at stable following distance, green light, crosswalk) with no hallucinated merge event. Sixth consecutive non-Gemini teacher/prompt combination to resolve this clip correctly.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "v7_verdict", "risk_score",
           "closing_risk", "lateral_risk", "intrusion_risk", "unreacted_risk",
           "conflict_source", "v7_caption", "gt_ego_manoeuvre", "v7_ego_path",
           "ego_path_correct", "apparent_vs_true", "score", "score_explanation"]
WIDTHS = [11, 11, 11, 15, 48, 10, 9, 8, 8, 8, 8, 15, 46, 26, 40, 10, 42, 8, 50]
WRAP_COLS = {"gt_reasoning_en", "v7_caption", "v7_ego_path", "apparent_vs_true",
             "gt_ego_manoeuvre", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    v7 = {r["video_id"]: r for r in _load_jsonl(V7_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t} - refusing to substitute a "
                            f"value from a disagreeing source (see module docstring).")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], v7[vid]
        gt_man, ego_ok = GT_EGO[vid]
        score, rationale, colour = SCORES.get(vid, (None, "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"],
            "gt_reasoning_en": v["gt_reasoning_en"],
            "v7_verdict": "YES" if q["verdict"] == 1 else "NO", "risk_score": q["risk_score"],
            "closing_risk": q["closing_risk"], "lateral_risk": q["lateral_risk"],
            "intrusion_risk": q["intrusion_risk"], "unreacted_risk": q["unreacted_risk"],
            "conflict_source": q["conflict_source"],
            "v7_caption": f"{q['caption_neutral']}, {q['risk_clause']}",
            "gt_ego_manoeuvre": gt_man, "v7_ego_path": q["ego_path"],
            "ego_path_correct": "YES" if ego_ok else "NO",
            "apparent_vs_true": q["apparent_vs_true"],
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["v7_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["v7_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["v7_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["v7_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["v7_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    cc = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}
    ego_ok = sum(1 for r in rows if r["ego_path_correct"] == "YES")

    from sklearn.metrics import average_precision_score, roc_auc_score
    y_true = [1 if r["gt_verdict"] == "YES" else 0 for r in rows]
    y_score = [r["risk_score"] for r in rows]

    metrics = [
        ("n", n), ("verdict_accuracy (score>=50 cut)", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP", tp), ("FP", fp), ("TN", tn), ("FN", fn),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("risk_score AUC", round(roc_auc_score(y_true, y_score), 3)),
        ("risk_score AP", round(average_precision_score(y_true, y_score), 3)),
        ("n_distinct_risk_scores", len(set(y_score))),
        ("ego_path correct vs GT", f"{ego_ok}/{n} ({ego_ok/n:.1%})"),
        ("mean_hand_score", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green", cc["green"]), ("n_orange", cc["orange"]), ("n_red", cc["red"]),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
