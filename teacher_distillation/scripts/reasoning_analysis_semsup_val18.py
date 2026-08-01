"""reasoning_analysis_semsup_val18.py -- per-clip qualitative scores for
PROMPT_SEMSUP_V2 and PROMPT_SEMSUP_V3_COT vs the 18-clip GT validation set,
modelled directly on reasoning_analysis_v6_debate.py's structure and scoring
rubric (0-10, hardcoded score+rationale per clip after reading the reasoning
against GT, same as that script's V6_SCORES/RECOVERY_SCORES pattern).

Reads:
- dataset/manifests/val_e3a.jsonl (video_id, gt_verdict, gt_reasoning_en)
- outputs/prompt_bakeoff/reasoning_analysis_v6_debate.xlsx (v6 final verdict +
  reasoning: col I unless the clip was debated - {00474,01153,01504,02104,02117}
  - in which case col P, per the debate rule)
- outputs/prompt_bakeoff/semsup_val18/raw_v2_native.jsonl
- outputs/prompt_bakeoff/semsup_val18/raw_v3_cot.jsonl
- outputs/prompt_bakeoff/semsup_val18/raw_v6_control_rerun.jsonl (PROMPT_G_OPT_v6_balanced
  re-run today at identical settings to the original run - see semsup_v6_control_rerun.py -
  a same-day/same-environment control, since the original v6 result turned out to not be
  reproducible today, most likely due to model drift on the OpenRouter model alias)

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_semsup_val18.xlsx
  (per_clip + summary sheets)

Scores below were assigned qualitatively (0-10) by Claude after reading each
caption against gt_reasoning_en. Disclosure: Claude wrote PROMPT_SEMSUP_V2 and
PROMPT_SEMSUP_V3_COT, so this is not an independent scorer - the per-clip
rationale is published for every score specifically so it can be audited.
BERTScore (apo_metric.score_one's 'alignment') is reported alongside as a
metric Claude did not influence, though it is length-biased in v6's favor (see
summary.md) and is secondary, not the headline number.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from apo_metric import score_one, warmup_bertscore  # noqa: E402

VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
V6_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "reasoning_analysis_v6_debate.xlsx"
OUT_DIR = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18"
V2_JSONL = OUT_DIR / "raw_v2_native.jsonl"
V3_JSONL = OUT_DIR / "raw_v3_cot.jsonl"
V6_RERUN_JSONL = OUT_DIR / "raw_v6_control_rerun.jsonl"
OUT_XLSX = OUT_DIR / "reasoning_analysis_semsup_val18.xlsx"

DEBATED = {"00474", "01153", "01504", "02104", "02117"}

# (score, rationale) per clip -- caption_neutral + risk_clause vs gt_reasoning_en.
# Same 0-10 rubric as reasoning_analysis_v6_debate.py's V6_SCORES.
V2_SCORES = {
    "00077": (8, "Braking sedan ahead, closing distance, rear-end imminent match GT; doesn't state the sedan merged in from the side"),
    "00147": (3, "Describes a car crossing 'in front' rather than EGO deviating into the other vehicle's lane -- the actual causal mechanism; wrong verdict"),
    "00283": (7, "Pickup+trailer merging from the right shoulder matches GT's agent and origin; doesn't capture the perpendicular/blocking nature of the turn"),
    "00319": (1, "Describes a static parked-vehicle scene; misses the moving crossing vehicle from the right entirely -- different scene"),
    "00372": (4, "Braking sedan ahead is roughly right, but invents a 'red light' cause GT never states and misses the turn+pedestrian-stop detail"),
    "00474": (2, "States the van is STOPPED; GT's entire causal event is the van turning INTO the ego lane -- same miss as v6"),
    "00493": (4, "Captures braking+closing (partial match) but misses EGO's own left-turn-merge context; wrong verdict despite own risk_clause implying danger"),
    "00529": (1, "Introduces a pedestrian/crosswalk scene not in GT; misses the SUV being forced into the ego lane by an obstruction entirely"),
    "00687": (6, "Correctly identifies the grey SUV merging left into the ego lane and closing distance -- but verdict (0) contradicts its own risk_clause ('moderate risk, close merge')"),
    "01153": (1, "Reproduces v6's exact hallucination: a crossing/turning sedan not present in GT's smooth-right-turn scene"),
    "01281": (3, "Invents a black-SUV merge event not in GT at all; verdict happens to be correct but the described event is fabricated"),
    "01504": (6, "Captures the following-distance/braking structure; swaps which SUV is 'red' vs 'dark' relative to GT"),
    "01550": (7, "Matches GT's controlled, maintained-distance framing; correct verdict"),
    "01552": (8, "Box truck ahead + minivan from gas station moving away closely matches GT's no-conflict resolution"),
    "01643": (5, "Correct empty-road verdict, but invents a dark sedan + white SUV in the adjacent lane not mentioned in GT"),
    "01737": (9, "Empty curving road, night, matches GT closely"),
    "02104": (7, "Sedan merging into ego lane at close distance matches GT's core event and verdict, though risk_clause ('moderate risk') is inconsistent with its own verdict=0"),
    "02117": (1, "Reproduces v6's exact hallucination: fabricated black-SUV merge; GT is a gray sedan at constant distance + a stopped van"),
}

V3_SCORES = {
    "00077": (7, "Captures braking+merge+closing correctly but states the sedan merges 'from the left' -- GT says the sedan starts to the right, a directional error"),
    "00147": (2, "States the lead sedan is 'stopped' (not in GT) and still frames the crossing car as passing by rather than EGO deviating into its lane; wrong verdict"),
    "00283": (9, "Explicitly names the trailer AND perpendicular crossing motion -- closely matches GT's 'left turn... blocking the lane'"),
    "00319": (9, "Matches GT closely: car from the right crossing into the ego path at close range; correct verdict"),
    "00372": (4, "Braking sedan ahead is right, but the turn+pedestrian-stop causal detail is missing, same gap as V2"),
    "00474": (2, "States the van is 'moving parallel' -- an improvement over V2's 'stopped' but still misses the critical left-turn-into-lane event entirely"),
    "00493": (7, "Braking+closing distance matches GT and verdict is correct; pickup truck framed as 'parallel' rather than the specific left-lane detail"),
    "00529": (2, "Frames the silver SUV as staying parallel -- the opposite of GT's claim that it drifts into the ego lane"),
    "00687": (8, "Grey SUV merging into ego path matches GT and verdict is correct; invents a 'from a driveway' detail not present in GT"),
    "01153": (1, "Reproduces the same hallucinated crossing-car scene as v6 and V2"),
    "01281": (1, "Same fabricated black-SUV merge as V2, here also flipping the verdict to wrong"),
    "01504": (7, "Braking+controlled-closing ('routine deceleration') matches GT's framing better than V2's; same red/dark SUV colour swap"),
    "01550": (2, "Reframes GT's explicitly 'controlled manner, maintaining distance' scene as 'rapidly closes... close proximity' -- an exaggeration not supported by GT; wrong verdict"),
    "01552": (7, "Box truck + minivan crossing matches the scene; 'crosses path' reads slightly more conflictual than GT's resolution but verdict is correct"),
    "01643": (6, "Correct empty-road verdict; invents a distant white car not in GT, softened by 'distant'"),
    "01737": (10, "Explicitly matches GT's 'bridge with lighting' via 'illuminated overpass' -- closest paraphrase in the set"),
    "02104": (1, "Hallucinates a stationary flatbed truck as the primary hazard; GT's actual event (sedan merging with reasonable distance) is not described at all"),
    "02117": (1, "Reproduces v6's and V2's exact hallucination, even more specifically ('executes a lane change... rapidly closing')"),
}

# V6-CONTROL-RERUN: PROMPT_G_OPT_v6_balanced re-run TODAY at identical settings
# to the original v6_hires_full18.py run (same model slug, temperature, image
# encoding, max_tokens=8192) - a same-day control to test whether the ORIGINAL
# recorded v6 baseline (72.2%/83.3% accuracy, mean 6.28/6.78) is still
# reproducible, or whether something in the calling environment (most likely
# model drift on the "google/gemini-3.1-pro-preview" OpenRouter alias) has
# changed since. See semsup_v6_control_rerun.py.
V6_RERUN_SCORES = {
    "00077": (8, "Braking lead vehicle, critical closing speed, rear-end imminent - matches GT; doesn't state the merge-in event"),
    "00147": (7, "Correct verdict; blames the other vehicle 'cutting across' rather than EGO's own deviation - same attribution error as the original v6 run"),
    "00283": (7, "Pickup+trailer from the right shoulder matches GT's agent and origin; misses the stationary-then-turns/blocking nuance"),
    "00319": (9, "Close paraphrase of GT: crossing vehicle from the right, rapid distance decrease, insufficient space"),
    "00372": (2, "Misses the stopped-sedan/pedestrian-crosswalk hazard entirely; wrong verdict, different failure mode than the original v6 run but equally wrong"),
    "00474": (1, "Same miss as the original v6 run, V2, and V3: 'no sudden maneuvers' when GT's entire event is the van's left turn"),
    "00493": (8, "Braking lead vehicle, EGO fails to maintain distance, critically short gap by frame 16 - matches GT's core mechanism and verdict"),
    "00529": (1, "Misses the forced-merge-due-to-obstruction event the ORIGINAL v6 run uniquely caught (score 10) - a clear regression on this specific clip"),
    "00687": (8, "Grey SUV merging into ego lane, converging trajectory, rapid closure - matches GT and verdict"),
    "01153": (1, "Reproduces the near-identical hallucinated left-turning white sedan as the original v6 run, V2, and V3"),
    "01281": (1, "Fabricates an abrupt black-SUV lane change not in GT at all - a NEW hallucination not present in the original v6 run (which got this one right)"),
    "01504": (1, "Same red/dark SUV colour confusion as V2/V3, but here it drives a wrong verdict where V2/V3 both got this one right - a regression"),
    "01550": (8, "Matches GT's calm, controlled-closing framing; correct verdict"),
    "01552": (1, "Fabricates a rear-end hazard from a box truck GT does not support - a NEW hallucination, regression vs. the original v6 run (which correctly said NO)"),
    "01643": (9, "Clean empty-road match to GT, no extraneous details"),
    "01737": (9, "Matches GT's empty curving night road closely"),
    "02104": (1, "Hallucinates a 'flatbed truck' rear-end hazard - the same fabricated object V3 also produced independently; opposite of GT's actual merging-sedan event"),
    "02117": (1, "Reproduces the same black-SUV-merge hallucination as the original v6 run, V2, and V3 - 4-way agreement on this exact wrong reading"),
}


def _load_jsonl(p: Path) -> dict:
    return {json.loads(l)["video_id"]: json.loads(l) for l in open(p, encoding="utf-8") if l.strip()}


def _load_v6_final() -> dict:
    wb = load_workbook(V6_XLSX)
    ws = wb["per_clip"]
    out = {}
    for r in range(2, ws.max_row + 1):
        vid = ws.cell(r, 1).value
        v6_reasoning = ws.cell(r, 9).value       # col I
        rec_reasoning = ws.cell(r, 16).value     # col P
        final_verdict = ws.cell(r, 17).value     # col Q
        final_score = ws.cell(r, 19).value       # col S (final_after_debate_score)
        reasoning = rec_reasoning if vid in DEBATED else v6_reasoning
        out[vid] = {"verdict": final_verdict, "reasoning": reasoning, "score": final_score}
    return out


def main():
    warmup_bertscore()

    val = _load_jsonl(VAL_MANIFEST)
    v2 = _load_jsonl(V2_JSONL)
    v3 = _load_jsonl(V3_JSONL)
    v6_final = _load_v6_final()
    v6_rerun = _load_jsonl(V6_RERUN_JSONL)

    rows = []
    for vid in sorted(val):
        gt_v = val[vid]["gt_verdict"]
        gt_reason = val[vid]["gt_reasoning_en"]

        v6 = v6_final[vid]
        v6_verdict = v6["verdict"]
        v6_correct = v6_verdict == gt_v

        v2r, v3r = v2[vid], v3[vid]
        v2_verdict = "YES" if v2r["verdict"] == 1 else "NO"
        v3_verdict = "YES" if v3r["verdict"] == 1 else "NO"
        v2_correct = v2_verdict == gt_v
        v3_correct = v3_verdict == gt_v

        v2_text = v2r["caption_neutral"] + ", " + v2r["risk_clause"]
        v3_text = v3r["caption_neutral"] + ", " + v3r["risk_clause"]
        v2_bert = round(score_one(v2_verdict, v2_text, gt_v, gt_reason).alignment, 3)
        v3_bert = round(score_one(v3_verdict, v3_text, gt_v, gt_reason).alignment, 3)

        v2_score, v2_rat = V2_SCORES.get(vid, (None, ""))
        v3_score, v3_rat = V3_SCORES.get(vid, (None, ""))

        v6r = v6_rerun[vid]
        v6r_verdict = v6r["verdict"]  # already "YES"/"NO"
        v6r_correct = v6r_verdict == gt_v
        v6r_bert = round(score_one(v6r_verdict, v6r["verdict_reasoning"], gt_v, gt_reason).alignment, 3)
        v6r_score, v6r_rat = V6_RERUN_SCORES.get(vid, (None, ""))

        rows.append({
            "video_id": vid,
            "gt_verdict": gt_v,
            "gt_reasoning_en": gt_reason,
            "v6_final_verdict": v6_verdict,
            "v6_final_correct": v6_correct,
            "v6_final_reasoning": v6["reasoning"],
            "v6_final_score": v6["score"],
            "v2__caption_neutral": v2r["caption_neutral"],
            "v2__risk_clause": v2r["risk_clause"],
            "v2__verdict": v2_verdict,
            "v2__correct": v2_correct,
            "v2__bert": v2_bert,
            "v2__score": v2_score,
            "v2__rationale": v2_rat,
            "v3__caption_neutral": v3r["caption_neutral"],
            "v3__risk_clause": v3r["risk_clause"],
            "v3__verdict": v3_verdict,
            "v3__correct": v3_correct,
            "v3__bert": v3_bert,
            "v3__score": v3_score,
            "v3__rationale": v3_rat,
            "v3__scene_context": v3r.get("scene_context", ""),
            "v3__dynamic_objects": v3r.get("dynamic_objects", ""),
            "v3__temporal_analysis": v3r.get("temporal_analysis", ""),
            "v6_rerun__verdict": v6r_verdict,
            "v6_rerun__correct": v6r_correct,
            "v6_rerun__reasoning": v6r["verdict_reasoning"],
            "v6_rerun__bert": v6r_bert,
            "v6_rerun__score": v6r_score,
            "v6_rerun__rationale": v6r_rat,
        })

    df = pd.DataFrame(rows)

    def stats(name, ss, corrects):
        ss = [s for s in ss if s is not None]
        return {
            "stage": name, "n": len(ss),
            "mean": round(sum(ss) / len(ss), 2) if ss else 0,
            "median": sorted(ss)[len(ss) // 2] if ss else 0,
            "n_ge8": sum(1 for s in ss if s >= 8),
            "n_le2": sum(1 for s in ss if s <= 2),
            "verdict_acc": f"{sum(corrects)}/{len(corrects)} ({sum(corrects)/len(corrects):.1%})",
        }

    df_summary = pd.DataFrame([
        stats("v6 ORIGINAL final (after debate, historical)", df["v6_final_score"].tolist(), df["v6_final_correct"].tolist()),
        stats("v6 RERUN TODAY (control, same env/day as V2/V3)", df["v6_rerun__score"].tolist(), df["v6_rerun__correct"].tolist()),
        stats("V2 (direct caption)", df["v2__score"].tolist(), df["v2__correct"].tolist()),
        stats("V3 (CoT-then-distill)", df["v3__score"].tolist(), df["v3__correct"].tolist()),
    ])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="per_clip", index=False)
        df_summary.to_excel(w, sheet_name="summary", index=False)

    print(f"Wrote {OUT_XLSX}: {len(df)} rows, {len(df.columns)} cols")
    print()
    print(df_summary.to_string(index=False))


if __name__ == "__main__":
    main()
