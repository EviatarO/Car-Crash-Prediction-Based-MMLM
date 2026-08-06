"""build_failures587_review_xlsx.py -- human-review sheet for V10-hybrid
captions on the train4500 failure windows (587 total; this run: any subset,
e.g. the 10-clip sanity check).

WHY NO AUTOMATED SCORING
---------------------------
Unlike val_e3a's 18 clips, the 587 failure windows have NO human
gt_reasoning_en (only 27/396 failure videos have any teacher reasoning at
all, and that's model-generated, not ground truth -- see the 2026-08-04
plan). So `slot_recall` and the fabrication check from
reasoning_analysis_v10_gt_val18.py CANNOT run here. This sheet is a plain
review instrument: one row per clip, the caption + supporting fields, a
blank `manual_review` column for hand annotation.

Reads whatever of these exist (works on a 10-clip check or all 587):
  outputs/semantic_captions/failures587/check10_gt_pos.jsonl / raw_v10_gemini_gt_pos.jsonl
  outputs/semantic_captions/failures587/check10_blind_neg.jsonl / raw_v10_gemini_blind_neg.jsonl

Writes: outputs/semantic_captions/failures587/review_<label>.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import cv2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
FAILURES_DIR = REPO_ROOT / "outputs" / "semantic_captions" / "failures587"
SRC_VIDEOS = Path(
    r"C:\Users\eviatar.ohayon\Ramon Space\PycharmProjects\Thesis"
    r"\Data-Centric-Crash-Prediction-Using-3LC-and-MViT\src\Nexar_DataSet\train"
)
# Matches build_train4500_manifest.py's NEG_BUCKETS and
# semsup_extract_promptbakeoff_frames.py's midpoint+offset/T_FLOOR formula
# EXACTLY - train4500_hires.jsonl deliberately leaves t_seconds=None for
# negatives ("depends on each video's own midpoint; filled by the
# extractor" - the extractor computes it at frame-extraction time but never
# writes it back to the manifest). Recomputed here from the raw video's own
# fps/frame-count so negatives can show a real time value instead of just
# the bucket label.
NEG_OFFSETS = {"MID-10": -10.0, "MID-4": -4.0, "MID-8": -8.0}
T_FLOOR = 2.0
_video_meta_cache: dict = {}


def compute_neg_t_seconds(video_id: str, horizon_label: str) -> float | None:
    if horizon_label not in NEG_OFFSETS:
        return None
    if video_id not in _video_meta_cache:
        mp4 = SRC_VIDEOS / f"{video_id}.mp4"
        if not mp4.exists():
            _video_meta_cache[video_id] = None
        else:
            cap = cv2.VideoCapture(str(mp4))
            fps = cap.get(cv2.CAP_PROP_FPS)
            total = cap.get(cv2.CAP_PROP_FRAME_COUNT)
            cap.release()
            _video_meta_cache[video_id] = (total / fps) / 2.0 if fps and total else None
    midpoint = _video_meta_cache[video_id]
    if midpoint is None:
        return None
    return round(max(T_FLOOR, midpoint + NEG_OFFSETS[horizon_label]), 3)


HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
POS_FILL = PatternFill("solid", fgColor="FFC6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFFFC7CE")
DOUBLE_FAIL_FILL = PatternFill("solid", fgColor="FFC00000")  # harder red
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    "video_id", "event_occurs", "horizon_label", "t_seconds", "gt_mode",
    "caption_neutral", "risk_clause",
    "hazard_agent", "hazard_motion", "hazard_position", "closing_dynamic",
    "evidence_frames", "mechanism_visible_or_scene", "risk_score", "verdict",
    "double_failure", "manual_review",
]
WIDTHS = [10, 10, 12, 10, 8, 55, 24, 22, 26, 20, 26, 16, 14, 10, 8, 14, 40]
WRAP_COLS = {"caption_neutral", "risk_clause", "hazard_agent", "hazard_motion",
             "hazard_position", "closing_dynamic", "manual_review"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--pos", nargs="+",
                     default=[str(FAILURES_DIR / "raw_v10_gemini_gt_pos.jsonl")],
                     help="one or more GT-mode (positive) raw caption JSONLs to merge")
    ap.add_argument("--neg", nargs="+",
                     default=[str(FAILURES_DIR / "raw_v10_gemini_blind_neg.jsonl")],
                     help="one or more blind-mode (negative) raw caption JSONLs to merge")
    ap.add_argument("--out", default=str(FAILURES_DIR / "review_failures587.xlsx"),
                     help="single output file, re-run this to refresh as more clips are captioned")
    args = ap.parse_args()

    # Dedup by frames_dir (the per-WINDOW unique key - see row_resume_key()
    # in semsup_caption_promptbakeoff.py) across the merged files, NOT by
    # video_id: a video can fail A0 at more than one TTE/MID bucket, so
    # video_id repeats legitimately within a single file (76/269 positive
    # and 70/318 negative video_ids do). Deduping by video_id here hit the
    # exact same bug class as the runner's original resume-key bug and
    # silently collapsed 587 rows down to 396 the first time this ran.
    # Last file wins so a full-run file, if present, takes priority over a
    # small check-batch file for the same window.
    pos_rows_by_key, neg_rows_by_key = {}, {}
    for p in args.pos:
        if Path(p).exists():
            for r in _load_jsonl(Path(p)):
                pos_rows_by_key[r.get("frames_dir") or r["video_id"]] = r
    for p in args.neg:
        if Path(p).exists():
            for r in _load_jsonl(Path(p)):
                neg_rows_by_key[r.get("frames_dir") or r["video_id"]] = r
    pos_rows = list(pos_rows_by_key.values())
    neg_rows = list(neg_rows_by_key.values())
    if not pos_rows and not neg_rows:
        raise FileNotFoundError(f"None of {args.pos + args.neg} exist.")

    # horizon_label/event_occurs/t_seconds are now written directly into
    # each caption row by semsup_caption_promptbakeoff.py (fixed 2026-08-04
    # alongside the frames_dir-based resume-key bug - see row_resume_key()'s
    # docstring there). No manifest re-join needed or wanted: the previous
    # version of this script joined on "video_id + requested_time_to_event",
    # but requested_time_to_event doesn't exist in the train4500 manifest
    # schema at all (it uses time_before_event_s), so EVERY row silently
    # matched on video_id alone and collapsed onto that video's first
    # manifest entry - every negative showed "MID-10" and several positives
    # showed the wrong TTE bucket, regardless of which was actually
    # captioned. Reading the fields straight off the caption row is both
    # simpler and correct.

    records = []
    for row in pos_rows:
        records.append({
            "video_id": row["video_id"],
            "event_occurs": row.get("event_occurs", 1),
            "horizon_label": row.get("horizon_label", ""),
            "t_seconds": round(row["t_seconds"], 3) if row.get("t_seconds") is not None else None,
            "gt_mode": "gt",
            "caption_neutral": row.get("caption_neutral", ""),
            "risk_clause": row.get("risk_clause", ""),
            "hazard_agent": row.get("hazard_agent", ""),
            "hazard_motion": row.get("hazard_motion", ""),
            "hazard_position": row.get("hazard_position", ""),
            "closing_dynamic": row.get("closing_dynamic", ""),
            "evidence_frames": row.get("evidence_frames", []),
            "mechanism_visible_or_scene": row.get("mechanism_visible"),
            "risk_score": None,
            "verdict": None,
            # GT mode is told the label, so "verdict != event_occurs" isn't a
            # concept that applies here - only blind-mode rows can double-fail.
            "double_failure": None,
            "manual_review": "",
            "_fill": POS_FILL,
        })
    for row in neg_rows:
        horizon_label = row.get("horizon_label", "")
        event_occurs = row.get("event_occurs", 0)
        verdict = row.get("verdict")
        # A0 (the frozen scorer) already got this window wrong - that's why
        # it's in the 587 in the first place. If the TEACHER's own blind
        # verdict *also* disagrees with event_occurs, that's a second,
        # independent model failing on the same window - a stronger signal
        # this is a genuinely hard/ambiguous clip, not an A0-specific quirk.
        double_failure = (verdict is not None and int(verdict) != int(event_occurs))
        records.append({
            "video_id": row["video_id"],
            "event_occurs": event_occurs,
            "horizon_label": horizon_label,
            "t_seconds": compute_neg_t_seconds(row["video_id"], horizon_label),
            "gt_mode": "blind",
            "caption_neutral": row.get("caption_neutral", ""),
            "risk_clause": row.get("risk_clause", ""),
            "hazard_agent": row.get("hazard_agent", ""),
            "hazard_motion": row.get("hazard_motion", ""),
            "hazard_position": row.get("hazard_position", ""),
            "closing_dynamic": row.get("closing_dynamic", ""),
            "evidence_frames": row.get("evidence_frames", []),
            "mechanism_visible_or_scene": row.get("mechanism_visible"),
            "risk_score": row.get("risk_score"),
            "verdict": verdict,
            "double_failure": double_failure,
            "manual_review": "",
            "_fill": DOUBLE_FAIL_FILL if double_failure else NEG_FILL,
        })

    n_double_fail = sum(1 for r in records if r["double_failure"])
    records.sort(key=lambda r: (not r["double_failure"] if r["double_failure"] is not None else True,
                                 r["gt_mode"], r["video_id"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, rec in enumerate(records, start=2):
        for c, h in enumerate(HEADERS, 1):
            val_ = rec.get(h)
            if isinstance(val_, list):
                val_ = ",".join(str(x) for x in val_)
            cell = ws.cell(row=i, column=c, value=val_)
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            if h != "manual_review":
                cell.fill = rec["_fill"]
                if rec["_fill"] is DOUBLE_FAIL_FILL:
                    cell.font = Font(color="FFFFFF", bold=(h in ("video_id", "verdict",
                                                                   "double_failure")))
        longest = max(len(str(rec.get(h) or "")) for h in
                       ("caption_neutral", "hazard_motion"))
        ws.row_dimensions[i].height = max(40, 13.5 * (longest // 55 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}  ({len(pos_rows)} GT/positive + {len(neg_rows)} blind/negative = "
          f"{len(records)} rows, {n_double_fail} double-failures)")


if __name__ == "__main__":
    main()
