"""compare_gemini_gt_vs_blind_val18.py -- side-by-side Gemini 3.6 Flash
GT-mode vs blind-mode on the 18 val clips, WITH the actual caption text.

Why this exists separately from reasoning_analysis_v10_gt_val18.py: that
script scores all 4 arms but writes only metrics (recall/label/agent) -- it
never included `caption_neutral`, which makes it impossible to eyeball what
the model actually produced. This one is the human-review sheet: two arms,
full caption text, one row per clip.

Columns:
  GT reference : video_id, gt_verdict, t_seconds, requested_time_to_event,
                 gt_reasoning_en
  Gemini GT    : label, hazard_agent, caption_neutral, fabrication_flag
  Gemini blind : label, hazard_agent, caption_neutral, verdict, fabrication_flag

Colour (per-arm cells, not whole row, so the two arms can be compared at a
glance on the same row):
  GREEN  = MATCH
  ORANGE = PARTIAL
  RED    = MISS / CONTRADICT / fabrication-flagged

Writes: outputs/prompt_bakeoff/semsup_val18_gt/compare_gemini_gt_vs_blind_val18.xlsx
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import reasoning_analysis_v10_gt_val18 as R

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
GT_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_gt" / "raw_v10_gemini_gt.jsonl"
BLIND_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_gt" / "raw_v10_gemini_blind.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_gt" / "compare_gemini_gt_vs_blind_val18.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
GT_GROUP_FILL = PatternFill("solid", fgColor="1F5C99")
BLIND_GROUP_FILL = PatternFill("solid", fgColor="7030A0")
FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    "video_id", "gt_verdict", "t_seconds", "requested_time_to_event", "gt_reasoning_en",
    "gemini_gt_label", "gemini_gt_hazard_agent", "gemini_gt_caption_neutral",
    "gemini_gt_fabrication",
    "gemini_blind_label", "gemini_blind_hazard_agent", "gemini_blind_caption_neutral",
    "gemini_blind_verdict", "gemini_blind_fabrication",
]
WIDTHS = [10, 10, 10, 14, 62, 13, 26, 62, 26, 13, 26, 62, 13, 26]
WRAP_COLS = {"gt_reasoning_en", "gemini_gt_caption_neutral", "gemini_blind_caption_neutral",
             "gemini_gt_fabrication", "gemini_blind_fabrication", "gemini_gt_hazard_agent",
             "gemini_blind_hazard_agent"}

LABEL_COLOUR = {"MATCH": "green", "PARTIAL": "orange", "MISS": "red",
                 "CONTRADICT": "red", "N/A": "orange"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        if not fp.exists():
            continue
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    gt_rows = {r["video_id"]: r for r in _load_jsonl(GT_JSONL)}
    bl_rows = {r["video_id"]: r for r in _load_jsonl(BLIND_JSONL)}
    slots = R.load_gt_slots()
    t_seconds = _load_t_seconds()

    records = []
    for vid in sorted(val):
        v = val[vid]
        target = int(v["target"])
        g, b = gt_rows[vid], bl_rows[vid]

        g_score = R.score_blob(R._blob_from_v10_row(g), slots[vid], target)
        b_score = R.score_blob(R._blob_from_v10_row(b), slots[vid], target)
        g_fab = R.fabrication_check(g, v["gt_reasoning_en"], target)
        b_fab = R.fabrication_check(b, v["gt_reasoning_en"], target)

        records.append({
            "video_id": vid,
            "gt_verdict": v.get("gt_verdict", "YES" if target else "NO"),
            "t_seconds": round(t_seconds[vid], 3) if vid in t_seconds else None,
            "requested_time_to_event": v.get("requested_time_to_event"),
            "gt_reasoning_en": v["gt_reasoning_en"],
            "gemini_gt_label": g_score["label"],
            "gemini_gt_hazard_agent": g.get("hazard_agent", ""),
            "gemini_gt_caption_neutral": g.get("caption_neutral", ""),
            "gemini_gt_fabrication": (f"{g_fab['rule']}: {g_fab['detail']}"
                                       if g_fab["flagged"] else ""),
            "gemini_blind_label": b_score["label"],
            "gemini_blind_hazard_agent": b.get("hazard_agent", ""),
            "gemini_blind_caption_neutral": b.get("caption_neutral", ""),
            "gemini_blind_verdict": b.get("verdict"),
            "gemini_blind_fabrication": (f"{b_fab['rule']}: {b_fab['detail']}"
                                          if b_fab["flagged"] else ""),
            "_gt_colour": "red" if g_fab["flagged"] else LABEL_COLOUR.get(g_score["label"], "orange"),
            "_bl_colour": "red" if b_fab["flagged"] else LABEL_COLOUR.get(b_score["label"], "orange"),
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "gt_vs_blind"

    # group banner row
    ws.cell(row=1, column=1, value="GROUND TRUTH REFERENCE")
    ws.cell(row=1, column=6, value="GEMINI 3.6 FLASH -- GT MODE")
    ws.cell(row=1, column=10, value="GEMINI 3.6 FLASH -- BLIND MODE")
    for col, fill in ((1, HEADER_FILL), (6, GT_GROUP_FILL), (10, BLIND_GROUP_FILL)):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=14)
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=col).border = BORDER
        if col in range(6, 10):
            ws.cell(row=1, column=col).fill = GT_GROUP_FILL
        elif col >= 10:
            ws.cell(row=1, column=col).fill = BLIND_GROUP_FILL
        else:
            ws.cell(row=1, column=col).fill = HEADER_FILL

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "B3"

    for i, rec in enumerate(records, start=3):
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            if h.startswith("gemini_gt"):
                cell.fill = FILL[rec["_gt_colour"]]
            elif h.startswith("gemini_blind"):
                cell.fill = FILL[rec["_bl_colour"]]
        longest = max(len(str(rec[h])) for h in
                       ("gt_reasoning_en", "gemini_gt_caption_neutral",
                        "gemini_blind_caption_neutral"))
        ws.row_dimensions[i].height = max(48, 13.5 * (longest // 60 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = w

    # ---- summary sheet ----
    ws2 = wb.create_sheet("summary")
    pos = [r for r in records if r["gt_verdict"] == "YES"]
    neg = [r for r in records if r["gt_verdict"] == "NO"]

    def _count(rows, key, label):
        return sum(1 for r in rows if r[key] == label)

    def _fab(rows, key):
        return sum(1 for r in rows if r[key])

    metrics = [
        ("n clips", len(records)),
        ("--- ALL CLIPS ---", ""),
        ("GT   MATCH / PARTIAL / MISS / CONTRADICT",
         f"{_count(records,'gemini_gt_label','MATCH')} / {_count(records,'gemini_gt_label','PARTIAL')} / "
         f"{_count(records,'gemini_gt_label','MISS')} / {_count(records,'gemini_gt_label','CONTRADICT')}"),
        ("BLIND MATCH / PARTIAL / MISS / CONTRADICT",
         f"{_count(records,'gemini_blind_label','MATCH')} / {_count(records,'gemini_blind_label','PARTIAL')} / "
         f"{_count(records,'gemini_blind_label','MISS')} / {_count(records,'gemini_blind_label','CONTRADICT')}"),
        ("--- POSITIVES (gt_verdict=YES) ---", ""),
        ("n", len(pos)),
        ("GT   CONTRADICT", _count(pos, "gemini_gt_label", "CONTRADICT")),
        ("BLIND CONTRADICT", _count(pos, "gemini_blind_label", "CONTRADICT")),
        ("GT   MATCH", _count(pos, "gemini_gt_label", "MATCH")),
        ("BLIND MATCH", _count(pos, "gemini_blind_label", "MATCH")),
        ("--- NEGATIVES (gt_verdict=NO) ---", ""),
        ("n", len(neg)),
        ("GT   fabrication-flagged", _fab(neg, "gemini_gt_fabrication")),
        ("BLIND fabrication-flagged", _fab(neg, "gemini_blind_fabrication")),
        ("GT   MATCH", _count(neg, "gemini_gt_label", "MATCH")),
        ("BLIND MATCH", _count(neg, "gemini_blind_label", "MATCH")),
        ("--- verdict (blind arm only) ---", ""),
        ("blind verdict correct",
         f"{sum(1 for r in records if (r['gemini_blind_verdict'] == 1) == (r['gt_verdict'] == 'YES'))}/{len(records)}"),
        ("model", "google/gemini-3.6-flash"),
        ("prompt", "PROMPT_SEMSUP_V10_GT (--gt-mode gt vs blind)"),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 34

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  ({len(records)} rows)")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
