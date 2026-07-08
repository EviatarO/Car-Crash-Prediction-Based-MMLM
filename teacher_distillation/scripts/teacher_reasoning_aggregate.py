"""
teacher_reasoning_aggregate.py
==============================
Aggregate per-stage teacher JSONL outputs into the two canonical documentation files:

  outputs/teacher_reasoning/Teacher_Reasoning_All_Clips.jsonl   (full records, all fields)
  outputs/teacher_reasoning/Teacher_Reasoning_All_Clips.xlsx    (10 concise review columns)

Idempotent UPSERT keyed by (dataset, video_id, requested_time_to_event) — re-running a clip
overwrites its row; the file never duplicates. Each stage you finish, drop its teacher output
JSONL into outputs/teacher_reasoning/stages/ and run:

  python teacher_distillation/scripts/teacher_reasoning_aggregate.py \
      --dataset train --stage outputs/teacher_reasoning/stages/train_stage1.jsonl

--dataset {test,train} tags the rows (test video_ids and train video_ids share the 00000 id
namespace, so we do NOT infer membership — you tell us which pool the stage came from).

Reuses the teacher record schema of Teacher_dataset_distill_v11.py (fields video_id, t_seconds,
requested_time_to_event, gt_verdict, collision_verdict, verdict_reasoning, p2_collision_verdict,
p2_verdict_reasoning, final_verdict, final_reasoning, target, ...). The xlsx renames the p2_*
debate fields to debate_* per the review spec.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "outputs" / "teacher_reasoning"
ALL_JSONL = OUT_DIR / "Teacher_Reasoning_All_Clips.jsonl"
ALL_XLSX = OUT_DIR / "Teacher_Reasoning_All_Clips.xlsx"

# Review-sheet columns, in the exact order requested (inputs then outputs).
XLSX_COLUMNS = [
    "video_id", "t_seconds", "requested_time_to_event", "gt_verdict",         # inputs
    "final_verdict", "final_reasoning",                                       # outputs
    "collision_verdict", "verdict_reasoning",                                 #  pass-1
    "debate_collision_verdict", "debate_verdict_reasoning",                   #  pass-2 (=p2_*)
]
WIDTHS = {
    "video_id": 10, "t_seconds": 11, "requested_time_to_event": 16, "gt_verdict": 10,
    "final_verdict": 12, "final_reasoning": 70,
    "collision_verdict": 14, "verdict_reasoning": 70,
    "debate_collision_verdict": 18, "debate_verdict_reasoning": 70,
}
WRAP_COLS = {"final_reasoning", "verdict_reasoning", "debate_verdict_reasoning"}

HEADER = PatternFill("solid", fgColor="2E75B6")
STRONG_GREEN = PatternFill("solid", fgColor="00B050")   # pass-1 correct
SOFT_GREEN = PatternFill("solid", fgColor="C6EFCE")     # pass-2 (debate) correct
RED = PatternFill("solid", fgColor="FFC7CE")            # still wrong
GREY = PatternFill("solid", fgColor="D9D9D9")           # dataset=test marker on split column


def _norm_vid(v) -> str:
    """Zero-pad numeric ids to 5 digits; leave non-numeric ids untouched."""
    s = str(v).strip()
    try:
        return f"{int(s):05d}"
    except ValueError:
        return s


def _norm_verdict(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().upper()
    if s in ("1", "YES", "TRUE"):
        return "YES"
    if s in ("0", "NO", "FALSE"):
        return "NO"
    return s


_MID_MAP = {"MID": "TTE_0.5", "MID-4": "TTE_1.0", "MID-8": "TTE_1.5"}


def _tte_label(req) -> str:
    """A single value -> canonical TTE_x.y label."""
    if req is None:
        return ""
    s = str(req).strip()
    try:
        return f"TTE_{float(s):.1f}"
    except ValueError:
        return _MID_MAP.get(s, s)


def _resolve_tte(rec: dict) -> str:
    """Prefer horizon_label (clean TTE_*/MID*) then requested_time_to_event.
    Live teacher records carry requested_time_to_event; the legacy combined seed
    carries horizon_label (+ a non-numeric requested_time_to_event for negatives)."""
    hz = str(rec.get("horizon_label") or "").strip()
    if hz.startswith("TTE_"):
        return _tte_label(hz.replace("TTE_", ""))
    if hz in _MID_MAP:
        return _MID_MAP[hz]
    return _tte_label(rec.get("requested_time_to_event"))


def _key(rec: dict) -> Tuple[str, str, str]:
    ds = rec.get("dataset", "")
    return (ds, _norm_vid(rec.get("video_id")), _resolve_tte(rec))


def _load_all() -> Dict[Tuple[str, str, str], dict]:
    if not ALL_JSONL.exists():
        return {}
    out: Dict[Tuple[str, str, str], dict] = {}
    for line in ALL_JSONL.read_text(encoding="utf-8").splitlines():
        if line.strip():
            r = json.loads(line)
            out[_key(r)] = r
    return out


def _load_stage(path: Path, dataset: str) -> List[dict]:
    recs = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        r["dataset"] = dataset
        r["video_id"] = _norm_vid(r.get("video_id"))
        recs.append(r)
    return recs


def _outcome(rec: dict) -> str:
    """pass1 / pass2 / wrong from teacher verdicts vs gt."""
    gt = _norm_verdict(rec.get("gt_verdict") if rec.get("gt_verdict") is not None
                       else rec.get("target"))
    p1 = _norm_verdict(rec.get("collision_verdict"))
    final = _norm_verdict(rec.get("final_verdict") or p1)
    if p1 and p1 == gt:
        return "pass1"
    if final and final == gt:
        return "pass2"
    return "wrong"


def write_xlsx(store: Dict[Tuple[str, str, str], dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "all_clips"

    for c, name in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(1, c, name)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # deterministic order: dataset, video_id, TTE
    tte_order = {"TTE_0.5": 0, "TTE_1.0": 1, "TTE_1.5": 2}
    rows = sorted(store.values(),
                  key=lambda r: (r.get("dataset", ""), _norm_vid(r.get("video_id")),
                                 tte_order.get(_resolve_tte(r), 9)))

    fills = {"pass1": STRONG_GREEN, "pass2": SOFT_GREEN, "wrong": RED}
    for i, r in enumerate(rows, start=2):
        vals = {
            "video_id": _norm_vid(r.get("video_id")),
            "t_seconds": r.get("t_seconds"),
            "requested_time_to_event": r.get("requested_time_to_event"),
            "gt_verdict": _norm_verdict(r.get("gt_verdict") if r.get("gt_verdict") is not None
                                        else r.get("target")),
            "final_verdict": _norm_verdict(r.get("final_verdict")),
            "final_reasoning": r.get("final_reasoning"),
            "collision_verdict": _norm_verdict(r.get("collision_verdict")),
            "verdict_reasoning": r.get("verdict_reasoning"),
            "debate_collision_verdict": _norm_verdict(r.get("p2_collision_verdict")),
            "debate_verdict_reasoning": r.get("p2_verdict_reasoning"),
        }
        fill = fills.get(_outcome(r))
        for c, name in enumerate(XLSX_COLUMNS, start=1):
            cell = ws.cell(i, c, vals[name])
            if name in WRAP_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", horizontal="center")
            if name == "final_verdict" and fill is not None:
                cell.fill = fill

    for name, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(XLSX_COLUMNS.index(name) + 1)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(ALL_XLSX)


def main() -> None:
    ap = argparse.ArgumentParser(description="Upsert teacher stage JSONL into the all-clips files")
    ap.add_argument("--dataset", required=True, choices=["test", "train"],
                    help="which pool the stage came from (test/train share the id namespace)")
    ap.add_argument("--stage", required=True, action="append",
                    help="stage JSONL path(s); repeatable")
    args = ap.parse_args()

    store = _load_all()
    before = len(store)
    added = updated = 0
    for sp in args.stage:
        p = Path(sp if Path(sp).is_absolute() else REPO_ROOT / sp)
        if not p.exists():
            raise SystemExit(f"stage file not found: {p}")
        for rec in _load_stage(p, args.dataset):
            k = _key(rec)
            if k in store:
                updated += 1
            else:
                added += 1
            store[k] = rec

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with ALL_JSONL.open("w", encoding="utf-8") as f:
        for r in sorted(store.values(),
                        key=lambda r: (r.get("dataset", ""), _norm_vid(r.get("video_id")),
                                       _tte_label(r.get("requested_time_to_event")))):
            f.write(json.dumps(r, ensure_ascii=False, default=str) + "\n")
    write_xlsx(store)

    print(f"Aggregate: {before} -> {len(store)} rows  (+{added} new, {updated} overwritten)")
    print(f"  JSONL: {ALL_JSONL}")
    print(f"  XLSX : {ALL_XLSX}")


if __name__ == "__main__":
    main()
