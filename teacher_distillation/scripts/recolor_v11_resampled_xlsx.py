"""One-off: restructure results_v11_resampled.xlsx per-clip sheet.

Rebuilds the per_clip sheet from final_combined.jsonl with:
- Trimmed column set (10 columns) in a new order
- t_sec(last_frame) = v6_resampled_t_new if resampled, else t_seconds
- v6_reasoning merged from v6_orig_reasoning / v6_resampled_reasoning
- 3-color (green/orange/red) on final_verdict + teacher_reasoning_final
- 2-color (green/red) on v6_reasoning
- source moved to last column (keeps its existing per-source color)

The summary sheet is left untouched.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
NEW_DIR   = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v11_100clips_resampled"
JSONL     = NEW_DIR / "final_combined.jsonl"
XLSX      = NEW_DIR / "results_v11_resampled.xlsx"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
HEADER = PatternFill("solid", fgColor="2E75B6")

COLUMNS = [
    "video_id",
    "t_sec(last_frame)",
    "requested_time_to_event",
    "gt_verdict",
    "final_verdict",
    "v7_1_recovery_verdict",
    "teacher_reasoning_final",
    "v6_reasoning",
    "v7_1_recovery_reasoning",
    "source",
]

WIDTHS = {
    "video_id": 10,
    "t_sec(last_frame)": 14,
    "requested_time_to_event": 20,
    "gt_verdict": 10,
    "final_verdict": 12,
    "v7_1_recovery_verdict": 14,
    "teacher_reasoning_final": 60,
    "v6_reasoning": 60,
    "v7_1_recovery_reasoning": 60,
    "source": 20,
}


def _classify_final(r: dict) -> str:
    """Return 'green' / 'orange' / 'red' for final_verdict and teacher_reasoning_final."""
    gt = r.get("gt_verdict")
    final = r.get("final_verdict")
    src = r.get("source")
    if final is None or final != gt:
        return "red"
    # Correct. Did pass-1 get it, or did recovery?
    if src == "v11_preserved":
        return "green"
    if src == "v11_resampled_fp":
        # pass-1 of THIS pipeline = v6_resampled (t-4s)
        if r.get("v6_resampled_verdict") == gt:
            return "green"
        return "orange"
    if src == "v11_fn_v7_1":
        # v6_orig (FN pass-1) was wrong; only recovery could have fixed it
        return "orange"
    return "red"


def _classify_v6(r: dict) -> str:
    """Return 'green' / 'red' based on the pass-1 reasoning shown."""
    gt = r.get("gt_verdict")
    src = r.get("source")
    if src == "v11_resampled_fp":
        return "green" if r.get("v6_resampled_verdict") == gt else "red"
    return "green" if r.get("v6_orig_verdict") == gt else "red"


def _merge_v6_reasoning(r: dict) -> str:
    resampled = (r.get("v6_resampled_reasoning") or "").strip()
    if resampled:
        return resampled
    return (r.get("v6_orig_reasoning") or "").strip()


def _t_sec_last_frame(r: dict):
    v_new = r.get("v6_resampled_t_new")
    if v_new is not None:
        return v_new
    return r.get("t_seconds") if r.get("t_seconds") is not None else r.get("t_original")


def _source_fill(src: str) -> PatternFill | None:
    return {
        "v11_preserved":    GREEN,
        "v11_resampled_fp": ORANGE,
        "v11_fn_v7_1":      BLUE,
        "v11_no_verdict":   RED,
    }.get(src)


def _color_fill(name: str) -> PatternFill:
    return {"green": GREEN, "orange": ORANGE, "red": RED}[name]


def main() -> None:
    records = [json.loads(l) for l in JSONL.read_text(encoding="utf-8").splitlines() if l.strip()]
    print(f"Loaded {len(records)} records from {JSONL.name}")

    wb = openpyxl.load_workbook(XLSX)
    # Remove old per_clip sheet, keep summary intact
    if "per_clip" in wb.sheetnames:
        del wb["per_clip"]
    ws = wb.create_sheet("per_clip", 0)  # place first

    # Header
    for col_idx, name in enumerate(COLUMNS, start=1):
        c = ws.cell(1, col_idx, value=name)
        c.fill = HEADER
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 32

    # Body
    for row_idx, r in enumerate(records, start=2):
        row_values = {
            "video_id": r.get("video_id"),
            "t_sec(last_frame)": _t_sec_last_frame(r),
            "requested_time_to_event": r.get("requested_time_to_event"),
            "gt_verdict": r.get("gt_verdict"),
            "final_verdict": r.get("final_verdict"),
            "v7_1_recovery_verdict": r.get("v7_1_recovery_verdict"),
            "teacher_reasoning_final": r.get("teacher_reasoning_final"),
            "v6_reasoning": _merge_v6_reasoning(r),
            "v7_1_recovery_reasoning": r.get("v7_1_recovery_reasoning"),
            "source": r.get("source"),
        }

        final_color = _classify_final(r)
        v6_color = _classify_v6(r)

        for col_idx, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row_idx, col_idx, value=row_values[name])
            if name in ("final_verdict", "teacher_reasoning_final"):
                cell.fill = _color_fill(final_color)
            elif name == "v6_reasoning":
                cell.fill = _color_fill(v6_color)
            elif name == "source":
                fill = _source_fill(row_values["source"])
                if fill is not None:
                    cell.fill = fill
            # Wrap text on the long reasoning columns
            if name in ("teacher_reasoning_final", "v6_reasoning", "v7_1_recovery_reasoning"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS[name]

    # Reorder sheets so summary remains second
    if "summary" in wb.sheetnames:
        order = ["per_clip", "summary"] + [s for s in wb.sheetnames if s not in ("per_clip", "summary")]
        wb._sheets = [wb[s] for s in order]

    wb.save(XLSX)

    # Quick sanity log
    n_green_final = sum(1 for r in records if _classify_final(r) == "green")
    n_orange_final = sum(1 for r in records if _classify_final(r) == "orange")
    n_red_final = sum(1 for r in records if _classify_final(r) == "red")
    print(f"final_verdict colors: green={n_green_final}  orange={n_orange_final}  red={n_red_final}")
    print(f"-> {XLSX}")


if __name__ == "__main__":
    main()
