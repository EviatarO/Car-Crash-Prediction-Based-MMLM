"""reasoning_analysis_v9_val18.py -- scores PROMPT_SEMSUP_V9_MINIMAL captions from
qwen/qwen3-vl-235b-a22b-thinking against the 18-clip GT validation set.

V9 is the "less is more" arm: 866 prompt tokens (~21% of V7's 4,072), no
scaffolding fields, testing whether 7 rounds of added structure were net-negative.
Same reasoning_match axis as reasoning_analysis_v8_val18.py (MATCH / PARTIAL /
MISS / CONTRADICT against gt_reasoning_en), reported alongside verdict metrics
which are unusually strong this round and need the reasoning check to interpret
correctly - see summary.md for the finding that ties them together.

Reads:
- dataset/manifests/val_e3a.jsonl
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds)
- outputs/prompt_bakeoff/semsup_val18/raw_v9_qwen3vl235b.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v9_val18.xlsx

Colour rule (whole row, same convention as V4-V8):
  GREEN  = verdict correct AND caption substantively matches GT's mechanism
  ORANGE = verdict correct but reasoning generic/partial, OR verdict wrong while
           the reasoning still substantively matches GT
  RED    = caption misses or CONTRADICTS GT's stated mechanism (regardless of
           verdict - a contradicted or fabricated mechanism is not a usable
           training signal even when the derived verdict happens to be right)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
V9_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v9_qwen3vl235b.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v9_val18.xlsx"

# (score 0-10, reasoning_match, rationale, colour) per clip, hand-scored against
# gt_reasoning_en. reasoning_match is independent of verdict correctness.
SCORES = {
    "00077": (7, "PARTIAL", "Correctly identifies the decisive mechanism -- braking sedan, closing gap, ego not slowing -- matching GT's 'EGO fails to brake in time' for the first time with a comfortably correct verdict (score 73, not a near-miss). Misses that the sedan merged into the lane before braking.", "orange"),
    "00147": (5, "PARTIAL", "Causality still inverted (attributes the closing to the other sedan moving in; GT has EGO deviating into that sedan's lane) but correctly flags 'ego not slowing' as decisive. Score 43, one clip short of the cut, the same near-miss pattern seen on this clip in V7 and V8.", "orange"),
    "00283": (5, "PARTIAL", "Right shape (a vehicle cuts in, ego doesn't react) but mirrored direction and wrong object (GT: stationary pickup+trailer in the right lane turns left; caption: SUV entering from the left). Verdict correct.", "orange"),
    "00319": (1, "CONTRADICT", "Detects an agent near the intersection but reports it EXITING left -- the direct opposite of GT's car ENTERING from the right without slowing. The critical detail is inverted, not merely missing.", "red"),
    "00372": (1, "MISS", "Fabricates a red truck crossing and exiting; misses GT's actual mechanism (lead sedan stops for crosswalk pedestrians during a right turn) entirely.", "red"),
    "00474": (1, "CONTRADICT", "States 'ego braking in response' -- directly contradicting GT's explicit 'the EGO vehicle continues driving at the same speed.' Also misses the van's sharp turn into ego's lane.", "red"),
    "00493": (8, "MATCH", "Clean match: braking sedan, gap closing, ego not slowing -- exactly GT's 'ego does not slow down' mechanism. Verdict correct and comfortably above the cut (73), unlike the near-misses on this clip in V7 (43) and V8 (47).", "green"),
    "00529": (1, "CONTRADICT", "Highest-confidence prediction in the run (score 82) and it is a fabrication: invents a pedestrian entering the crosswalk into ego's path, a hazard GT does not describe, while missing the actual cause (a silver SUV drifting into ego's lane after a lane obstruction). Verdict happens to be correct.", "red"),
    "00687": (8, "MATCH", "Best-aligned reasoning on this clip across every round tested: correctly identifies the gray SUV merging into ego's lane with closing distance and no ego reaction, matching GT's mechanism directly. Verdict correct.", "green"),
    "01153": (0, "CONTRADICT", "The same fabricated crossing-sedan hallucination that broke V5, V6 and V8 on this identical clip reappears here, now confirmed present in 4 of 6 non-Gemini rounds (absent only in V7). GT explicitly states all vehicles remain in their own lanes.", "red"),
    "01281": (4, "PARTIAL", "Correctly identifies the lead truck and closing distance, but 'ego not slowing' contradicts GT's 'controlled closing distance' framing (GT implies ego is managing the approach). Verdict happens to land correctly regardless.", "orange"),
    "01504": (3, "MISS", "Misses the decisive mechanism GT describes -- that ego itself notices the braking ahead and brakes in time -- reporting plain green-light proceeding instead. Verdict correct via a generic low-risk read, not because the mechanism was understood.", "orange"),
    "01550": (8, "MATCH", "Genuine fix versus V8's regression on this exact clip: correctly captures the gentle, controlled reaction ('gap closing then steady with ego braking in response'), matching GT's 'closes the gap in a controlled manner while maintaining distance' closely. Verdict correct.", "green"),
    "01552": (6, "PARTIAL", "Reasonable match (an agent clearing ego's path, truck ahead, no fabricated hazard); the 'school bus' fabrication seen on this clip in V4-V7 does not reproduce. Verdict correct.", "green"),
    "01643": (7, "MATCH", "Clean, accurate, no fabricated hazard (parked cars noted only in passing, no invented road-work sign as in V6/V7). Verdict correct.", "green"),
    "01737": (6, "PARTIAL", "Accurate but unusually thin ('No changes in ego path with ego maintaining speed') -- loses the descriptive detail other rounds captured (the interchange turn, the lit overpass) though it does not contradict GT. Verdict correct.", "orange"),
    "02104": (2, "CONTRADICT", "States 'No changes observed' -- directly contradicting GT's explicit description of a white sedan merging into ego's lane in front of a tow truck. Harmless for the verdict here, but the reasoning is factually wrong about what GT says happened.", "red"),
    "02117": (4, "PARTIAL", "Reports the following distance as CLOSING, a mild contradiction of GT's 'constant distance... reasonable distance maintained' framing. Verdict correct.", "orange"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "v9_verdict", "risk_score", "reasoning_match",
           "v9_caption", "score", "score_explanation"]
WIDTHS = [11, 11, 11, 15, 55, 10, 9, 14, 55, 8, 60]
WRAP_COLS = {"gt_reasoning_en", "v9_caption", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    v9 = {r["video_id"]: r for r in _load_jsonl(V9_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t}")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], v9[vid]
        score, match, rationale, colour = SCORES.get(vid, (None, "MISS", "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"],
            "gt_reasoning_en": v["gt_reasoning_en"],
            "v9_verdict": "YES" if q["verdict"] == 1 else "NO", "risk_score": q["risk_score"],
            "reasoning_match": match,
            "v9_caption": f"{q['caption_neutral']}, {q['risk_clause']}",
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["v9_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["v9_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["v9_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["v9_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["v9_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    cc = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}
    match_counts = {m: sum(1 for r in rows if r["reasoning_match"] == m)
                     for m in ("MATCH", "PARTIAL", "MISS", "CONTRADICT")}

    from sklearn.metrics import average_precision_score, roc_auc_score
    y_true = [1 if r["gt_verdict"] == "YES" else 0 for r in rows]
    y_score = [r["risk_score"] for r in rows]

    metrics = [
        ("n", n),
        ("--- REASONING ALIGNMENT ---", ""),
        ("reasoning MATCH (substantively correct)", match_counts["MATCH"]),
        ("reasoning PARTIAL (right shape, wrong detail)", match_counts["PARTIAL"]),
        ("reasoning MISS (misses GT's mechanism)", match_counts["MISS"]),
        ("reasoning CONTRADICT (states the opposite of / fabricates beyond GT)", match_counts["CONTRADICT"]),
        ("mean_hand_score (0-10)", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green / n_orange / n_red", f"{cc['green']} / {cc['orange']} / {cc['red']}"),
        ("--- verdict metrics ---", ""),
        ("verdict_accuracy (score>=50 cut)", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP / FP / TN / FN", f"{tp} / {fp} / {tn} / {fn}"),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("risk_score AUC", round(roc_auc_score(y_true, y_score), 3)),
        ("risk_score AP", round(average_precision_score(y_true, y_score), 3)),
        ("n_distinct_risk_scores", len(set(y_score))),
        ("prompt length (tokens, approx)", 866),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 22

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
