"""APO v11scale data loader.

Reads `dataset/teacher_labels/teacher_dataset_v11.jsonl` and builds three splits:
  - Train: 31 v11 Pass-1 failures (clean, no GT overlap)
  - Regression: 67 v11 Pass-1 successes (clean, no GT overlap)
  - Val: 18 GT clips from teacher_dataset_GT_self_imply.xlsx

Each clip dict has the fields needed for evaluation:
    video_id, target, gt_verdict, t_seconds (or end_frame_idx),
    frames_dir (absolute path to per-video frame folder)

For VAL clips also: gt_reasoning_en (English-translated Hebrew reasoning).
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_V11_JSONL = REPO_ROOT / "outputs" / "teacher_dataset_v11.jsonl"
DEFAULT_GT_XLSX = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
DEFAULT_GT_FRAMES_ROOT = REPO_ROOT / "dataset" / "train"
DEFAULT_V11_FRAMES_ROOT = (
    Path(r"C:\Users\eviatar.ohayon\Ramon Space\PycharmProjects\Thesis")
    / "Data-Centric-Crash-Prediction-Using-3LC-and-MViT"
    / "src" / "Nexar_DataSet" / "train_frames256"
)


@dataclass
class APOClip:
    """Unified clip record across train / regression / val splits."""
    video_id: str
    target: int                    # 0 or 1
    gt_verdict: str                # "YES" or "NO"
    end_frame_idx: int             # last frame index in the 16-frame clip
    frame_indices: List[int]       # 16 indices, stride=4
    frames_dir: Path               # path to per-video frame folder
    split: str                     # "train" / "regression" / "val"
    # Val-only fields
    gt_reasoning_en: Optional[str] = None
    # Diagnostic / metadata (from v11 if available)
    v11_pass1_verdict: Optional[str] = None
    t_seconds: Optional[float] = None


def _normalize_video_id(value) -> str:
    return f"{int(float(str(value).strip())):05d}"


def _normalize_verdict(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


# ---------------------------------------------------------------------------
# v11 -> train + regression
# ---------------------------------------------------------------------------

def load_v11_records(path: Path) -> List[Dict]:
    """Load all records from v11 JSONL."""
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except Exception:
                continue
    return records


def classify_v11_records(
    records: List[Dict],
    exclude_video_ids: set,
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """Split v11 records into (failures, successes, dropped).

    Filters: exclude any record whose video_id is in `exclude_video_ids`,
             exclude records with parse errors or missing collision_verdict.
    """
    failures, successes, dropped = [], [], []
    for r in records:
        vid = r.get("video_id")
        if vid in exclude_video_ids:
            dropped.append(r)
            continue
        if r.get("error") is not None:
            dropped.append(r)
            continue
        verdict = _normalize_verdict(r.get("collision_verdict"))
        if verdict is None:
            dropped.append(r)
            continue
        target = r.get("target")
        gt_v = "YES" if target == 1 else "NO"
        is_correct = (verdict == gt_v)
        if is_correct:
            successes.append(r)
        else:
            failures.append(r)
    return failures, successes, dropped


def v11_record_to_clip(r: Dict, split: str, frames_root: Path) -> APOClip:
    vid = r["video_id"]
    target = int(r["target"])
    gt_verdict = "YES" if target == 1 else "NO"
    end_frame_idx = int(r["end_frame_idx"])
    frame_indices = list(r.get("frame_indices") or [])
    return APOClip(
        video_id=vid,
        target=target,
        gt_verdict=gt_verdict,
        end_frame_idx=end_frame_idx,
        frame_indices=frame_indices,
        frames_dir=frames_root / vid,
        split=split,
        v11_pass1_verdict=_normalize_verdict(r.get("collision_verdict")),
        t_seconds=r.get("t_seconds"),
    )


# ---------------------------------------------------------------------------
# GT excel -> val
# ---------------------------------------------------------------------------

def load_gt_clips(
    xlsx_path: Path,
    frames_root: Path,
    fps: float = 30.0,
    window_size: int = 16,
    stride: int = 4,
) -> List[APOClip]:
    """Read the 18 GT clips from the Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["video_id", "target", "gt_verdict", "t_seconds", "verdict_reasoning_en"]
    idx = {}
    for col in required:
        if col not in header:
            raise RuntimeError(f"Missing column '{col}' in {xlsx_path}. Headers: {header}")
        idx[col] = header.index(col)

    clips = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["video_id"]] is None:
            continue
        vid = _normalize_video_id(row[idx["video_id"]])
        target = int(row[idx["target"]]) if row[idx["target"]] is not None else None
        gt_verdict = str(row[idx["gt_verdict"]]).strip().upper() if row[idx["gt_verdict"]] else None
        t_seconds = float(row[idx["t_seconds"]]) if row[idx["t_seconds"]] is not None else None
        gt_en = (row[idx["verdict_reasoning_en"]] or "").strip()

        # Try to derive end_frame_idx from t_seconds; fall back to actual frame files on disk.
        # (Some Excel rows have formula-based t_seconds that lost their cached value when saved.)
        if t_seconds is not None:
            end_frame_idx = round(t_seconds * fps)
        else:
            # Fallback: use the highest-numbered frame in the per-video folder
            video_dir = frames_root / vid
            frame_files = sorted(video_dir.glob("frame_*.jpg"))
            if not frame_files:
                raise RuntimeError(f"No frames found for {vid} and t_seconds is None — cannot derive end_frame_idx")
            # Parse highest frame index from filename like 'frame_00587.jpg'
            indices_on_disk = sorted(int(f.stem.split("_")[1]) for f in frame_files)
            end_frame_idx = indices_on_disk[-1]
            t_seconds = end_frame_idx / fps  # back-fill for record

        frame_indices = [end_frame_idx - (window_size - 1 - i) * stride for i in range(window_size)]

        clips.append(APOClip(
            video_id=vid,
            target=target,
            gt_verdict=gt_verdict,
            end_frame_idx=end_frame_idx,
            frame_indices=frame_indices,
            frames_dir=frames_root / vid,
            split="val",
            gt_reasoning_en=gt_en,
            t_seconds=t_seconds,
        ))
    return clips


# ---------------------------------------------------------------------------
# Top-level loader
# ---------------------------------------------------------------------------

def build_v11scale_splits(
    v11_jsonl: Path = DEFAULT_V11_JSONL,
    gt_xlsx: Path = DEFAULT_GT_XLSX,
    v11_frames_root: Path = DEFAULT_V11_FRAMES_ROOT,
    gt_frames_root: Path = DEFAULT_GT_FRAMES_ROOT,
) -> Dict[str, List[APOClip]]:
    """Build all three splits.

    Returns:
        {"train": [...], "regression": [...], "val": [...]}
    """
    # 1. Load val (GT clips first — we need their video_ids to filter v11)
    val_clips = load_gt_clips(gt_xlsx, gt_frames_root)
    gt_vids = {c.video_id for c in val_clips}

    # 2. Load v11 and split into failures / successes (excluding GT overlap)
    v11_records = load_v11_records(v11_jsonl)
    failures, successes, dropped = classify_v11_records(v11_records, exclude_video_ids=gt_vids)

    train_clips = [v11_record_to_clip(r, "train", v11_frames_root) for r in failures]
    regression_clips = [v11_record_to_clip(r, "regression", v11_frames_root) for r in successes]

    print(f"  [v11scale loader] v11 total: {len(v11_records)}; "
          f"GT overlap dropped: {sum(1 for r in v11_records if r['video_id'] in gt_vids)}; "
          f"errors/parse-fail dropped: {len(dropped) - sum(1 for r in v11_records if r['video_id'] in gt_vids)}")
    print(f"  [v11scale loader] train (failures): {len(train_clips)}")
    print(f"  [v11scale loader]    FN (target=YES, pred=NO): {sum(1 for c in train_clips if c.target == 1)}")
    print(f"  [v11scale loader]    FP (target=NO, pred=YES): {sum(1 for c in train_clips if c.target == 0)}")
    print(f"  [v11scale loader] regression (successes): {len(regression_clips)}")
    print(f"  [v11scale loader] val (GT clips): {len(val_clips)}")

    return {
        "train": train_clips,
        "regression": regression_clips,
        "val": val_clips,
    }


def verify_frame_paths(splits: Dict[str, List[APOClip]]) -> List[str]:
    """Verify that all required frame files exist on disk.

    Returns list of error messages (empty if all OK).
    """
    errors = []
    for split_name, clips in splits.items():
        missing_count = 0
        for c in clips:
            if not c.frames_dir.exists():
                errors.append(f"  [{split_name}] missing folder: {c.frames_dir}")
                missing_count += 1
                continue
            # Check first and last frame exist as a quick sanity check
            for idx in [c.frame_indices[0], c.frame_indices[-1]]:
                fp = c.frames_dir / f"frame_{idx:05d}.jpg"
                if not fp.exists():
                    errors.append(f"  [{split_name}] missing frame: {fp}")
                    missing_count += 1
        if missing_count == 0:
            print(f"  [verify] {split_name}: all {len(clips)} clip frame paths exist")
    return errors


# ---------------------------------------------------------------------------
# CLI sanity
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    splits = build_v11scale_splits()
    print()
    print("Verifying frame paths exist on disk...")
    errors = verify_frame_paths(splits)
    if errors:
        print(f"\n{len(errors)} ERRORS:")
        for e in errors[:10]:
            print(e)
    else:
        print("All frame paths verified OK.")
