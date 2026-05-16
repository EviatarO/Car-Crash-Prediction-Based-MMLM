"""Teacher Bake-Off: Claude Opus 4.7 vs Gemini 3.1 Pro Preview 0219.

Runs PROMPT_G2 on the 18 GT clips with both models, side-by-side, and writes
a comparison Excel for manual review against Hebrew GT reasoning.

Usage:
    py teacher_distillation/scripts/teacher_bakeoff.py
    py teacher_distillation/scripts/teacher_bakeoff.py --skip_smoke_test
    py teacher_distillation/scripts/teacher_bakeoff.py --opus_model anthropic/claude-opus-4.7
"""

import argparse
import base64
import json
import os
import re
import sys
import time
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image

# Add repo root to sys.path so prompts/ is importable
REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))

# Import PROMPT_G2 — note that the file uses variable name PROMPT_G inside
import importlib.util
_spec = importlib.util.spec_from_file_location(
    "prompt_g2_module", REPO_ROOT / "prompts" / "PROMPT_G2.py"
)
_pg2_module = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_pg2_module)
PROMPT_G2 = _pg2_module.PROMPT_G  # the file defines PROMPT_G inside

try:
    from openai import (
        APIConnectionError as _OAIConnectionError,
        APITimeoutError as _OAITimeoutError,
        OpenAI,
        RateLimitError as _OAIRateLimitError,
    )
    _HAS_OPENAI_ERRORS = True
except ImportError:
    from openai import OpenAI  # type: ignore[no-redef]
    _HAS_OPENAI_ERRORS = False
    _OAIConnectionError = Exception
    _OAITimeoutError = Exception
    _OAIRateLimitError = Exception

try:
    import httpx as _httpx
except ImportError:
    _httpx = None

# ---------------------------------------------------------------------------
# Models & pricing (USD per 1M tokens unless noted)
# ---------------------------------------------------------------------------

MODELS = {
    "opus": {
        "slug": "anthropic/claude-opus-4.7",
        "label": "Opus 4.7",
        "price_in":  5.00,
        "price_out": 25.00,
    },
    "gemini": {
        "slug": "google/gemini-3.1-pro-preview",
        "label": "Gemini 3.1 Pro",
        "price_in":  2.00,
        "price_out": 12.00,
    },
}

WINDOW_SIZE = 16
STRIDE = 4
FPS = 30
FRAME_SIZE = 256
DEFAULT_TIMEOUT = 120.0
DEFAULT_MAX_RETRIES = 3
DEFAULT_RETRY_DELAY = 3.0

DEFAULT_GT_XLSX = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
DEFAULT_FRAMES_ROOT = REPO_ROOT / "dataset" / "train"
DEFAULT_OUT_JSONL = REPO_ROOT / "outputs" / "apo" / "bakeoff_results.jsonl"
DEFAULT_OUT_XLSX = REPO_ROOT / "outputs" / "apo" / "bakeoff_results.xlsx"


# ---------------------------------------------------------------------------
# Helpers (adapted from Teacher_dataset_distill_v11.py)
# ---------------------------------------------------------------------------

def _normalize_video_id(value) -> str:
    return f"{int(float(str(value).strip())):05d}"


def _normalize_verdict(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


def _normalize_confidence(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"HIGH", "MEDIUM", "LOW"} else None


def _encode_image(path: Path, frame_size: int) -> str:
    if path.exists():
        img = Image.open(path).convert("RGB")
    else:
        img = Image.new("RGB", (frame_size, frame_size), color=(0, 0, 0))
    if frame_size and img.size != (frame_size, frame_size):
        img = img.resize((frame_size, frame_size))
    buf = BytesIO()
    img.save(buf, format="JPEG")
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"


def _build_messages(prompt: str, image_b64s: Sequence[str], detail: str = "low") -> List[Dict]:
    content: List[Dict] = [{"type": "text", "text": prompt}]
    for b64 in image_b64s:
        image_url: Dict = {"url": b64}
        if detail:
            image_url["detail"] = detail
        content.append({"type": "image_url", "image_url": image_url})
    return [{"role": "user", "content": content}]


def _extract_json_object(raw: str) -> Optional[Dict]:
    if not raw:
        return None
    try:
        obj = json.loads(raw)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass
    fenced = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", raw, flags=re.IGNORECASE)
    if fenced:
        try:
            obj = json.loads(fenced.group(1))
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            obj = json.loads(raw[start : end + 1])
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    return None


def _parse_response(raw: str) -> Tuple[Optional[Dict], Optional[str]]:
    parsed = _extract_json_object(raw)
    if parsed is None:
        return None, None
    parsed["collision_verdict"] = _normalize_verdict(parsed.get("collision_verdict"))
    parsed["confidence"] = _normalize_confidence(parsed.get("confidence"))
    return parsed, parsed["collision_verdict"]


def _call_model(
    client: OpenAI,
    model: str,
    messages: List[Dict],
    timeout: float,
    max_retries: int,
    retry_delay: float,
    temperature: float = 0.1,
    max_tokens: int = 8192,  # caps OpenRouter's upfront reservation; Gemini uses ~2k reasoning tokens internally
) -> Tuple[str, Dict]:
    last_exc: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model, messages=messages, temperature=temperature, timeout=timeout,
                max_tokens=max_tokens,
            )
            text = response.choices[0].message.content if response.choices else ""
            usage = (
                response.usage.model_dump()
                if hasattr(response, "usage") and response.usage else {}
            )
            return text or "", usage
        except _OAITimeoutError as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] timeout -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries: time.sleep(wait)
        except _OAIRateLimitError as exc:
            last_exc = exc
            wait = retry_delay * 2 * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] rate-limit -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries: time.sleep(wait)
        except _OAIConnectionError as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] connection -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries: time.sleep(wait)
        except Exception as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] error: {exc!r} -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries: time.sleep(wait)
    raise RuntimeError(f"OpenRouter call failed after {max_retries} attempts: {last_exc}") from last_exc


def _calc_cost(usage: Dict, price_in: float, price_out: float) -> float:
    """Compute USD cost from usage dict + pricing per 1M tokens."""
    in_tok = usage.get("prompt_tokens", 0) or 0
    out_tok = usage.get("completion_tokens", 0) or 0
    return in_tok * price_in / 1_000_000 + out_tok * price_out / 1_000_000


# ---------------------------------------------------------------------------
# Clip loading
# ---------------------------------------------------------------------------

def _read_gt_excel(path: Path) -> List[Dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {name: header.index(name) for name in
           ["video_id", "target", "gt_verdict", "t_seconds", "verdict_reasoning"]}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["video_id"]] is None:
            continue
        rows.append({
            "video_id": _normalize_video_id(row[idx["video_id"]]),
            "target": int(row[idx["target"]]) if row[idx["target"]] is not None else None,
            "gt_verdict": str(row[idx["gt_verdict"]]).strip().upper() if row[idx["gt_verdict"]] else None,
            "t_seconds": float(row[idx["t_seconds"]]) if row[idx["t_seconds"]] is not None else None,
            "hebrew_gt": row[idx["verdict_reasoning"]] or "",
        })
    return rows


def _frame_indices(t_seconds: float, fps: float, window: int, stride: int) -> List[int]:
    end = round(t_seconds * fps)
    return [end - (window - 1 - i) * stride for i in range(window)]


def _load_clip_frames(frames_root: Path, video_id: str, indices: List[int], frame_size: int) -> List[str]:
    folder = frames_root / video_id
    return [_encode_image(folder / f"frame_{i:05d}.jpg", frame_size) for i in indices]


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _write_excel(records: List[Dict], path: Path) -> None:
    """Write side-by-side comparison Excel.

    Column order (per approved plan):
      Metadata first, headline columns next (verdict + reasoning side-by-side),
      then comparison flags + word counts, then full detail (later).
    """
    rows = []
    for r in records:
        row = {
            # --- METADATA FIRST ---
            "video_id": r["video_id"],
            "target": r["target"],
            "gt_verdict": r["gt_verdict"],
            "t_seconds": r["t_seconds"],
            "hebrew_gt": r["hebrew_gt"],
            # --- HEADLINE COLUMNS (at-a-glance review) ---
            "opus_verdict": r["opus"]["verdict"],
            "opus_reasoning": r["opus"]["reasoning"],
            "gemini_verdict": r["gemini"]["verdict"],
            "gemini_reasoning": r["gemini"]["reasoning"],
            # --- COMPARISON ---
            "opus_match_gt": r["opus"]["verdict"] == r["gt_verdict"] if r["opus"]["verdict"] else None,
            "gemini_match_gt": r["gemini"]["verdict"] == r["gt_verdict"] if r["gemini"]["verdict"] else None,
            "word_count_opus": len((r["opus"]["reasoning"] or "").split()),
            "word_count_gemini": len((r["gemini"]["reasoning"] or "").split()),
            # --- DETAIL (later columns) ---
            "opus_confidence": r["opus"]["confidence"],
            "opus_temporal": r["opus"]["temporal_analysis"],
            "opus_ttc": r["opus"]["time_to_contact"],
            "opus_full_json": json.dumps(r["opus"]["full"], ensure_ascii=False),
            "gemini_confidence": r["gemini"]["confidence"],
            "gemini_temporal": r["gemini"]["temporal_analysis"],
            "gemini_ttc": r["gemini"]["time_to_contact"],
            "gemini_full_json": json.dumps(r["gemini"]["full"], ensure_ascii=False),
            # --- DIAGNOSTICS ---
            "opus_latency_s": r["opus"]["latency_s"],
            "gemini_latency_s": r["gemini"]["latency_s"],
            "opus_cost_usd": r["opus"]["cost_usd"],
            "gemini_cost_usd": r["gemini"]["cost_usd"],
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)

    # Style: highlight verdict mismatches
    wb = load_workbook(path)
    ws = wb.active
    red = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    yellow = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
    headers = [c.value for c in ws[1]]
    col_opus_match = headers.index("opus_match_gt") + 1
    col_gemini_match = headers.index("gemini_match_gt") + 1
    col_opus_verdict = headers.index("opus_verdict") + 1
    col_gemini_verdict = headers.index("gemini_verdict") + 1

    for excel_row in range(2, ws.max_row + 1):
        opus_match = ws.cell(excel_row, col_opus_match).value
        gemini_match = ws.cell(excel_row, col_gemini_match).value
        if opus_match is False:
            ws.cell(excel_row, col_opus_verdict).fill = red
        if gemini_match is False:
            ws.cell(excel_row, col_gemini_verdict).fill = red
        if opus_match is True and gemini_match is False:
            ws.cell(excel_row, col_gemini_verdict).fill = yellow
        if gemini_match is True and opus_match is False:
            ws.cell(excel_row, col_opus_verdict).fill = yellow

    # Wrap reasoning columns
    wrap_cols = ["hebrew_gt", "opus_reasoning", "gemini_reasoning",
                 "opus_temporal", "gemini_temporal"]
    for col_name in wrap_cols:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 60
            for excel_row in range(2, ws.max_row + 1):
                ws.cell(excel_row, col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    # Reasonable widths for other columns
    for col_idx, name in enumerate(headers, start=1):
        if name not in wrap_cols:
            col_letter = get_column_letter(col_idx)
            current = ws.column_dimensions[col_letter].width
            if not current:
                ws.column_dimensions[col_letter].width = max(12, min(25, len(name) + 4))

    wb.save(path)


# ---------------------------------------------------------------------------
# Per-clip evaluation
# ---------------------------------------------------------------------------

def _empty_model_result() -> Dict:
    return {
        "verdict": None, "confidence": None, "reasoning": None,
        "temporal_analysis": None, "time_to_contact": None,
        "full": {}, "raw": "", "usage": {},
        "cost_usd": 0.0, "latency_s": 0.0, "error": None,
    }


def _call_one_model(
    client: OpenAI,
    model_key: str,
    messages: List[Dict],
    timeout: float,
    max_retries: int,
    retry_delay: float,
) -> Dict:
    """Run a single model on prepared messages; return result dict."""
    cfg = MODELS[model_key]
    t0 = time.time()
    try:
        raw, usage = _call_model(
            client, cfg["slug"], messages,
            timeout=timeout, max_retries=max_retries, retry_delay=retry_delay,
        )
        latency = time.time() - t0
        parsed, verdict = _parse_response(raw)
        cost = _calc_cost(usage, cfg["price_in"], cfg["price_out"])
        result = {
            "verdict": verdict,
            "confidence": parsed.get("confidence") if parsed else None,
            "reasoning": parsed.get("verdict_reasoning") if parsed else None,
            "temporal_analysis": parsed.get("temporal_analysis") if parsed else None,
            "time_to_contact": parsed.get("time_to_contact") if parsed else None,
            "full": parsed or {},
            "raw": raw,
            "usage": usage,
            "cost_usd": cost,
            "latency_s": round(latency, 2),
            "error": None,
        }
        print(f"    {cfg['label']:20s}: verdict={verdict} "
              f"words={len((result['reasoning'] or '').split())} "
              f"cost=${cost:.4f} latency={latency:.1f}s", flush=True)
        return result
    except Exception as exc:
        latency = time.time() - t0
        print(f"    {cfg['label']:20s}: ERROR -- {exc}", flush=True)
        return {
            "verdict": None, "confidence": None, "reasoning": None,
            "temporal_analysis": None, "time_to_contact": None,
            "full": {}, "raw": "", "usage": {},
            "cost_usd": 0.0, "latency_s": round(latency, 2),
            "error": str(exc),
        }


def _evaluate_clip(
    client: OpenAI,
    clip: Dict,
    existing_record: Optional[Dict],
    image_b64s: List[str],
    models_to_run: List[str],
    timeout: float,
    max_retries: int,
    retry_delay: float,
) -> Dict:
    """Run requested models on one clip; merge with existing record if present."""
    record = existing_record or {
        "video_id": clip["video_id"],
        "target": clip["target"],
        "gt_verdict": clip["gt_verdict"],
        "t_seconds": clip["t_seconds"],
        "hebrew_gt": clip["hebrew_gt"],
    }
    # Refresh metadata in case it changed
    record.update({
        "video_id": clip["video_id"],
        "target": clip["target"],
        "gt_verdict": clip["gt_verdict"],
        "t_seconds": clip["t_seconds"],
        "hebrew_gt": clip["hebrew_gt"],
    })
    # Ensure both model slots exist
    for key in ("opus", "gemini"):
        if key not in record:
            record[key] = _empty_model_result()

    messages = _build_messages(PROMPT_G2, image_b64s, detail="low")
    for key in models_to_run:
        record[key] = _call_one_model(
            client, key, messages, timeout, max_retries, retry_delay,
        )
    return record


def _load_existing_records(path: Path) -> Dict[str, Dict]:
    """Load existing JSONL into a dict keyed by video_id, or empty dict if missing."""
    out: Dict[str, Dict] = {}
    if not path.exists():
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                out[rec["video_id"]] = rec
            except Exception:
                continue
    return out


def _write_jsonl_full(records_dict: Dict[str, Dict], video_order: List[str], path: Path) -> None:
    """Write the complete current state to JSONL (atomic-ish via tmp file)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        for vid in video_order:
            if vid in records_dict:
                f.write(json.dumps(records_dict[vid], ensure_ascii=False) + "\n")
    tmp.replace(path)


def _has_model_done(rec: Optional[Dict], model_key: str) -> bool:
    """Return True if a record has a successful result for the given model."""
    if rec is None:
        return False
    m = rec.get(model_key)
    if not m:
        return False
    # Successful = has a verdict OR has a parsed reasoning (not just an error)
    return (m.get("verdict") is not None) or bool(m.get("reasoning"))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Teacher Bake-Off: Opus vs Gemini on 18 GT clips")
    parser.add_argument("--gt_xlsx", default=str(DEFAULT_GT_XLSX))
    parser.add_argument("--frames_root", default=str(DEFAULT_FRAMES_ROOT))
    parser.add_argument("--out_jsonl", default=str(DEFAULT_OUT_JSONL))
    parser.add_argument("--out_xlsx", default=str(DEFAULT_OUT_XLSX))
    parser.add_argument("--opus_model", default=MODELS["opus"]["slug"],
                        help="Override Opus model slug if needed")
    parser.add_argument("--gemini_model", default=MODELS["gemini"]["slug"],
                        help="Override Gemini model slug if needed")
    parser.add_argument("--only_model", choices=["opus", "gemini", "both"], default="both",
                        help="Run only one model (resume-friendly) or both")
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT)
    parser.add_argument("--max_retries", type=int, default=DEFAULT_MAX_RETRIES)
    parser.add_argument("--retry_delay", type=float, default=DEFAULT_RETRY_DELAY)
    parser.add_argument("--smoke_test_threshold", type=float, default=0.05,
                        help="Abort if first-clip cost exceeds this (USD)")
    parser.add_argument("--skip_smoke_test", action="store_true",
                        help="Skip the cost guard and run all 18 clips immediately")
    parser.add_argument("--inter_call_delay", type=float, default=1.5)
    args = parser.parse_args()

    # Apply slug overrides
    MODELS["opus"]["slug"] = args.opus_model
    MODELS["gemini"]["slug"] = args.gemini_model

    models_to_run = ["opus", "gemini"] if args.only_model == "both" else [args.only_model]

    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY not set in environment")

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "MMLM_Teacher_Bakeoff"),
        },
    )

    clips = _read_gt_excel(Path(args.gt_xlsx))
    video_order = [c["video_id"] for c in clips]
    print(f"Loaded {len(clips)} clips from {args.gt_xlsx}")
    print(f"Mode: only_model={args.only_model} (will run: {models_to_run})")
    print(f"Models:")
    print(f"  Opus  : {MODELS['opus']['slug']}  (${MODELS['opus']['price_in']}/${MODELS['opus']['price_out']} per 1M)")
    print(f"  Gemini: {MODELS['gemini']['slug']}  (${MODELS['gemini']['price_in']}/${MODELS['gemini']['price_out']} per 1M)")

    frames_root = Path(args.frames_root)
    out_jsonl = Path(args.out_jsonl)

    # ---- LOAD EXISTING (resume) ----
    existing = _load_existing_records(out_jsonl)
    if existing:
        already_opus = sum(1 for r in existing.values() if _has_model_done(r, "opus"))
        already_gemini = sum(1 for r in existing.values() if _has_model_done(r, "gemini"))
        print(f"Resume: existing JSONL has {len(existing)} records "
              f"(opus_done={already_opus}, gemini_done={already_gemini})")
    print()

    # ---- SMOKE TEST: first clip only, abort if too expensive ----
    if not args.skip_smoke_test:
        print("=== SMOKE TEST (1 clip) ===")
        smoke_clip = clips[0]
        # Skip smoke test for models already done on this clip
        smoke_models = [m for m in models_to_run
                        if not _has_model_done(existing.get(smoke_clip["video_id"]), m)]
        if not smoke_models:
            print(f"[smoke] All requested models already done for {smoke_clip['video_id']} -- skipping smoke test\n", flush=True)
        else:
            smoke_indices = _frame_indices(smoke_clip["t_seconds"], FPS, WINDOW_SIZE, STRIDE)
            smoke_b64s = _load_clip_frames(frames_root, smoke_clip["video_id"], smoke_indices, FRAME_SIZE)
            print(f"[smoke] video={smoke_clip['video_id']} target={smoke_clip['target']} "
                  f"gt={smoke_clip['gt_verdict']}  models={smoke_models}", flush=True)
            smoke_rec = _evaluate_clip(
                client, smoke_clip, existing.get(smoke_clip["video_id"]),
                smoke_b64s, smoke_models, args.timeout, args.max_retries, args.retry_delay,
            )
            smoke_cost = sum(smoke_rec[m]["cost_usd"] for m in smoke_models)
            print(f"\n[smoke] Cost for clip 1: ${smoke_cost:.4f}", flush=True)
            if smoke_cost > args.smoke_test_threshold:
                extrapolated = smoke_cost * len(clips)
                print(f"\nABORTED: smoke-test cost ${smoke_cost:.4f} exceeds threshold "
                      f"${args.smoke_test_threshold:.4f}. Full run would cost ~${extrapolated:.2f}.",
                      flush=True)
                print("Re-run with --skip_smoke_test to proceed anyway, or --smoke_test_threshold to raise.",
                      flush=True)
                sys.exit(2)
            existing[smoke_clip["video_id"]] = smoke_rec
            _write_jsonl_full(existing, video_order, out_jsonl)
            print(f"[smoke] OK -- written incrementally to {out_jsonl}\n", flush=True)

    # ---- FULL RUN ----
    for i, clip in enumerate(clips, start=1):
        # Determine which models still need to run for this clip
        rec_existing = existing.get(clip["video_id"])
        models_needed = [m for m in models_to_run if not _has_model_done(rec_existing, m)]
        if not models_needed:
            print(f"[{i}/{len(clips)}] video={clip['video_id']} -- already done for {models_to_run}, skipping",
                  flush=True)
            continue

        if args.inter_call_delay > 0 and i > 1:
            time.sleep(args.inter_call_delay)
        print(f"[{i}/{len(clips)}] video={clip['video_id']} target={clip['target']} "
              f"gt={clip['gt_verdict']}  running={models_needed}", flush=True)
        indices = _frame_indices(clip["t_seconds"], FPS, WINDOW_SIZE, STRIDE)
        b64s = _load_clip_frames(frames_root, clip["video_id"], indices, FRAME_SIZE)
        rec = _evaluate_clip(
            client, clip, rec_existing, b64s, models_needed,
            args.timeout, args.max_retries, args.retry_delay,
        )
        existing[clip["video_id"]] = rec
        # Per-clip incremental save (no data loss on crash)
        _write_jsonl_full(existing, video_order, out_jsonl)

    # ---- FINAL OUTPUTS ----
    records = [existing[v] for v in video_order if v in existing]
    print(f"\nFinal JSONL: {out_jsonl} ({len(records)} records)")

    out_xlsx = Path(args.out_xlsx)
    _write_excel(records, out_xlsx)
    print(f"Wrote {out_xlsx}")

    # ---- SUMMARY ----
    n = len(records)
    opus_correct = sum(1 for r in records if r["opus"]["verdict"] and r["opus"]["verdict"] == r["gt_verdict"])
    gemini_correct = sum(1 for r in records if r["gemini"]["verdict"] and r["gemini"]["verdict"] == r["gt_verdict"])
    opus_words = [len((r["opus"]["reasoning"] or "").split()) for r in records if r["opus"]["reasoning"]]
    gemini_words = [len((r["gemini"]["reasoning"] or "").split()) for r in records if r["gemini"]["reasoning"]]
    opus_len_ok = sum(1 for w in opus_words if w <= 150)
    gemini_len_ok = sum(1 for w in gemini_words if w <= 150)
    opus_cost_total = sum(r["opus"]["cost_usd"] for r in records)
    gemini_cost_total = sum(r["gemini"]["cost_usd"] for r in records)

    print(f"\n=== BAKE-OFF SUMMARY ===")
    print(f"  Opus    verdict accuracy: {opus_correct}/{n} ({opus_correct/n*100:.1f}%)")
    print(f"  Gemini  verdict accuracy: {gemini_correct}/{n} ({gemini_correct/n*100:.1f}%)")
    if opus_words:
        print(f"  Opus    mean reasoning length: {sum(opus_words)/len(opus_words):.0f} words")
    if gemini_words:
        print(f"  Gemini  mean reasoning length: {sum(gemini_words)/len(gemini_words):.0f} words")
    print(f"  Length compliance (<=150 words): Opus={opus_len_ok}/{n}, Gemini={gemini_len_ok}/{n}")
    print(f"  Cost: Opus=${opus_cost_total:.4f}, Gemini=${gemini_cost_total:.4f}, "
          f"Total=${opus_cost_total + gemini_cost_total:.4f}")
    print(f"\nReview side-by-side reasoning vs Hebrew GT in: {out_xlsx}")


if __name__ == "__main__":
    main()
