"""Build results_v6_debate_v11.xlsx and leaderboard_v6_debate_v11.md.

Sources:
  - outputs/prompt_bakeoff/v11_100clips/v6_hires_v11.jsonl   (100 clips)
  - outputs/prompt_bakeoff/v11_100clips/v6_debate_v11.jsonl  (34 debate clips)
  - dataset/teacher_labels/teacher_dataset_v11.xlsx                          (GT + final_reasoning)

Key differences vs. build_debate_outputs.py (18-clip version):
  - No BERTScore columns (no GT reasoning to compare against)
  - Adds color-coded 'passes_for_student_training' column:
      pass_v6     = correct on first prompt          -> green  row
      pass_debate = wrong on v6, fixed by debate     -> orange row
      fail        = wrong after debate (or no debate)-> red    row
  - Summary sheet includes student-training yield counts
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
V11_DIR   = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v11_100clips"
V6_JSONL  = V11_DIR / "v6_hires_v11.jsonl"
DEB_JSONL = V11_DIR / "v6_debate_v11.jsonl"
V11_XLSX  = REPO_ROOT / "outputs" / "teacher_dataset_v11.xlsx"

OUT_XLSX  = V11_DIR / "results_v6_debate_v11.xlsx"
OUT_MD    = V11_DIR / "leaderboard_v6_debate_v11.md"

# ── Colours ──────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")   # pass_v6
ORANGE = PatternFill("solid", fgColor="FFEB9C")   # pass_debate
RED    = PatternFill("solid", fgColor="FFC7CE")   # fail
HEADER = PatternFill("solid", fgColor="4472C4")   # header row
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
# ─────────────────────────────────────────────────────────────────────────────

BUCKET_FILL = {"pass_v6": GREEN, "pass_debate": ORANGE, "fail": RED}


def _load_jsonl(path: Path) -> List[Dict]:
    if not path.exists():
        return []
    out = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            try:
                out.append(json.loads(line))
            except Exception:
                pass
    return out


def _norm_vid(v) -> str:
    s = str(v).strip().split(".")[0]
    return f"{int(s):05d}" if s.isdigit() else s


def _confusion(df: pd.DataFrame, pred_col: str) -> Dict[str, int]:
    tp = int(((df[pred_col] == "YES") & (df["gt_verdict"] == "YES")).sum())
    fp = int(((df[pred_col] == "YES") & (df["gt_verdict"] == "NO")).sum())
    tn = int(((df[pred_col] == "NO")  & (df["gt_verdict"] == "NO")).sum())
    fn = int(((df[pred_col] == "NO")  & (df["gt_verdict"] == "YES")).sum())
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn}


def _apply_sheet_formatting(ws) -> None:
    """Style header row + auto-width + freeze pane."""
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def _color_rows(ws, bucket_col_idx: int) -> None:
    """Color every data row according to its 'passes_for_student_training' value."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        bucket_cell = row[bucket_col_idx - 1]
        fill = BUCKET_FILL.get(str(bucket_cell.value or ""), None)
        if fill:
            for cell in row:
                cell.fill = fill


def main() -> None:
    # ── Load data ─────────────────────────────────────────────────────────────
    v6_recs  = {r["video_id"]: r for r in _load_jsonl(V6_JSONL)}
    deb_recs = {r["video_id"]: r for r in _load_jsonl(DEB_JSONL)}

    gt_df  = pd.read_excel(V11_XLSX)
    gt_map = {}
    for _, row in gt_df.iterrows():
        vid = _norm_vid(row["video_id"])
        gt_map[vid] = {
            "gt_verdict":  str(row["gt_verdict"]).strip(),
            "gt_reasoning": str(row.get("final_reasoning", "")).strip(),
        }

    # ── Build per-clip rows ───────────────────────────────────────────────────
    rows = []
    for vid in sorted(v6_recs):
        h   = v6_recs[vid]
        gt  = gt_map.get(vid, {}).get("gt_verdict") or h.get("gt_verdict", "")
        gt_reasoning = gt_map.get(vid, {}).get("gt_reasoning", "")

        v6_verdict = h.get("verdict")          # may be None if API error
        v6_correct = (v6_verdict == gt)
        v6_reason  = (h.get("reasoning") or "").strip()

        d = deb_recs.get(vid)
        if d is None:
            # No debate — passed v6 or was skipped
            recovery_prompt  = ""
            recovery_verdict = ""
            recovery_correct = ""
            recovery_reason  = ""
            final_verdict    = v6_verdict or ""
            flipped          = ""
        else:
            recovery_prompt  = d.get("recovery_prompt", "")
            recovery_verdict = d.get("recovery_verdict") or ""
            recovery_correct = (recovery_verdict == gt) if recovery_verdict else ""
            recovery_reason  = (d.get("recovery_reasoning") or "").strip()
            final_verdict    = recovery_verdict if recovery_verdict else (v6_verdict or "")
            if not v6_correct and recovery_verdict == gt:
                flipped = "FIXED"
            elif v6_correct and recovery_verdict and recovery_verdict != gt:
                flipped = "BROKE"
            else:
                flipped = "still-wrong"

        final_correct = (final_verdict == gt)

        # Bucket for student training
        if v6_correct:
            bucket = "pass_v6"
        elif flipped == "FIXED":
            bucket = "pass_debate"
        else:
            bucket = "fail"

        rows.append({
            "video_id":                    vid,
            "gt_verdict":                  gt,
            "gt_reasoning":                gt_reasoning,
            "v6_hires__verdict":           v6_verdict or "",
            "v6_hires__correct":           v6_correct,
            "v6_hires__reasoning":         v6_reason,
            "recovery_prompt":             recovery_prompt,
            "recovery__verdict":           recovery_verdict,
            "recovery__correct":           recovery_correct,
            "recovery__reasoning":         recovery_reason,
            "final_after_debate":          final_verdict,
            "final_correct":               final_correct,
            "flipped_by_debate":           flipped,
            "passes_for_student_training": bucket,
        })

    df = pd.DataFrame(rows)
    n  = len(df)

    # ── Metrics ───────────────────────────────────────────────────────────────
    # v6 accuracy: count non-None predictions only
    v6_valid   = df["v6_hires__verdict"].isin(["YES", "NO"])
    v6_correct_n = int(df.loc[v6_valid, "v6_hires__correct"].sum())
    v6_total     = int(v6_valid.sum())
    v6_acc       = v6_correct_n / v6_total if v6_total else 0.0

    final_correct_n = int(df["final_correct"].sum())
    final_acc       = final_correct_n / n if n else 0.0

    # Temp column for confusion on final
    df["_final_for_cm"] = df["final_after_debate"].apply(
        lambda x: x if x in ("YES", "NO") else None
    )
    cm_v6    = _confusion(df[v6_valid], "v6_hires__verdict")
    cm_final = _confusion(df[df["_final_for_cm"].notna()], "_final_for_cm")
    df.drop(columns=["_final_for_cm"], inplace=True)

    n_debated    = int((df["recovery_prompt"] != "").sum())
    n_fixed      = int((df["flipped_by_debate"] == "FIXED").sum())
    n_broke      = int((df["flipped_by_debate"] == "BROKE").sum())
    n_still_wrong = int((df["flipped_by_debate"] == "still-wrong").sum())
    fn_recovered = int(((df["flipped_by_debate"] == "FIXED") &
                        (df["recovery_prompt"] == "PROMPT_G_OPT_v6_TP_RECOVERY")).sum())
    fp_recovered = int(((df["flipped_by_debate"] == "FIXED") &
                        (df["recovery_prompt"] == "PROMPT_G_OPT_v6_TN_RECOVERY")).sum())

    n_pass_v6     = int((df["passes_for_student_training"] == "pass_v6").sum())
    n_pass_debate = int((df["passes_for_student_training"] == "pass_debate").sum())
    n_fail        = int((df["passes_for_student_training"] == "fail").sum())
    n_valid_train = n_pass_v6 + n_pass_debate

    summary_rows = [
        {"metric": "n_clips",                            "value": n},
        {"metric": "v6@hires accuracy",                  "value": f"{v6_acc:.1%}  ({v6_correct_n}/{v6_total})"},
        {"metric": "after-debate accuracy",              "value": f"{final_acc:.1%}  ({final_correct_n}/{n})"},
        {"metric": "v6@hires confusion (TP/FP/TN/FN)",  "value": f"{cm_v6['TP']}/{cm_v6['FP']}/{cm_v6['TN']}/{cm_v6['FN']}"},
        {"metric": "after-debate confusion (TP/FP/TN/FN)","value": f"{cm_final['TP']}/{cm_final['FP']}/{cm_final['TN']}/{cm_final['FN']}"},
        {"metric": "clips debated",                      "value": n_debated},
        {"metric": "debate FIXED (wrong->right)",         "value": n_fixed},
        {"metric": "debate BROKE (right->wrong)",        "value": n_broke},
        {"metric": "debate still-wrong",                 "value": n_still_wrong},
        {"metric": "FN recovered by TP_RECOVERY",        "value": fn_recovered},
        {"metric": "FP recovered by TN_RECOVERY",        "value": fp_recovered},
        {"metric": "--- Student training yield ---",      "value": ""},
        {"metric": "pass_v6 (correct on first pass)",     "value": n_pass_v6},
        {"metric": "pass_debate (fixed by recovery)",     "value": n_pass_debate},
        {"metric": "fail (wrong after debate)",           "value": n_fail},
        {"metric": "TOTAL valid for student training",    "value": n_valid_train},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # ── Write xlsx ────────────────────────────────────────────────────────────
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="per_clip", index=False)
        df_summary.to_excel(w, sheet_name="summary", index=False)

    # Now open with openpyxl to apply colors
    wb = load_workbook(OUT_XLSX)

    # per_clip sheet: color rows + header
    ws_clip = wb["per_clip"]
    _apply_sheet_formatting(ws_clip)
    bucket_col_idx = df.columns.get_loc("passes_for_student_training") + 1
    _color_rows(ws_clip, bucket_col_idx)

    # summary sheet: header only
    ws_sum = wb["summary"]
    _apply_sheet_formatting(ws_sum)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")

    # ── Console summary ───────────────────────────────────────────────────────
    print()
    print(df_summary.to_string(index=False))

    # ── Write leaderboard.md ──────────────────────────────────────────────────
    md = []
    md.append("# Leaderboard — PROMPT_G_OPT_v6_balanced @ Hi-Res + Debate Recovery (100 clips)\n")
    md.append("**Setup:** 100 clips from `teacher_dataset_v11.xlsx` (50 YES / 50 NO), "
              "frames @ NATIVE 1280×720, Gemini `detail=\"high\"`, "
              "`temperature=0.1`, 16-frame window, stride=4.\n")
    md.append("**Debate:** Every clip where v6 disagreed with GT was retested with a "
              "targeted recovery prompt "
              "(TP_RECOVERY for FN cases, TN_RECOVERY for FP cases).\n")
    md.append("")

    md.append("## Headline metrics\n")
    md.append("| Stage | Accuracy | TP | FP | TN | FN | n |")
    md.append("|-------|----------|----|----|----|----|---|")
    md.append(f"| v6@hires (single-pass) | **{v6_acc:.1%}** ({v6_correct_n}/{v6_total}) "
              f"| {cm_v6['TP']} | {cm_v6['FP']} | {cm_v6['TN']} | {cm_v6['FN']} | {v6_total} |")
    md.append(f"| after debate | **{final_acc:.1%}** ({final_correct_n}/{n}) "
              f"| {cm_final['TP']} | {cm_final['FP']} | {cm_final['TN']} | {cm_final['FN']} | {n} |")
    md.append("")

    md.append("## Debate yield\n")
    md.append(f"- Clips debated: **{n_debated}**")
    md.append(f"- FIXED (wrong→right): **{n_fixed}**")
    md.append(f"  - FN recovered by `TP_RECOVERY`: {fn_recovered} / {cm_v6['FN']} FN clips")
    md.append(f"  - FP recovered by `TN_RECOVERY`: {fp_recovered} / {cm_v6['FP']} FP clips")
    md.append(f"- BROKE (right→wrong): **{n_broke}**")
    md.append(f"- Still-wrong after debate: **{n_still_wrong}**")
    md.append("")

    md.append("## Student-training yield\n")
    md.append(f"| Bucket | Count | Meaning |")
    md.append(f"|--------|-------|---------|")
    md.append(f"| 🟢 pass_v6 | **{n_pass_v6}** | Correct on first pass — high-confidence teacher label |")
    md.append(f"| 🟠 pass_debate | **{n_pass_debate}** | Fixed by recovery — usable but lower confidence |")
    md.append(f"| 🔴 fail | **{n_fail}** | Wrong after debate — exclude from training |")
    md.append(f"| **✅ Total valid** | **{n_valid_train} / {n}** | Available for student distillation |")
    md.append("")

    md.append("## Per-clip table\n")
    md.append("| Clip | GT | v6@hires | recovery | final | flip | bucket |")
    md.append("|------|----|----------|----------|-------|------|--------|")
    for _, row in df.iterrows():
        rv   = row["recovery__verdict"] or "—"
        rp   = row["recovery_prompt"] or ""
        rp_s = rp.replace("PROMPT_G_OPT_v6_", "").replace("_RECOVERY", "") if rp else ""
        rec_cell  = f"{rv} ({rp_s})" if rp else "—"
        flip      = row["flipped_by_debate"] or "—"
        v6_mark   = "✓" if row["v6_hires__correct"] else "✗"
        fin_mark  = "✓" if row["final_correct"] else "✗"
        bkt_emoji = {"pass_v6": "🟢", "pass_debate": "🟠", "fail": "🔴"}.get(
            row["passes_for_student_training"], "")
        md.append(
            f"| {row['video_id']} | {row['gt_verdict']} "
            f"| {row['v6_hires__verdict'] or '??'} {v6_mark} "
            f"| {rec_cell} "
            f"| {row['final_after_debate']} {fin_mark} "
            f"| {flip} "
            f"| {bkt_emoji} {row['passes_for_student_training']} |"
        )
    md.append("")

    md.append("## Files\n")
    md.append(f"- Per-clip data: `outputs/prompt_bakeoff/v11_100clips/{OUT_XLSX.name}`")
    md.append(f"- Leaderboard: `outputs/prompt_bakeoff/v11_100clips/{OUT_MD.name}`")
    md.append(f"- v6 records: `outputs/prompt_bakeoff/v11_100clips/v6_hires_v11.jsonl`")
    md.append(f"- Debate records: `outputs/prompt_bakeoff/v11_100clips/v6_debate_v11.jsonl`")
    md.append("")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")
    print(f"\nWrote {OUT_MD}")


if __name__ == "__main__":
    main()
