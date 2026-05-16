"""Teacher Prompt Bake-Off: 5 prompts on the 18 GT clips, Gemini 3.1 Pro only.

Compares: PROMPT_G, PROMPT_G2, PROMPT_G_BASE, PROMPT_G_CRITIC, PROMPT_G_OPT
Metrics:  verdict accuracy + confusion matrix + BERTScore F1 + length compliance
          + composite score (0.30 verdict + 0.45 alignment + 0.25 length, per apo_metric)

Outputs:
    outputs/prompt_bakeoff/results.jsonl    -- per-(clip,prompt) records (resumable)
    outputs/prompt_bakeoff/results.xlsx     -- 3 sheets: per_clip, summary, failures
    outputs/prompt_bakeoff/leaderboard.md   -- ranked summary

Usage:
    py teacher_distillation/scripts/teacher_prompt_bakeoff.py
    py teacher_distillation/scripts/teacher_prompt_bakeoff.py --skip_smoke_test
    py teacher_distillation/scripts/teacher_prompt_bakeoff.py --prompts PROMPT_G,PROMPT_G_OPT
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Path setup --------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "teacher_distillation" / "scripts"))

# Reuse heavy lifting from teacher_bakeoff (frame loading, OpenRouter calls, JSON parse)
from teacher_bakeoff import (  # noqa: E402
    FPS,
    FRAME_SIZE,
    STRIDE,
    WINDOW_SIZE,
    _build_messages,
    _calc_cost,
    _call_model,
    _frame_indices,
    _load_clip_frames,
    _normalize_video_id,
    _parse_response,
)

# Reuse scoring from apo_metric (BERTScore + composite)
from apo_metric import score_one, warmup_bertscore  # noqa: E402

# Reuse safety pre-run credit check
from apo_safety import assert_sufficient_credit  # noqa: E402

from openai import OpenAI  # noqa: E402


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MODEL_SLUG = "google/gemini-3.1-pro-preview"
MODEL_LABEL = "Gemini 3.1 Pro"
PRICE_IN = 2.00   # USD / 1M input tokens (verified via apo_safety pre-run query)
PRICE_OUT = 12.00  # USD / 1M output tokens

TEMPERATURE = 0.1
MAX_TOKENS = 6000  # PROMPT_G2's 8-step output needs headroom
                   # NOTE: not passed to _call_model (reused from teacher_bakeoff).
                   # Gemini's OpenRouter default (~8192 out) is sufficient.
DEFAULT_TIMEOUT = 120.0
DEFAULT_MAX_RETRIES = 3
DEFAULT_RETRY_DELAY = 3.0
DEFAULT_INTER_CALL_DELAY = 1.5

# Prompt registry: name -> (filename in prompts/, variable name inside that file)
PROMPT_REGISTRY: Dict[str, tuple] = {
    "PROMPT_G":        ("PROMPT_G.py",        "PROMPT_G"),
    "PROMPT_G2":       ("PROMPT_G2.py",       "PROMPT_G"),   # this file's variable is PROMPT_G
    "PROMPT_G_BASE":   ("PROMPT_G_BASE.py",   "PROMPT_G_BASE"),
    "PROMPT_G_CRITIC": ("PROMPT_G_CRITIC.py", "PROMPT_G_CRITIC"),
    "PROMPT_G_OPT":    ("PROMPT_G_OPT.py",    "PROMPT_G_OPT"),
    "PROMPT_G_OPT_v2":    ("PROMPT_G_OPT.py",    "PROMPT_G_OPT"),     # v2: removed 150-token constraint (restored since)
    "PROMPT_G_OPT_rerun": ("PROMPT_G_OPT.py",    "PROMPT_G_OPT"),     # rerun: same as original (150 tok), determinism check
    "PROMPT_G_OPT_v3":    ("PROMPT_G_OPT_v3.py", "PROMPT_G_OPT_v3"),  # v3: 120-token constraint
    "PROMPT_G_OPT_v4":    ("PROMPT_G_OPT_v4.py", "PROMPT_G_OPT_v4"),  # v4: 180-token constraint
    "PROMPT_G_OPT_v5_balanced": ("PROMPT_G_OPT_v5_balanced.py", "PROMPT_G_OPT_v5_balanced"),  # v5: symmetric + GATES A/B/C
    "PROMPT_G_OPT_v6_balanced": ("PROMPT_G_OPT_v6_balanced.py", "PROMPT_G_OPT_v6_balanced"),  # v6: base-rate first, ambiguous→NO
}

DEFAULT_GT_XLSX = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
DEFAULT_FRAMES_ROOT = REPO_ROOT / "dataset" / "train"
OUT_DIR = REPO_ROOT / "outputs" / "prompt_bakeoff"
DEFAULT_OUT_JSONL = OUT_DIR / "results.jsonl"
DEFAULT_OUT_XLSX = OUT_DIR / "results.xlsx"
DEFAULT_LEADERBOARD = OUT_DIR / "leaderboard.md"


# ---------------------------------------------------------------------------
# Prompt loading
# ---------------------------------------------------------------------------

def _load_prompt(filename: str, variable: str) -> str:
    """Import a prompt module via importlib and return its prompt-string constant."""
    path = REPO_ROOT / "prompts" / filename
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    spec = importlib.util.spec_from_file_location(f"_prompt_{path.stem}", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not import {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    if not hasattr(module, variable):
        raise RuntimeError(f"{filename} does not export variable '{variable}'")
    prompt = getattr(module, variable)
    if not isinstance(prompt, str):
        raise RuntimeError(f"{filename}.{variable} is not a string (got {type(prompt)})")
    return prompt


# ---------------------------------------------------------------------------
# GT loading (extends teacher_bakeoff's reader to include verdict_reasoning_en)
# ---------------------------------------------------------------------------

def _read_gt_excel_with_en(path: Path) -> List[Dict]:
    """Read the 18-clip GT xlsx, including the verdict_reasoning_en column (col G)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["video_id", "target", "gt_verdict", "t_seconds",
                "verdict_reasoning", "verdict_reasoning_en"]
    idx = {}
    for name in required:
        if name not in header:
            raise RuntimeError(f"Column '{name}' missing from {path}. Found: {header}")
        idx[name] = header.index(name)

    rows: List[Dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["video_id"]] is None:
            continue
        rows.append({
            "video_id": _normalize_video_id(row[idx["video_id"]]),
            "target": int(row[idx["target"]]) if row[idx["target"]] is not None else None,
            "gt_verdict": str(row[idx["gt_verdict"]]).strip().upper() if row[idx["gt_verdict"]] else None,
            "t_seconds": float(row[idx["t_seconds"]]) if row[idx["t_seconds"]] is not None else None,
            "hebrew_gt": row[idx["verdict_reasoning"]] or "",
            "gt_reasoning_en": (row[idx["verdict_reasoning_en"]] or "").strip(),
        })
    return rows


# ---------------------------------------------------------------------------
# Resume / persistence
# ---------------------------------------------------------------------------

def _record_key(video_id: str, prompt_name: str) -> str:
    return f"{video_id}::{prompt_name}"


def _load_existing(path: Path) -> Dict[str, Dict]:
    out: Dict[str, Dict] = {}
    if not path.exists():
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                key = _record_key(rec["video_id"], rec["prompt_name"])
                out[key] = rec
            except Exception:
                continue
    return out


def _write_jsonl_full(records: Dict[str, Dict], path: Path,
                      clip_order: List[str], prompt_order: List[str]) -> None:
    """Write ALL records to JSONL, not just the current prompt_order.

    Iterates clip_order x all_prompt_names (discovered from records dict)
    so partial runs (--prompts subset) don't drop other prompts' data.
    """
    # Discover all prompt names present in records
    all_prompts_in_records = sorted(set(
        r["prompt_name"] for r in records.values() if "prompt_name" in r
    ))
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        for vid in clip_order:
            for pname in all_prompts_in_records:
                k = _record_key(vid, pname)
                if k in records:
                    f.write(json.dumps(records[k], ensure_ascii=False) + "\n")
    tmp.replace(path)


def _is_done(rec: Optional[Dict]) -> bool:
    """A record is 'done' if it has either a parsed verdict or a parsed reasoning."""
    if rec is None:
        return False
    return (rec.get("verdict") is not None) or bool(rec.get("reasoning"))


# ---------------------------------------------------------------------------
# Per-(clip, prompt) evaluation
# ---------------------------------------------------------------------------

def _evaluate_one(
    client: OpenAI,
    prompt_name: str,
    prompt_text: str,
    clip: Dict,
    image_b64s: List[str],
    timeout: float,
    max_retries: int,
    retry_delay: float,
) -> Dict:
    """Call Gemini with one prompt on one clip; parse + score."""
    messages = _build_messages(prompt_text, image_b64s, detail="low")
    t0 = time.time()
    try:
        raw, usage = _call_model(
            client, MODEL_SLUG, messages,
            timeout=timeout, max_retries=max_retries, retry_delay=retry_delay,
            temperature=TEMPERATURE,
        )
        latency = time.time() - t0
        parsed, verdict = _parse_response(raw)
        cost = _calc_cost(usage, PRICE_IN, PRICE_OUT)
        reasoning = parsed.get("verdict_reasoning") if parsed else None

        # Score with BERTScore + composite (apo_metric)
        sb = score_one(verdict, reasoning, clip["gt_verdict"], clip["gt_reasoning_en"])

        rec = {
            "video_id": clip["video_id"],
            "prompt_name": prompt_name,
            "gt_verdict": clip["gt_verdict"],
            "target": clip["target"],
            "t_seconds": clip["t_seconds"],
            "verdict": verdict,
            "reasoning": reasoning,
            "confidence": parsed.get("confidence") if parsed else None,
            "temporal_analysis": parsed.get("temporal_analysis") if parsed else None,
            "time_to_contact": parsed.get("time_to_contact") if parsed else None,
            "full_json": parsed or {},
            "scores": sb.to_dict(),
            "raw": raw,
            "usage": usage,
            "cost_usd": cost,
            "latency_s": round(latency, 2),
            "error": None,
        }
        verdict_ok = "[OK]" if verdict == clip["gt_verdict"] else "[XX]"
        print(
            f"    {prompt_name:18s} | verdict={verdict or '??':3s} {verdict_ok} | "
            f"BERT={sb.alignment:.3f} | words={sb.word_count:3d} | "
            f"composite={sb.composite:.3f} | cost=${cost:.4f} | {latency:.1f}s",
            flush=True,
        )
        return rec
    except Exception as exc:
        latency = time.time() - t0
        print(f"    {prompt_name:18s} | ERROR: {exc}", flush=True)
        return {
            "video_id": clip["video_id"],
            "prompt_name": prompt_name,
            "gt_verdict": clip["gt_verdict"],
            "target": clip["target"],
            "t_seconds": clip["t_seconds"],
            "verdict": None, "reasoning": None,
            "confidence": None, "temporal_analysis": None, "time_to_contact": None,
            "full_json": {},
            "scores": {"composite": 0.0, "verdict": 0.0, "alignment": 0.0,
                       "length": 0.0, "word_count": 0},
            "raw": "", "usage": {}, "cost_usd": 0.0,
            "latency_s": round(latency, 2),
            "error": str(exc),
        }


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def _confusion_matrix(records: List[Dict]) -> Dict[str, int]:
    """Compute TP/FP/FN/TN treating None verdict as 'wrong'."""
    tp = fp = fn = tn = parse_err = 0
    for r in records:
        gt = r["gt_verdict"]
        pred = r["verdict"]
        if pred is None:
            parse_err += 1
            # Count as wrong: misses on positives become FN, on negatives become FP
            if gt == "YES":
                fn += 1
            else:
                fp += 1
            continue
        if gt == "YES" and pred == "YES":
            tp += 1
        elif gt == "YES" and pred == "NO":
            fn += 1
        elif gt == "NO" and pred == "YES":
            fp += 1
        elif gt == "NO" and pred == "NO":
            tn += 1
    return {"TP": tp, "FP": fp, "FN": fn, "TN": tn, "parse_err": parse_err}


def _summarize(records_by_prompt: Dict[str, List[Dict]]) -> List[Dict]:
    rows = []
    for prompt_name, recs in records_by_prompt.items():
        n = len(recs)
        cm = _confusion_matrix(recs)
        acc = (cm["TP"] + cm["TN"]) / n if n else 0.0
        bert = sum(r["scores"]["alignment"] for r in recs) / n if n else 0.0
        comp = sum(r["scores"]["composite"] for r in recs) / n if n else 0.0
        wc = sum(r["scores"]["word_count"] for r in recs) / n if n else 0.0
        len_ok = sum(1 for r in recs if r["scores"]["length"] >= 1.0)
        cost = sum(r["cost_usd"] for r in recs)
        rows.append({
            "prompt": prompt_name,
            "n": n,
            "TP": cm["TP"], "FP": cm["FP"], "FN": cm["FN"], "TN": cm["TN"],
            "parse_err": cm["parse_err"],
            "accuracy": round(acc, 4),
            "bertscore_mean": round(bert, 4),
            "composite_mean": round(comp, 4),
            "mean_word_count": round(wc, 1),
            "length_ok": f"{len_ok}/{n}",
            "cost_usd": round(cost, 4),
        })
    rows.sort(key=lambda r: (r["composite_mean"], r["accuracy"]), reverse=True)
    return rows


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def _write_excel(
    records: Dict[str, Dict],
    clips: List[Dict],
    prompt_order: List[str],
    summary: List[Dict],
    out_path: Path,
) -> None:
    """Write a 3-sheet Excel: per_clip, summary, failures."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # ---- Sheet 1: per_clip ----
    per_clip_rows = []
    for clip in clips:
        row = {
            "video_id": clip["video_id"],
            "target": clip["target"],
            "gt_verdict": clip["gt_verdict"],
            "t_seconds": clip["t_seconds"],
            "gt_reasoning_en": clip["gt_reasoning_en"],
        }
        for pname in prompt_order:
            r = records.get(_record_key(clip["video_id"], pname), {})
            row[f"{pname}__verdict"] = r.get("verdict")
            row[f"{pname}__BERT"] = round(r.get("scores", {}).get("alignment", 0.0), 4)
            row[f"{pname}__composite"] = round(r.get("scores", {}).get("composite", 0.0), 4)
            row[f"{pname}__words"] = r.get("scores", {}).get("word_count", 0)
            row[f"{pname}__reasoning"] = r.get("reasoning")
        per_clip_rows.append(row)
    df_per_clip = pd.DataFrame(per_clip_rows)

    # ---- Sheet 2: summary ----
    df_summary = pd.DataFrame(summary)

    # ---- Sheet 3: failures ----
    failure_rows = []
    for clip in clips:
        verdicts = {
            p: records.get(_record_key(clip["video_id"], p), {}).get("verdict")
            for p in prompt_order
        }
        any_wrong = any(v != clip["gt_verdict"] for v in verdicts.values())
        if not any_wrong:
            continue
        row = {
            "video_id": clip["video_id"],
            "gt_verdict": clip["gt_verdict"],
            "gt_reasoning_en": clip["gt_reasoning_en"],
        }
        for p in prompt_order:
            r = records.get(_record_key(clip["video_id"], p), {})
            row[f"{p}__verdict"] = r.get("verdict")
            row[f"{p}__correct"] = (r.get("verdict") == clip["gt_verdict"])
            row[f"{p}__reasoning"] = r.get("reasoning")
        failure_rows.append(row)
    df_failures = pd.DataFrame(failure_rows) if failure_rows else pd.DataFrame(
        columns=["(no failures - all prompts correct on all clips)"]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_per_clip.to_excel(writer, sheet_name="per_clip", index=False)
        df_summary.to_excel(writer, sheet_name="summary", index=False)
        df_failures.to_excel(writer, sheet_name="failures", index=False)

    # Style: color verdict-wrong cells red on per_clip; wrap reasoning columns
    wb = load_workbook(out_path)
    red = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    green = PatternFill(start_color="FFBBFFBB", end_color="FFBBFFBB", fill_type="solid")
    bold = Font(bold=True)

    # per_clip styling
    ws = wb["per_clip"]
    headers = [c.value for c in ws[1]]
    for h in headers:
        ws.cell(1, headers.index(h) + 1).font = bold
    gt_col_idx = headers.index("gt_verdict") + 1
    for pname in prompt_order:
        v_col = headers.index(f"{pname}__verdict") + 1
        for r_idx in range(2, ws.max_row + 1):
            v = ws.cell(r_idx, v_col).value
            gt = ws.cell(r_idx, gt_col_idx).value
            if v is None:
                continue
            if v == gt:
                ws.cell(r_idx, v_col).fill = green
            else:
                ws.cell(r_idx, v_col).fill = red
    # Width + wrap for reasoning + gt_reasoning_en
    wrap_cols = [h for h in headers if h.endswith("__reasoning")] + ["gt_reasoning_en"]
    for col_name in wrap_cols:
        if col_name in headers:
            c_idx = headers.index(col_name) + 1
            ws.column_dimensions[get_column_letter(c_idx)].width = 55
            for r_idx in range(2, ws.max_row + 1):
                ws.cell(r_idx, c_idx).alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, name in enumerate(headers, start=1):
        if name not in wrap_cols:
            cl = get_column_letter(col_idx)
            if not ws.column_dimensions[cl].width:
                ws.column_dimensions[cl].width = max(12, min(22, len(str(name)) + 4))

    # summary styling
    ws2 = wb["summary"]
    headers2 = [c.value for c in ws2[1]]
    for h in headers2:
        ws2.cell(1, headers2.index(h) + 1).font = bold
    for col_idx in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    # failures styling
    ws3 = wb["failures"]
    headers3 = [c.value for c in ws3[1]]
    if "gt_verdict" in headers3:
        for h in headers3:
            ws3.cell(1, headers3.index(h) + 1).font = bold
        wrap3 = [h for h in headers3 if h.endswith("__reasoning")] + ["gt_reasoning_en"]
        for col_name in wrap3:
            if col_name in headers3:
                c_idx = headers3.index(col_name) + 1
                ws3.column_dimensions[get_column_letter(c_idx)].width = 55
                for r_idx in range(2, ws3.max_row + 1):
                    ws3.cell(r_idx, c_idx).alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, name in enumerate(headers3, start=1):
            if name not in wrap3:
                cl = get_column_letter(col_idx)
                if not ws3.column_dimensions[cl].width:
                    ws3.column_dimensions[cl].width = max(12, min(22, len(str(name)) + 4))

    wb.save(out_path)


def _write_leaderboard(summary: List[Dict], out_path: Path, total_cost: float) -> None:
    """Write a Markdown leaderboard with rankings + caveats."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("# Teacher Prompt Bake-Off — Leaderboard\n")
    lines.append(f"**Model:** {MODEL_LABEL} (`{MODEL_SLUG}`)  ")
    lines.append(f"**Temperature:** {TEMPERATURE}  ")
    lines.append(f"**Clips:** 18 (Hebrew GT translated, col G `verdict_reasoning_en`)  ")
    lines.append(f"**Total cost:** ${total_cost:.2f}  ")
    lines.append(f"**Ranking key:** `composite_mean` (0.30 verdict + 0.45 BERTScore + 0.25 length)\n")

    lines.append("## Ranking\n")
    lines.append("| Rank | Prompt | Acc | TP/FP/FN/TN | BERT F1 | Composite | Words (mean) | Length OK | Cost |")
    lines.append("|------|--------|-----|-------------|---------|-----------|--------------|-----------|------|")
    for i, row in enumerate(summary, 1):
        lines.append(
            f"| {i} | **{row['prompt']}** | "
            f"{row['accuracy']*100:.1f}% ({row['TP'] + row['TN']}/{row['n']}) | "
            f"{row['TP']}/{row['FP']}/{row['FN']}/{row['TN']} | "
            f"{row['bertscore_mean']:.3f} | "
            f"**{row['composite_mean']:.3f}** | "
            f"{row['mean_word_count']:.0f} | "
            f"{row['length_ok']} | "
            f"${row['cost_usd']:.2f} |"
        )
    lines.append("")

    lines.append("## Caveats\n")
    lines.append("1. **Small sample.** 18 clips → 1-clip swing = 5.6%. "
                 "Differences smaller than ~11% (2 clips) are within noise.")
    lines.append("2. **No Pass-2 debate.** This isolates the prompt's Pass-1 quality; "
                 "the production pipeline still uses debate to correct mismatches.")
    lines.append("3. **BERTScore against translated GT.** Translation drift can confound "
                 "the alignment score (same caveat as APO).")
    lines.append("4. **YES-bias known.** PROMPT_G historically over-predicts YES on TN clips. "
                 "Check the FP column, not just bottom-line accuracy.")
    lines.append("")

    lines.append("## Files")
    lines.append(f"- Raw records: `outputs/prompt_bakeoff/results.jsonl`")
    lines.append(f"- Excel (3 sheets): `outputs/prompt_bakeoff/results.xlsx`")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Teacher Prompt Bake-Off (5 prompts, Gemini 3.1 Pro)")
    parser.add_argument("--gt_xlsx", default=str(DEFAULT_GT_XLSX))
    parser.add_argument("--frames_root", default=str(DEFAULT_FRAMES_ROOT))
    parser.add_argument("--out_jsonl", default=str(DEFAULT_OUT_JSONL))
    parser.add_argument("--out_xlsx", default=str(DEFAULT_OUT_XLSX))
    parser.add_argument("--out_leaderboard", default=str(DEFAULT_LEADERBOARD))
    parser.add_argument("--prompts", default=",".join(PROMPT_REGISTRY.keys()),
                        help="Comma-separated prompt names from the registry")
    parser.add_argument("--max_clips", type=int, default=0,
                        help="If >0, run only the first N clips (debug)")
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT)
    parser.add_argument("--max_retries", type=int, default=DEFAULT_MAX_RETRIES)
    parser.add_argument("--retry_delay", type=float, default=DEFAULT_RETRY_DELAY)
    parser.add_argument("--inter_call_delay", type=float, default=DEFAULT_INTER_CALL_DELAY)
    parser.add_argument("--smoke_test_threshold", type=float, default=0.10,
                        help="Abort if first (clip,prompt) cost exceeds this (USD)")
    parser.add_argument("--skip_smoke_test", action="store_true")
    parser.add_argument("--skip_credit_check", action="store_true")
    parser.add_argument("--projected_cost", type=float, default=2.50,
                        help="Projected total cost for the pre-run credit check")
    args = parser.parse_args()

    # ---- Resolve prompts ----
    prompt_names = [p.strip() for p in args.prompts.split(",") if p.strip()]
    for p in prompt_names:
        if p not in PROMPT_REGISTRY:
            raise SystemExit(f"Unknown prompt: {p}. Valid: {list(PROMPT_REGISTRY)}")
    prompts: Dict[str, str] = {}
    print(f"Loading {len(prompt_names)} prompts...")
    for name in prompt_names:
        filename, var = PROMPT_REGISTRY[name]
        prompts[name] = _load_prompt(filename, var)
        print(f"  {name:18s} <- prompts/{filename}::{var} ({len(prompts[name])} chars)")
    print()

    # ---- Environment / client ----
    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise SystemExit("OPENROUTER_API_KEY not set in environment")

    # ---- Pre-run credit check ----
    if not args.skip_credit_check:
        assert_sufficient_credit(api_key, args.projected_cost, safety_margin=1.5, label="bakeoff/pre-run")
        print()

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "MMLM_Prompt_Bakeoff"),
        },
    )

    # ---- Clips ----
    clips = _read_gt_excel_with_en(Path(args.gt_xlsx))
    if args.max_clips and args.max_clips > 0:
        clips = clips[: args.max_clips]
    clip_order = [c["video_id"] for c in clips]
    print(f"Loaded {len(clips)} clips from {args.gt_xlsx}")
    print(f"Model: {MODEL_SLUG} @ temp={TEMPERATURE}")
    print(f"Prompts to test: {prompt_names}\n")

    # ---- Warmup BERTScore model (one-time) ----
    print("Warming up BERTScore model (downloads ~1.4GB on first use)...")
    warmup_bertscore()
    print("BERTScore ready.\n")

    # ---- Resume existing JSONL ----
    out_jsonl = Path(args.out_jsonl)
    existing = _load_existing(out_jsonl)
    if existing:
        done_count = sum(1 for r in existing.values() if _is_done(r))
        total_targets = len(clips) * len(prompt_names)
        print(f"Resume: found {len(existing)} existing records "
              f"({done_count}/{total_targets} (clip,prompt) pairs done)\n")

    # ---- Smoke test: first (clip, first-prompt) only ----
    if not args.skip_smoke_test:
        print("=== SMOKE TEST (1 clip × 1 prompt) ===")
        sclip = clips[0]
        spname = prompt_names[0]
        key = _record_key(sclip["video_id"], spname)
        if _is_done(existing.get(key)):
            print(f"[smoke] {sclip['video_id']} x {spname} already done -- skipping smoke test\n")
        else:
            indices = _frame_indices(sclip["t_seconds"], FPS, WINDOW_SIZE, STRIDE)
            b64s = _load_clip_frames(Path(args.frames_root), sclip["video_id"], indices, FRAME_SIZE)
            rec = _evaluate_one(client, spname, prompts[spname], sclip, b64s,
                                args.timeout, args.max_retries, args.retry_delay)
            cost = rec["cost_usd"]
            print(f"\n[smoke] Cost: ${cost:.4f}")
            if cost > args.smoke_test_threshold:
                projected = cost * len(clips) * len(prompt_names)
                print(f"\nABORTED: smoke-test cost ${cost:.4f} > threshold "
                      f"${args.smoke_test_threshold:.4f}. Full run would cost ~${projected:.2f}.")
                print("Re-run with --skip_smoke_test or raise --smoke_test_threshold.")
                sys.exit(2)
            existing[key] = rec
            _write_jsonl_full(existing, out_jsonl, clip_order, prompt_names)
            print(f"[smoke] OK -- written to {out_jsonl}\n")

    # ---- Full run: clip-major loop (so frames are loaded once per clip) ----
    total_pairs = len(clips) * len(prompt_names)
    pair_idx = 0
    for clip in clips:
        # Load frames once per clip (reused across prompts)
        indices = _frame_indices(clip["t_seconds"], FPS, WINDOW_SIZE, STRIDE)
        b64s_loaded = False
        b64s: List[str] = []
        for pname in prompt_names:
            pair_idx += 1
            key = _record_key(clip["video_id"], pname)
            if _is_done(existing.get(key)):
                print(f"[{pair_idx}/{total_pairs}] {clip['video_id']} x {pname} -- already done")
                continue
            print(f"[{pair_idx}/{total_pairs}] video={clip['video_id']} target={clip['target']} "
                  f"gt={clip['gt_verdict']}  prompt={pname}")
            if not b64s_loaded:
                b64s = _load_clip_frames(Path(args.frames_root), clip["video_id"], indices, FRAME_SIZE)
                b64s_loaded = True
            if args.inter_call_delay > 0:
                time.sleep(args.inter_call_delay)
            rec = _evaluate_one(client, pname, prompts[pname], clip, b64s,
                                args.timeout, args.max_retries, args.retry_delay)
            existing[key] = rec
            _write_jsonl_full(existing, out_jsonl, clip_order, prompt_names)

    # ---- Aggregate & write summary outputs ----
    records_by_prompt: Dict[str, List[Dict]] = {p: [] for p in prompt_names}
    for clip in clips:
        for p in prompt_names:
            rec = existing.get(_record_key(clip["video_id"], p))
            if rec is not None:
                records_by_prompt[p].append(rec)

    summary = _summarize(records_by_prompt)
    total_cost = sum(row["cost_usd"] for row in summary)

    out_xlsx = Path(args.out_xlsx)
    _write_excel(existing, clips, prompt_names, summary, out_xlsx)
    print(f"\nWrote Excel: {out_xlsx}")

    out_md = Path(args.out_leaderboard)
    _write_leaderboard(summary, out_md, total_cost)
    print(f"Wrote leaderboard: {out_md}")

    # ---- Console summary ----
    print(f"\n=== LEADERBOARD ===")
    print(f"{'rank':<5}{'prompt':<20}{'acc':>8}{'BERT':>8}{'comp':>8}{'words':>8}{'cost':>9}")
    for i, row in enumerate(summary, 1):
        print(f"{i:<5}{row['prompt']:<20}"
              f"{row['accuracy']*100:>7.1f}%"
              f"{row['bertscore_mean']:>8.3f}"
              f"{row['composite_mean']:>8.3f}"
              f"{row['mean_word_count']:>8.0f}"
              f"   ${row['cost_usd']:>5.2f}")
    print(f"\nTotal cost: ${total_cost:.2f}")
    print(f"Records: {sum(len(r) for r in records_by_prompt.values())}/{total_pairs}")


if __name__ == "__main__":
    main()
