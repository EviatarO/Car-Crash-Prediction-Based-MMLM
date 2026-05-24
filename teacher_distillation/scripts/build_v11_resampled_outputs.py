"""v11-resampled Stage 4: build final outputs.

Assembles three artifacts from the v11-resampled pipeline:

  outputs/prompt_bakeoff/v11_100clips_resampled/
    - final_combined.jsonl        (100 records, one per clip, with provenance)
    - leaderboard_v11_resampled.md (human-readable summary)
    - results_v11_resampled.xlsx   (mirror of results_v6_debate_v11.xlsx,
                                    BUT no gt_reasoning column —
                                    teacher_reasoning_final replaces it)

Provenance per clip:
  - 66 preserved (TP/TN from v11): source = "v11_preserved"
  - 18 FP resampled: source = "v11_resampled_fp"
  - 15 FN: source = "v11_fn_v7_1"
  - 1 undetermined (00040 in v11): source = "v11_no_verdict"
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
V11_DIR  = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v11_100clips"
NEW_DIR  = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v11_100clips_resampled"
V11_XLSX = REPO_ROOT / "outputs" / "teacher_dataset_v11.xlsx"

V11_HIRES_JSONL = V11_DIR / "v6_hires_v11.jsonl"
PASS1_JSONL     = NEW_DIR / "v6_balanced_resampled_pass1.jsonl"
RECOVERY_JSONL  = NEW_DIR / "v7_1_recovery.jsonl"
RESAMPLE_LOG    = NEW_DIR / "resample_log.json"

OUT_JSONL = NEW_DIR / "final_combined.jsonl"
OUT_MD    = NEW_DIR / "leaderboard_v11_resampled.md"
OUT_XLSX  = NEW_DIR / "results_v11_resampled.xlsx"

# Sets from leaderboard_v6_debate_v11.md
TP_CLIPS = set("""00068 00093 00104 00254 00307 00332 00341 00357 00382 00424 00427 00435 00469 00488 00503 00586 00587 00592 00597 00598 00634 00655 00663 00667 00707 00757 00766 00839 00842 00858 00900 00932 00968 00977""".split())
TN_CLIPS = set("""01103 01115 01136 01209 01221 01282 01315 01485 01514 01534 01552 01563 01583 01599 01606 01691 01704 01749 01768 01800 01806 01822 01842 01893 01950 01969 01980 01995 02019 02039 02092 02129""".split())
FP_CLIPS = set("""01045 01144 01225 01261 01305 01307 01400 01420 01470 01508 01539 01569 01614 01655 01771 01817 01904 02064""".split())
FN_CLIPS = set("""00065 00089 00097 00401 00428 00573 00590 00604 00621 00670 00741 00832 00876 01013 01024""".split())


# Colors
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
HEADER = PatternFill("solid", fgColor="2E75B6")


def _load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return [json.loads(l) for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


def _load_v11_map() -> dict[str, dict]:
    return {r["video_id"]: r for r in _load_jsonl(V11_HIRES_JSONL)}


def _load_pass1_map() -> dict[str, dict]:
    return {r["video_id"]: r for r in _load_jsonl(PASS1_JSONL)}


def _load_recovery_map() -> dict[tuple, dict]:
    """Keyed by (vid, recovery_prompt)."""
    return {(r["video_id"], r["recovery_prompt"]): r for r in _load_jsonl(RECOVERY_JSONL)}


def _load_v11_xlsx_map() -> dict[str, dict]:
    """Load teacher_dataset_v11.xlsx and return {video_id: {t_seconds, requested_time_to_event}}."""
    if not V11_XLSX.exists():
        return {}
    df = pd.read_excel(V11_XLSX)
    out: dict[str, dict] = {}
    for _, row in df.iterrows():
        raw_id = row.get("video_id")
        if pd.isna(raw_id):
            continue
        vid = f"{int(raw_id):05d}"
        t_sec = row.get("t_seconds")
        rtte = row.get("requested_time_to_event")
        # t_seconds is numeric; requested_time_to_event may be numeric or string (e.g. 'TN_MIDPOINT')
        try:
            t_sec_val = float(t_sec) if not pd.isna(t_sec) else None
        except (TypeError, ValueError):
            t_sec_val = None
        try:
            rtte_val = (float(rtte) if not pd.isna(rtte) else None) if isinstance(rtte, (int, float)) else (str(rtte) if not pd.isna(rtte) else None)
        except (TypeError, ValueError):
            rtte_val = str(rtte) if rtte is not None else None
        out[vid] = {
            "t_seconds": t_sec_val,
            "requested_time_to_event": rtte_val,
        }
    return out


def _build_combined() -> list[dict]:
    v11 = _load_v11_map()
    pass1 = _load_pass1_map()
    recovery = _load_recovery_map()
    xlsx_map = _load_v11_xlsx_map()

    out: list[dict] = []
    for vid in sorted(v11.keys()):
        v11_rec = v11[vid]
        gt = v11_rec.get("gt_verdict")
        t_orig = v11_rec.get("t_seconds")
        v6_orig_verdict = v11_rec.get("verdict")
        v6_orig_reasoning = (v11_rec.get("reasoning") or "").strip()

        xlsx_extra = xlsx_map.get(vid, {})
        t_seconds_xlsx = xlsx_extra.get("t_seconds")
        requested_time_to_event = xlsx_extra.get("requested_time_to_event")

        rec = {
            "video_id": vid,
            "gt_verdict": gt,
            "t_original": t_orig,
            "t_seconds": t_seconds_xlsx if t_seconds_xlsx is not None else t_orig,
            "requested_time_to_event": requested_time_to_event,
            "v6_orig_verdict": v6_orig_verdict,
            "v6_orig_reasoning": v6_orig_reasoning,
        }

        if vid in TP_CLIPS or vid in TN_CLIPS:
            rec["source"] = "v11_preserved"
            rec["v6_resampled_verdict"] = None
            rec["v6_resampled_reasoning"] = ""
            rec["v6_resampled_t_new"] = None
            rec["v7_1_recovery_prompt"] = None
            rec["v7_1_recovery_verdict"] = None
            rec["v7_1_recovery_reasoning"] = ""
            rec["final_verdict"] = v6_orig_verdict
            rec["teacher_reasoning_final"] = v6_orig_reasoning

        elif vid in FP_CLIPS:
            p1 = pass1.get(vid, {})
            p1_verdict = p1.get("verdict")
            p1_reasoning = (p1.get("reasoning") or "").strip()
            p1_t_new = p1.get("t_new")

            tn_rec = recovery.get((vid, "PROMPT_G_OPT_v7_1_TN_RECOVERY"))
            rec["source"] = "v11_resampled_fp"
            rec["v6_resampled_verdict"] = p1_verdict
            rec["v6_resampled_reasoning"] = p1_reasoning
            rec["v6_resampled_t_new"] = p1_t_new

            if tn_rec is not None:
                tn_v = tn_rec.get("recovery_verdict")
                tn_r = (tn_rec.get("recovery_reasoning") or "").strip()
                rec["v7_1_recovery_prompt"] = "PROMPT_G_OPT_v7_1_TN_RECOVERY"
                rec["v7_1_recovery_verdict"] = tn_v
                rec["v7_1_recovery_reasoning"] = tn_r
                rec["final_verdict"] = tn_v if tn_v else p1_verdict
                rec["teacher_reasoning_final"] = tn_r if tn_v else p1_reasoning
            else:
                # pass-1 flipped to NO; no recovery needed
                rec["v7_1_recovery_prompt"] = None
                rec["v7_1_recovery_verdict"] = None
                rec["v7_1_recovery_reasoning"] = ""
                rec["final_verdict"] = p1_verdict
                rec["teacher_reasoning_final"] = p1_reasoning

        elif vid in FN_CLIPS:
            tp_rec = recovery.get((vid, "PROMPT_G_OPT_v7_1_TP_RECOVERY"))
            tp_v = (tp_rec or {}).get("recovery_verdict")
            tp_r = ((tp_rec or {}).get("recovery_reasoning") or "").strip()
            rec["source"] = "v11_fn_v7_1"
            rec["v6_resampled_verdict"] = None
            rec["v6_resampled_reasoning"] = ""
            rec["v6_resampled_t_new"] = None
            rec["v7_1_recovery_prompt"] = "PROMPT_G_OPT_v7_1_TP_RECOVERY"
            rec["v7_1_recovery_verdict"] = tp_v
            rec["v7_1_recovery_reasoning"] = tp_r
            rec["final_verdict"] = tp_v if tp_v else v6_orig_verdict
            rec["teacher_reasoning_final"] = tp_r if tp_v else v6_orig_reasoning

        else:
            # 00040: no verdict in v11 first-pass
            rec["source"] = "v11_no_verdict"
            rec["v6_resampled_verdict"] = None
            rec["v6_resampled_reasoning"] = ""
            rec["v6_resampled_t_new"] = None
            rec["v7_1_recovery_prompt"] = None
            rec["v7_1_recovery_verdict"] = None
            rec["v7_1_recovery_reasoning"] = ""
            rec["final_verdict"] = v6_orig_verdict
            rec["teacher_reasoning_final"] = v6_orig_reasoning

        rec["final_correct"] = (rec["final_verdict"] == gt) if rec["final_verdict"] else False
        rec["v6_orig_correct"] = (v6_orig_verdict == gt) if v6_orig_verdict else False
        rec["passes_for_student_training"] = rec["final_correct"]
        out.append(rec)

    return out


def _write_jsonl(records: list[dict]) -> None:
    OUT_JSONL.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def _confusion(records: list[dict], verdict_field: str) -> dict:
    tp = fp = tn = fn = none = 0
    for r in records:
        v = r.get(verdict_field)
        g = r.get("gt_verdict")
        if v is None:
            none += 1
        elif v == "YES" and g == "YES":
            tp += 1
        elif v == "YES" and g == "NO":
            fp += 1
        elif v == "NO" and g == "NO":
            tn += 1
        elif v == "NO" and g == "YES":
            fn += 1
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn, "none": none}


def _write_md(records: list[dict]) -> None:
    n = len(records)
    cm_orig  = _confusion(records, "v6_orig_verdict")
    cm_final = _confusion(records, "final_verdict")

    acc_orig  = (cm_orig["TP"] + cm_orig["TN"]) / n if n else 0.0
    acc_final = (cm_final["TP"] + cm_final["TN"]) / n if n else 0.0

    n_preserved = sum(1 for r in records if r["source"] == "v11_preserved")
    n_fp_done   = sum(1 for r in records if r["source"] == "v11_resampled_fp")
    n_fn_done   = sum(1 for r in records if r["source"] == "v11_fn_v7_1")
    n_novr      = sum(1 for r in records if r["source"] == "v11_no_verdict")

    # How many FP clips were corrected by the t-4s resample pass-1 alone (verdict flipped to NO)
    fp_fixed_pass1 = sum(1 for r in records
                         if r["source"] == "v11_resampled_fp"
                         and r["v6_resampled_verdict"] == r["gt_verdict"])
    # How many of the remaining (still-YES after pass-1) were then fixed by TN_RECOVERY
    fp_fixed_by_rec = sum(1 for r in records
                          if r["source"] == "v11_resampled_fp"
                          and r.get("v7_1_recovery_verdict") == r["gt_verdict"])
    fp_total_fixed  = fp_fixed_pass1 + fp_fixed_by_rec   # e.g. 10 + 6 = 16

    fn_fixed_rec   = sum(1 for r in records
                         if r["source"] == "v11_fn_v7_1"
                         and r.get("v7_1_recovery_verdict") == r["gt_verdict"])

    # Intermediate confusion after FP resample pass-1 only (before any recovery)
    # Preserved clips keep their v6_orig verdict; FP clips use v6_resampled; FN clips keep v6_orig
    def _intermediate_verdict(r: dict):
        src = r["source"]
        if src in ("v11_preserved", "v11_fn_v7_1", "v11_no_verdict"):
            return r.get("v6_orig_verdict")
        if src == "v11_resampled_fp":
            return r.get("v6_resampled_verdict")
        return None

    cm_mid = _confusion([{**r, "_mid": _intermediate_verdict(r)} for r in records], "_mid")
    # _confusion uses the field name, but we patched a dict — call directly:
    cm_mid = {"TP": 0, "FP": 0, "TN": 0, "FN": 0, "none": 0}
    for r in records:
        v = _intermediate_verdict(r)
        g = r.get("gt_verdict")
        if v is None:
            cm_mid["none"] += 1
        elif v == "YES" and g == "YES":
            cm_mid["TP"] += 1
        elif v == "YES" and g == "NO":
            cm_mid["FP"] += 1
        elif v == "NO" and g == "NO":
            cm_mid["TN"] += 1
        elif v == "NO" and g == "YES":
            cm_mid["FN"] += 1
    acc_mid = (cm_mid["TP"] + cm_mid["TN"]) / n if n else 0.0

    md: list[str] = []
    md.append("# v11-resampled Leaderboard\n")
    md.append("**Context:** Rerun of the 33 v11 first-pass failures (18 FP + 15 FN) at "
              "earlier timestamps (FP only) and with the corrected v7.1 recovery prompts. "
              "The 34 TP + 32 TN clips that passed in v11 are preserved without re-running.\n")
    md.append(f"**Total clips:** {n}\n")
    md.append("")

    md.append("## Headline numbers\n")
    md.append("| Stage | Accuracy | TP | FP | TN | FN |")
    md.append("|-------|---------|----|----|----|----|")
    md.append(f"| v6 first-pass (v11, baseline)  | "
              f"**{acc_orig:.1%}** ({cm_orig['TP']+cm_orig['TN']}/{n}) "
              f"| {cm_orig['TP']} | {cm_orig['FP']} | {cm_orig['TN']} | {cm_orig['FN']} |")
    md.append(f"| After FP resample only (v6_balanced at t − 4 s, no recovery yet) | "
              f"**{acc_mid:.1%}** ({cm_mid['TP']+cm_mid['TN']}/{n}) "
              f"| {cm_mid['TP']} | {cm_mid['FP']} | {cm_mid['TN']} | {cm_mid['FN']} |")
    md.append(f"| Final after v11-resampled pipeline (+ v7.1 recovery) | "
              f"**{acc_final:.1%}** ({cm_final['TP']+cm_final['TN']}/{n}) "
              f"| {cm_final['TP']} | {cm_final['FP']} | {cm_final['TN']} | {cm_final['FN']} |")
    md.append("")
    md.append("(v11 had 78% after the inverted-prompt v6 debate. This pipeline replaces "
              "that debate with resampled FP frames + corrected v7.1 prompts.)\n")

    md.append("## Pipeline breakdown\n")
    md.append(f"- **Preserved (no re-run):** {n_preserved} clips (34 TP + 32 TN from v11)")
    md.append(f"- **FP clips re-sampled to t − 4 seconds (temporal offset, not stride) + "
              f"pass-1 (v6_balanced):** {n_fp_done} "
              f"→ {fp_fixed_pass1}/{n_fp_done} fixed by the earlier timestamp alone")
    md.append(f"- **FP clips still predicting YES after resample → v7.1 TN_RECOVERY:** "
              f"{n_fp_done - fp_fixed_pass1} routed, "
              f"{fp_fixed_by_rec} additionally fixed")
    md.append(f"  - **Total FP fixed:** {fp_total_fixed}/{n_fp_done}")
    md.append(f"- **FN clips → v7.1 TP_RECOVERY (at original midpoint):** {n_fn_done} "
              f"→ {fn_fixed_rec}/{n_fn_done} fixed")
    md.append(f"- **No-verdict clips (preserved as-is):** {n_novr}")
    md.append("")

    md.append("## Per-clip table (33 reworked + no-verdict)\n")
    md.append("| video_id | gt | v6_orig | v6_resampled | v7.1_recovery | final | source | t_orig→t_new |")
    md.append("|----------|----|---------|--------------|---------------|-------|--------|--------------|")
    for r in records:
        if r["source"] == "v11_preserved":
            continue
        vid = r["video_id"]
        gt = r["gt_verdict"]
        v6o = r["v6_orig_verdict"] or "—"
        v6r = r["v6_resampled_verdict"] or "—"
        v71 = r["v7_1_recovery_verdict"] or "—"
        fv  = r["final_verdict"] or "—"
        ok = "✓" if r["final_correct"] else "✗"
        src = r["source"]
        ts = f"{r['t_original']:.2f}→{r['v6_resampled_t_new']:.2f}" if r["v6_resampled_t_new"] is not None else f"{r['t_original']:.2f} (unchanged)"
        md.append(f"| {vid} | {gt} | {v6o} | {v6r} | {v71} | **{fv}** {ok} | {src} | {ts} |")
    md.append("")

    md.append("## Still-wrong after this pipeline\n")
    md.append("| video_id | gt | final | source | note |")
    md.append("|----------|----|-------|--------|------|")
    for r in records:
        if r["final_correct"]:
            continue
        if r["final_verdict"] is None:
            note = "no verdict from any pass"
        elif r["source"] == "v11_resampled_fp":
            note = "FP persistent: resample + v7.1 TN_RECOVERY both predict YES"
        elif r["source"] == "v11_fn_v7_1":
            note = "FN persistent: v7.1 TP_RECOVERY also predicts NO"
        elif r["source"] == "v11_preserved":
            note = "preserved from v11 but somehow doesn't match (audit)"
        else:
            note = "—"
        md.append(f"| {r['video_id']} | {r['gt_verdict']} | "
                  f"{r['final_verdict'] or '—'} | {r['source']} | {note} |")
    md.append("")

    md.append("## Notes\n")
    md.append("- **GT reasoning:** v11 had a `gt_reasoning` column populated with model output. "
              "This pipeline drops that column. The xlsx now has `teacher_reasoning_final` — "
              "the model-generated reasoning chosen as the final teacher signal "
              "(recovery output if available, otherwise pass-1 output). This is what the student "
              "model will train on; calling it \"GT reasoning\" would be misleading.\n")
    md.append("- **FP timestamp rule:** `t_new = max(2.0, t_original − 4.0)` — this is a "
              "**4-second temporal offset** (earlier in the video), not related to the frame "
              "stride (which is always 4 frames). 18/18 FP clips extracted successfully; "
              "1 clip (01144) was floored to t=2.0s.\n")
    md.append("- **Preserved clips:** 66 TP/TN from v11 were NOT re-run to save cost. Their "
              "verdicts and reasoning come directly from `v6_hires_v11.jsonl`.\n")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")


def _write_xlsx(records: list[dict]) -> None:
    """Build the xlsx mirroring results_v6_debate_v11.xlsx, but WITHOUT gt_reasoning."""
    columns = [
        "video_id",
        "gt_verdict",
        "source",
        "v6_orig_verdict",
        "v6_orig_correct",
        "v6_resampled_verdict",
        "v6_resampled_t_new",
        "v7_1_recovery_prompt",
        "v7_1_recovery_verdict",
        "final_verdict",
        "final_correct",
        "passes_for_student_training",
        "t_original",
        "t_seconds",
        "requested_time_to_event",
        "teacher_reasoning_final",
        "v6_orig_reasoning",
        "v6_resampled_reasoning",
        "v7_1_recovery_reasoning",
    ]

    df_rows = []
    for r in records:
        df_rows.append({c: r.get(c) for c in columns})
    df = pd.DataFrame(df_rows, columns=columns)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUT_XLSX, sheet_name="per_clip", index=False)

    # Color-coding & summary sheet
    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb["per_clip"]

    # Header styling
    for col in range(1, ws.max_column + 1):
        c = ws.cell(1, col)
        c.fill = HEADER
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(wrap_text=True, vertical="center")

    # Build column index map
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(hdr)}

    # Color rows by final_correct
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fc = row[idx["final_correct"] - 1].value
        if fc is True:
            row[idx["final_correct"] - 1].fill = GREEN
            row[idx["passes_for_student_training"] - 1].fill = GREEN
        elif fc is False:
            row[idx["final_correct"] - 1].fill = RED
            row[idx["passes_for_student_training"] - 1].fill = RED

        # v6_orig_correct
        v6c = row[idx["v6_orig_correct"] - 1].value
        if v6c is True:
            row[idx["v6_orig_correct"] - 1].fill = GREEN
        elif v6c is False:
            row[idx["v6_orig_correct"] - 1].fill = RED

        # source coloring
        src = row[idx["source"] - 1].value
        if src == "v11_preserved":
            row[idx["source"] - 1].fill = GREEN
        elif src == "v11_resampled_fp":
            row[idx["source"] - 1].fill = ORANGE
        elif src == "v11_fn_v7_1":
            row[idx["source"] - 1].fill = BLUE
        else:
            row[idx["source"] - 1].fill = RED

        # t_seconds and requested_time_to_event: informational BLUE
        for fld in ("t_seconds", "requested_time_to_event"):
            if fld in idx:
                row[idx[fld] - 1].fill = BLUE

        # Wrap text for long fields
        for fld in ["teacher_reasoning_final", "v6_orig_reasoning",
                    "v6_resampled_reasoning", "v7_1_recovery_reasoning"]:
            row[idx[fld] - 1].alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = {
        "video_id": 10, "gt_verdict": 10, "source": 20,
        "v6_orig_verdict": 12, "v6_orig_correct": 10,
        "v6_resampled_verdict": 14, "v6_resampled_t_new": 14,
        "v7_1_recovery_prompt": 30, "v7_1_recovery_verdict": 14,
        "final_verdict": 12, "final_correct": 10,
        "passes_for_student_training": 18, "t_original": 10,
        "t_seconds": 12, "requested_time_to_event": 20,
        "teacher_reasoning_final": 60, "v6_orig_reasoning": 60,
        "v6_resampled_reasoning": 60, "v7_1_recovery_reasoning": 60,
    }
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(idx[name])].width = w
    ws.row_dimensions[1].height = 32

    # Summary sheet
    ws_s = wb.create_sheet("summary")
    cm_orig = _confusion(records, "v6_orig_verdict")
    cm_final = _confusion(records, "final_verdict")
    n = len(records)
    rows = [
        ["metric", "v6 first-pass (baseline)", "final after v11-resampled"],
        ["accuracy",
         f"{(cm_orig['TP']+cm_orig['TN'])/n:.1%} ({cm_orig['TP']+cm_orig['TN']}/{n})",
         f"{(cm_final['TP']+cm_final['TN'])/n:.1%} ({cm_final['TP']+cm_final['TN']}/{n})"],
        ["TP", cm_orig["TP"], cm_final["TP"]],
        ["FP", cm_orig["FP"], cm_final["FP"]],
        ["TN", cm_orig["TN"], cm_final["TN"]],
        ["FN", cm_orig["FN"], cm_final["FN"]],
        ["no_verdict", cm_orig["none"], cm_final["none"]],
        ["", "", ""],
        ["source breakdown", "count", ""],
        ["v11_preserved",   sum(1 for r in records if r["source"] == "v11_preserved"), ""],
        ["v11_resampled_fp", sum(1 for r in records if r["source"] == "v11_resampled_fp"), ""],
        ["v11_fn_v7_1",     sum(1 for r in records if r["source"] == "v11_fn_v7_1"), ""],
        ["v11_no_verdict",  sum(1 for r in records if r["source"] == "v11_no_verdict"), ""],
    ]
    for r_idx, vals in enumerate(rows, start=1):
        for c_idx, v in enumerate(vals, start=1):
            cell = ws_s.cell(r_idx, c_idx, value=v)
            if r_idx == 1 or vals[0] == "source breakdown":
                cell.fill = HEADER
                cell.font = Font(bold=True, color="FFFFFF")
    ws_s.column_dimensions["A"].width = 22
    ws_s.column_dimensions["B"].width = 28
    ws_s.column_dimensions["C"].width = 28

    wb.save(OUT_XLSX)


def main() -> None:
    print("Building combined records...")
    records = _build_combined()
    print(f"  {len(records)} records")

    print("Writing final_combined.jsonl...")
    _write_jsonl(records)
    print(f"  -> {OUT_JSONL}")

    print("Writing leaderboard_v11_resampled.md...")
    _write_md(records)
    print(f"  -> {OUT_MD}")

    print("Writing results_v11_resampled.xlsx...")
    _write_xlsx(records)
    print(f"  -> {OUT_XLSX}")

    # Summary
    n_correct = sum(1 for r in records if r["final_correct"])
    print()
    print(f"Final accuracy: {n_correct}/{len(records)} = {n_correct/len(records):.1%}")
    cm = _confusion(records, "final_verdict")
    print(f"Confusion: TP={cm['TP']}  FP={cm['FP']}  TN={cm['TN']}  FN={cm['FN']}  none={cm['none']}")


if __name__ == "__main__":
    main()
