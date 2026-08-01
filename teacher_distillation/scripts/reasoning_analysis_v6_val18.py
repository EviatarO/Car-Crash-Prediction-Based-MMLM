"""reasoning_analysis_v6_val18.py -- scores PROMPT_SEMSUP_V6_KINEMATIC captions
from qwen/qwen3-vl-235b-a22b-thinking against the 18-clip GT validation set,
formatted like the other semsup_val18 workbooks. Includes the 4 decomposed
sub-scores and the raw lateral_watch text as columns, because V6's headline
risk (see summary.md) is template-confabulation of lateral-intrusion events on
clips that never had one -- lateral_watch is the field where that shows up.

Reads:
- dataset/manifests/val_e3a.jsonl
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds)
- outputs/prompt_bakeoff/semsup_val18/raw_v6_qwen3vl235b.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v6_val18.xlsx

Colour rule (whole row, same convention as the V4/V5 workbooks):
  GREEN  = verdict correct AND caption/lateral_watch matches GT reasoning
  ORANGE = verdict correct but caption is middling/generic or has a minor
           fabricated detail, OR verdict wrong but the caption still describes
           the scene well
  RED    = verdict wrong AND caption wrong/hallucinated, OR verdict happens to
           be right but lateral_watch/caption fabricate a mechanism GT does
           not describe (a fabricated observation is not a usable training
           signal even when the mechanical threshold lands on the right side)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
V6_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v6_qwen3vl235b.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v6_val18.xlsx"

# (score 0-10, rationale, colour) per clip, hand-scored against gt_reasoning_en.
SCORES = {
    "00077": (5, "Captures the core mechanism (braking sedan ahead, ego closing distance) but lateral_watch misses that the sedan MERGED into ego's lane from the side - GT's actual trigger event. Wrong verdict.", "orange"),
    "00147": (6, "Correctly flags a lateral conflict with the right-side sedan, but inverts causality (GT: ego deviates into the sedan's lane; caption: sedan merges into ego's lane) - the same inversion flagged on this clip in the V4 round. Correct verdict.", "orange"),
    "00283": (5, "Flags a lateral merge conflict correctly in spirit, but lane/direction and object are wrong (GT: pickup+trailer in the RIGHT lane turns left into ego's lane; caption: SUV in the LEFT lane moves right). Correct verdict by a mirrored mechanism.", "orange"),
    "00319": (1, "Misses GT's crossing vehicle from the right entirely; caption invents a stationary parked truck instead. Wrong verdict.", "red"),
    "00372": (2, "Verdict correct but the caption describes an entirely different, unsupported mechanism (a white SUV turning left across the intersection) - GT's actual event is a lead sedan stopping for crosswalk pedestrians. Not a usable caption despite the right-side threshold.", "orange"),
    "00474": (1, "Describes 'all maintaining positions' - directly misses GT's van-turns-into-ego-lane event, the same miss recorded for every teacher/prompt tested on this clip so far. Wrong verdict.", "red"),
    "00493": (4, "Improvement over the V5 caption on this clip: now correctly reports the sedan braking (V5 said only 'maintaining following distance'), but still frames it as passive rather than GT's ego-fails-to-react mechanism. Wrong verdict.", "orange"),
    "00529": (1, "Falls back to 'maintaining left lane position' - the exact V5 failure mode this prompt targeted, on the exact same clip. Misses GT's SUV-drifts-into-ego-lane mechanism. Wrong verdict.", "red"),
    "00687": (8, "Best result recorded on this clip across every round: correctly identifies the gray SUV turning across the lane line into ego's path (V5 had regressed this to 'parked'), and gets the verdict right. Clean match to GT's exact mechanism.", "green"),
    "01153": (1, "Hallucinates a 'white sedan turning left across intersection into ego's path'. GT explicitly states all vehicles remain in their lanes and ego performs its own uncontested right turn. Same false mechanism independently invented in the V5 round on this identical clip.", "red"),
    "01281": (1, "New false positive. Hallucinates 'black SUV drifting toward container truck' - GT describes zero lateral events on this clip, only a controlled longitudinal closing on a braking pickup truck ahead. This clip was a clean TN in every prior round.", "red"),
    "01504": (1, "New false positive. Hallucinates 'dark SUV turning right across path from side street' - GT has no turning vehicle or side street at all; the actual scenario is two vehicles braking ahead with ego reacting in time. This clip was a clean TN in every prior round.", "red"),
    "01550": (8, "Matches GT's controlled-closing-on-braking-vehicle description closely, correctly reports no lateral movement. Correct verdict.", "green"),
    "01552": (6, "Captures the gas-station scene reasonably but invents a 'yellow school bus' - the same fabrication independently produced by V4 and V5 on this identical clip, now a 3-for-3 pattern worth flagging as a persistent artifact rather than noise.", "orange"),
    "01643": (5, "Correct empty-road verdict, but introduces parked cars and a 'Road Work Ahead' sign not in GT's 'no vehicles around it' - same parked-car fabrication seen on this clip in the V4/V5 rounds, now with an additional invented detail.", "orange"),
    "01737": (9, "Clean match to GT's empty, curving, night-time interchange-under-a-bridge description. Correct verdict.", "green"),
    "02104": (6, "'Silver sedan now in left lane' captures state-change language (an improvement over V5's fully static phrasing) but doesn't specify the sedan merging into ego's lane as GT describes. Correct verdict, avoids fabricating danger.", "orange"),
    "02117": (8, "Matches GT's core mechanism (vehicle ahead at constant distance, green light, no crossing pedestrians) without hallucinating a merge event. Fifth consecutive non-Gemini teacher/prompt combination to solve this clip correctly.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "v6_verdict", "risk_score",
           "closing_risk", "lateral_risk", "intrusion_risk", "unreacted_risk",
           "v6_caption", "lateral_watch", "counter_evidence",
           "score", "score_explanation"]
WIDTHS = [11, 11, 11, 15, 50, 10, 9, 8, 8, 8, 8, 48, 42, 38, 8, 50]
WRAP_COLS = {"gt_reasoning_en", "v6_caption", "lateral_watch", "counter_evidence", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    v6 = {r["video_id"]: r for r in _load_jsonl(V6_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t} - refusing to substitute a "
                            f"value from a disagreeing source (see module docstring).")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], v6[vid]
        verdict_str = "YES" if q["verdict"] == 1 else "NO"
        score, rationale, colour = SCORES.get(vid, (None, "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"], "gt_reasoning_en": v["gt_reasoning_en"],
            "v6_verdict": verdict_str, "risk_score": q["risk_score"],
            "closing_risk": q["closing_risk"], "lateral_risk": q["lateral_risk"],
            "intrusion_risk": q["intrusion_risk"], "unreacted_risk": q["unreacted_risk"],
            "v6_caption": q["caption_neutral"], "lateral_watch": q["lateral_watch"],
            "counter_evidence": q.get("counter_evidence", ""),
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["v6_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["v6_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["v6_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["v6_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["v6_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    colour_counts = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}

    try:
        from sklearn.metrics import average_precision_score, roc_auc_score
        y_true = [1 if r["gt_verdict"] == "YES" else 0 for r in rows]
        y_score = [r["risk_score"] for r in rows]
        auc = round(roc_auc_score(y_true, y_score), 3)
        ap = round(average_precision_score(y_true, y_score), 3)
    except ImportError:
        auc = ap = "sklearn not available"

    metrics = [
        ("n", n), ("verdict_accuracy (score>=50 cut)", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP", tp), ("FP", fp), ("TN", tn), ("FN", fn),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("risk_score AUC", auc), ("risk_score AP", ap),
        ("mean_hand_score", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green", colour_counts["green"]), ("n_orange", colour_counts["orange"]),
        ("n_red", colour_counts["red"]),
        ("n_distinct_risk_scores", len(set(r["risk_score"] for r in rows))),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
