"""
build_teacher_monitor.py
========================
Rebuild outputs/teacher_reasoning/monitor_teacher_coverage.xlsx from the aggregate
(Teacher_Reasoning_All_Clips.jsonl) + the fixed test manifests + the 1,500-clip train pool
(train.csv). Two sheets, each a full pre-listed grid so blanks show what is still undone:

  test  : 1,344 rows (677 private + 667 public), one clip = one TTE. Cols: video_id, split,
          GT_verdict(0/1), TTE(colored).
  train : 4,500 rows = 1,500 train.csv video_ids x {TTE_0.5, TTE_1.0, TTE_1.5}. Cols:
          video_id, GT_verdict(0/1), TTE(colored).

Cell colors (the coverage status):
  strong green (00B050) pass-1 correct | soft green (C6EFCE) pass-2/debate correct
  soft red   (FFC7CE) still wrong      | blank not yet run

Each sheet also carries distribution panels (TP/TN done, TTE-of-TP, totals) and two embedded
matplotlib histograms (colors-per-TTE, overall pass/fail). Reuses the styling approach of
build_e3a_tte_fill_outputs.py.
"""
from __future__ import annotations

import csv
import io
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "outputs" / "teacher_reasoning"
ALL_JSONL = OUT_DIR / "Teacher_Reasoning_All_Clips.jsonl"
OUT_MONITOR = OUT_DIR / "monitor_teacher_coverage.xlsx"

TEST_PRIV = REPO_ROOT / "dataset" / "manifests" / "test_manifest_hires.jsonl"
TEST_PUB = REPO_ROOT / "dataset" / "manifests" / "test_manifest_public_hires.jsonl"
TRAIN_CSV = Path(r"C:\Users\eviatar.ohayon\Ramon Space\PycharmProjects\Thesis"
                 r"\Data-Centric-Crash-Prediction-Using-3LC-and-MViT\src\Nexar_DataSet\train.csv")

TTES = ["TTE_0.5", "TTE_1.0", "TTE_1.5"]

HEADER = PatternFill("solid", fgColor="2E75B6")
STRONG_GREEN = PatternFill("solid", fgColor="00B050")
SOFT_GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
PANEL = PatternFill("solid", fgColor="FCE4D6")
OUTCOME_FILL = {"pass1": STRONG_GREEN, "pass2": SOFT_GREEN, "wrong": RED}


def _norm_vid(v) -> str:
    s = str(v).strip()
    try:
        return f"{int(s):05d}"
    except ValueError:
        return s


def _norm_verdict(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().upper()
    if s in ("1", "YES", "TRUE"):
        return "YES"
    if s in ("0", "NO", "FALSE"):
        return "NO"
    return s


_MID_MAP = {"MID": "TTE_0.5", "MID-4": "TTE_1.0", "MID-8": "TTE_1.5"}


def _tte_label(req) -> str:
    if req is None:
        return ""
    s = str(req).strip()
    try:
        return f"TTE_{float(s):.1f}"
    except ValueError:
        return _MID_MAP.get(s, s)


def _resolve_tte(rec: dict) -> str:
    hz = str(rec.get("horizon_label") or "").strip()
    if hz.startswith("TTE_"):
        return _tte_label(hz.replace("TTE_", ""))
    if hz in _MID_MAP:
        return _MID_MAP[hz]
    return _tte_label(rec.get("requested_time_to_event"))


def _outcome(rec: dict) -> str:
    gt = _norm_verdict(rec.get("gt_verdict") if rec.get("gt_verdict") is not None
                       else rec.get("target"))
    p1 = _norm_verdict(rec.get("collision_verdict"))
    final = _norm_verdict(rec.get("final_verdict") or p1)
    if p1 and p1 == gt:
        return "pass1"
    if final and final == gt:
        return "pass2"
    return "wrong"


def _load_outcomes() -> Dict[Tuple[str, str, str], str]:
    """(dataset, video_id, TTE_label) -> outcome."""
    out: Dict[Tuple[str, str, str], str] = {}
    if not ALL_JSONL.exists():
        return out
    for line in ALL_JSONL.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        k = (r.get("dataset", ""), _norm_vid(r.get("video_id")), _resolve_tte(r))
        out[k] = _outcome(r)
    return out


def _load_jsonl(p: Path) -> List[dict]:
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]


# ---------------- grid rows ----------------
def _test_rows() -> List[dict]:
    rows = []
    for path, split in ((TEST_PRIV, "private"), (TEST_PUB, "public")):
        for r in _load_jsonl(path):
            rows.append({
                "video_id": _norm_vid(r["video_id"]),
                "split": split,
                "gt": int(r.get("event_occurs", 0)),
                "tte": _tte_label(r.get("time_before_event_s")),
            })
    rows.sort(key=lambda x: (x["split"], x["video_id"]))
    return rows


def _train_rows() -> List[dict]:
    rows = []
    with TRAIN_CSV.open(encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            vid = _norm_vid(r["id"])
            gt = int(r["target"])
            for tte in TTES:
                rows.append({"video_id": vid, "gt": gt, "tte": tte})
    rows.sort(key=lambda x: (x["video_id"], x["tte"]))
    return rows


# ---------------- histogram PNGs ----------------
def _hist_colors_per_tte(done_by_tte: Dict[str, Counter], title: str) -> XLImage:
    fig, ax = plt.subplots(figsize=(6.4, 3.6))
    x = range(len(TTES))
    p1 = [done_by_tte[t].get("pass1", 0) for t in TTES]
    p2 = [done_by_tte[t].get("pass2", 0) for t in TTES]
    wr = [done_by_tte[t].get("wrong", 0) for t in TTES]
    b1 = ax.bar(x, p1, color="#00B050", label="pass-1")
    b2 = ax.bar(x, p2, bottom=p1, color="#A9D08E", label="pass-2")
    b3 = ax.bar(x, wr, bottom=[a + b for a, b in zip(p1, p2)], color="#FF7C80", label="wrong")
    ax.set_xticks(list(x)); ax.set_xticklabels(TTES)
    ax.set_ylabel("clips done"); ax.set_title(title, fontsize=10)
    ax.legend(fontsize=8, loc="upper right")
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=120); buf.seek(0); plt.close(fig)
    return XLImage(buf)


def _hist_overall(pass_n: int, wrong_n: int, blank_n: int, title: str) -> XLImage:
    fig, ax = plt.subplots(figsize=(6.4, 3.6))
    labels = ["passed", "wrong", "not-run"]
    vals = [pass_n, wrong_n, blank_n]
    bars = ax.bar(labels, vals, color=["#00B050", "#FF7C80", "#D9D9D9"])
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(vals) * 0.01 + 0.1,
                str(v), ha="center", va="bottom", fontsize=10, fontweight="bold")
    ax.set_ylabel("clips"); ax.set_title(title, fontsize=10)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=120); buf.seek(0); plt.close(fig)
    return XLImage(buf)


# ---------------- sheet writer ----------------
def _write_sheet(wb, name: str, dataset: str, rows: List[dict], extra_cols: List[str]) -> None:
    outcomes = OUTCOMES
    ws = wb.create_sheet(name)

    cols = ["video_id"] + extra_cols + ["GT_verdict", "TTE"]
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER; cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tte_col = len(cols)
    done_by_tte: Dict[str, Counter] = defaultdict(Counter)
    tp_done = tn_done = 0
    tp_by_tte = Counter()

    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r["video_id"]).alignment = Alignment(horizontal="center")
        cc = 2
        for ec in extra_cols:
            ws.cell(i, cc, r.get(ec, "")).alignment = Alignment(horizontal="center")
            cc += 1
        ws.cell(i, cc, r["gt"]).alignment = Alignment(horizontal="center"); cc += 1
        tte_cell = ws.cell(i, tte_col, r["tte"])
        tte_cell.alignment = Alignment(horizontal="center")

        oc = outcomes.get((dataset, r["video_id"], r["tte"]))
        if oc:
            tte_cell.fill = OUTCOME_FILL[oc]
            done_by_tte[r["tte"]][oc] += 1
            if r["gt"] == 1:
                tp_done += 1; tp_by_tte[r["tte"]] += 1
            else:
                tn_done += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 11
    for c in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # ---- distribution panel (to the right of the grid) ----
    p = tte_col + 2  # first panel column
    total = len(rows)
    done = tp_done + tn_done
    pass_n = sum(done_by_tte[t].get("pass1", 0) + done_by_tte[t].get("pass2", 0) for t in done_by_tte)
    wrong_n = sum(done_by_tte[t].get("wrong", 0) for t in done_by_tte)
    blank_n = total - done

    def _panel(r0, title, pairs):
        ws.cell(r0, p, title).font = Font(bold=True)
        ws.cell(r0, p, title).fill = PANEL
        for j, (k, v) in enumerate(pairs, start=1):
            ws.cell(r0 + j, p, k)
            ws.cell(r0 + j, p + 1, v)

    _panel(2, "Coverage summary", [
        ("total clips", total), ("done", done), ("not-run", blank_n),
        ("passed", pass_n), ("wrong", wrong_n),
        ("pass-rate (of done)", f"{pass_n/done:.1%}" if done else "-"),
    ])
    _panel(10, "TP vs TN (done)  [target 50/50]", [
        ("TP done (collision)", tp_done), ("TN done (safe)", tn_done),
    ])
    _panel(14, "TP done by TTE  [target equal]", [(t, tp_by_tte.get(t, 0)) for t in TTES])
    for col in (p, p + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26 if col == p else 12

    # ---- histograms (embedded PNGs, below the panels) ----
    img1 = _hist_colors_per_tte(done_by_tte, f"{name}: outcome by TTE")
    img1.anchor = f"{get_column_letter(p)}20"
    ws.add_image(img1)
    img2 = _hist_overall(pass_n, wrong_n, blank_n, f"{name}: overall coverage")
    img2.anchor = f"{get_column_letter(p)}40"
    ws.add_image(img2)


OUTCOMES: Dict[Tuple[str, str, str], str] = {}


def main() -> None:
    global OUTCOMES
    OUTCOMES = _load_outcomes()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "test", "test", _test_rows(), extra_cols=["split"])
    _write_sheet(wb, "train", "train", _train_rows(), extra_cols=[])
    wb.save(OUT_MONITOR)

    n_done = sum(1 for _ in OUTCOMES)
    print(f"Monitor written -> {OUT_MONITOR}")
    print(f"  aggregate rows keyed: {n_done}")
    print(f"  test grid: {len(_test_rows())} rows | train grid: {len(_train_rows())} rows")


if __name__ == "__main__":
    main()
