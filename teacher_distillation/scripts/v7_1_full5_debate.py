"""v7.1 debate on all 5 stride-4 first-pass failures.

Loads first-pass results from BOTH:
  outputs/prompt_bakeoff/highres_test.jsonl      (6 clips: incl. 01153, 01504)
  outputs/prompt_bakeoff/v6_hires_full18.jsonl   (12 clips: incl. 00474, 02104, 02117)

Identifies all 5 failures and applies v7.1 recovery prompts (2s, correct alignment):
  GT=YES (FN) -> PROMPT_G_OPT_v7_1_TP_RECOVERY
  GT=NO  (FP) -> PROMPT_G_OPT_v7_1_TN_RECOVERY

Resume-safe: skips clips already present in the output JSONL.
Uses stride-4 hires frames: dataset/train/<vid>_hires/

Self-contained: no imports from teacher_bakeoff or teacher_prompt_bakeoff.

Output (append-safe):
  outputs/prompt_bakeoff/v7_1_s4_ab/v7_1_debate.jsonl
"""
from __future__ import annotations

import base64
import concurrent.futures
import importlib.util
import json
import os
import re
import sys
import time
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
from dotenv import load_dotenv
from PIL import Image

try:
    from openai import (
        APIConnectionError as _OAIConnectionError,
        APITimeoutError as _OAITimeoutError,
        OpenAI,
        RateLimitError as _OAIRateLimitError,
    )
except ImportError:
    from openai import OpenAI  # type: ignore[no-redef]
    _OAIConnectionError = Exception
    _OAITimeoutError = Exception
    _OAIRateLimitError = Exception

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "teacher_distillation" / "scripts"))

from apo_metric import score_one, warmup_bertscore  # noqa: E402

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

MODEL_SLUG  = "google/gemini-3.1-pro-preview"
PRICE_IN    = 2.00
PRICE_OUT   = 12.00
TEMPERATURE = 0.1

DEFAULT_TIMEOUT       = 90.0
MAX_RETRIES           = 2
RETRY_DELAY           = 5.0
CLIP_BUDGET_SECS      = 300
INTER_CALL_DELAY      = 1.5
PROMPT_TOKEN_HARD_CAP = 100_000

DEFAULT_GT_XLSX     = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
DEFAULT_FRAMES_ROOT = REPO_ROOT / "dataset" / "train"

# Both first-pass JSONLs
FIRST_JSONL_A = REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl"
FIRST_JSONL_B = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl"

OUT_DIR      = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v7_1_s4_ab"
DEBATE_JSONL = OUT_DIR / "v7_1_debate.jsonl"

PROMPT_TP_FILE = REPO_ROOT / "prompts" / "PROMPT_G_OPT_v7_1_TP_RECOVERY.py"
PROMPT_TN_FILE = REPO_ROOT / "prompts" / "PROMPT_G_OPT_v7_1_TN_RECOVERY.py"


# ---------------------------------------------------------------------------
# Inlined utility functions (from teacher_bakeoff.py / teacher_prompt_bakeoff.py)
# ---------------------------------------------------------------------------

def _normalize_video_id(value) -> str:
    return f"{int(float(str(value).strip())):05d}"


def _normalize_verdict(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


def _normalize_confidence(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"HIGH", "MEDIUM", "LOW"} else None


def _encode_image(path: Path, frame_size: int) -> str:
    if path.exists():
        img = Image.open(path).convert("RGB")
    else:
        img = Image.new("RGB", (frame_size or 64, frame_size or 64), color=(0, 0, 0))
    if frame_size and img.size != (frame_size, frame_size):
        img = img.resize((frame_size, frame_size))
    buf = BytesIO()
    img.save(buf, format="JPEG")
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"


def _build_messages(prompt: str, image_b64s: Sequence[str], detail: str = "low") -> List[Dict]:
    content: List[Dict] = [{"type": "text", "text": prompt}]
    for b64 in image_b64s:
        image_url: Dict = {"url": b64}
        if detail:
            image_url["detail"] = detail
        content.append({"type": "image_url", "image_url": image_url})
    return [{"role": "user", "content": content}]


def _extract_json_object(raw: str) -> Optional[Dict]:
    if not raw:
        return None
    try:
        obj = json.loads(raw)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass
    fenced = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", raw, flags=re.IGNORECASE)
    if fenced:
        try:
            obj = json.loads(fenced.group(1))
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            obj = json.loads(raw[start : end + 1])
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    return None


def _parse_response(raw: str) -> Tuple[Optional[Dict], Optional[str]]:
    parsed = _extract_json_object(raw)
    if parsed is None:
        return None, None
    parsed["collision_verdict"] = _normalize_verdict(parsed.get("collision_verdict"))
    parsed["confidence"] = _normalize_confidence(parsed.get("confidence"))
    return parsed, parsed["collision_verdict"]


def _call_model(
    client: OpenAI,
    model: str,
    messages: List[Dict],
    timeout: float,
    max_retries: int,
    retry_delay: float,
    temperature: float = 0.1,
    max_tokens: int = 8192,
) -> Tuple[str, Dict]:
    last_exc: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model, messages=messages, temperature=temperature,
                timeout=timeout, max_tokens=max_tokens,
            )
            text = response.choices[0].message.content if response.choices else ""
            usage = (
                response.usage.model_dump()
                if hasattr(response, "usage") and response.usage else {}
            )
            return text or "", usage
        except _OAITimeoutError as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] timeout -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries:
                time.sleep(wait)
        except _OAIRateLimitError as exc:
            last_exc = exc
            wait = retry_delay * 2 * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] rate-limit -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries:
                time.sleep(wait)
        except _OAIConnectionError as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] connection -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries:
                time.sleep(wait)
        except Exception as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(f"  [retry {attempt}/{max_retries}] error: {exc!r} -- waiting {wait:.1f}s", flush=True)
            if attempt < max_retries:
                time.sleep(wait)
    raise RuntimeError(f"OpenRouter call failed after {max_retries} attempts: {last_exc}") from last_exc


def _calc_cost(usage: Dict, price_in: float, price_out: float) -> float:
    in_tok = usage.get("prompt_tokens", 0) or 0
    out_tok = usage.get("completion_tokens", 0) or 0
    return in_tok * price_in / 1_000_000 + out_tok * price_out / 1_000_000


def _load_clip_frames(frames_root: Path, video_id: str, indices: List[int], frame_size: int) -> List[str]:
    folder = frames_root / video_id
    return [_encode_image(folder / f"frame_{i:05d}.jpg", frame_size) for i in indices]


def _read_gt_excel_with_en(path: Path) -> List[Dict]:
    """Read GT xlsx including verdict_reasoning_en column."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["video_id", "target", "gt_verdict", "t_seconds",
                "verdict_reasoning", "verdict_reasoning_en"]
    idx = {}
    for name in required:
        if name not in header:
            raise RuntimeError(f"Column '{name}' missing from {path}. Found: {header}")
        idx[name] = header.index(name)
    rows: List[Dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["video_id"]] is None:
            continue
        rows.append({
            "video_id": _normalize_video_id(row[idx["video_id"]]),
            "target": int(row[idx["target"]]) if row[idx["target"]] is not None else None,
            "gt_verdict": str(row[idx["gt_verdict"]]).strip().upper() if row[idx["gt_verdict"]] else None,
            "t_seconds": float(row[idx["t_seconds"]]) if row[idx["t_seconds"]] is not None else None,
            "hebrew_gt": row[idx["verdict_reasoning"]] or "",
            "gt_reasoning_en": (row[idx["verdict_reasoning_en"]] or "").strip(),
        })
    return rows


# ---------------------------------------------------------------------------
# Script-specific helpers
# ---------------------------------------------------------------------------

def _load_prompt(path: Path, var_name: str) -> str:
    spec = importlib.util.spec_from_file_location(f"_p_{path.stem}", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return getattr(mod, var_name)


def _load_jsonl_dict(path: Path) -> Dict[str, Dict]:
    if not path.exists():
        return {}
    out: Dict[str, Dict] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
            out[rec["video_id"]] = rec
        except Exception:
            pass
    return out


def _append(path: Path, rec: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def _call_with_budget(client, prompt: str, vid: str, frames_subdir: str) -> Tuple[str, Dict, float]:
    def _inner():
        indices = list(range(1, 17))
        b64s = _load_clip_frames(DEFAULT_FRAMES_ROOT, frames_subdir, indices, frame_size=0)
        messages = _build_messages(prompt, b64s, detail="high")
        t0 = time.time()
        raw, usage = _call_model(
            client, MODEL_SLUG, messages,
            timeout=DEFAULT_TIMEOUT, max_retries=MAX_RETRIES,
            retry_delay=RETRY_DELAY, temperature=TEMPERATURE,
        )
        latency = time.time() - t0
        prompt_tok = usage.get("prompt_tokens", 0) if usage else 0
        if prompt_tok and prompt_tok > PROMPT_TOKEN_HARD_CAP:
            raise RuntimeError(f"Prompt token count {prompt_tok} > hard cap {PROMPT_TOKEN_HARD_CAP}")
        return raw, usage, latency

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(_inner)
        try:
            return future.result(timeout=CLIP_BUDGET_SECS)
        except concurrent.futures.TimeoutError:
            future.cancel()
            raise RuntimeError(f"Wall-clock budget {CLIP_BUDGET_SECS}s exceeded for {vid}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    tp_prompt = _load_prompt(PROMPT_TP_FILE, "PROMPT_G_OPT_v7_1_TP_RECOVERY")
    tn_prompt = _load_prompt(PROMPT_TN_FILE, "PROMPT_G_OPT_v7_1_TN_RECOVERY")
    print(f"Loaded v7.1 TP_RECOVERY ({len(tp_prompt)} chars)")
    print(f"Loaded v7.1 TN_RECOVERY ({len(tn_prompt)} chars)")

    # Merge both first-pass JSONLs
    first: Dict[str, Dict] = {}
    for jsonl_path in (FIRST_JSONL_A, FIRST_JSONL_B):
        for vid, rec in _load_jsonl_dict(jsonl_path).items():
            first[vid] = rec
    print(f"\nFirst-pass records loaded: {len(first)} clips total")
    print(f"  from {FIRST_JSONL_A.name}: {len(_load_jsonl_dict(FIRST_JSONL_A))} clips")
    print(f"  from {FIRST_JSONL_B.name}: {len(_load_jsonl_dict(FIRST_JSONL_B))} clips")

    # Identify failures
    failures = [
        (vid, rec) for vid, rec in sorted(first.items())
        if rec.get("verdict") is not None and rec["verdict"] != rec["gt_verdict"]
    ]
    print(f"\nAll first-pass failures: {len(failures)}")
    for vid, rec in failures:
        kind = "FN" if rec["gt_verdict"] == "YES" else "FP"
        print(f"  {vid}  GT={rec['gt_verdict']}  pred={rec['verdict']}  [{kind}]")

    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise SystemExit("OPENROUTER_API_KEY missing")
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title":      os.environ.get("OPENROUTER_APP_TITLE", "MMLM_v7_1_full5"),
        },
    )

    clips_all = _read_gt_excel_with_en(DEFAULT_GT_XLSX)
    clip_map = {c["video_id"]: c for c in clips_all}

    print("\nWarming up BERTScore...")
    warmup_bertscore()
    print("BERTScore ready.\n")

    existing = _load_jsonl_dict(DEBATE_JSONL)
    total_cost = 0.0
    n_fixed = n_still = n_err = 0

    print(f"=== v7.1 FULL-5 DEBATE (stride-4) ===\n")
    print(f"Already done in {DEBATE_JSONL.name}: {list(existing.keys())}\n")

    for idx, (vid, first_rec) in enumerate(failures, start=1):
        clip = clip_map.get(vid)
        if not clip:
            print(f"[{idx}/{len(failures)}] {vid} -- skip (not in GT xlsx)")
            continue

        gt = first_rec["gt_verdict"]

        if vid in existing and existing[vid].get("recovery_verdict") is not None:
            r = existing[vid]
            flipped = "FIXED" if r["recovery_verdict"] == gt else "still-wrong"
            print(f"[{idx}/{len(failures)}] {vid} -- skip (done: recovery={r['recovery_verdict']} {flipped})")
            total_cost += r.get("cost_usd", 0.0)
            if flipped == "FIXED":
                n_fixed += 1
            else:
                n_still += 1
            continue

        recovery_name = "PROMPT_G_OPT_v7_1_TP_RECOVERY" if gt == "YES" else "PROMPT_G_OPT_v7_1_TN_RECOVERY"
        prompt = tp_prompt if gt == "YES" else tn_prompt
        frames_subdir = f"{vid}_hires"

        print(f"[{idx}/{len(failures)}] {vid}  GT={gt}  v6={first_rec['verdict']}  "
              f"-> {recovery_name}", end="  ", flush=True)

        if INTER_CALL_DELAY > 0:
            time.sleep(INTER_CALL_DELAY)

        clip_start = time.time()
        try:
            raw, usage, latency = _call_with_budget(client, prompt, vid, frames_subdir)
            parsed, recovery_verdict = _parse_response(raw)
            cost = _calc_cost(usage, PRICE_IN, PRICE_OUT)
            reasoning = parsed.get("verdict_reasoning") if parsed else None
            sb = score_one(recovery_verdict, reasoning, gt, clip["gt_reasoning_en"])

            flipped = "FIXED" if recovery_verdict == gt else "still-wrong"
            print(f"recovery={recovery_verdict or '??':3s}  {flipped}  "
                  f"BERT={sb.alignment:.3f}  cost=${cost:.4f}  "
                  f"in_tok={usage.get('prompt_tokens')}  "
                  f"out_tok={usage.get('completion_tokens')}  {latency:.1f}s")
            total_cost += cost
            if flipped == "FIXED":
                n_fixed += 1
            else:
                n_still += 1

            rec = {
                "video_id": vid,
                "recovery_prompt": recovery_name,
                "resolution": "native_1280x720",
                "stride": 4,
                "detail": "high",
                "gt_verdict": gt,
                "target": clip["target"],
                "t_seconds": clip["t_seconds"],
                "v6_verdict": first_rec["verdict"],
                "recovery_verdict": recovery_verdict,
                "recovery_reasoning": reasoning,
                "full_json": parsed or {},
                "scores": sb.to_dict(),
                "raw": raw,
                "usage": usage,
                "cost_usd": cost,
                "latency_s": round(latency, 2),
                "error": None,
            }
        except Exception as exc:
            elapsed = time.time() - clip_start
            msg = f"{type(exc).__name__}: {exc}"
            print(f"[ERR] {msg}")
            n_err += 1
            rec = {
                "video_id": vid, "recovery_prompt": recovery_name,
                "resolution": "native_1280x720", "stride": 4, "detail": "high",
                "gt_verdict": gt, "target": clip.get("target"), "t_seconds": clip.get("t_seconds"),
                "v6_verdict": first_rec["verdict"],
                "recovery_verdict": None, "recovery_reasoning": None,
                "full_json": {}, "scores": {"composite": 0, "verdict": 0, "alignment": 0, "length": 0, "word_count": 0},
                "raw": "", "usage": {}, "cost_usd": 0.0,
                "latency_s": round(elapsed, 2), "error": msg,
            }
        _append(DEBATE_JSONL, rec)

    print()
    print("-" * 65)
    print(f"Done. FIXED={n_fixed}  still-wrong={n_still}  errors={n_err}")
    print(f"Total cost (this run): ${total_cost:.4f}")
    print(f"Output: {DEBATE_JSONL}")
    print("-" * 65)

    # Print all results for review
    final = _load_jsonl_dict(DEBATE_JSONL)
    print("\n=== ALL RESULTS — review for manual scoring ===")
    for vid, rec in sorted(final.items()):
        gt = rec["gt_verdict"]
        rv = rec.get("recovery_verdict", "??")
        flipped = "FIXED" if rv == gt else "still-wrong"
        bert = rec.get("scores", {}).get("alignment", 0)
        print(f"\n--- {vid}  GT={gt}  v6={rec.get('v6_verdict')}  v7.1={rv}  {flipped}  BERT={bert:.3f} ---")
        print(f"REASONING: {rec.get('recovery_reasoning', '')}")


if __name__ == "__main__":
    main()
