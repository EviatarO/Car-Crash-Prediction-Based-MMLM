"""Generate reasoning_analysis_v7_debate.{xlsx,md} with per-clip qualitative
scores for v6_balanced_s8 (stride-8 first pass) + v7 recovery reasonings vs GT.

Reads:
- outputs/prompt_bakeoff/v7_stride8/v6_s8_hires.jsonl  (18 first-pass records)
- outputs/prompt_bakeoff/v7_stride8/v7_debate.jsonl    (recovery records)
- dataset/teacher_dataset_GT_self_imply.xlsx column G (verdict_reasoning_en)

Writes:
- outputs/prompt_bakeoff/v7_stride8/reasoning_analysis_v7_debate.xlsx
- outputs/prompt_bakeoff/v7_stride8/reasoning_analysis_v7_debate.md

Scores below are assigned qualitatively by Claude after reading each clip's
reasoning vs the GT narrative. NO BERTScore (user requested manual only).
Scoring rubric (0-10):
  10  Matches GT in verdict, agents, causal chain, AND outcome. Reads like a paraphrase.
  8-9 Same verdict + same agent + same mechanism. Minor detail mismatches.
  6-7 Same verdict + correct agent but causal mechanism partially off.
  4-5 Wrong verdict but partial scene match; OR right verdict with wrong reasoning.
  2-3 Major hallucination but right verdict by coincidence.
  0-1 Wrong verdict + wrong scene + hallucinated agents.
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
V7_DIR    = REPO_ROOT / "outputs" / "prompt_bakeoff" / "v7_stride8"
FIRST     = V7_DIR / "v6_s8_hires.jsonl"
DEBATE    = V7_DIR / "v7_debate.jsonl"
GT_XLSX   = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"

OUT_XLSX = V7_DIR / "reasoning_analysis_v7_debate.xlsx"
OUT_MD   = V7_DIR / "reasoning_analysis_v7_debate.md"

# ── colours ─────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
HEADER = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BUCKET_FILL = {"pass_v6": GREEN, "pass_debate": ORANGE, "fail": RED}

# ── (score, rationale) per clip for v6_balanced_s8 (first pass) ────────────
# Populated by Claude after reading v6_s8__reasoning vs gt_reasoning_en
# DEFAULT: each clip set to None — will be populated post-run.
V7_S8_SCORES = {
    # YES clips (GT=YES) — first-pass
    "00077": (7, "Correct verdict+agent (black sedan merge+brake→rear-end). Wrong lane direction: model says 'adjacent left' but GT places sedan to the right. Mechanism otherwise correct."),
    "00147": (2, "Same YES verdict but entirely different mechanism. GT: EGO deviates laterally into white car's lane. Model: stopped lead car at intersection → rear-end. Scene misread."),
    "00283": (9, "Near-perfect. Correct agent (white pickup+trailer), same night highway, same 'left turn blocking lane, EGO too fast to brake'. Minor: model says 'from shoulder' vs GT 'stationary then turns'."),
    "00319": (1, "Wrong verdict (FN). Model describes clear intersection; completely misses car entering from right at constant speed."),
    "00372": (8, "Same verdict, same agent (sedan ahead brakes), same rear-end mechanism. GT: stops for pedestrians in crosswalk; model: 'traffic at intersection' — small causal detail difference."),
    "00474": (8, "Same verdict, correct agent (white van sharp left into EGO lane), same EGO-continues-at-speed mechanism. Hallucinated 'Brinks' brand detail, otherwise solid."),
    "00493": (5, "Same YES but inverted interaction: GT has EGO turn-left then merge behind braking sedan; model has sedan merge right to avoid parked truck. Core 'sedan brakes, EGO hits' correct."),
    "00529": (1, "Wrong verdict (FN). Model reads scene as normal heavy traffic; misses silver SUV drifting into EGO lane entirely."),
    "00687": (6, "Same verdict, correct agent (gray SUV into EGO lane). EGO motion wrong: GT says EGO turns left; model says EGO goes straight through green light. Collision mechanism otherwise same."),
    # NO clips (GT=NO) — first-pass
    "01153": (0, "Wrong verdict (FP). Hallucinated left-turning white sedan blocking EGO. GT: EGO performs smooth right turn; all vehicles in lanes; no conflict."),
    "01281": (0, "Wrong verdict (FP). Hallucinated black SUV aggressive merge. GT: blue pickup ahead brakes, EGO closes in controlled manner; no accident."),
    "01504": (1, "Wrong verdict (FP). Partial scene match (vehicles braking ahead), but model calls it dangerous merge; GT says EGO also brakes and avoids safely."),
    "01550": (2, "Wrong verdict (FP). Identifies lead vehicle with brake lights (correct). Interprets as high-speed approach; GT says controlled following with stable gap."),
    "01552": (6, "Correct NO, similar low-speed driveway scenario. GT: gas station with black SUV; model: parking lot with box truck. Correct verdict and scenario type, different specific agents."),
    "01643": (9, "Near-perfect. Both: empty road ahead, no vehicles nearby, no danger. Very close paraphrase of GT."),
    "01737": (9, "Near-perfect. Both: single lane, curved road, no other vehicles, no accident. Essentially same narrative."),
    "02104": (2, "Wrong verdict (FP). Correctly identifies tow truck and slow traffic (partial scene match), but interprets gap as insufficient when GT says reasonable distances maintained."),
    "02117": (1, "Wrong verdict (FP). Identifies black van/SUV on right (correct agent), but hallucinates it pulling out; GT says it is stopped at crosswalk and does not interfere."),
}

# ── (score, rationale) per clip for v7 recovery reasoning ───────────────────
RECOVERY_SCORES = {
    # FN clips rescued by TP_RECOVERY
    "00319": (7, "Correctly identifies vehicle from right entering intersection. GT: constant speed, no slowing. Model: 'suddenly enters in frames 14-16'. Same agent+direction, verdict FIXED."),
    "00529": (3, "Verdict FIXED (YES) but wrong mechanism: model focuses on pedestrian in yellow shirt stepping off curb; GT says silver SUV drifts into EGO lane. Pedestrian not mentioned in GT."),
    # FP clips — TN_RECOVERY results
    "01153": (0, "Still wrong (FP). Maintains hallucinated left-turning car scenario. TN_RECOVERY failed to correct."),
    "01281": (0, "Still wrong (FP). Maintains black SUV merge hallucination through TN_RECOVERY."),
    "01504": (1, "Still wrong (FP). Red SUV braking ahead described as dangerous high-speed rear-end; GT says EGO noticed and braked safely."),
    "01550": (9, "Verdict FIXED (NO). Near-perfect: 'following a red car, brake lights on, spacing stable, controlled following, no collision.' Very close to GT 'controlled manner while maintaining distance.'"),
    "02104": (2, "Still wrong (FP). Now identifies silver sedan merging (GT mentions this too), but interprets gap as insufficient; GT says distances maintained throughout."),
    "02117": (1, "Still wrong (FP). Still interprets black SUV/van as aggressively pulling out; GT says it is stopped at crosswalk and vehicles continue safely."),
}


def _load(p: Path):
    if not p.exists():
        return []
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]


def _norm_vid(v):
    s = str(v).strip()
    return f"{int(s):05d}" if s.isdigit() else s


def _style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def _color_rows(ws, bucket_col_idx: int) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        bucket_cell = row[bucket_col_idx - 1]
        fill = BUCKET_FILL.get(str(bucket_cell.value or ""))
        if fill:
            for cell in row:
                cell.fill = fill


def main():
    first  = {r["video_id"]: r for r in _load(FIRST)}
    debate = {r["video_id"]: r for r in _load(DEBATE)}

    gt_df = pd.read_excel(GT_XLSX)
    gt_map = {
        _norm_vid(row[gt_df.columns[0]]): str(row[gt_df.columns[6]]).strip()
        for _, row in gt_df.iterrows()
    }

    rows = []
    for vid in sorted(first):
        h = first[vid]
        gt_v       = h["gt_verdict"]
        gt_reason  = gt_map.get(vid, "")
        s8_verdict = h.get("verdict") or ""
        s8_correct = (s8_verdict == gt_v)
        s8_score, s8_rat = V7_S8_SCORES.get(vid, (None, ""))
        s8_reason  = (h.get("reasoning") or "").strip()

        d = debate.get(vid)
        if d:
            rec_prompt   = d.get("recovery_prompt", "")
            rec_verdict  = d.get("recovery_verdict") or ""
            rec_correct  = (rec_verdict == gt_v) if rec_verdict else False
            rec_score, rec_rat = RECOVERY_SCORES.get(vid, (None, ""))
            rec_reason   = (d.get("recovery_reasoning") or "").strip()
            final_v      = rec_verdict or s8_verdict
            final_correct = (final_v == gt_v)
            if rec_correct and not s8_correct:
                final_score = rec_score
                flip = "FIXED"
            elif s8_correct and rec_verdict and not rec_correct:
                final_score = s8_score
                flip = "BROKE"
            else:
                final_score = s8_score
                flip = "still-wrong" if not s8_correct else "no-flip"
        else:
            rec_prompt = rec_verdict = rec_rat = rec_reason = ""
            rec_correct = ""
            rec_score = None
            final_v = s8_verdict
            final_correct = s8_correct
            final_score = s8_score
            flip = ""

        # Bucket for color coding
        if s8_correct:
            bucket = "pass_v6"
        elif flip == "FIXED":
            bucket = "pass_debate"
        else:
            bucket = "fail"

        rows.append({
            "video_id": vid,
            "gt_verdict": gt_v,
            "gt_reasoning_en": gt_reason,
            "v6_s8__verdict": s8_verdict,
            "v6_s8__correct": s8_correct,
            "v6_s8__score": s8_score,
            "v6_s8__rationale": s8_rat,
            "v6_s8__reasoning": s8_reason,
            "recovery_prompt": rec_prompt,
            "recovery__verdict": rec_verdict,
            "recovery__correct": rec_correct,
            "recovery__score": rec_score,
            "recovery__rationale": rec_rat,
            "recovery__reasoning": rec_reason,
            "final_after_debate_verdict": final_v,
            "final_after_debate_correct": final_correct,
            "final_after_debate_score": final_score,
            "flipped_by_debate": flip,
            "passes_for_student_training": bucket,
        })

    df = pd.DataFrame(rows)

    # ── Summary stats ────────────────────────────────────────────────────────
    s8_scores  = [s for s, _ in V7_S8_SCORES.values() if s is not None]
    rec_scores = [s for s, _ in RECOVERY_SCORES.values() if s is not None]
    final_scores = [r["final_after_debate_score"] for r in rows if r["final_after_debate_score"] is not None]

    def stats(name, ss):
        return {
            "stage": name,
            "n": len(ss),
            "mean": round(sum(ss) / len(ss), 2) if ss else 0,
            "median": sorted(ss)[len(ss) // 2] if ss else 0,
            "n_ge8": sum(1 for s in ss if s >= 8),
            "n_le2": sum(1 for s in ss if s <= 2),
        }

    df_summary = pd.DataFrame([
        stats("v6_balanced_s8 (first pass, all 18)", s8_scores),
        stats(f"v7 recovery (debated {len(rec_scores)})", rec_scores),
        stats("final after debate (18)", final_scores),
    ])

    # ── Write xlsx with color coding ────────────────────────────────────────
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="per_clip", index=False)
        df_summary.to_excel(w, sheet_name="summary", index=False)

    wb = load_workbook(OUT_XLSX)
    _style_sheet(wb["per_clip"])
    bucket_col_idx = df.columns.get_loc("passes_for_student_training") + 1
    _color_rows(wb["per_clip"], bucket_col_idx)
    _style_sheet(wb["summary"])
    wb.save(OUT_XLSX)

    print(f"Wrote {OUT_XLSX.name}: {len(df)} rows, {len(df.columns)} cols")
    print()
    print(df_summary.to_string(index=False))

    # ── markdown ─────────────────────────────────────────────────────────────
    n = len(df)
    s8_acc = df["v6_s8__correct"].sum() / n if n else 0
    final_acc = df["final_after_debate_correct"].sum() / n if n else 0
    n_debated = sum(1 for r in rows if r["recovery_prompt"])
    n_fixed   = sum(1 for r in rows if r["flipped_by_debate"] == "FIXED")
    n_broke   = sum(1 for r in rows if r["flipped_by_debate"] == "BROKE")
    n_still   = sum(1 for r in rows if r["flipped_by_debate"] == "still-wrong")

    md = []
    md.append("# Reasoning Quality Analysis - v6_balanced_s8 + v7 Debate Recovery\n")
    md.append("**Setup:** 18 GT clips, frames @ NATIVE 1280x720 with **stride=8 (16 frames over ~4 seconds, same final-frame anchor as the stride-4 test)**, Gemini `detail=\"high\"`.")
    md.append("**Compared against:** `verdict_reasoning_en` (column G of GT Excel).")
    md.append(f"**Records:** 18 first-pass + {n_debated} v7 recovery reasonings.")
    md.append("**Method:** Qualitative manual 0-10 scoring per clip and stage against GT narrative (no BERTScore).")
    md.append("")

    md.append("## Scoring rubric (0-10)\n")
    md.append("| Score | Meaning |")
    md.append("|-------|---------|")
    md.append("| 10  | Matches GT in verdict, agents, causal chain, AND outcome. Reads like a paraphrase. |")
    md.append("| 8-9 | Same verdict + same agent + same mechanism. Minor detail mismatches. |")
    md.append("| 6-7 | Same verdict + correct agent but causal mechanism partially off. |")
    md.append("| 4-5 | Wrong verdict but partial scene match; OR right verdict with wrong reasoning. |")
    md.append("| 2-3 | Major hallucination but right verdict by coincidence. |")
    md.append("| 0-1 | Wrong verdict + wrong scene + hallucinated agents. |")
    md.append("")

    md.append("## Summary\n")
    md.append("| Stage | n | Mean | Median | #>=8 | #<=2 | Verdict accuracy |")
    md.append("|-------|---|------|--------|------|------|------------------|")
    if s8_scores:
        md.append(f"| v6_balanced_s8 | 18 | {round(sum(s8_scores)/len(s8_scores), 2)} "
                  f"| {sorted(s8_scores)[len(s8_scores)//2]} "
                  f"| {sum(1 for s in s8_scores if s>=8)} "
                  f"| {sum(1 for s in s8_scores if s<=2)} "
                  f"| **{s8_acc:.1%}** ({df['v6_s8__correct'].sum()}/18) |")
    if rec_scores:
        md.append(f"| v7 recovery (debated) | {len(rec_scores)} | {round(sum(rec_scores)/len(rec_scores), 2)} "
                  f"| {sorted(rec_scores)[len(rec_scores)//2]} "
                  f"| {sum(1 for s in rec_scores if s>=8)} "
                  f"| {sum(1 for s in rec_scores if s<=2)} "
                  f"| {n_fixed}/{n_debated} correct |")
    if final_scores:
        md.append(f"| **Final after debate** | 18 | **{round(sum(final_scores)/len(final_scores), 2)}** "
                  f"| {sorted(final_scores)[len(final_scores)//2]} "
                  f"| {sum(1 for s in final_scores if s>=8)} "
                  f"| {sum(1 for s in final_scores if s<=2)} "
                  f"| **{final_acc:.1%}** ({df['final_after_debate_correct'].sum()}/18) |")
    md.append("")
    md.append(f"**Debate outcomes:** {n_debated} clips debated -> "
              f"**{n_fixed} FIXED**, **{n_broke} BROKE**, {n_still} still-wrong.")
    md.append("")

    md.append("## Per-clip table\n")
    md.append("| Clip | GT | v6_s8 verdict | v6_s8 score | recovery | rec verdict | rec score | final score | bucket |")
    md.append("|------|----|---------------|-------------|----------|-------------|-----------|-------------|--------|")
    for r in rows:
        rp = r["recovery_prompt"].replace("PROMPT_G_OPT_v7_", "").replace("_RECOVERY", "") if r["recovery_prompt"] else "-"
        rv = r["recovery__verdict"] or "-"
        rs = r["recovery__score"] if r["recovery__score"] is not None else "-"
        v6_mark = "OK" if r["v6_s8__correct"] else "XX"
        bkt = {"pass_v6": "GREEN", "pass_debate": "ORANGE", "fail": "RED"}.get(
            r["passes_for_student_training"], "")
        md.append(f"| {r['video_id']} | {r['gt_verdict']} "
                  f"| {r['v6_s8__verdict']} {v6_mark} | {r['v6_s8__score']} "
                  f"| {rp} | {rv} | {rs} | {r['final_after_debate_score']} | {bkt} |")
    md.append("")

    md.append("## Verdict accuracy: confusion matrices\n")
    md.append("**v6_balanced_s8 (first pass):**")
    tp = ((df["v6_s8__verdict"]=="YES") & (df["gt_verdict"]=="YES")).sum()
    fp = ((df["v6_s8__verdict"]=="YES") & (df["gt_verdict"]=="NO")).sum()
    tn = ((df["v6_s8__verdict"]=="NO") & (df["gt_verdict"]=="NO")).sum()
    fn = ((df["v6_s8__verdict"]=="NO") & (df["gt_verdict"]=="YES")).sum()
    md.append(f"  TP={tp}  FP={fp}  TN={tn}  FN={fn}")
    md.append("")
    md.append("**After v7 debate:**")
    tp_f = ((df["final_after_debate_verdict"]=="YES") & (df["gt_verdict"]=="YES")).sum()
    fp_f = ((df["final_after_debate_verdict"]=="YES") & (df["gt_verdict"]=="NO")).sum()
    tn_f = ((df["final_after_debate_verdict"]=="NO") & (df["gt_verdict"]=="NO")).sum()
    fn_f = ((df["final_after_debate_verdict"]=="NO") & (df["gt_verdict"]=="YES")).sum()
    md.append(f"  TP={tp_f}  FP={fp_f}  TN={tn_f}  FN={fn_f}")
    md.append("")

    md.append("## Files\n")
    md.append(f"- xlsx: `outputs/prompt_bakeoff/v7_stride8/{OUT_XLSX.name}`")
    md.append(f"- md:   `outputs/prompt_bakeoff/v7_stride8/{OUT_MD.name}`")
    md.append("")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")
    print(f"\nWrote {OUT_MD.name}")


if __name__ == "__main__":
    main()
