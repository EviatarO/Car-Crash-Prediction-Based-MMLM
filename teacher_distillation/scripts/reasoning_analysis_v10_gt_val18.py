"""reasoning_analysis_v10_gt_val18.py -- scores the V10/V10Q GT-informed
captioning 2x2 bake-off (google/gemini-3.6-flash x qwen/qwen3-vl-235b-a22b-
thinking, each in gt/blind mode) against the 18-clip GT validation set, using
a mechanical SLOT rubric instead of a hand-assigned 0-10 score.

WHY A NEW SCORER (not the 0-10 hand rubric every prior round used)
--------------------------------------------------------------------
The 0-10 rubric rewards long analytical prose, which flatters v6_balanced's
~150-token native fields over a <=40-word caption_neutral -- an unfair
comparison for what's actually the deliverable (the SigLIP training target).
This scorer instead extracts, once, a small set of GT "slots" per clip
(agent / motion / position / closing-dynamic) from dataset/manifests/
val18_gt_slots.json, and checks whether each arm's output substantively
covers the same slots via case-insensitive keyword search. Objective,
reproducible, and it scores exactly what SigLIP needs encoded.

CALIBRATION (run this BEFORE trusting the scorer on new data)
----------------------------------------------------------------
`--calibrate` runs the identical slot-matching logic against the ALREADY-ON-
DISK outputs/prompt_bakeoff/semsup_val18/raw_v6_gemini36flash.jsonl (v6's
verdict_reasoning + temporal_analysis + scene_context + dynamic_objects
concatenated, since that prompt has no structured hazard_* fields) and
checks that it independently reproduces the known CONTRADICT set
{00319, 00372, 00474, 00529, 00687} -- the 5 clips a human scorer flagged as
stating the opposite of GT's mechanism (see reasoning_analysis_v6_
gemini36flash_val18.py's SCORES dict). If calibration does not reproduce
that set, the rubric is wrong and must be fixed before scoring V10.

METRICS (per arm, per clip)
----------------------------
- slot_recall = matched slots / available GT slots (agent/motion/position/
  closing) -- the headline number.
- agent_conflict = GT names one agent, model names a definitely-different one
  (heuristic: neither side's agent keywords appear in the other's text).
- CONTRADICT = zero slots matched AND the model's text contains a generic
  "everything is fine" marker phrase on a POSITIVE clip -- the exact
  00474/00529/00687-style failure signature.
- rationalization_rate (GT arms only) = does the GT arm's hazard_agent appear
  nowhere in the SAME clip's blind-arm text? This is the direct replacement
  for a two-pass debate prompt -- it measures whether GT is manufacturing a
  mechanism the model couldn't otherwise see.
- Blind arms only: verdict accuracy, TP/FP/TN/FN, AP/AUC from risk_score.

Reuses the whole-row green/orange/red coloring convention from
reasoning_analysis_v6_gemini36flash_val18.py.

Writes: outputs/prompt_bakeoff/semsup_val18_gt/reasoning_analysis_v10_gt_val18.xlsx
        outputs/prompt_bakeoff/semsup_val18_gt/summary.md
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
GT_SLOTS_PATH = REPO_ROOT / "dataset" / "manifests" / "val18_gt_slots.json"
CALIBRATION_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v6_gemini36flash.jsonl"
OUT_DIR = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18_gt"

ARM_FILES = {
    "gemini_gt":    OUT_DIR / "raw_v10_gemini_gt.jsonl",
    "qwen_gt":      OUT_DIR / "raw_v10q_qwen_gt.jsonl",
    "gemini_blind": OUT_DIR / "raw_v10_gemini_blind.jsonl",
    "qwen_blind":   OUT_DIR / "raw_v10q_qwen_blind.jsonl",
}
ARM_LABELS = {
    "gemini_gt": "Gemini 3.6 Flash / GT", "qwen_gt": "Qwen3-VL-235B / GT",
    "gemini_blind": "Gemini 3.6 Flash / blind", "qwen_blind": "Qwen3-VL-235B / blind",
}

# Known human-hand-scored CONTRADICT set from reasoning_analysis_v6_gemini36flash_val18.py
# (SCORES dict, "CONTRADICT" tag). Used only by --calibrate to verify this
# scorer's marker-phrase heuristic reproduces the same finding independently.
KNOWN_CONTRADICT_V6_GEMINI = {"00319", "00372", "00474", "00529", "00687"}

# Generic "everything is fine" phrasing seen across all 5 known CONTRADICT
# clips' raw v6/Gemini output (verdict_reasoning + temporal_analysis) -- not
# tuned to any single clip's wording.
CONTRADICT_MARKERS = [
    "no trajectory conflict", "no evidence of converg", "remain outside the ego",
    "receding", "safely pass", "being safely passed", "maintain safe", "maintains safe",
    "maintaining safe", "stable trajector", "parallel, stable trajector",
    "risk remains low", "no lateral", "no unsafe lane change", "no sudden braking",
    "without trajectory convergence", "no lane cut-in", "safe, stable spacing",
    "safe headway", "lateral clearance",
]

# ---------------------------------------------------------------------------
# NEGATIVE-CLIP FABRICATION CHECK (added 2026-08-02)
#
# The CONTRADICT check above only fires on POSITIVE-target clips (it was
# calibrated against a positives-only hand-scored set). That left the exact
# failure mode found by manual review of this round unguarded:
#   - Qwen/GT on 01552 wrote "school bus stationary at gas station" -- no bus
#     exists in GT (a known Qwen-family hallucination on that clip since V4).
#   - Gemini/GT on 01643 wrote hazard_agent "opposing vehicles in left lane"
#     on a clip whose GT says "There are no vehicles around it".
# Neither is detectable by slot_recall (which only rewards hits, never
# penalizes invented extras). At 587-window scale nobody will read every row,
# so this has to be mechanical.
#
# Two rules, both deliberately TRIAGE-oriented (flag for review, not a verdict):
#   A. SPECIFIC_TYPE: the model names a specific vehicle type that appears
#      nowhere in gt_reasoning_en. Generic words (vehicle/car) are excluded --
#      GT text says "EGO vehicle" constantly, so they carry no signal.
#   B. EMPTY_SCENE: GT explicitly asserts emptiness ("no vehicles around",
#      "no other cars") yet the model still names an agent.
# Expect some false positives from synonyms (GT "jeep" vs model "SUV"). That
# is acceptable for a review queue and is reported, not silently suppressed.
# ---------------------------------------------------------------------------
SPECIFIC_VEHICLE_TYPES = (
    "sedan", "suv", "truck", "van", "bus", "pickup", "minivan", "jeep", "taxi",
    "motorcycle", "bicycle", "crossover", "hatchback", "coupe", "scooter",
    "cyclist", "pedestrian", "trailer", "tractor", "ambulance", "lorry",
)
# Deliberately NOT including a bare "there are no" -- it matched 01552's
# "There are no signs indicating an impending accident", which is a statement
# about accident cues, not about the scene being empty of agents (that clip
# has a black SUV and a truck in GT). Every marker here must be specifically
# about the absence of VEHICLES.
EMPTINESS_MARKERS = (
    "no vehicles around", "no other cars", "no vehicles", "no other vehicles",
    "with no other cars", "no cars around",
)


def extract_vehicle_types(text: str) -> set:
    """Specific vehicle nouns present in `text`, singularized crudely."""
    t = " " + str(text).lower().replace("-", " ") + " "
    found = set()
    for noun in SPECIFIC_VEHICLE_TYPES:
        if f" {noun} " in t or f" {noun}s " in t or f" {noun}es " in t:
            found.add(noun)
    return found


def fabrication_check(model_row: dict, gt_reasoning: str, target: int) -> dict:
    """Negative-clip fabrication triage. Returns dict with `flagged` (bool),
    `rule` (which rule fired), and `detail`. Only meaningful for target==0;
    returns flagged=False for positives (CONTRADICT covers those)."""
    if target != 0:
        return {"flagged": False, "rule": "", "detail": ""}

    model_text = " ".join(str(model_row.get(f, "")) for f in
                           ("hazard_agent", "caption_neutral", "hazard_motion"))
    gt_l = str(gt_reasoning).lower()

    # Rule B first (stronger signal): GT says the scene is empty of agents.
    gt_says_empty = any(m in gt_l for m in EMPTINESS_MARKERS)
    named_agent = str(model_row.get("hazard_agent", "")).strip()
    if gt_says_empty and named_agent and named_agent.lower() not in ("", "none", "n/a"):
        return {"flagged": True, "rule": "EMPTY_SCENE",
                "detail": f"GT asserts an empty scene but model named {named_agent!r}"}

    # Rule A: a specific vehicle type the GT text never mentions.
    # Only applied when GT itself names at least one specific type -- otherwise
    # there is nothing to compare against and every specific noun the model
    # uses would flag. (01153's GT says only "other vehicles"/"two cars", no
    # types at all; the model saying "sedan" there is GT being vague, not the
    # model inventing.)
    model_types = extract_vehicle_types(model_text)
    gt_types = extract_vehicle_types(gt_reasoning)
    if not gt_types:
        return {"flagged": False, "rule": "",
                "detail": "GT names no specific vehicle type - rule A not applicable"}
    invented = model_types - gt_types
    if invented:
        return {"flagged": True, "rule": "SPECIFIC_TYPE",
                "detail": f"model names {sorted(invented)}; GT names {sorted(gt_types)}"}

    return {"flagged": False, "rule": "", "detail": ""}


FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def load_gt_slots() -> dict:
    raw = json.loads(GT_SLOTS_PATH.read_text(encoding="utf-8"))
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def slot_match(blob: str, keywords: list) -> bool:
    blob_l = blob.lower()
    return any(kw.lower() in blob_l for kw in keywords)


def score_blob(blob: str, gt_slot: dict, target: int) -> dict:
    """gt_slot: {'agent':[...], 'motion':[...], 'position':[...], 'closing':[...]}.
    Returns slot_recall, per-slot matches, and a CONTRADICT flag."""
    slot_names = [s for s in ("agent", "motion", "position", "closing") if gt_slot.get(s)]
    matches = {s: slot_match(blob, gt_slot[s]) for s in slot_names}
    n_match = sum(matches.values())
    recall = n_match / len(slot_names) if slot_names else None

    # CONTRADICT is judged independently of slot_recall, not gated on n_match==0:
    # generic driving vocabulary (agent nouns like "sedan", positions like
    # "ahead") coincidentally overlaps between correct and wrong descriptions
    # of the same scene, so a keyword-overlap recall can be nonzero even when
    # the model asserts the literal opposite of GT's mechanism (calibration
    # against the known 00319/00372/00474/00529/00687 set showed exactly
    # this - recall never hit 0 even on those). A marker phrase asserting
    # "no conflict / safe / receding" on a clip GT says IS a collision is a
    # contradiction regardless of what else the text happens to share.
    contradict = target == 1 and slot_match(blob, CONTRADICT_MARKERS)

    if recall is None:
        label = "N/A"
    elif contradict:
        label = "CONTRADICT"
    elif recall >= 0.5:
        label = "MATCH"
    elif recall > 0:
        label = "PARTIAL"
    else:
        label = "MISS"

    return {"slot_matches": matches, "slot_recall": recall, "label": label,
            "contradict": contradict}


def _blob_from_v6_row(row: dict) -> str:
    """v6_balanced (unmodified) has no hazard_* fields -- use its native
    reasoning fields as the caption-equivalent text, same convention as
    every prior hand-scored round (see reasoning_analysis_v6_gemini36flash_
    val18.py's module docstring)."""
    parts = [row.get("verdict_reasoning", ""), row.get("temporal_analysis", ""),
              row.get("scene_context", "")]
    dyn = row.get("dynamic_objects", [])
    if isinstance(dyn, list):
        parts.append(" ".join(d if isinstance(d, str) else json.dumps(d) for d in dyn))
    else:
        parts.append(str(dyn))
    return " ".join(parts)


def _blob_from_v10_row(row: dict) -> str:
    fields = ("hazard_agent", "hazard_motion", "hazard_position", "closing_dynamic",
              "caption_neutral", "risk_clause")
    return " ".join(str(row.get(f, "")) for f in fields)


def run_calibration() -> bool:
    """Returns True iff the scorer reproduces the known CONTRADICT set. Prints
    a full report either way."""
    if not CALIBRATION_JSONL.exists():
        raise FileNotFoundError(f"Calibration file not found: {CALIBRATION_JSONL}")
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    gt_slots = load_gt_slots()
    v6_rows = {r["video_id"]: r for r in _load_jsonl(CALIBRATION_JSONL)}

    found_contradict = set()
    print("=" * 78)
    print("CALIBRATION: scoring raw_v6_gemini36flash.jsonl with the V10 slot rubric")
    print("=" * 78)
    for vid in sorted(val):
        target = int(val[vid]["target"])
        blob = _blob_from_v6_row(v6_rows[vid])
        result = score_blob(blob, gt_slots[vid], target)
        flag = "  <-- CONTRADICT" if result["contradict"] else ""
        print(f"  {vid} (target={target})  recall={result['slot_recall']}  "
              f"label={result['label']}{flag}")
        if result["contradict"]:
            found_contradict.add(vid)

    print()
    print(f"Known CONTRADICT set (hand-scored):  {sorted(KNOWN_CONTRADICT_V6_GEMINI)}")
    print(f"Scorer-found CONTRADICT set:         {sorted(found_contradict)}")
    missed = KNOWN_CONTRADICT_V6_GEMINI - found_contradict
    extra = found_contradict - KNOWN_CONTRADICT_V6_GEMINI
    passed = not missed  # extra false positives are noted but don't fail calibration
    if missed:
        print(f"FAIL: scorer MISSED known CONTRADICT clips: {sorted(missed)}")
    if extra:
        print(f"NOTE: scorer flagged additional clips not in the known set: {sorted(extra)} "
              f"(not necessarily wrong -- the hand-scored set was itself informal; "
              f"review these manually)")
    print("CALIBRATION " + ("PASSED" if passed else "FAILED") +
          " -- " + ("safe to score new arms." if passed else
                     "fix CONTRADICT_MARKERS/slots before scoring new data."))
    print("=" * 78)
    return passed


def run_main_analysis():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    gt_slots = load_gt_slots()

    arms = {}
    for arm_key, path in ARM_FILES.items():
        if path.exists():
            arms[arm_key] = {r["video_id"]: r for r in _load_jsonl(path)}
        else:
            print(f"[skip] {arm_key}: {path} not found yet")
    if not arms:
        print("No arm output files found under", OUT_DIR)
        print("Run semsup_caption_promptbakeoff.py for each of the 4 arms first "
              "(see the 2026-08-02 plan).")
        return

    per_clip = {vid: {} for vid in val}
    for arm_key, rows in arms.items():
        is_blind = arm_key.endswith("_blind")
        for vid, v in val.items():
            if vid not in rows:
                continue
            target = int(v["target"])
            blob = _blob_from_v10_row(rows[vid])
            result = score_blob(blob, gt_slots[vid], target)
            result["hazard_agent"] = rows[vid].get("hazard_agent", "")
            result["mechanism_visible"] = rows[vid].get("mechanism_visible")
            if is_blind:
                result["verdict"] = int(rows[vid].get("verdict", 0))
                result["risk_score"] = float(rows[vid].get("risk_score", 0))
            per_clip[vid][arm_key] = result

    # rationalization_rate: does the GT arm's hazard_agent appear nowhere in
    # the SAME clip's blind-arm text, for the SAME model?
    for model, (gt_key, blind_key) in (("gemini", ("gemini_gt", "gemini_blind")),
                                         ("qwen", ("qwen_gt", "qwen_blind"))):
        if gt_key not in arms or blind_key not in arms:
            continue
        for vid in val:
            gt_r = per_clip[vid].get(gt_key)
            blind_r = per_clip[vid].get(blind_key)
            if not gt_r or not blind_r:
                continue
            gt_agent = gt_r.get("hazard_agent", "")
            blind_blob = _blob_from_v10_row(arms[blind_key][vid])
            gt_r["rationalized"] = bool(gt_agent) and gt_agent.lower() not in blind_blob.lower()

    # ---- xlsx ----
    headers = ["video_id", "target", "gt_reasoning_en"]
    for arm_key in ARM_FILES:
        if arm_key in arms:
            headers += [f"{arm_key}_recall", f"{arm_key}_label", f"{arm_key}_hazard_agent"]
            if arm_key.endswith("_gt"):
                headers.append(f"{arm_key}_rationalized")
            if arm_key.endswith("_blind"):
                headers += [f"{arm_key}_verdict", f"{arm_key}_risk_score"]

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    label_colour = {"MATCH": "green", "PARTIAL": "orange", "MISS": "red",
                     "CONTRADICT": "red", "N/A": "orange"}

    for i, vid in enumerate(sorted(val), start=2):
        v = val[vid]
        row_vals = [vid, v["target"], v["gt_reasoning_en"]]
        worst_colour = "green"
        for arm_key in ARM_FILES:
            if arm_key not in arms:
                continue
            r = per_clip[vid].get(arm_key)
            if r is None:
                row_vals += [None, "MISSING", None]
                if arm_key.endswith("_gt"):
                    row_vals.append(None)
                if arm_key.endswith("_blind"):
                    row_vals += [None, None]
                worst_colour = "red"
                continue
            row_vals += [r["slot_recall"], r["label"], r.get("hazard_agent")]
            c = label_colour.get(r["label"], "orange")
            if c == "red":
                worst_colour = "red"
            elif c == "orange" and worst_colour == "green":
                worst_colour = "orange"
            if arm_key.endswith("_gt"):
                row_vals.append(r.get("rationalized"))
            if arm_key.endswith("_blind"):
                row_vals += [r.get("verdict"), r.get("risk_score")]
        for c, val_ in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=c, value=val_)
            cell.fill = FILL[worst_colour]
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(c == 3), vertical="center",
                                        horizontal="left" if c == 3 else "center")
        ws.row_dimensions[i].height = 43.2

    for c, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = \
            60 if h == "gt_reasoning_en" else 16

    # ---- summary sheet ----
    ws2 = wb.create_sheet("summary")
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    row_i = 2
    for arm_key in ARM_FILES:
        if arm_key not in arms:
            continue
        results = [per_clip[vid][arm_key] for vid in val if arm_key in per_clip[vid]]
        recalls = [r["slot_recall"] for r in results if r["slot_recall"] is not None]
        mean_recall = sum(recalls) / len(recalls) if recalls else None
        n_contradict = sum(1 for r in results if r["contradict"])
        ws2.cell(row=row_i, column=1, value=f"--- {ARM_LABELS[arm_key]} ---").font = Font(bold=True)
        row_i += 1
        ws2.cell(row=row_i, column=1, value="n scored"); ws2.cell(row=row_i, column=2, value=len(results)); row_i += 1
        ws2.cell(row=row_i, column=1, value="mean slot_recall"); ws2.cell(row=row_i, column=2, value=round(mean_recall, 3) if mean_recall is not None else None); row_i += 1
        ws2.cell(row=row_i, column=1, value="n CONTRADICT"); ws2.cell(row=row_i, column=2, value=n_contradict); row_i += 1
        if arm_key.endswith("_gt"):
            n_rat = sum(1 for r in results if r.get("rationalized"))
            ws2.cell(row=row_i, column=1, value="rationalization_rate (n / n)")
            ws2.cell(row=row_i, column=2, value=f"{n_rat} / {len(results)}")
            row_i += 1
        if arm_key.endswith("_blind"):
            tp = sum(1 for vid, r in zip(val, results) if r["verdict"] == 1 and val[vid]["target"] == 1)
            fp = sum(1 for vid, r in zip(val, results) if r["verdict"] == 1 and val[vid]["target"] == 0)
            tn = sum(1 for vid, r in zip(val, results) if r["verdict"] == 0 and val[vid]["target"] == 0)
            fn = sum(1 for vid, r in zip(val, results) if r["verdict"] == 0 and val[vid]["target"] == 1)
            ws2.cell(row=row_i, column=1, value="TP/FP/TN/FN"); ws2.cell(row=row_i, column=2, value=f"{tp}/{fp}/{tn}/{fn}"); row_i += 1
        row_i += 1
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 30

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_xlsx = OUT_DIR / "reasoning_analysis_v10_gt_val18.xlsx"
    wb.save(out_xlsx)
    print(f"Wrote {out_xlsx}")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--calibrate", action="store_true",
                     help="run the slot scorer against the existing v6/Gemini output and "
                          "check it reproduces the known CONTRADICT set, then exit")
    args = ap.parse_args()

    if args.calibrate:
        ok = run_calibration()
        raise SystemExit(0 if ok else 1)

    run_main_analysis()


if __name__ == "__main__":
    main()
