"""Frame-density test: PROMPT_G_OPT on 18 GT clips × {32f, 64f}.

Compares against the existing 16f baseline (already in outputs/prompt_bakeoff/token_sweep.jsonl
under PROMPT_G_OPT_rerun). Same prompt logic, same clips — only the frame count changes.

Outputs:
    outputs/prompt_bakeoff/frame_density.jsonl     -- per-(clip, density) records (resumable)
    outputs/prompt_bakeoff/frame_density.xlsx      -- 3 sheets: per_clip, summary, failures
    outputs/prompt_bakeoff/frame_density.md        -- CMs + per-clip diff vs 16f baseline

Usage:
    py teacher_distillation/scripts/frame_density_test.py
    py teacher_distillation/scripts/frame_density_test.py --densities 32f
    py teacher_distillation/scripts/frame_density_test.py --max_clips 1
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "teacher_distillation" / "scripts"))

from teacher_bakeoff import (  # noqa: E402
    _build_messages, _calc_cost, _call_model, _load_clip_frames, _parse_response,
)
from teacher_prompt_bakeoff import _read_gt_excel_with_en  # noqa: E402
from apo_metric import score_one, warmup_bertscore  # noqa: E402

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MODEL_SLUG = "google/gemini-3.1-pro-preview"
MODEL_LABEL = "Gemini 3.1 Pro"
PRICE_IN = 2.00
PRICE_OUT = 12.00

TEMPERATURE = 0.1
FRAME_SIZE = 256
DEFAULT_TIMEOUT = 180.0  # longer than bakeoff: 64-frame calls may be slower
DEFAULT_MAX_RETRIES = 3
DEFAULT_RETRY_DELAY = 3.0
DEFAULT_INTER_CALL_DELAY = 1.5

# Density configs: (tag, prompt_file, prompt_var, n_frames)
DENSITY_CONFIGS = {
    "8f": ("PROMPT_G_OPT_8f.py", "PROMPT_G_OPT_8f", 8),
    "32f": ("PROMPT_G_OPT_32f.py", "PROMPT_G_OPT_32f", 32),
    "64f": ("PROMPT_G_OPT_64f.py", "PROMPT_G_OPT_64f", 64),
}

DEFAULT_GT_XLSX = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
DEFAULT_FRAMES_ROOT = REPO_ROOT / "dataset" / "train"
OUT_DIR = REPO_ROOT / "outputs" / "prompt_bakeoff"
DEFAULT_OUT_JSONL = OUT_DIR / "frame_density.jsonl"
DEFAULT_OUT_XLSX = OUT_DIR / "frame_density.xlsx"
DEFAULT_OUT_MD = OUT_DIR / "frame_density.md"

# Token-sweep results to merge as the 16f baseline
TOKEN_SWEEP_JSONL = OUT_DIR / "token_sweep.jsonl"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_prompt(filename: str, variable: str) -> str:
    path = REPO_ROOT / "prompts" / filename
    spec = importlib.util.spec_from_file_location(f"_prompt_{path.stem}", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return getattr(module, variable)


def _record_key(video_id: str, density: str) -> str:
    return f"{video_id}::{density}"


def _load_existing(path: Path) -> Dict[str, Dict]:
    out: Dict[str, Dict] = {}
    if not path.exists():
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                key = _record_key(rec["video_id"], rec["density"])
                out[key] = rec
            except Exception:
                continue
    return out


def _append_jsonl(rec: Dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def _is_done(rec) -> bool:
    return rec is not None and (rec.get("verdict") is not None or bool(rec.get("reasoning")))


def _evaluate_one(
    client: OpenAI,
    density: str,
    prompt_text: str,
    clip: Dict,
    image_b64s: List[str],
    timeout: float,
    max_retries: int,
    retry_delay: float,
) -> Dict:
    messages = _build_messages(prompt_text, image_b64s, detail="low")
    t0 = time.time()
    try:
        raw, usage = _call_model(
            client, MODEL_SLUG, messages,
            timeout=timeout, max_retries=max_retries, retry_delay=retry_delay,
            temperature=TEMPERATURE,
        )
        latency = time.time() - t0
        parsed, verdict = _parse_response(raw)
        cost = _calc_cost(usage, PRICE_IN, PRICE_OUT)
        reasoning = parsed.get("verdict_reasoning") if parsed else None
        sb = score_one(verdict, reasoning, clip["gt_verdict"], clip["gt_reasoning_en"])

        rec = {
            "video_id": f"{clip['video_id']}_{density}",
            "base_video_id": clip["video_id"],
            "density": density,
            "prompt_name": f"PROMPT_G_OPT_{density}",
            "gt_verdict": clip["gt_verdict"],
            "target": clip["target"],
            "t_seconds": clip["t_seconds"],
            "verdict": verdict,
            "reasoning": reasoning,
            "confidence": parsed.get("confidence") if parsed else None,
            "temporal_analysis": parsed.get("temporal_analysis") if parsed else None,
            "full_json": parsed or {},
            "scores": sb.to_dict(),
            "raw": raw,
            "usage": usage,
            "cost_usd": cost,
            "latency_s": round(latency, 2),
            "error": None,
        }
        ok = "[OK]" if verdict == clip["gt_verdict"] else "[XX]"
        print(
            f"    {density} | verdict={verdict or '??':3s} {ok} | "
            f"BERT={sb.alignment:.3f} | words={sb.word_count:3d} | "
            f"composite={sb.composite:.3f} | cost=${cost:.4f} | {latency:.1f}s",
            flush=True,
        )
        return rec
    except Exception as exc:
        latency = time.time() - t0
        print(f"    {density} | ERROR: {exc}", flush=True)
        return {
            "video_id": f"{clip['video_id']}_{density}",
            "base_video_id": clip["video_id"],
            "density": density,
            "prompt_name": f"PROMPT_G_OPT_{density}",
            "gt_verdict": clip["gt_verdict"],
            "target": clip["target"],
            "t_seconds": clip["t_seconds"],
            "verdict": None, "reasoning": None,
            "confidence": None, "temporal_analysis": None, "full_json": {},
            "scores": {"composite": 0.0, "verdict": 0.0, "alignment": 0.0,
                       "length": 0.0, "word_count": 0},
            "raw": "", "usage": {}, "cost_usd": 0.0,
            "latency_s": round(latency, 2),
            "error": str(exc),
        }


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def _confusion_matrix(records: List[Dict]) -> Dict[str, int]:
    tp = fp = fn = tn = parse_err = 0
    for r in records:
        gt = r["gt_verdict"]
        pred = r["verdict"]
        if pred is None:
            parse_err += 1
            if gt == "YES":
                fn += 1
            else:
                fp += 1
            continue
        if gt == "YES" and pred == "YES": tp += 1
        elif gt == "YES" and pred == "NO": fn += 1
        elif gt == "NO" and pred == "YES": fp += 1
        elif gt == "NO" and pred == "NO": tn += 1
    return {"TP": tp, "FP": fp, "FN": fn, "TN": tn, "parse_err": parse_err}


def _load_baseline_16f(path: Path) -> Dict[str, Dict]:
    """Read PROMPT_G_OPT_rerun records (the 16f baseline) from token_sweep.jsonl."""
    out: Dict[str, Dict] = {}
    if not path.exists():
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                if rec.get("prompt_name") == "PROMPT_G_OPT_rerun":
                    out[rec["video_id"]] = rec
            except Exception:
                continue
    return out


def _summarize(
    records_32: List[Dict],
    records_64: List[Dict],
    baseline_16: Dict[str, Dict],
) -> List[Dict]:
    rows = []
    for name, recs in [
        ("16f (PROMPT_G_OPT_rerun)", list(baseline_16.values())),
        ("32f (PROMPT_G_OPT_32f)", records_32),
        ("64f (PROMPT_G_OPT_64f)", records_64),
    ]:
        n = len(recs)
        if n == 0:
            continue
        cm = _confusion_matrix(recs)
        acc = (cm["TP"] + cm["TN"]) / n
        bert = sum(r["scores"]["alignment"] for r in recs) / n
        comp = sum(r["scores"]["composite"] for r in recs) / n
        wc = sum(r["scores"]["word_count"] for r in recs) / n
        len_ok = sum(1 for r in recs if r["scores"]["length"] >= 1.0)
        cost = sum(r.get("cost_usd", 0.0) for r in recs)
        rows.append({
            "density": name, "n": n,
            "TP": cm["TP"], "FP": cm["FP"], "FN": cm["FN"], "TN": cm["TN"],
            "parse_err": cm["parse_err"],
            "accuracy": round(acc, 4),
            "bertscore_mean": round(bert, 4),
            "composite_mean": round(comp, 4),
            "mean_word_count": round(wc, 1),
            "length_ok": f"{len_ok}/{n}",
            "cost_usd": round(cost, 4),
        })
    return rows


def _write_excel(
    clips: List[Dict],
    baseline_16: Dict[str, Dict],
    by_density: Dict[str, Dict[str, Dict]],  # density -> {video_id -> rec}
    summary: List[Dict],
    out_path: Path,
) -> None:
    """3 sheets: per_clip, summary, failures."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    densities = list(by_density.keys())  # e.g. ["32f", "64f"]

    # --- per_clip ---
    rows = []
    for clip in clips:
        vid = clip["video_id"]
        row = {
            "video_id": vid,
            "target": clip["target"],
            "gt_verdict": clip["gt_verdict"],
            "t_seconds": clip["t_seconds"],
            "gt_reasoning_en": clip["gt_reasoning_en"],
        }
        # 16f baseline
        r16 = baseline_16.get(vid, {})
        row["16f__verdict"] = r16.get("verdict")
        row["16f__BERT"] = round(r16.get("scores", {}).get("alignment", 0.0) or 0.0, 4)
        row["16f__composite"] = round(r16.get("scores", {}).get("composite", 0.0) or 0.0, 4)
        row["16f__words"] = r16.get("scores", {}).get("word_count", 0) or 0
        row["16f__reasoning"] = r16.get("reasoning")
        # 32f / 64f
        for d in densities:
            r = by_density[d].get(vid, {})
            row[f"{d}__verdict"] = r.get("verdict")
            row[f"{d}__BERT"] = round(r.get("scores", {}).get("alignment", 0.0) or 0.0, 4)
            row[f"{d}__composite"] = round(r.get("scores", {}).get("composite", 0.0) or 0.0, 4)
            row[f"{d}__words"] = r.get("scores", {}).get("word_count", 0) or 0
            row[f"{d}__reasoning"] = r.get("reasoning")
        rows.append(row)
    df_per_clip = pd.DataFrame(rows)

    # --- summary ---
    df_summary = pd.DataFrame(summary)

    # --- failures: clips where >=1 density was wrong ---
    failure_rows = []
    for clip in clips:
        vid = clip["video_id"]
        gt = clip["gt_verdict"]
        v16 = baseline_16.get(vid, {}).get("verdict")
        v32 = by_density.get("32f", {}).get(vid, {}).get("verdict")
        v64 = by_density.get("64f", {}).get(vid, {}).get("verdict")
        if any(v != gt for v in [v16, v32, v64] if v is not None):
            failure_rows.append({
                "video_id": vid,
                "gt_verdict": gt,
                "16f__verdict": v16,
                "32f__verdict": v32,
                "64f__verdict": v64,
                "16f__correct": v16 == gt,
                "32f__correct": v32 == gt,
                "64f__correct": v64 == gt,
                "gt_reasoning_en": clip["gt_reasoning_en"],
                "16f__reasoning": baseline_16.get(vid, {}).get("reasoning"),
                "32f__reasoning": by_density.get("32f", {}).get(vid, {}).get("reasoning"),
                "64f__reasoning": by_density.get("64f", {}).get(vid, {}).get("reasoning"),
            })
    df_fail = pd.DataFrame(failure_rows) if failure_rows else pd.DataFrame(
        columns=["(no failures)"]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df_per_clip.to_excel(w, sheet_name="per_clip", index=False)
        df_summary.to_excel(w, sheet_name="summary", index=False)
        df_fail.to_excel(w, sheet_name="failures", index=False)

    # styling
    wb = load_workbook(out_path)
    red = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    green = PatternFill(start_color="FFBBFFBB", end_color="FFBBFFBB", fill_type="solid")
    bold = Font(bold=True)

    for sheet in ["per_clip", "failures"]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        for h in headers:
            if h is not None:
                ws.cell(1, headers.index(h) + 1).font = bold
        if "gt_verdict" in headers:
            gt_col = headers.index("gt_verdict") + 1
            for pname in ["16f", "32f", "64f"]:
                v_name = f"{pname}__verdict"
                if v_name in headers:
                    v_col = headers.index(v_name) + 1
                    for r_idx in range(2, ws.max_row + 1):
                        v = ws.cell(r_idx, v_col).value
                        gtv = ws.cell(r_idx, gt_col).value
                        if v is None: continue
                        ws.cell(r_idx, v_col).fill = green if v == gtv else red
        wrap_cols = [h for h in headers if h and (h.endswith("__reasoning") or h == "gt_reasoning_en")]
        for cname in wrap_cols:
            c_idx = headers.index(cname) + 1
            ws.column_dimensions[get_column_letter(c_idx)].width = 55
            for r_idx in range(2, ws.max_row + 1):
                ws.cell(r_idx, c_idx).alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, name in enumerate(headers, start=1):
            if name in wrap_cols: continue
            cl = get_column_letter(col_idx)
            ws.column_dimensions[cl].width = max(12, min(22, len(str(name or "")) + 4))

    ws2 = wb["summary"]
    for c in ws2[1]:
        c.font = bold
    for col_idx in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(out_path)


def _write_markdown(summary: List[Dict], out_path: Path, total_cost: float,
                    baseline_16: Dict[str, Dict],
                    by_density: Dict[str, Dict[str, Dict]]) -> None:
    lines = []
    lines.append("# Frame-Density Test — PROMPT_G_OPT @ 16 / 32 / 64 frames\n")
    lines.append(f"**Model:** {MODEL_LABEL} (`{MODEL_SLUG}`)  ")
    lines.append(f"**Temperature:** {TEMPERATURE}  ")
    lines.append(f"**Clips:** 18 (Hebrew GT translated, col G `verdict_reasoning_en`)  ")
    lines.append(f"**Total cost (32f + 64f runs):** ${total_cost:.2f}  \n")

    lines.append("## Summary\n")
    lines.append("| Density | n | Acc | TP/FP/FN/TN | BERT F1 | Composite | Words (mean) | Length OK | Cost |")
    lines.append("|---------|---|-----|-------------|---------|-----------|--------------|-----------|------|")
    for row in summary:
        lines.append(
            f"| **{row['density']}** | {row['n']} | "
            f"{row['accuracy']*100:.1f}% ({row['TP']+row['TN']}/{row['n']}) | "
            f"{row['TP']}/{row['FP']}/{row['FN']}/{row['TN']} | "
            f"{row['bertscore_mean']:.3f} | **{row['composite_mean']:.3f}** | "
            f"{row['mean_word_count']:.0f} | {row['length_ok']} | ${row['cost_usd']:.2f} |"
        )

    # Per-clip flip analysis
    densities = list(by_density.keys())
    lines.append("\n## Per-clip flip analysis (vs 16f baseline)\n")
    lines.append("| video_id | gt | 16f | " + " | ".join(densities) + " | flip_summary |")
    lines.append("|---|---|---|" + "|".join(["---"] * len(densities)) + "|---|")
    for vid in sorted(baseline_16):
        gt = baseline_16[vid]["gt_verdict"]
        v16 = baseline_16[vid]["verdict"]
        vs = {d: by_density.get(d, {}).get(vid, {}).get("verdict") for d in densities}
        flip = []
        for d, v in vs.items():
            if v == gt and v16 != gt:
                flip.append(f"{d}:FIXED")
            elif v != gt and v16 == gt:
                flip.append(f"{d}:BROKE")
            elif v != gt and v16 != gt:
                flip.append(f"{d}:still-wrong")
        lines.append(
            f"| {vid} | {gt} | {v16 or '?'} | " +
            " | ".join(str(vs[d] or "?") for d in densities) +
            f" | {', '.join(flip) if flip else '-'} |"
        )

    lines.append("\n## Caveats\n")
    lines.append("1. **Same physical window.** All 3 densities cover ~2 sec at FPS=30. Denser sampling adds no NEW visual info, only finer temporal resolution.")
    lines.append("2. **18 clips is small.** A 1-2 clip difference is within noise.")
    lines.append("3. **Not deterministic.** PROMPT_G_OPT_rerun at temp=0.1 gave 11/18 vs the original PROMPT_G_OPT's 12/18 — same prompt, different runs.")
    lines.append("")
    lines.append("## Files\n")
    lines.append("- Raw records: `outputs/prompt_bakeoff/frame_density.jsonl`")
    lines.append("- Excel: `outputs/prompt_bakeoff/frame_density.xlsx`")
    lines.append("- 16f baseline: `outputs/prompt_bakeoff/token_sweep.jsonl` (PROMPT_G_OPT_rerun)")

    out_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--gt_xlsx", default=str(DEFAULT_GT_XLSX))
    parser.add_argument("--frames_root", default=str(DEFAULT_FRAMES_ROOT))
    parser.add_argument("--out_jsonl", default=str(DEFAULT_OUT_JSONL))
    parser.add_argument("--out_xlsx", default=str(DEFAULT_OUT_XLSX))
    parser.add_argument("--out_md", default=str(DEFAULT_OUT_MD))
    parser.add_argument("--densities", default="32f,64f",
                        help="Comma-separated densities to run (32f or 64f)")
    parser.add_argument("--max_clips", type=int, default=0)
    parser.add_argument("--inter_call_delay", type=float, default=DEFAULT_INTER_CALL_DELAY)
    parser.add_argument("--skip_smoke_test", action="store_true")
    parser.add_argument("--smoke_test_threshold", type=float, default=0.20)
    args = parser.parse_args()

    densities = [d.strip() for d in args.densities.split(",") if d.strip()]
    for d in densities:
        if d not in DENSITY_CONFIGS:
            raise SystemExit(f"Unknown density: {d}. Valid: {list(DENSITY_CONFIGS)}")

    # Load prompts
    prompts: Dict[str, str] = {}
    for d in densities:
        pfile, pvar, _n = DENSITY_CONFIGS[d]
        prompts[d] = _load_prompt(pfile, pvar)
        print(f"Loaded prompt {d}: prompts/{pfile}::{pvar} ({len(prompts[d])} chars)")
    print()

    # Env
    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise SystemExit("OPENROUTER_API_KEY missing")
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "MMLM_FrameDensity"),
        },
    )

    # Clips
    clips = _read_gt_excel_with_en(Path(args.gt_xlsx))
    if args.max_clips and args.max_clips > 0:
        clips = clips[: args.max_clips]
    print(f"Loaded {len(clips)} clips")
    print(f"Densities to run: {densities}\n")

    # BERTScore
    print("Warming up BERTScore model...")
    warmup_bertscore()
    print("BERTScore ready.\n")

    # Resume
    out_jsonl = Path(args.out_jsonl)
    existing = _load_existing(out_jsonl)
    done = sum(1 for r in existing.values() if _is_done(r))
    total = len(clips) * len(densities)
    print(f"Resume: {done}/{total} (clip, density) pairs already done\n")

    # Smoke test
    if not args.skip_smoke_test:
        first_d = densities[0]
        first_c = clips[0]
        key = _record_key(first_c["video_id"], first_d)
        if not _is_done(existing.get(key)):
            print(f"=== SMOKE TEST: {first_c['video_id']} x {first_d} ===")
            n_frames = DENSITY_CONFIGS[first_d][2]
            frames_dir = Path(args.frames_root) / f"{first_c['video_id']}_{first_d}"
            indices = list(range(1, n_frames + 1))  # already pre-sampled, sequential
            b64s = _load_clip_frames(frames_dir.parent, frames_dir.name, indices, FRAME_SIZE)
            rec = _evaluate_one(client, first_d, prompts[first_d], first_c, b64s,
                                DEFAULT_TIMEOUT, DEFAULT_MAX_RETRIES, DEFAULT_RETRY_DELAY)
            if rec["cost_usd"] > args.smoke_test_threshold:
                proj = rec["cost_usd"] * len(clips) * len(densities)
                print(f"ABORT: smoke cost ${rec['cost_usd']:.4f} > ${args.smoke_test_threshold:.4f}. "
                      f"Full run would be ~${proj:.2f}.")
                sys.exit(2)
            existing[key] = rec
            _append_jsonl(rec, out_jsonl)
            print()

    # Full run
    pair_idx = 0
    for clip in clips:
        for d in densities:
            pair_idx += 1
            key = _record_key(clip["video_id"], d)
            if _is_done(existing.get(key)):
                print(f"[{pair_idx}/{total}] {clip['video_id']} x {d} -- already done")
                continue
            n_frames = DENSITY_CONFIGS[d][2]
            frames_dir_name = f"{clip['video_id']}_{d}"
            indices = list(range(1, n_frames + 1))  # pre-sampled, sequential
            print(f"[{pair_idx}/{total}] video={clip['video_id']} target={clip['target']} "
                  f"gt={clip['gt_verdict']}  density={d}  frames={n_frames}")
            if args.inter_call_delay > 0:
                time.sleep(args.inter_call_delay)
            b64s = _load_clip_frames(Path(args.frames_root), frames_dir_name, indices, FRAME_SIZE)
            rec = _evaluate_one(client, d, prompts[d], clip, b64s,
                                DEFAULT_TIMEOUT, DEFAULT_MAX_RETRIES, DEFAULT_RETRY_DELAY)
            existing[key] = rec
            _append_jsonl(rec, out_jsonl)

    # Aggregate
    baseline_16 = _load_baseline_16f(TOKEN_SWEEP_JSONL)
    by_density: Dict[str, Dict[str, Dict]] = {}
    for d in densities:
        d_map: Dict[str, Dict] = {}
        for rec in existing.values():
            if rec.get("density") == d:
                d_map[rec["base_video_id"]] = rec
        by_density[d] = d_map

    summary = _summarize(
        records_32=list(by_density.get("32f", {}).values()),
        records_64=list(by_density.get("64f", {}).values()),
        baseline_16=baseline_16,
    )
    total_cost = sum(r.get("cost_usd", 0.0) for r in existing.values())

    _write_excel(clips, baseline_16, by_density, summary, Path(args.out_xlsx))
    print(f"\nWrote Excel: {args.out_xlsx}")

    _write_markdown(summary, Path(args.out_md), total_cost, baseline_16, by_density)
    print(f"Wrote Markdown: {args.out_md}")

    print("\n=== SUMMARY ===")
    for row in summary:
        print(f"  {row['density']:30s} acc={row['accuracy']*100:>5.1f}%  "
              f"TP/FP/FN/TN={row['TP']}/{row['FP']}/{row['FN']}/{row['TN']}  "
              f"BERT={row['bertscore_mean']:.3f}  comp={row['composite_mean']:.3f}")
    print(f"\nTotal cost (32f+64f runs): ${total_cost:.2f}")


if __name__ == "__main__":
    main()
