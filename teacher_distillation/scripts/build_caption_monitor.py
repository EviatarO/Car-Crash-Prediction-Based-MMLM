"""
build_caption_monitor.py
=========================
Deliverable 5 of ~/.claude/plans/but-if-b-a1-it-woolly-metcalfe.md - the
coverage monitor specified in the original Stage-0 plan
(~/.claude/plans/CCP based BADAS/2026-07-07_Plan Semantic-Supervision Crash
Prediction...md:55-58) that was never built (confirmed absent by exploration:
no caption_next_batch.py/caption_generate.py/caption_aggregate.py/
build_caption_monitor.py, no monitor_caption_coverage.xlsx).

Clones build_teacher_monitor.py's TRAIN SHEET ONLY (drop the test sheet, per
the Stage-0 plan's explicit instruction), with two differences:

1. Grid rows come from dataset/manifests/train4500_hires.jsonl (the actual
   4,446-row train4500 pool this plan built), not train.csv x {TTE_0.5,1.0,
   1.5} unconditionally. build_teacher_monitor.py's _train_rows() assigns
   TTE_0.5/1.0/1.5 to EVERY video including negatives, which does not match
   this project's real negative bucket naming (MID/MID-4/MID-8) - this
   version uses each row's actual horizon_label instead.

2. NEW model_verdict column with its OWN color pass, showing whether the
   frozen A0 scorer got that (video_id, bucket) cell right - separate from
   the TTE cell's caption-coverage color. This is what makes the failure set
   navigable in the same view as caption progress, per the plan's stated
   reason for adding this column.

Palette reused verbatim from teacher_reasoning_aggregate.py (HEADER 2E75B6,
STRONG_GREEN 00B050, RED FFC7CE, PANEL FCE4D6) - unlike the teacher monitor,
this grid has no pass1/pass2 distinction (captioning has no debate/pass2
step here), so only STRONG_GREEN/RED/blank are used per column.

Caption coverage will be near-empty on the first run of this script: this
plan is inference-only (captioning the mined failures is explicitly out of
scope, deferred to the next plan), so the TTE cell's color pass mostly
reflects whichever windows already exist in --captions (default
Caption_Train_All_Clips.jsonl, the pre-existing 267-row set, which was NOT
excluded from the train4500 pool - see build_train4500_manifest.py, so some
overlap is expected and legitimate, not a bug).

model_verdict coverage requires --scores (one or more scorer-output JSONL
files) - blank until the pod inference step (deliverable 2b/3) has produced
real output. Uses the exact same (video_id, group) join-key logic as
mine_train_failures.py - see that script's docstring for why position-based
joining would be unsafe across chunked scoring.
"""
from __future__ import annotations

import argparse
import io
import json
from collections import Counter, defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TRAIN4500_MANIFEST = PROJECT_ROOT / "dataset" / "manifests" / "train4500_hires.jsonl"
DEFAULT_CAPTIONS = PROJECT_ROOT / "outputs" / "semantic_captions" / "Caption_Train_All_Clips.jsonl"
OUT_XLSX = PROJECT_ROOT / "outputs" / "train4500_inference" / "monitor_train4500_coverage.xlsx"

# MID moved from offset 0.0 to -10.0 (renamed MID-10) after real chunk-0 scoring found
# the clip-midpoint window produced 42.8% high-confidence false positives, isolated to
# that one bucket - see build_train4500_manifest.py's NEG_BUCKETS comment.
BUCKETS = ["TTE_0.5", "TTE_1.0", "TTE_1.5", "MID-10", "MID-4", "MID-8"]

HEADER = PatternFill("solid", fgColor="2E75B6")
STRONG_GREEN = PatternFill("solid", fgColor="00B050")
RED = PatternFill("solid", fgColor="FFC7CE")
PANEL = PatternFill("solid", fgColor="FCE4D6")


def _norm_vid(v) -> str:
    s = str(v).strip()
    try:
        return f"{int(s):05d}"
    except ValueError:
        return s


def load_jsonl(path: Path) -> list:
    if not path.exists():
        return []
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def load_grid_rows(manifest_path: Path) -> list:
    rows = []
    for r in load_jsonl(manifest_path):
        rows.append({"video_id": _norm_vid(r["video_id"]), "gt": int(r["event_occurs"]),
                     "bucket": r["horizon_label"], "group": r["group"], "frames_dir": r["frames_dir"]})
    rows.sort(key=lambda x: (x["video_id"], x["bucket"]))
    return rows


def _resolve_caption_bucket(req) -> str | None:
    """Map a caption row's requested_time_to_event to one of the 6 canonical
    train4500 bucket labels, or None if it can't be resolved unambiguously.

    Caption_Train_All_Clips.jsonl (the pre-existing 267-row set) predates
    train4500's naming and is messier than expected: positives use clean
    numeric strings ("0.5"/"1"/"1.5"), but its 141 negatives span 7
    heterogeneous offset labels (-4.0_offset, -8.0_offset, TN_MIDPOINT,
    MID-4s_resampled, -5.0_offset, 4.0_offset, -9.0_offset - measured by
    inspection, not assumed). Only offsets that exactly match train4500's own
    0.0/-4.0/-8.0 scheme (build_train4500_manifest.py's NEG_BUCKETS) resolve;
    off-grid values (-5.0_offset, 4.0_offset, -9.0_offset - remnants of an
    older/different windowing scheme) are deliberately left UNRESOLVED rather
    than snapped to the nearest bucket, which would misrepresent real
    coverage. This mirrors, but does not silently paper over, the same
    ambiguity build_teacher_monitor.py's narrower _MID_MAP already has for
    the literal "MID"/"MID-4"/"MID-8" strings.

    MID moved from offset 0.0 to -10.0 (renamed MID-10) after real chunk-0
    scoring found the clip-midpoint window produced 42.8% high-confidence
    false positives. TN_MIDPOINT (legacy captions at offset ~0) therefore no
    longer corresponds to any current bucket and is deliberately left
    UNRESOLVED rather than mapped to MID-10, which is a genuinely different
    window."""
    s = str(req).strip()
    try:
        v = float(s)
        return {0.5: "TTE_0.5", 1.0: "TTE_1.0", 1.5: "TTE_1.5"}.get(v)
    except ValueError:
        pass
    if s.startswith("MID-4"):
        return "MID-4"
    if s.startswith("MID-8"):
        return "MID-8"
    if s.endswith("_offset"):
        try:
            off = float(s.replace("_offset", ""))
            return {-10.0: "MID-10", -4.0: "MID-4", -8.0: "MID-8"}.get(off)
        except ValueError:
            pass
    return None


def load_caption_outcomes(captions_path: Path) -> dict:
    """(video_id, bucket) -> "correct"|"wrong". Caption_Train_All_Clips.jsonl
    has no horizon_label field (verified by inspection) - resolved from
    requested_time_to_event instead, and uses final_verdict, not verdict."""
    out = {}
    n_unresolved = 0
    for r in load_jsonl(captions_path):
        vid = _norm_vid(r.get("video_id"))
        bucket = _resolve_caption_bucket(r.get("requested_time_to_event"))
        if bucket is None:
            n_unresolved += 1
            continue
        gt = str(r.get("gt_verdict", "")).strip().upper()
        verdict = str(r.get("verdict") if r.get("verdict") is not None else r.get("final_verdict", "")).strip().upper()
        verdict = "YES" if verdict in ("1", "YES", "TRUE") else ("NO" if verdict in ("0", "NO", "FALSE") else verdict)
        out[(vid, bucket)] = "correct" if verdict and verdict == gt else "wrong"
    if n_unresolved:
        print(f"  NOTE: {n_unresolved} caption rows had a requested_time_to_event that could not "
              f"be resolved to a train4500 bucket (off-grid legacy offsets) - left uncolored.")
    return out


def load_model_outcomes(score_paths: list, manifest_by_key: dict) -> dict:
    """(video_id, bucket) -> "correct"|"wrong", joined via (video_id, group) -
    the same safe key mine_train_failures.py uses (group is 0/1/2 per video,
    distinct per bucket by construction, safe across chunk concatenation
    order - unlike row position)."""
    out = {}
    for p in score_paths:
        for s in load_jsonl(Path(p)):
            key = (_norm_vid(s["video_id"]), s.get("group"))
            m = manifest_by_key.get(key)
            if m is None:
                continue
            pred = 1 if float(s["score"]) >= 0.5 else 0
            out[(m["video_id"], m["bucket"])] = "correct" if pred == m["gt"] else "wrong"
    return out


def _hist_by_bucket(outcomes_by_bucket: dict, title: str) -> XLImage:
    fig, ax = plt.subplots(figsize=(6.4, 3.6))
    x = range(len(BUCKETS))
    correct = [outcomes_by_bucket[b].get("correct", 0) for b in BUCKETS]
    wrong = [outcomes_by_bucket[b].get("wrong", 0) for b in BUCKETS]
    ax.bar(x, correct, color="#00B050", label="correct")
    ax.bar(x, wrong, bottom=correct, color="#FF7C80", label="wrong")
    ax.set_xticks(list(x))
    ax.set_xticklabels(BUCKETS, rotation=30, ha="right", fontsize=8)
    ax.set_ylabel("windows"); ax.set_title(title, fontsize=10)
    ax.legend(fontsize=8, loc="upper right")
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=120); buf.seek(0); plt.close(fig)
    return XLImage(buf)


def _hist_overall(correct_n: int, wrong_n: int, blank_n: int, title: str) -> XLImage:
    fig, ax = plt.subplots(figsize=(6.4, 3.6))
    labels = ["correct", "wrong", "not-scored"]
    vals = [correct_n, wrong_n, blank_n]
    bars = ax.bar(labels, vals, color=["#00B050", "#FF7C80", "#D9D9D9"])
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(vals, default=1) * 0.01 + 0.1,
                str(v), ha="center", va="bottom", fontsize=10, fontweight="bold")
    ax.set_ylabel("windows"); ax.set_title(title, fontsize=10)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=120); buf.seek(0); plt.close(fig)
    return XLImage(buf)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--manifest", default=str(TRAIN4500_MANIFEST))
    ap.add_argument("--captions", default=str(DEFAULT_CAPTIONS))
    ap.add_argument("--scores", nargs="*", default=[],
                     help="scorer-output JSONL file(s); omit if the pod step hasn't run yet")
    ap.add_argument("--out", default=str(OUT_XLSX))
    args = ap.parse_args()

    rows = load_grid_rows(Path(args.manifest))
    manifest_by_key = {(r["video_id"], r["group"]): {"video_id": r["video_id"], "bucket": r["bucket"],
                        "gt": r["gt"]} for r in rows}
    caption_outcomes = load_caption_outcomes(Path(args.captions))
    model_outcomes = load_model_outcomes(args.scores, manifest_by_key) if args.scores else {}

    print(f"Grid rows: {len(rows)}  |  caption outcomes matched: "
          f"{sum(1 for r in rows if (r['video_id'], r['bucket']) in caption_outcomes)}  |  "
          f"model outcomes matched: {sum(1 for r in rows if (r['video_id'], r['bucket']) in model_outcomes)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("train4500")

    cols = ["video_id", "GT_verdict", "TTE", "model_verdict"]
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    caption_by_bucket = defaultdict(Counter)
    model_by_bucket = defaultdict(Counter)
    tp_captioned = tn_captioned = 0
    tp_scored = tn_scored = 0

    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r["video_id"]).alignment = Alignment(horizontal="center")
        ws.cell(i, 2, r["gt"]).alignment = Alignment(horizontal="center")

        tte_cell = ws.cell(i, 3, r["bucket"])
        tte_cell.alignment = Alignment(horizontal="center")
        cap_oc = caption_outcomes.get((r["video_id"], r["bucket"]))
        if cap_oc:
            tte_cell.fill = STRONG_GREEN if cap_oc == "correct" else RED
            caption_by_bucket[r["bucket"]][cap_oc] += 1
            if r["gt"] == 1:
                tp_captioned += 1
            else:
                tn_captioned += 1

        mv_cell = ws.cell(i, 4, "")
        mv_cell.alignment = Alignment(horizontal="center")
        model_oc = model_outcomes.get((r["video_id"], r["bucket"]))
        if model_oc:
            mv_cell.value = model_oc
            mv_cell.fill = STRONG_GREEN if model_oc == "correct" else RED
            model_by_bucket[r["bucket"]][model_oc] += 1
            if r["gt"] == 1:
                tp_scored += 1
            else:
                tn_scored += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 11
    for c in range(2, 5):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ---- panels ----
    p = 6
    total = len(rows)
    cap_done = tp_captioned + tn_captioned
    cap_correct = sum(caption_by_bucket[b].get("correct", 0) for b in caption_by_bucket)
    cap_wrong = sum(caption_by_bucket[b].get("wrong", 0) for b in caption_by_bucket)
    scored_done = tp_scored + tn_scored
    model_correct = sum(model_by_bucket[b].get("correct", 0) for b in model_by_bucket)
    model_wrong = sum(model_by_bucket[b].get("wrong", 0) for b in model_by_bucket)

    def _panel(r0, title, pairs):
        ws.cell(r0, p, title).font = Font(bold=True)
        ws.cell(r0, p, title).fill = PANEL
        for j, (k, v) in enumerate(pairs, start=1):
            ws.cell(r0 + j, p, k)
            ws.cell(r0 + j, p + 1, v)

    _panel(2, "Caption coverage", [
        ("total windows", total), ("captioned", cap_done), ("not captioned", total - cap_done),
        ("verdict correct", cap_correct), ("verdict wrong", cap_wrong),
        ("TP captioned", tp_captioned), ("TN captioned", tn_captioned),
    ])
    _panel(11, "Model (frozen A0) verdict coverage", [
        ("scored", scored_done), ("not scored", total - scored_done),
        ("model correct", model_correct), ("model wrong", model_wrong),
        ("model error rate", f"{model_wrong/scored_done:.1%}" if scored_done else "-"),
        ("TP scored", tp_scored), ("TN scored", tn_scored),
    ])
    _panel(20, "By bucket (model verdict)", [(b, f"{model_by_bucket[b].get('wrong',0)}/"
                                              f"{sum(model_by_bucket[b].values())}") for b in BUCKETS])
    for col in (p, p + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28 if col == p else 14

    # ---- histograms: model-verdict coverage (the populated signal once the
    # pod has run; caption coverage is expected near-empty this run, so a
    # histogram for it would mostly be noise - the numeric panel above covers it) ----
    img1 = _hist_by_bucket(model_by_bucket, "train4500: model verdict by bucket")
    img1.anchor = f"{get_column_letter(p)}30"
    ws.add_image(img1)
    img2 = _hist_overall(model_correct, model_wrong, total - scored_done,
                          "train4500: overall model-verdict coverage")
    img2.anchor = f"{get_column_letter(p)}50"
    ws.add_image(img2)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Monitor written -> {args.out}")
    print(f"  grid: {total} rows | captioned: {cap_done} ({cap_done/total:.1%}) | "
          f"scored: {scored_done} ({scored_done/total:.1%})")


if __name__ == "__main__":
    main()
