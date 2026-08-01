"""reasoning_analysis_v8_val18.py -- scores PROMPT_SEMSUP_V8_NARRATIVE captions
from qwen/qwen3-vl-235b-a22b-thinking against the 18-clip GT validation set.

Unlike the V4-V7 analyses, the primary axis here is REASONING ALIGNMENT, not
verdict correctness: does delta/cause/ego_response match what gt_reasoning_en
actually describes? The user asked explicitly to focus this round's review on
whether the reasoning aligns with GT, not just the score. reasoning_match is a
separate column/metric from the legacy 0-10 score for that reason.

Reads:
- dataset/manifests/val_e3a.jsonl
- outputs/prompt_bakeoff/highres_test.jsonl + v6_hires_full18.jsonl (t_seconds)
- outputs/prompt_bakeoff/semsup_val18/raw_v8_qwen3vl235b.jsonl

Writes:
- outputs/prompt_bakeoff/semsup_val18/reasoning_analysis_v8_val18.xlsx

Colour rule (whole row):
  GREEN  = verdict correct AND delta/cause/ego_response substantively match GT
  ORANGE = verdict correct but reasoning generic/partial, OR verdict wrong while
           the reasoning still substantively matches GT's mechanism
  RED    = reasoning misses or CONTRADICTS GT's stated mechanism (regardless of
           verdict - a contradicted mechanism is not a usable training signal)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[2]
VAL_MANIFEST = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
T_SECONDS_SOURCES = [
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "highres_test.jsonl",
    REPO_ROOT / "outputs" / "prompt_bakeoff" / "v6_hires_full18.jsonl",
]
V8_JSONL = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "raw_v8_qwen3vl235b.jsonl"
OUT_XLSX = REPO_ROOT / "outputs" / "prompt_bakeoff" / "semsup_val18" / "reasoning_analysis_v8_val18.xlsx"

# (score 0-10, reasoning_match, rationale, colour). reasoning_match is the
# headline judgment: does delta+cause+ego_response substantively match GT's
# stated mechanism, independent of whether the binary verdict landed correctly.
SCORES = {
    "00077": (7, "PARTIAL", "First round to correctly detect ego's non-reaction ('no nose dip, gap continues closing') matching GT's 'EGO fails to brake in time' exactly. Misses that the sedan MERGED into the lane before braking - reports it as already ahead. Score 49, one point under the cut.", "orange"),
    "00147": (5, "PARTIAL", "Correctly avoids a wrong absolute-direction claim, but true_movers still attributes the closing to the OTHER sedan moving into ego's path, when GT says EGO deviates into that sedan's lane - the causality-inversion problem the STEP 4 corollary was meant to fix persists here. Verdict correct.", "orange"),
    "00283": (5, "PARTIAL", "Captures the right shape (a vehicle cuts in, lead vehicle brakes, ego doesn't react) but misidentifies the object (SUV vs stationary pickup+trailer) and the lane/direction (GT: right lane truck turns left; V8: SUV enters from left). Verdict correct.", "orange"),
    "00319": (1, "MISS", "Reports a van moving OUT of view; completely misses GT's car entering the intersection from the right without slowing - the single clip missed by every model and every prompt across all 7 rounds tested to date.", "red"),
    "00372": (1, "MISS", "Reports stable following traffic under a green signal; misses GT's actual mechanism entirely (lead sedan stops for crosswalk pedestrians during a right turn).", "red"),
    "00474": (1, "MISS", "Reports normal green-light proceeding; misses GT's van-performs-a-sharp-left-turn-into-ego-lane event, the same miss recorded on this clip in every round to date.", "red"),
    "00493": (8, "MATCH", "Best-aligned reasoning in this run: true_movers correctly applies the STEP 4 corollary (sedan holds position, ego is the one converging), cause plausibly infers braking traffic, and ego_response correctly reports no reaction - matching GT's 'ego does not slow down' exactly. Score 47, one point under the cut.", "orange"),
    "00529": (1, "CONTRADICT", "true_movers explicitly states the silver SUV 'maintains position relative to ego' - a direct contradiction of GT's claim that the SUV drifts into ego's lane after the left lane is obstructed. Reasoning is wrong, not merely incomplete.", "red"),
    "00687": (1, "CONTRADICT", "Regression vs V6 and V7, both of which read this clip correctly. V8 reports the gray SUV merging AWAY (into the right lane, gap OPENING) - the exact opposite of GT's SUV-drifts-left-into-ego-lane, gap-closing-rapidly mechanism.", "red"),
    "01153": (0, "CONTRADICT", "Worst result in the run: the highest-confidence prediction (score 88) fabricates a sedan 'turning across intersection into ego path.' GT explicitly states all vehicles remain in their own lanes and ego performs an uncontested right turn. This is the same false-crossing-sedan template that broke V5 and V6 on this identical clip - it resurfaces here despite the true_movers grounding mechanism that fixed it in V7.", "red"),
    "01281": (7, "MATCH", "Correctly reports controlled, unchanged following distance to the lead pickup truck with no fabricated hazard; matches GT's 'controlled closing distance... no accident expected' framing.", "green"),
    "01504": (3, "PARTIAL", "Correct verdict via a low score, but misses the actual decisive mechanism GT describes - that ego itself notices the braking ahead and brakes in time. ego_response reports plain following, not ego braking.", "orange"),
    "01550": (2, "CONTRADICT", "New false positive, directly caused by this round's own ego_response cue list. GT explicitly says ego closes the gap 'in a controlled manner' (a real, gentle reaction) - the cue list (nose dip, drop in expansion rate) is tuned for hard braking and misses gradual deceleration, so the model reports 'no reaction' when GT says ego was in fact controlling the approach. TN in V5, V6 and V7; FP here.", "red"),
    "01552": (7, "MATCH", "Reasonable match to GT (minivan crossing and clearing the path, truck ahead) with no fabricated hazard. Notably, the 'school bus' fabrication reproduced on this clip in every one of V4/V5/V6/V7 does NOT appear here.", "green"),
    "01643": (6, "PARTIAL", "Correct verdict, correctly reports no agents entering ego's path, but still introduces parked cars and a 'road work' sign - the same minor fabrication seen on this clip in V6 and V7.", "orange"),
    "01737": (8, "MATCH", "Clean match to GT's empty road / overpass description. Avoids V7's wrong turn-direction claim entirely by using path-relative language rather than asserting a direction.", "green"),
    "02104": (6, "PARTIAL", "Real improvement over V5/V6/V7 on this clip: for the first time, a version detects an actual merge event (a sedan entering the lane) rather than a purely static inventory - though it places the merge in the wrong lane and omits the tow truck GT describes.", "orange"),
    "02117": (7, "MATCH", "Clean match to GT's calm, controlled-following scenario with no fabricated merge event - the hallucination that broke every Gemini run on this clip. Sixth consecutive non-Gemini teacher/prompt combination to resolve it correctly.", "green"),
}

FILL = {
    "green": PatternFill("solid", fgColor="FFC6EFCE"),
    "orange": PatternFill("solid", fgColor="FFFFEB9C"),
    "red": PatternFill("solid", fgColor="FFFFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["video_id", "gt_verdict", "t_seconds", "requested_time_to_event",
           "gt_reasoning_en", "v8_verdict", "risk_score", "reasoning_match",
           "delta", "cause", "ego_response", "v8_caption",
           "score", "score_explanation"]
WIDTHS = [11, 11, 11, 15, 48, 10, 9, 14, 42, 26, 40, 44, 8, 55]
WRAP_COLS = {"gt_reasoning_en", "delta", "ego_response", "v8_caption", "score_explanation"}


def _load_jsonl(p: Path) -> list:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def _load_t_seconds() -> dict:
    out = {}
    for fp in T_SECONDS_SOURCES:
        for r in _load_jsonl(fp):
            vid = str(r.get("video_id"))
            if r.get("t_seconds") is not None and vid not in out:
                out[vid] = r["t_seconds"]
    return out


def main():
    val = {r["video_id"]: r for r in _load_jsonl(VAL_MANIFEST)}
    v8 = {r["video_id"]: r for r in _load_jsonl(V8_JSONL)}
    t_seconds = _load_t_seconds()

    missing_t = set(val) - set(t_seconds)
    if missing_t:
        raise RuntimeError(f"t_seconds missing for {missing_t}")

    rows = []
    for vid in sorted(val):
        v, q = val[vid], v8[vid]
        score, match, rationale, colour = SCORES.get(vid, (None, "MISS", "", "red"))
        rows.append({
            "video_id": vid, "gt_verdict": v["gt_verdict"], "t_seconds": round(t_seconds[vid], 3),
            "requested_time_to_event": v["requested_time_to_event"],
            "gt_reasoning_en": v["gt_reasoning_en"],
            "v8_verdict": "YES" if q["verdict"] == 1 else "NO", "risk_score": q["risk_score"],
            "reasoning_match": match,
            "delta": q["delta"], "cause": q["cause"], "ego_response": q["ego_response"],
            "v8_caption": f"{q['caption_neutral']}, {q['risk_clause']}",
            "score": score, "score_explanation": rationale, "_colour": colour,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "per_clip"
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, rec in enumerate(rows, start=2):
        fill = FILL[rec["_colour"]]
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=i, column=c, value=rec[h])
            wrap = h in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                        horizontal="left" if wrap else "center")
            cell.border = BORDER
            cell.fill = fill
        longest = max(len(str(rec[h])) for h in WRAP_COLS)
        ws.row_dimensions[i].height = max(43.2, 14.4 * (longest // 90 + 1))

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    ws2 = wb.create_sheet("summary")
    n = len(rows)
    correct = sum(1 for r in rows if r["v8_verdict"] == r["gt_verdict"])
    tp = sum(1 for r in rows if r["v8_verdict"] == "YES" and r["gt_verdict"] == "YES")
    fp = sum(1 for r in rows if r["v8_verdict"] == "YES" and r["gt_verdict"] == "NO")
    tn = sum(1 for r in rows if r["v8_verdict"] == "NO" and r["gt_verdict"] == "NO")
    fn = sum(1 for r in rows if r["v8_verdict"] == "NO" and r["gt_verdict"] == "YES")
    scores = [r["score"] for r in rows if r["score"] is not None]
    cc = {c: sum(1 for r in rows if r["_colour"] == c) for c in ("green", "orange", "red")}
    match_counts = {m: sum(1 for r in rows if r["reasoning_match"] == m)
                     for m in ("MATCH", "PARTIAL", "MISS", "CONTRADICT")}

    from sklearn.metrics import average_precision_score, roc_auc_score
    y_true = [1 if r["gt_verdict"] == "YES" else 0 for r in rows]
    y_score = [r["risk_score"] for r in rows]

    metrics = [
        ("n", n),
        ("--- REASONING ALIGNMENT (headline this round) ---", ""),
        ("reasoning MATCH (substantively correct)", match_counts["MATCH"]),
        ("reasoning PARTIAL (right shape, wrong detail)", match_counts["PARTIAL"]),
        ("reasoning MISS (misses GT's mechanism)", match_counts["MISS"]),
        ("reasoning CONTRADICT (states the opposite of GT)", match_counts["CONTRADICT"]),
        ("mean_hand_score (0-10)", round(sum(scores) / len(scores), 2)),
        ("median_hand_score", sorted(scores)[len(scores) // 2]),
        ("n_green / n_orange / n_red", f"{cc['green']} / {cc['orange']} / {cc['red']}"),
        ("--- verdict metrics (secondary) ---", ""),
        ("verdict_accuracy (score>=50 cut)", f"{correct}/{n} ({correct/n:.1%})"),
        ("TP / FP / TN / FN", f"{tp} / {fp} / {tn} / {fn}"),
        ("recall", round(tp / (tp + fn), 3) if (tp + fn) else None),
        ("precision", round(tp / (tp + fp), 3) if (tp + fp) else None),
        ("risk_score AUC", round(roc_auc_score(y_true, y_score), 3)),
        ("risk_score AP", round(average_precision_score(y_true, y_score), 3)),
        ("n_distinct_risk_scores", len(set(y_score))),
    ]
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
    for i, (k, v) in enumerate(metrics, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 22

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}: {n} rows")
    for k, v in metrics:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
