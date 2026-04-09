"""Teacher Dataset Distill v11 -- 100-clip scaled run (50 TP + 50 TN).

Same settings as v9 (PROMPT_G Pass-1 + PROMPT_DEBATE_G Pass-2 on mismatches,
google/gemini-3.1-pro-preview, temperature=0.1).

Key differences from the 18-clip v9 test:
  - 100 clips: 50 TP (3 TTE buckets x ~17 videos) + 50 TN (random frames)
  - TN clips use random frame selection across the full video (not midpoint)
  - TP clips use TTE-bucket selection (0.5s, 1.0s, 1.5s before event)
  - Manifest built by build_teacher_manifest.py with --tn_random_sampling

Inputs:
  --replay_jsonl outputs/manifest_v11_100clips.jsonl
  --frames_root  <local frames path>

Outputs:
  outputs/teacher_dataset_v11.jsonl
  outputs/teacher_dataset_v11.xlsx  (mismatch rows highlighted red)
"""

import argparse
import base64
import json
import os
import re
import time
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from prompts.templates import PROMPT_DEBATE_G, PROMPT_G

try:
    from openai import (
        APIConnectionError as _OAIConnectionError,
        APITimeoutError as _OAITimeoutError,
        OpenAI,
        RateLimitError as _OAIRateLimitError,
    )
    _HAS_OPENAI_ERRORS = True
except ImportError:
    from openai import OpenAI  # type: ignore[no-redef]
    _HAS_OPENAI_ERRORS = False
    _OAIConnectionError = Exception  # type: ignore[assignment,misc]
    _OAITimeoutError = Exception      # type: ignore[assignment,misc]
    _OAIRateLimitError = Exception    # type: ignore[assignment,misc]

try:
    import httpx as _httpx
except ImportError:
    _httpx = None  # type: ignore[assignment]


DEFAULT_MODEL = "google/gemini-3.1-pro-preview"
PROMPT_ID = "SEMI_ORACLE_DEBATE_V11"

# ---------------------------------------------------------------------------
# Helpers: normalisation
# ---------------------------------------------------------------------------

def _normalize_video_id(value: object) -> str:
    return f"{int(float(value)):05d}"


def _normalize_verdict(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


def _normalize_confidence(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"HIGH", "MEDIUM", "LOW"} else None


# ---------------------------------------------------------------------------
# Image encoding
# ---------------------------------------------------------------------------

def _encode_image(path: Path, frame_size: int) -> str:
    if path.exists():
        img = Image.open(path).convert("RGB")
    else:
        img = Image.new("RGB", (frame_size, frame_size), color=(0, 0, 0))
    if frame_size and img.size != (frame_size, frame_size):
        img = img.resize((frame_size, frame_size))
    buf = BytesIO()
    img.save(buf, format="JPEG")
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"


# ---------------------------------------------------------------------------
# Message building
# ---------------------------------------------------------------------------

def _build_messages(prompt: str, image_b64s: Sequence[str], detail: Optional[str]) -> List[Dict]:
    content: List[Dict] = [{"type": "text", "text": prompt}]
    for b64 in image_b64s:
        image_url: Dict = {"url": b64}
        if detail:
            image_url["detail"] = detail
        content.append({"type": "image_url", "image_url": image_url})
    return [{"role": "user", "content": content}]


# ---------------------------------------------------------------------------
# JSON extraction
# ---------------------------------------------------------------------------

def _extract_json_object(raw: str) -> Optional[Dict]:
    """Try three strategies to extract a JSON object from a model response."""
    if not raw:
        return None
    # Strategy 1: entire response is valid JSON.
    try:
        obj = json.loads(raw)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass
    # Strategy 2: fenced code block ```json ... ``` or ``` ... ```.
    fenced = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", raw, flags=re.IGNORECASE)
    if fenced:
        try:
            obj = json.loads(fenced.group(1))
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    # Strategy 3: first { to last } in the entire string.
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            obj = json.loads(raw[start : end + 1])
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    return None


def _parse_prompt_g_response(raw: str) -> Tuple[Optional[Dict], Optional[str]]:
    """Return (parsed_dict, collision_verdict) from a PROMPT_G / PROMPT_DEBATE_G response."""
    parsed = _extract_json_object(raw)
    if parsed is None:
        return None, None
    parsed["collision_verdict"] = _normalize_verdict(parsed.get("collision_verdict"))
    parsed["confidence"] = _normalize_confidence(parsed.get("confidence"))
    return parsed, parsed["collision_verdict"]


# ---------------------------------------------------------------------------
# Model call with per-type retry / backoff
# ---------------------------------------------------------------------------

def _call_model(
    client: OpenAI,
    model: str,
    messages: List[Dict],
    timeout: float,
    max_retries: int,
    retry_delay: float,
    temperature: float = 0.1,
) -> Tuple[str, Dict]:
    """Call the chat completions endpoint with robust retry logic.

    Retry strategy (all retried up to max_retries):
      - APITimeoutError     : exponential back-off (was previously raising immediately)
      - RateLimitError      : 2x back-off to respect rate windows
      - APIConnectionError  : exponential back-off for transient network failures
      - Other exceptions    : exponential back-off

    A RuntimeError is raised only after all retries are exhausted.
    """
    last_exc: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                timeout=timeout,
            )
            text = response.choices[0].message.content if response.choices else ""
            usage = (
                response.usage.model_dump()
                if hasattr(response, "usage") and response.usage
                else {}
            )
            return text or "", usage

        except _OAITimeoutError as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(
                f"  [retry {attempt}/{max_retries}] timeout after {timeout}s -- "
                f"waiting {wait:.1f}s before retry",
                flush=True,
            )
            if attempt < max_retries:
                time.sleep(wait)

        except _OAIRateLimitError as exc:
            last_exc = exc
            wait = retry_delay * 2 * (2 ** (attempt - 1))
            print(
                f"  [retry {attempt}/{max_retries}] rate-limit -- "
                f"waiting {wait:.1f}s before retry",
                flush=True,
            )
            if attempt < max_retries:
                time.sleep(wait)

        except _OAIConnectionError as exc:
            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            print(
                f"  [retry {attempt}/{max_retries}] connection error -- "
                f"waiting {wait:.1f}s before retry",
                flush=True,
            )
            if attempt < max_retries:
                time.sleep(wait)

        except Exception as exc:
            # Catch httpx-level timeouts when raised before openai wraps them.
            is_http_timeout = False
            if _httpx is not None and isinstance(exc, _httpx.TimeoutException):
                is_http_timeout = True
            if not is_http_timeout and "timeout" in str(exc).lower():
                is_http_timeout = True

            last_exc = exc
            wait = retry_delay * (2 ** (attempt - 1))
            label = "timeout" if is_http_timeout else "error"
            print(
                f"  [retry {attempt}/{max_retries}] {label}: {exc!r} -- "
                f"waiting {wait:.1f}s before retry",
                flush=True,
            )
            if attempt < max_retries:
                time.sleep(wait)

    raise RuntimeError(
        f"OpenRouter call failed after {max_retries} attempts: {last_exc}"
    ) from last_exc


# ---------------------------------------------------------------------------
# Clip loading & materialisation
# ---------------------------------------------------------------------------

def _load_replay_clips(path: Path) -> List[Dict]:
    clips: List[Dict] = []
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            raw = line.strip()
            if raw:
                clips.append(json.loads(raw))
    if not clips:
        raise RuntimeError(f"No clips found in replay JSONL: {path}")
    return clips


def _materialize_clips(
    replay_clips: List[Dict],
    frames_root: Path,
    frame_size: int,
    fps: float,
) -> List[Dict]:
    out: List[Dict] = []
    for rec in replay_clips:
        video_id = _normalize_video_id(rec.get("video_id"))
        frame_indices = rec.get("frame_indices")
        if not isinstance(frame_indices, list) or not frame_indices:
            raise RuntimeError(f"Replay clip missing frame_indices for video {video_id}")

        frames_dir = frames_root / video_id
        image_b64s: List[str] = []
        for idx in frame_indices:
            frame_path = frames_dir / f"frame_{int(idx):05d}.jpg"
            image_b64s.append(_encode_image(frame_path, frame_size=frame_size))

        end_frame_idx = int(rec.get("end_frame_idx", frame_indices[-1]))
        t_seconds = rec.get("t_seconds")
        if t_seconds is None:
            t_seconds = float(end_frame_idx) / max(1e-9, fps)

        target = int(rec.get("target", 0))
        gt_verdict: str = rec.get("gt_verdict", "YES" if target == 1 else "NO")

        out.append(
            {
                "video_id": video_id,
                "end_frame_idx": end_frame_idx,
                "frame_indices": frame_indices,
                "window_size": int(rec.get("window_size", 16)),
                "stride": int(rec.get("stride", 4)),
                "target": target,
                "gt_verdict": gt_verdict,
                "time_to_event": rec.get("time_to_event"),
                "time_of_alert": rec.get("time_of_alert"),
                "time_of_event": rec.get("time_of_event"),
                "t_seconds": t_seconds,
                "target_risk": rec.get("target_risk"),
                "requested_time_to_event": rec.get("requested_time_to_event"),
                "image_b64s": image_b64s,
            }
        )
    return out


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _dynamic_objects_to_str(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, list):
        parts: List[str] = []
        for item in value:
            if isinstance(item, dict):
                t = item.get("type", "?")
                pos = item.get("position", "?")
                feat = item.get("feature", "")
                parts.append(f"{t} ({pos}{', ' + feat if feat else ''})")
            else:
                parts.append(str(item))
        return " | ".join(parts) if parts else None
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=True)
    return str(value)


def _write_excel(records: List[Dict], output_xlsx: Path) -> None:
    rows: List[Dict] = []
    for rec in records:
        target = rec.get("target")
        requested_tte = rec.get("requested_time_to_event")
        if target == 0 and requested_tte is None:
            requested_tte = "TN_MIDPOINT"

        rows.append(
            {
                # Metadata
                "video_id": rec.get("video_id"),
                "target": target,
                "gt_verdict": rec.get("gt_verdict"),
                "t_seconds": rec.get("t_seconds"),
                "requested_time_to_event": requested_tte,
                # Pass 1
                "collision_verdict": rec.get("collision_verdict"),
                "confidence": rec.get("confidence"),
                "verdict_reasoning": rec.get("verdict_reasoning"),
                "scene_context": rec.get("scene_context"),
                "dynamic_objects": _dynamic_objects_to_str(rec.get("dynamic_objects")),
                "temporal_analysis": rec.get("temporal_analysis"),
                "occlusion_check": rec.get("occlusion_check"),
                "time_to_contact": rec.get("time_to_contact"),
                # Comparison
                "mismatch": rec.get("mismatch"),
                # Pass 2
                "p2_collision_verdict": rec.get("p2_collision_verdict"),
                "p2_confidence": rec.get("p2_confidence"),
                "p2_verdict_reasoning": rec.get("p2_verdict_reasoning"),
                "p2_scene_context": rec.get("p2_scene_context"),
                "p2_dynamic_objects": _dynamic_objects_to_str(rec.get("p2_dynamic_objects")),
                "p2_temporal_analysis": rec.get("p2_temporal_analysis"),
                "p2_occlusion_check": rec.get("p2_occlusion_check"),
                "p2_time_to_contact": rec.get("p2_time_to_contact"),
                # Final (uniform PROMPT_G schema for student SFT)
                "final_verdict": rec.get("final_verdict"),
                "final_reasoning": rec.get("final_reasoning"),
                # Diagnostics
                "error": rec.get("error"),
                "review": "",
            }
        )

    df = pd.DataFrame(rows)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_xlsx, index=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active

    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    n_cols = len(df.columns)

    # Highlight mismatch rows in red.
    for row_idx, rec in enumerate(records, start=2):
        if rec.get("mismatch"):
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

    # Auto-fit column widths (cap at 60).
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, len(df) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_xlsx)


# ---------------------------------------------------------------------------
# Summary helpers
# ---------------------------------------------------------------------------

def _confusion_counts(records: List[Dict], verdict_key: str) -> Dict[str, int]:
    tp = fp = tn = fn = 0
    for rec in records:
        target = rec.get("target")
        pred = _normalize_verdict(rec.get(verdict_key))
        if target == 1 and pred == "YES":
            tp += 1
        elif target == 1 and pred == "NO":
            fn += 1
        elif target == 0 and pred == "NO":
            tn += 1
        elif target == 0 and pred == "YES":
            fp += 1
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn}


def _accuracy(cm: Dict[str, int]) -> float:
    total = cm["TP"] + cm["FP"] + cm["TN"] + cm["FN"]
    return (cm["TP"] + cm["TN"]) / total if total > 0 else 0.0


def _load_v2_accuracy(path: Path) -> Optional[float]:
    if not path.exists():
        return None
    records: List[Dict] = []
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            raw = line.strip()
            if raw:
                try:
                    records.append(json.loads(raw))
                except Exception:
                    pass
    if not records:
        return None
    cm = _confusion_counts(records, verdict_key="collision_verdict")
    return _accuracy(cm)


# ---------------------------------------------------------------------------
# Checkpoint
# ---------------------------------------------------------------------------

def _load_checkpoint(output_jsonl: Path) -> Dict[str, Dict]:
    """Return {video_id: record} for clips already fully processed (no error, verdict present)."""
    checkpoint: Dict[str, Dict] = {}
    if not output_jsonl.exists():
        return checkpoint
    with open(output_jsonl, "r", encoding="utf-8") as fh:
        for line in fh:
            raw = line.strip()
            if not raw:
                continue
            try:
                rec = json.loads(raw)
            except Exception:
                continue
            vid = rec.get("video_id")
            if not vid:
                continue
            # Only checkpoint when the record is fully complete:
            #   - no error logged
            #   - Pass-1 produced a parseable verdict
            #   - final_verdict is populated (covers mismatch rows where Pass-2
            #     may have succeeded or been correctly skipped)
            is_complete = (
                rec.get("error") is None
                and rec.get("collision_verdict") is not None
                and rec.get("final_verdict") is not None
            )
            if is_complete:
                checkpoint[vid] = rec
    return checkpoint


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Semi-Oracle with Debate v11 (100-clip scaled run)")
    parser.add_argument("--replay_jsonl", required=True, help="Path to replay JSONL (e.g. teacher_dataset_v8.jsonl)")
    parser.add_argument("--frames_root", required=True, help="Root directory containing per-video frame folders")
    parser.add_argument("--output_jsonl", required=True, help="Path to write output JSONL")
    parser.add_argument("--output_xlsx", default="", help="Path to write output XLSX (default: same stem as --output_jsonl)")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--fps", type=float, default=30.0)
    parser.add_argument("--frame_size", type=int, default=256)
    parser.add_argument("--timeout", type=float, default=90.0, help="Per-call HTTP timeout in seconds (default 90)")
    parser.add_argument("--max_retries", type=int, default=3, help="Max attempts per API call (default 3)")
    parser.add_argument("--retry_delay", type=float, default=3.0, help="Base retry back-off in seconds (default 3)")
    parser.add_argument("--detail", default="low", choices=["low", "high", "auto"])
    parser.add_argument("--temperature", type=float, default=0.1)
    parser.add_argument(
        "--inter_clip_delay",
        type=float,
        default=2.0,
        help="Seconds to wait between clips (default 2); set 0 to disable",
    )
    parser.add_argument(
        "--inter_pass_delay",
        type=float,
        default=1.0,
        help="Seconds to wait between Pass 1 and Pass 2 of the same clip (default 1)",
    )
    parser.add_argument(
        "--v2_jsonl",
        default="outputs/teacher_dataset_v2.jsonl",
        help="Path to v2 JSONL for accuracy comparison (optional)",
    )
    args = parser.parse_args()

    load_dotenv()
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY environment variable is not set")

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": os.environ.get("OPENROUTER_HTTP_REFERER", "http://localhost"),
            "X-Title": os.environ.get("OPENROUTER_APP_TITLE", "MMLM_Teacher_Distill_V11"),
        },
    )

    replay_path = Path(args.replay_jsonl)
    frames_root = Path(args.frames_root)
    output_jsonl = Path(args.output_jsonl)
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx = Path(args.output_xlsx) if args.output_xlsx else output_jsonl.with_suffix(".xlsx")

    # Load clips.
    replay_clips = _load_replay_clips(replay_path)
    clips = _materialize_clips(
        replay_clips,
        frames_root=frames_root,
        frame_size=args.frame_size,
        fps=args.fps,
    )

    # Load checkpoint from any previous partial run.
    checkpoint = _load_checkpoint(output_jsonl)
    if checkpoint:
        print(
            f"Checkpoint: {len(checkpoint)} clip(s) already done -- will skip: "
            f"{', '.join(sorted(checkpoint))}",
            flush=True,
        )

    print(
        f"Loaded {len(clips)} clips from {replay_path}\n"
        f"  model={args.model}  temp={args.temperature}  "
        f"frame_size={args.frame_size}  detail={args.detail}  "
        f"timeout={args.timeout}s  max_retries={args.max_retries}",
        flush=True,
    )

    run_records: List[Dict] = []
    total = len(clips)
    pass2_calls = 0
    mismatch_video_ids: List[str] = []

    # Write output file; prepend checkpoint records so file is always valid on interrupt.
    with open(output_jsonl, "w", encoding="utf-8") as fout:
        for ck_rec in checkpoint.values():
            fout.write(json.dumps(ck_rec, ensure_ascii=True) + "\n")
            run_records.append(ck_rec)
            if ck_rec.get("mismatch"):
                pass2_calls += 1
                mismatch_video_ids.append(ck_rec["video_id"])

        for i, clip in enumerate(clips, start=1):
            if clip["video_id"] in checkpoint:
                print(
                    f"[{i}/{total}] SKIP video={clip['video_id']} (checkpoint)",
                    flush=True,
                )
                continue

            # Polite inter-clip delay (skip before very first real clip).
            active_count = sum(
                1 for c in clips[:i-1] if c["video_id"] not in checkpoint
            )
            if active_count > 0 and args.inter_clip_delay > 0:
                time.sleep(args.inter_clip_delay)

            # ------------------------------------------------------------------
            # Pass 1 -- blind prediction with PROMPT_G
            # ------------------------------------------------------------------
            print(
                f"[{i}/{total}] PASS-1 video={clip['video_id']} "
                f"target={clip['target']} gt={clip['gt_verdict']}",
                flush=True,
            )
            messages_p1 = _build_messages(PROMPT_G, clip["image_b64s"], detail=args.detail)
            raw_p1 = ""
            usage_p1: Dict = {}
            err_p1: Optional[str] = None
            t0 = time.time()
            try:
                raw_p1, usage_p1 = _call_model(
                    client=client,
                    model=args.model,
                    messages=messages_p1,
                    timeout=args.timeout,
                    max_retries=args.max_retries,
                    retry_delay=args.retry_delay,
                    temperature=args.temperature,
                )
            except Exception as exc:
                err_p1 = str(exc)
            latency_p1 = time.time() - t0

            parsed_p1, verdict_p1 = _parse_prompt_g_response(raw_p1)

            # ------------------------------------------------------------------
            # Mismatch check (CRITICAL FIX: verdict_p1 must be non-None)
            # Only trigger Pass 2 when:
            #   (a) Pass 1 completed without error,
            #   (b) Pass 1 returned a parseable YES/NO verdict, AND
            #   (c) that verdict disagrees with the ground-truth label.
            # ------------------------------------------------------------------
            mismatch = (
                err_p1 is None
                and verdict_p1 is not None
                and verdict_p1 != clip["gt_verdict"]
            )

            # ------------------------------------------------------------------
            # Pass 2 -- debate advocate call (only on mismatch)
            # ------------------------------------------------------------------
            raw_p2 = ""
            usage_p2: Dict = {}
            err_p2: Optional[str] = None
            latency_p2: Optional[float] = None
            parsed_p2: Optional[Dict] = None
            verdict_p2: Optional[str] = None
            debate_position: Optional[str] = None

            if mismatch:
                pass2_calls += 1
                mismatch_video_ids.append(clip["video_id"])
                debate_position = "WILL" if clip["gt_verdict"] == "YES" else "WILL NOT"
                print(
                    f"[{i}/{total}] PASS-2 video={clip['video_id']} "
                    f"debate_position=\"{debate_position}\"",
                    flush=True,
                )
                if args.inter_pass_delay > 0:
                    time.sleep(args.inter_pass_delay)
                debate_prompt = PROMPT_DEBATE_G.replace("{debate_position}", debate_position)
                messages_p2 = _build_messages(debate_prompt, clip["image_b64s"], detail=args.detail)
                t1 = time.time()
                try:
                    raw_p2, usage_p2 = _call_model(
                        client=client,
                        model=args.model,
                        messages=messages_p2,
                        timeout=args.timeout,
                        max_retries=args.max_retries,
                        retry_delay=args.retry_delay,
                        temperature=args.temperature,
                    )
                except Exception as exc:
                    err_p2 = str(exc)
                latency_p2 = time.time() - t1
                parsed_p2, verdict_p2 = _parse_prompt_g_response(raw_p2)

            # ------------------------------------------------------------------
            # Final verdict: Pass-1 if match; Pass-2 if mismatch
            # ------------------------------------------------------------------
            if not mismatch:
                final_verdict = verdict_p1
                final_reasoning = parsed_p1.get("verdict_reasoning") if parsed_p1 else None
            else:
                final_verdict = verdict_p2
                final_reasoning = parsed_p2.get("verdict_reasoning") if parsed_p2 else None

            # Compose error string.
            error_parts: List[str] = []
            if err_p1:
                error_parts.append(f"pass1={err_p1}")
            if err_p2:
                error_parts.append(f"pass2={err_p2}")
            # Flag the silent case: mismatch triggered, Pass-2 call succeeded (no err_p2)
            # but the response couldn't be parsed into a valid verdict.
            if mismatch and err_p2 is None and verdict_p2 is None:
                error_parts.append("pass2=parse_fail")
            error_value: Optional[str] = " | ".join(error_parts) if error_parts else None

            # ------------------------------------------------------------------
            # Build record
            # ------------------------------------------------------------------
            record: Dict = {
                # Provenance
                "video_id": clip["video_id"],
                "end_frame_idx": clip["end_frame_idx"],
                "frame_indices": clip["frame_indices"],
                "window_size": clip["window_size"],
                "stride": clip["stride"],
                "target": clip["target"],
                "gt_verdict": clip["gt_verdict"],
                "time_to_event": clip["time_to_event"],
                "time_of_alert": clip["time_of_alert"],
                "time_of_event": clip["time_of_event"],
                "t_seconds": clip["t_seconds"],
                "target_risk": clip["target_risk"],
                "requested_time_to_event": clip["requested_time_to_event"],
                "model_id": args.model,
                "prompt_id": PROMPT_ID,
                # Pass-1 fields
                "latency_p1_s": round(latency_p1, 3),
                "input_tokens_p1": usage_p1.get("prompt_tokens"),
                "output_tokens_p1": usage_p1.get("completion_tokens"),
                "scene_context": parsed_p1.get("scene_context") if parsed_p1 else None,
                "dynamic_objects": parsed_p1.get("dynamic_objects") if parsed_p1 else None,
                "temporal_analysis": parsed_p1.get("temporal_analysis") if parsed_p1 else None,
                "occlusion_check": parsed_p1.get("occlusion_check") if parsed_p1 else None,
                "time_to_contact": parsed_p1.get("time_to_contact") if parsed_p1 else None,
                "collision_verdict": verdict_p1,
                "confidence": parsed_p1.get("confidence") if parsed_p1 else None,
                "verdict_reasoning": parsed_p1.get("verdict_reasoning") if parsed_p1 else None,
                # Comparison
                "mismatch": mismatch,
                "debate_position": debate_position,
                # Pass-2 fields (None when no mismatch)
                "latency_p2_s": round(latency_p2, 3) if latency_p2 is not None else None,
                "input_tokens_p2": usage_p2.get("prompt_tokens"),
                "output_tokens_p2": usage_p2.get("completion_tokens"),
                "p2_scene_context": parsed_p2.get("scene_context") if parsed_p2 else None,
                "p2_dynamic_objects": parsed_p2.get("dynamic_objects") if parsed_p2 else None,
                "p2_temporal_analysis": parsed_p2.get("temporal_analysis") if parsed_p2 else None,
                "p2_occlusion_check": parsed_p2.get("occlusion_check") if parsed_p2 else None,
                "p2_time_to_contact": parsed_p2.get("time_to_contact") if parsed_p2 else None,
                "p2_collision_verdict": verdict_p2,
                "p2_confidence": parsed_p2.get("confidence") if parsed_p2 else None,
                "p2_verdict_reasoning": parsed_p2.get("verdict_reasoning") if parsed_p2 else None,
                # Final (uniform PROMPT_G schema for student SFT)
                "final_verdict": final_verdict,
                "final_reasoning": final_reasoning,
                # Raw responses for audit
                "raw_response_pass1": raw_p1,
                "raw_response_pass2": raw_p2,
                # Diagnostics
                "error": error_value,
            }

            run_records.append(record)
            fout.write(json.dumps(record, ensure_ascii=True) + "\n")
            fout.flush()

            status_tag = "ok" if error_value is None else "error"
            parse_note = "" if verdict_p1 is not None else " [parse-fail]"
            print(
                f"[{i}/{total}] video={clip['video_id']} target={clip['target']} "
                f"p1={verdict_p1}{parse_note} mismatch={mismatch} p2={verdict_p2} "
                f"final={final_verdict} latency_p1={latency_p1:.1f}s "
                f"status={status_tag}",
                flush=True,
            )

    # --------------------------------------------------------------------------
    # Excel output
    # --------------------------------------------------------------------------
    _write_excel(run_records, output_xlsx=output_xlsx)

    # --------------------------------------------------------------------------
    # Summary statistics
    # --------------------------------------------------------------------------
    cm_p1 = _confusion_counts(run_records, verdict_key="collision_verdict")
    p1_acc = _accuracy(cm_p1)
    v2_acc = _load_v2_accuracy(Path(args.v2_jsonl))
    err_count = sum(1 for rec in run_records if rec.get("error"))
    parse_fail_count = sum(
        1 for rec in run_records
        if rec.get("error") is None and rec.get("collision_verdict") is None
    )

    print("\n" + "=" * 60, flush=True)
    print("Semi-Oracle with Debate Summary  (v11 / SEMI_ORACLE_DEBATE_V11)", flush=True)
    print("=" * 60, flush=True)
    print("Pass-1 Confusion Matrix (PROMPT_G blind, 100 clips):", flush=True)
    print(
        f"  TP={cm_p1['TP']}  FP={cm_p1['FP']}  TN={cm_p1['TN']}  FN={cm_p1['FN']}",
        flush=True,
    )
    print(f"  Pass-1 Accuracy: {p1_acc:.1%}", flush=True)
    if v2_acc is not None:
        print(f"  v2   Accuracy:   {v2_acc:.1%}", flush=True)
        print(f"  Delta (v11 p1 - v2): {p1_acc - v2_acc:+.1%}", flush=True)
    else:
        print(f"  v2 Accuracy: N/A (file not found: {args.v2_jsonl})", flush=True)
    print(f"\nMismatches (Pass-2 used): {pass2_calls}/{total}", flush=True)
    if mismatch_video_ids:
        print(f"  Mismatch video_ids: {', '.join(mismatch_video_ids)}", flush=True)
    print(f"Total API calls: {total + pass2_calls}", flush=True)
    print(f"\nErrors:       {err_count}/{total}", flush=True)
    print(f"Parse fails:  {parse_fail_count}/{total}", flush=True)
    print(f"\nWrote JSONL: {output_jsonl}", flush=True)
    print(f"Wrote XLSX:  {output_xlsx}", flush=True)
    print("=" * 60, flush=True)


if __name__ == "__main__":
    main()
