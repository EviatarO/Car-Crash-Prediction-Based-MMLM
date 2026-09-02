"""
build_semtest200_comparison.py
================================
Per-clip comparison workbook for the SemTest-200(-v2) arms (vision / v10 / v12 /
v12shuf), modeled on build_pool1761_comparison.py: same header styling, conditional-
formatting convention (bgColor for dxf fills - fgColor is silently invisible in
Excel's CF renderer, see that script's note), widths, freeze panes.

Key difference from the pool1761 workbook: there is only ONE pre-FT column (A0), not
per-arm baselines - every arm starts from A0 exactly (LoRA B-matrix zero-init at step 0
means the untrained model IS A0, bit-for-bit on the forward pass). Each arm's column is
its POST-FT score, plus a signed gap column (post - A0).

Sheets:
  per_clip            - captions, A0 (pre-FT), per-arm (post-FT + gap), source tier
  summary_vs_A0       - fixed/broken/still_wrong/net vs the pre-FT baseline
  summary_vs_vision   - same, vs the vision-only arm (did captions add anything)
  metrics             - AP/AUC/acc@0.5 per arm per split (train/val/all), n, epoch
  metrics_stratified  - AP/AUC/acc@0.5 per arm x {hard / easy / all}, where "easy"
                        is every row whose source tier starts with "easy_" (v2's
                        +100 A0-correct anchor clips - see select_semtest200_easy.py).
                        This is what answers "did the easy clips just inflate the
                        number" - the hard-subset row is the SAME readout the v1
                        workbook already reported, unpooled.

Works for both v1 (200 rows, no easy_* tier, --epoch-label an int per arm) and v2
(300 rows, easy_TN/easy_TP present, --epoch-label "CV" when scores come from pooled
cross-validation rather than one selected epoch).

Usage:
  python build_semtest200_comparison.py                     # v1, defaults
  python build_semtest200_comparison.py \
      --st-dir ../../outputs/semtest200_v2 \
      --scores-dir ../../outputs/semtest200_v2/scores \
      --selection ../../outputs/semtest200_v2/selection_v2.jsonl \
      --n-expected 300 --epoch-label CV \
      --v10-captions None --v12-captions None   # omit if not yet generated
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from sklearn.metrics import average_precision_score, roc_auc_score

ROOT = Path(__file__).resolve().parents[2]

DEFAULT_ARMS = ["vision", "v10", "v12", "v12shuf"]
DEFAULT_SELECTED_EPOCH = {"vision": 8, "v10": 8, "v12": 10, "v12shuf": 10}

# NOTE: conditional-formatting (dxf) fills render from bgColor, not fgColor - the
# reverse of a normal cell fill (see build_pool1761_comparison.py's identical note,
# confirmed 2026-08 via Excel COM screenshot).
GREEN = PatternFill(patternType="solid", bgColor="C6EFCE")
RED = PatternFill(patternType="solid", bgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BASE_FILL = PatternFill("solid", fgColor="4A5A80")
NOTE_FONT = Font(italic=True, color="666666", size=9)


def load_jsonl(path):
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def is_ok(score, gt):
    return (score >= 0.5) == (gt == "YES")


def is_easy(source_tier):
    return source_tier.startswith("easy_")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--st-dir", default=str(ROOT / "outputs" / "semtest200"))
    ap.add_argument("--scores-dir", default=None,
                     help="default: <st-dir>/scores")
    ap.add_argument("--selection", default=None,
                     help="default: <st-dir>/selection.jsonl")
    ap.add_argument("--v10-captions", default="__default__",
                     help="V10 caption jsonl for the per_clip sheet, or 'None' to "
                          "omit the column (e.g. captions not yet generated for a "
                          "v2 pool). Default: <st-dir>/Caption_semtest200_V10.jsonl")
    ap.add_argument("--v12-captions", default="__default__",
                     help="same as --v10-captions, for V12")
    ap.add_argument("--out", default=None, help="default: <st-dir>/semtest200_arm_comparison.xlsx")
    ap.add_argument("--n-expected", type=int, default=200)
    ap.add_argument("--arms", nargs="+", default=DEFAULT_ARMS)
    ap.add_argument("--epoch-label", default=None,
                     help="shown in the metrics sheet's selected_epoch column, e.g. "
                          "'CV' for a pooled-cross-validation run. Default: the "
                          "per-arm epoch used for the original single-split v1 run.")
    args = ap.parse_args()

    st_dir = Path(args.st_dir)
    scores_dir = Path(args.scores_dir) if args.scores_dir else st_dir / "scores"
    selection_path = Path(args.selection) if args.selection else st_dir / "selection.jsonl"
    out_path = Path(args.out) if args.out else st_dir / "semtest200_arm_comparison.xlsx"
    arms = args.arms
    epoch_of = ({a: args.epoch_label for a in arms} if args.epoch_label
                else DEFAULT_SELECTED_EPOCH)

    def resolve_captions(flag, default_name):
        if flag == "__default__":
            p = st_dir / default_name
            return p if p.exists() else None
        if flag in ("None", "none", ""):
            return None
        return Path(flag)

    v10cap_path = resolve_captions(args.v10_captions, "Caption_semtest200_V10.jsonl")
    v12cap_path = resolve_captions(args.v12_captions, "Caption_semtest200_V12.jsonl")

    sel = {r["frames_dir"]: r for r in load_jsonl(selection_path)}
    v10cap = ({r["frames_dir"]: r["caption"] for r in load_jsonl(v10cap_path)}
              if v10cap_path else {})
    v12cap = ({r["frames_dir"]: r["caption"] for r in load_jsonl(v12cap_path)}
              if v12cap_path else {})

    scores = {}
    for arm in ["A0"] + arms:
        scores[arm] = {r["frames_dir"]: r["score"] for r in load_jsonl(scores_dir / f"{arm}.jsonl")}

    # ---- hard gates ----
    if len(sel) != args.n_expected:
        raise SystemExit(f"expected {args.n_expected} selection rows, got {len(sel)}")
    for arm in ["A0"] + arms:
        if len(scores[arm]) != args.n_expected:
            raise SystemExit(f"{arm}.jsonl has {len(scores[arm])} rows, expected {args.n_expected}")
    for fd in sel:
        for arm in ["A0"] + arms:
            if fd not in scores[arm]:
                raise SystemExit(f"{arm} missing score for {fd}")
    # A0 consistency: scores/A0.jsonl must match selection's own a0_score
    mismatches = [fd for fd in sel if abs(scores["A0"][fd] - sel[fd]["a0_score"]) > 1e-3]
    if mismatches:
        raise SystemExit(f"A0 re-score mismatch on {len(mismatches)} rows, e.g. {mismatches[:3]}")
    print(f"[verify] gates passed: {args.n_expected} rows, all arms present, "
          f"A0 re-score consistent")

    rows = []
    for fd, s in sel.items():
        row = {
            "video_id": s["video_id"], "window": s["horizon_label"], "split": s["split"],
            "source_tier": s["source"], "gt": s["gt_verdict"], "A0": scores["A0"][fd],
        }
        if v10cap:
            row["caption_V10"] = f"V10: {v10cap.get(fd, '(missing)')}"
        if v12cap:
            row["caption_V12"] = f"V12: {v12cap.get(fd, '(missing)')}"
        for arm in arms:
            row[arm] = scores[arm][fd]
            row[f"{arm}_gap"] = scores[arm][fd] - scores["A0"][fd]
        rows.append(row)

    wb = Workbook()

    # ================================================================ per_clip
    ws = wb.active
    ws.title = "per_clip"
    ws.append(["A0 = pre-FT baseline (identical for every arm - LoRA B-matrix is "
               "zero-init, so the untrained model IS A0). Each arm's *_gap = post-FT "
               "score minus A0. Red/green at threshold 0.5 (a convention here, not any "
               "arm's calibrated optimum - see findings doc)."])
    ws.cell(row=1, column=1).font = NOTE_FONT
    cap_headers = (["caption_V10"] if v10cap else []) + (["caption_V12"] if v12cap else [])
    headers = (["video_id", "window", "split", "source_tier"] + cap_headers + ["gt", "A0"]
               + [c for arm in arms for c in (arm, f"{arm}_gap")])
    ws.append(headers)
    header_row = 2
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r in rows:
        line = [r["video_id"], r["window"], r["split"], r["source_tier"]]
        if v10cap:
            line.append(r["caption_V10"])
        if v12cap:
            line.append(r["caption_V12"])
        line += [r["gt"], round(r["A0"], 4)]
        line += [v for arm in arms for v in (round(r[arm], 4), round(r[f"{arm}_gap"], 4))]
        ws.append(line)

    n_rows = len(rows) + header_row
    gt_col_idx = 5 + len(cap_headers)
    gt_col_letter = get_column_letter(gt_col_idx)
    a0_col = gt_col_idx + 1
    score_cols = [a0_col] + [a0_col + 1 + 2 * i for i in range(len(arms))]
    for col_idx in score_cols:
        col = get_column_letter(col_idx)
        rng = f"{col}{header_row + 1}:{col}{n_rows}"
        pass_f = (f'OR(AND({col}{header_row+1}>=0.5,${gt_col_letter}{header_row+1}="YES"),'
                  f'AND({col}{header_row+1}<0.5,${gt_col_letter}{header_row+1}="NO"))')
        fail_f = (f'OR(AND({col}{header_row+1}>=0.5,${gt_col_letter}{header_row+1}="NO"),'
                  f'AND({col}{header_row+1}<0.5,${gt_col_letter}{header_row+1}="YES"))')
        ws.conditional_formatting.add(rng, FormulaRule(formula=[pass_f], fill=GREEN))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[fail_f], fill=RED))

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{n_rows}"
    widths = {"A": 10, "B": 9, "C": 8, "D": 17}
    for i, _ in enumerate(cap_headers):
        widths[get_column_letter(5 + i)] = 55
    widths[gt_col_letter] = 6
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i in range(a0_col, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 9

    # ================================================================ summary builder
    def build_summary(sheet_name, baseline_arm, compare_arms):
        ws2 = wb.create_sheet(sheet_name)
        ws2.append([f"{baseline_arm} baseline (reference for every row below)"])
        base_headers = ["subset", "n", f"{baseline_arm}_correct", f"{baseline_arm}_wrong"]
        ws2.append(base_headers)
        for c in range(1, len(base_headers) + 1):
            cell = ws2.cell(row=2, column=c)
            cell.fill = BASE_FILL
            cell.font = HEADER_FONT

        subset_defs = [("train", lambda r: r["split"] == "train"),
                       ("val", lambda r: r["split"] == "val"),
                       ("all", lambda r: True)]

        def base_score(r):
            return r["A0"] if baseline_arm == "A0" else r[baseline_arm]

        baseline_wrong = {}
        for name, pred in subset_defs:
            pool = [r for r in rows if pred(r)]
            ok = sum(1 for r in pool if is_ok(base_score(r), r["gt"]))
            wrong = len(pool) - ok
            baseline_wrong[name] = wrong
            ws2.append([name, len(pool), ok, wrong])

        ws2.append([])
        header_row2 = ws2.max_row + 1
        headers2 = ["arm", "subset", "n", "fixed_FP", "fixed_FN", "broken_FP",
                    "broken_FN", "still_wrong", "net"]
        ws2.append(headers2)
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=header_row2, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for name, pred in subset_defs:
            pool = [r for r in rows if pred(r)]
            base_fp = [r for r in pool if not is_ok(base_score(r), r["gt"]) and r["gt"] == "NO"]
            base_fn = [r for r in pool if not is_ok(base_score(r), r["gt"]) and r["gt"] == "YES"]
            base_ok_neg = [r for r in pool if is_ok(base_score(r), r["gt"]) and r["gt"] == "NO"]
            base_ok_pos = [r for r in pool if is_ok(base_score(r), r["gt"]) and r["gt"] == "YES"]
            assert len(base_fp) + len(base_fn) == baseline_wrong[name]
            for arm in compare_arms:
                ffp = sum(1 for r in base_fp if is_ok(r[arm], r["gt"]))
                ffn = sum(1 for r in base_fn if is_ok(r[arm], r["gt"]))
                bfp = sum(1 for r in base_ok_neg if not is_ok(r[arm], r["gt"]))
                bfn = sum(1 for r in base_ok_pos if not is_ok(r[arm], r["gt"]))
                broke = bfp + bfn
                still = (len(base_fp) + len(base_fn)) - ffp - ffn + broke
                ws2.append([arm, name, len(pool), ffp, ffn, bfp, bfn, still, ffp + ffn - broke])
        for col, w in {"A": 9, "B": 8, "C": 6}.items():
            ws2.column_dimensions[col].width = w
        return ws2

    build_summary("summary_vs_A0", "A0", arms)
    if "vision" in arms:
        build_summary("summary_vs_vision", "vision", [a for a in arms if a != "vision"])

    # ================================================================ metrics
    ws4 = wb.create_sheet("metrics")
    ws4.append(["arm", "split", "n", "AP", "AUC", "acc@0.5", "selected_epoch"])
    for c in range(1, 8):
        cell = ws4.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for arm in ["A0"] + arms:
        for split_name, pred in [("train", lambda r: r["split"] == "train"),
                                  ("val", lambda r: r["split"] == "val"),
                                  ("all", lambda r: True)]:
            pool = [r for r in rows if pred(r)]
            y = [1 if r["gt"] == "YES" else 0 for r in pool]
            s = [r["A0"] if arm == "A0" else r[arm] for r in pool]
            ap_ = average_precision_score(y, s)
            auc = roc_auc_score(y, s)
            acc = sum(1 for yy, ss in zip(y, s) if (ss >= 0.5) == bool(yy)) / len(y)
            ep = epoch_of.get(arm, "-")
            ws4.append([arm, split_name, len(pool), round(ap_, 4), round(auc, 4),
                        round(acc, 4), ep])
    for col, w in {"A": 9, "B": 7}.items():
        ws4.column_dimensions[col].width = w

    # ================================================================ metrics_stratified
    # Answers "did the +100 easy anchor clips just inflate the number": the hard-subset
    # row here is the exact same readout the original 200-clip pool already reported,
    # unpooled with the easy clips. A v1 workbook (no easy_* tier at all) still produces
    # this sheet - hard == all, easy has n=0 and is skipped.
    ws5 = wb.create_sheet("metrics_stratified")
    ws5.append(["note: 'hard' = every original SemTest-200 tier (FN_near/FN_wide/"
               "TP_fill/FP_near_boundary/FP_fill); 'easy' = the +100 A0-correct anchor "
               "clips (easy_TN/easy_TP); 'all' = both pooled."])
    ws5.cell(row=1, column=1).font = NOTE_FONT
    ws5.append(["arm", "tier", "n", "AP", "AUC", "acc@0.5"])
    for c in range(1, 7):
        cell = ws5.cell(row=2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    tier_defs = [("hard", lambda r: not is_easy(r["source_tier"])),
                 ("easy", lambda r: is_easy(r["source_tier"])),
                 ("all", lambda r: True)]
    for arm in ["A0"] + arms:
        for tier_name, pred in tier_defs:
            pool = [r for r in rows if pred(r)]
            if not pool:
                continue
            y = [1 if r["gt"] == "YES" else 0 for r in pool]
            s = [r["A0"] if arm == "A0" else r[arm] for r in pool]
            if len(set(y)) < 2:
                ws5.append([arm, tier_name, len(pool), "n/a (single class)", "n/a", "n/a"])
                continue
            ap_ = average_precision_score(y, s)
            auc = roc_auc_score(y, s)
            acc = sum(1 for yy, ss in zip(y, s) if (ss >= 0.5) == bool(yy)) / len(y)
            ws5.append([arm, tier_name, len(pool), round(ap_, 4), round(auc, 4), round(acc, 4)])
    for col, w in {"A": 9, "B": 7}.items():
        ws5.column_dimensions[col].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[wrote] {out_path}")


if __name__ == "__main__":
    main()
