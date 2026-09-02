"""
add_vs_a1_summary_sheet.py
============================
Adds a 'summary_vs_A1' sheet to pool1761_arm_comparison.xlsx: the same
fixed_FP/fixed_FN/broken/still_wrong/net accounting the existing 'summary' sheet
computes against A0, but with A1 (the vision-only control) as the baseline instead.

WHY THIS IS A DIFFERENT QUESTION FROM THE A0 SHEET: A0 never trained on any of this
data, so the A0-relative sheet answers "did training help at all". A1 is itself a
trained arm (vision-only), so the A1-relative sheet answers the actually-pre-registered
thesis question: does adding the semantic-aux loss help or hurt RELATIVE TO an
identically-trained vision-only control. B-v1/B-v2/B-v3/P1 are compared; A1 is not
compared to itself.

Adds mined_failure as a SECOND stratification dimension alongside subset (train/val/all),
per user request 2026-08-26 - the original A0 sheet only split by subset.

Reads per_clip (unchanged), writes/replaces a 'summary_vs_A1' sheet - does not touch
any existing sheet's data.

Usage:
  python add_vs_a1_summary_sheet.py --xlsx ../../outputs/e4_vjepa_reason/pool1761_arm_comparison.xlsx
"""
import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ARMS = ["B-v1", "B-v2", "B-v3", "P1"]
HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    path = Path(args.xlsx)
    wb = load_workbook(path, data_only=True)
    ws = wb["per_clip"]
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[col["video_id"]] is None:
            continue
        rows.append({
            "split": r[col["split"]],
            "mined_failure": bool(r[col["mined_failure"]]),
            "gt": r[col["gt"]],
            "A1": r[col["A1"]],
            **{arm: r[col[arm]] for arm in ARMS},
        })

    def is_ok(row, arm):
        return (row[arm] >= 0.5) == (row["gt"] == "YES")

    if "summary_vs_A1" in wb.sheetnames:
        del wb["summary_vs_A1"]
    ws2 = wb.create_sheet("summary_vs_A1")

    # ---- A1's own baseline first (2026-08-26): n / A1-correct / A1-wrong per subset x
    # mined_failure, so the arm rows below are readable without cross-referencing chat. ----
    BASE_FILL = PatternFill("solid", fgColor="4A5A80")
    base_headers = ["subset", "mined_failure", "n", "A1_correct", "A1_wrong"]
    ws2.append(["A1 baseline on this pool (reference for every row below)"])
    ws2.append(base_headers)
    for c in range(1, len(base_headers) + 1):
        cell = ws2.cell(row=2, column=c)
        cell.fill = BASE_FILL
        cell.font = HEADER_FONT

    subset_defs = [
        ("train", lambda r: r["split"] == "train"),
        ("val", lambda r: r["split"] == "val"),
        ("all", lambda r: True),
    ]
    mined_defs = [
        ("mined_only", lambda r: r["mined_failure"] is True),
        ("easy_only", lambda r: r["mined_failure"] is False),
        ("both", lambda r: True),
    ]

    baseline_wrong = {}   # (subset_name, mined_name) -> A1_wrong count, reused below
    for subset_name, subset_pred in subset_defs:
        for mined_name, mined_pred in mined_defs:
            pool = [r for r in rows if subset_pred(r) and mined_pred(r)]
            ok = sum(1 for r in pool if is_ok(r, "A1"))
            wrong = len(pool) - ok
            baseline_wrong[(subset_name, mined_name)] = wrong
            ws2.append([subset_name, mined_name, len(pool), ok, wrong])

    ws2.append([])  # blank separator row
    header_row = ws2.max_row + 1
    # broken split into broken_FP/broken_FN (2026-08-26, user request): "broken" alone
    # can't distinguish a semantic arm that destabilizes correct crash calls (A1's TPs
    # turning into FNs) from one that destabilizes correct no-crash calls (A1's TNs
    # turning into FPs) - exactly the bias question motivating this split (does the
    # semantic loss make TN prediction more reliable at the cost of TP reliability, or
    # vice versa).
    headers_out = ["arm", "subset", "mined_failure", "n", "fixed_FP", "fixed_FN",
                   "broken_FP", "broken_FN", "still_wrong", "net"]
    ws2.append(headers_out)
    for c in range(1, len(headers_out) + 1):
        cell = ws2.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for subset_name, subset_pred in subset_defs:
        for mined_name, mined_pred in mined_defs:
            pool = [r for r in rows if subset_pred(r) and mined_pred(r)]
            a1_fp = [r for r in pool if not is_ok(r, "A1") and r["gt"] == "NO"]
            a1_fn = [r for r in pool if not is_ok(r, "A1") and r["gt"] == "YES"]
            # A1-correct clips, split by class: a1_ok_neg (A1 correctly said NO, a true
            # negative) can only be BROKEN into a false positive; a1_ok_pos (A1 correctly
            # said YES, a true positive) can only be broken into a false negative.
            a1_ok_neg = [r for r in pool if is_ok(r, "A1") and r["gt"] == "NO"]
            a1_ok_pos = [r for r in pool if is_ok(r, "A1") and r["gt"] == "YES"]
            assert len(a1_fp) + len(a1_fn) == baseline_wrong[(subset_name, mined_name)]
            for arm in ARMS:
                ffp = sum(1 for r in a1_fp if is_ok(r, arm))
                ffn = sum(1 for r in a1_fn if is_ok(r, arm))
                broke_fp = sum(1 for r in a1_ok_neg if not is_ok(r, arm))  # TN -> FP
                broke_fn = sum(1 for r in a1_ok_pos if not is_ok(r, arm))  # TP -> FN
                broke = broke_fp + broke_fn
                # still_wrong = TOTAL wrong under this arm (2026-08-26 fix) - broken clips
                # are wrong too and were missing from this count before.
                still = (len(a1_fp) + len(a1_fn)) - ffp - ffn + broke
                ws2.append([arm, subset_name, mined_name, len(pool), ffp, ffn, broke_fp,
                            broke_fn, still, ffp + ffn - broke])

    widths = {"A": 8, "B": 8, "C": 12, "D": 6}
    for col_letter, w in widths.items():
        ws2.column_dimensions[col_letter].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)
    print(f"[wrote] summary_vs_A1 sheet -> {path}")


if __name__ == "__main__":
    main()
