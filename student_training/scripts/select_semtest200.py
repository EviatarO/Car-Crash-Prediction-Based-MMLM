"""
select_semtest200.py
=====================
SemTest-200 clip selection (2026-08-26 plan, Phase 1b).

Selects 200 windows (100 TP + 100 TN) from the full 4,446-window pool, using A0's
full-pool re-score (Phase 1a's score_arms_on_pool1761.py run against
Pool_Train4500_Full_4446.jsonl) plus dataset/train.xlsx's response-time column, so the
result is a clean, verified-headroom set for the caption-effect experiment:

  - TP: A0-correct (score > 0.5), response_time > TTE (dataset/train.xlsx column E is
    response_time = time_of_event - time_of_alert; NULL for every negative row - so this
    filter applies to TP only), ranked by |score - 0.6| ascending. (2026-08-26: the
    original +0.5s safety margin was dropped after measurement showed it was the binding
    constraint on TP availability at TTE_1.0/1.5 - RT > TTE is the minimum needed to keep
    the crash itself out of the 16-frame window; the margin was extra conservatism the
    data doesn't support at n=200. User decision, not a silent relaxation.)
  - TN: A0-correct (score < 0.5), no response-time filter possible (see above), ranked
    by |score - 0.4| ascending.
  - Quotas: 33/33/34 per TTE_0.5/1.0/1.5 (TP) and per MID-4/MID-8/MID-10 (TN).
  - One window per video_id, enforced GLOBALLY per class by filling buckets in a fixed
    order and skipping any video already used - a video with e.g. both a TTE_0.5 and a
    TTE_1.0 candidate window can only land in one bucket. (No video spans both classes -
    verified against the manifest: 1,482 unique video_ids, 0 with mixed event_occurs.)
  - A fixed, deterministic, stratified 40-window (20 TP + 20 TN, 7/7/6 per bucket) val
    split is written into the output as a `split` column - NOT semsup_train.py's
    clip_level_split (video-id-shuffle only, not label- or TTE-stratified). Feed it back
    into training via --val-video-ids val_vids.txt.
  - A reserve list (~10 next-closest candidates per bucket) is written alongside, for
    swapping out any clip whose caption fails manual verification (Phase 1d).

If any bucket can't fill its quota, this script reports the shortfall and does NOT write
selection.jsonl/xlsx - silently relaxing the RT filter would defeat the point of the test.

Usage:
  python select_semtest200.py \
    --a0-scores ../../outputs/semtest200/A0_full4446.jsonl \
    --manifest ../../dataset/manifests/train4500_hires.jsonl \
    --train-xlsx ../../dataset/train.xlsx \
    --out-dir ../../outputs/semtest200
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

TTE_RT_MIN = {"TTE_0.5": 0.5, "TTE_1.0": 1.0, "TTE_1.5": 1.5}  # response_time > TTE (no margin)
TP_BUCKETS = ["TTE_0.5", "TTE_1.0", "TTE_1.5"]
TN_BUCKETS = ["MID-4", "MID-8", "MID-10"]
QUOTA = {"TTE_0.5": 33, "TTE_1.0": 33, "TTE_1.5": 34,
         "MID-4": 33, "MID-8": 33, "MID-10": 34}
VAL_QUOTA = {"TTE_0.5": 7, "TTE_1.0": 7, "TTE_1.5": 6,
             "MID-4": 7, "MID-8": 7, "MID-10": 6}
RESERVE_PER_BUCKET = 10

HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def load_a0_scores(path):
    out = {}
    for line in open(path, encoding="utf-8"):
        if not line.strip():
            continue
        r = json.loads(line)
        out[r["frames_dir"]] = {"score": float(r["score"]), "gt_verdict": r["gt_verdict"]}
    return out


def load_manifest(path):
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def load_response_times(xlsx_path):
    """video_id (5-digit zero-padded str) -> response_time (float) or None.

    dataset/train.xlsx columns: A id (bare int) | B time_of_event | C time_of_alert |
    D target | E response time (= B - C, seconds; NULL for all 750 negatives)."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["train"]
    out = {}
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        if row[0] is None:
            continue
        vid = f"{int(row[0]):05d}"
        rt = row[4]
        out[vid] = float(rt) if rt is not None else None
    return out


def fill_buckets(candidates_by_bucket, quota, center):
    """candidates_by_bucket: {bucket: [row, ...]} already carrying an 'a0_score' key.
    Returns (selected {bucket: [row,...]}, reserve {bucket: [row,...]}, shortfall {bucket:int})."""
    used_videos = set()
    selected, reserve, shortfall = {}, {}, {}
    for bucket, rows in candidates_by_bucket.items():
        ranked = sorted(rows, key=lambda r: abs(r["a0_score"] - center))
        picked = []
        spare = []
        for r in ranked:
            if r["video_id"] in used_videos:
                continue
            if len(picked) < quota[bucket]:
                picked.append(r)
                used_videos.add(r["video_id"])
            elif len(spare) < RESERVE_PER_BUCKET:
                spare.append(r)
            else:
                break
        selected[bucket] = picked
        reserve[bucket] = spare
        if len(picked) < quota[bucket]:
            shortfall[bucket] = quota[bucket] - len(picked)
    return selected, reserve, shortfall


def assign_split(selected, val_quota):
    """Deterministic stratified val split: within each bucket's closeness-ranked list
    (already the fill order), take an evenly-strided sample of size val_quota[bucket]."""
    for bucket, rows in selected.items():
        n = len(rows)
        vq = val_quota[bucket]
        stride = max(1, n // vq) if vq else n + 1
        val_idx = {i * stride for i in range(vq) if i * stride < n}
        for i, r in enumerate(rows):
            r["split"] = "val" if i in val_idx else "train"


def write_jsonl(rows, path):
    with open(path, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r) + "\n")


def write_xlsx(rows, path, sheet_name="selection"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ["video_id", "frames_dir", "gt_verdict", "horizon_label", "a0_score",
               "response_time", "split"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in rows:
        ws.append([r["video_id"], r["frames_dir"], r["gt_verdict"], r["horizon_label"],
                   round(r["a0_score"], 4), r.get("response_time"), r.get("split", "")])
    widths = {"A": 10, "B": 20, "C": 10, "D": 13, "E": 10, "F": 14, "G": 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(rows) + 1}"
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--a0-scores", required=True)
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--train-xlsx", required=True)
    ap.add_argument("--out-dir", required=True)
    args = ap.parse_args()

    a0 = load_a0_scores(args.a0_scores)
    manifest = load_manifest(args.manifest)
    rt = load_response_times(args.train_xlsx)

    print(f"[load] a0_scores={len(a0)}  manifest={len(manifest)}  "
          f"response_times={len(rt)} ({sum(1 for v in rt.values() if v is not None)} non-null)")

    missing_score = 0
    tp_by_bucket = {b: [] for b in TP_BUCKETS}
    tn_by_bucket = {b: [] for b in TN_BUCKETS}
    for m in manifest:
        fd = m["frames_dir"]
        s = a0.get(fd)
        if s is None:
            missing_score += 1
            continue
        vid = m["video_id"]
        row = {"video_id": vid, "frames_dir": fd, "horizon_label": m["horizon_label"],
               "gt_verdict": s["gt_verdict"], "a0_score": s["score"],
               "response_time": rt.get(vid)}
        if m["event_occurs"] == 1:
            bucket = m["horizon_label"]
            if bucket not in tp_by_bucket:
                continue
            if s["score"] <= 0.5:
                continue
            r_time = rt.get(vid)
            if r_time is None or r_time <= TTE_RT_MIN[bucket]:
                continue
            tp_by_bucket[bucket].append(row)
        else:
            bucket = m["horizon_label"]
            if bucket not in tn_by_bucket:
                continue
            if s["score"] >= 0.5:
                continue
            tn_by_bucket[bucket].append(row)

    if missing_score:
        print(f"[warn] {missing_score} manifest windows have no A0 score - "
              f"check the Phase 1a re-score covers the full 4,446 pool")

    for b in TP_BUCKETS:
        print(f"[candidates] TP {b}: {len(tp_by_bucket[b])} eligible (need {QUOTA[b]})")
    for b in TN_BUCKETS:
        print(f"[candidates] TN {b}: {len(tn_by_bucket[b])} eligible (need {QUOTA[b]})")

    tp_sel, tp_reserve, tp_short = fill_buckets(tp_by_bucket, QUOTA, center=0.6)
    tn_sel, tn_reserve, tn_short = fill_buckets(tn_by_bucket, QUOTA, center=0.4)
    shortfall = {**tp_short, **tn_short}

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if shortfall:
        print("\n[STOP] quota shortfall - selection.jsonl/xlsx NOT written:")
        for b, n in shortfall.items():
            print(f"  {b}: short by {n}")
        with open(out_dir / "selection_shortfall_report.json", "w", encoding="utf-8") as f:
            json.dump({
                "shortfall": shortfall,
                "candidates_available": {
                    **{b: len(v) for b, v in tp_by_bucket.items()},
                    **{b: len(v) for b, v in tn_by_bucket.items()},
                },
                "quota": QUOTA,
            }, f, indent=2)
        print(f"[wrote] {out_dir / 'selection_shortfall_report.json'} - "
              f"needs a user call before relaxing any filter.")
        return

    assign_split(tp_sel, VAL_QUOTA)
    assign_split(tn_sel, VAL_QUOTA)

    all_selected = [r for b in TP_BUCKETS for r in tp_sel[b]] + \
                   [r for b in TN_BUCKETS for r in tn_sel[b]]
    all_reserve = [r for b in TP_BUCKETS for r in tp_reserve[b]] + \
                  [r for b in TN_BUCKETS for r in tn_reserve[b]]

    n_val = sum(1 for r in all_selected if r["split"] == "val")
    n_train = len(all_selected) - n_val
    print(f"\n[selection] {len(all_selected)} windows "
          f"({sum(len(v) for v in tp_sel.values())} TP / {sum(len(v) for v in tn_sel.values())} TN)  "
          f"train={n_train}  val={n_val}")

    write_jsonl(all_selected, out_dir / "selection.jsonl")
    write_xlsx(all_selected, out_dir / "selection.xlsx")
    write_jsonl(all_reserve, out_dir / "selection_reserve.jsonl")
    with open(out_dir / "val_vids.txt", "w", encoding="utf-8") as f:
        for r in all_selected:
            if r["split"] == "val":
                f.write(r["video_id"] + "\n")

    print(f"[wrote] {out_dir / 'selection.jsonl'}")
    print(f"[wrote] {out_dir / 'selection.xlsx'}")
    print(f"[wrote] {out_dir / 'selection_reserve.jsonl'} ({len(all_reserve)} reserve rows)")
    print(f"[wrote] {out_dir / 'val_vids.txt'} ({n_val} ids)")


if __name__ == "__main__":
    main()
