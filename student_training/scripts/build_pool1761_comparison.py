"""Build the per-clip A0-vs-arm comparison workbook over the full 1,761 training pool.

Inputs:
  - Caption_Train4500_Mixed_1761.jsonl        (V10, used by B-v1)
  - Caption_V12_Neutral_1761_fortrain.jsonl   (V12, used by A1's caption bank / B-v2 / B-v3 / P1)
  - Caption_Train4500_Failures_587.jsonl      (flags the 587 mined A0 failures)
  - one JSONL per arm from score_arms_on_pool1761.py, all keyed by frames_dir

The train/val split is NOT re-derived independently - it is recomputed by replicating
semsup_common.clip_level_split(val_frac=0.2, seed=0) exactly, over the same 1,761-window
example list every arm actually trained on. This must match what training used, or the
train/val column is simply wrong.

Usage:
  python build_pool1761_comparison.py --scores-dir /path/to/pool1761_scores \
      --out outputs/e4_vjepa_reason/pool1761_arm_comparison.xlsx
"""
import argparse
import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parents[2]
CAP = ROOT / "outputs" / "semantic_captions"

ARMS = ["A0", "A1", "B-v1", "B-v2", "B-v3", "P1"]
ARM_CAPTION = {  # which caption version each arm actually trained the semantic loss on
    "A0": None, "A1": None,
    "B-v1": "V10", "B-v2": "V12", "B-v3": "V12", "P1": "V12",
}

# NOTE: conditional-formatting (dxf) fills render from bgColor, not fgColor - the reverse
# of a normal cell fill. Using fgColor here silently produces an invisible fill in Excel
# (confirmed: openpyxl round-trips it, but Excel's dxf renderer ignores fgColor for solid
# patternType). bgColor is the fix for CF fills specifically.
GREEN = PatternFill(patternType="solid", bgColor="C6EFCE")  # soft green
RED = PatternFill(patternType="solid", bgColor="FFC7CE")    # soft red
HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def clip_level_split(video_ids, val_frac=0.2, seed=0):
    """Exact replica of semsup_common.clip_level_split's video_id partition."""
    vids = sorted(set(video_ids))
    random.Random(seed).shuffle(vids)
    n_val = max(1, int(len(vids) * val_frac))
    return set(vids[:n_val])  # val_vids


def load_jsonl(path):
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--scores-dir", required=True,
                     help="dir containing A0.jsonl, A1.jsonl, B-v1.jsonl, B-v2.jsonl, "
                          "B-v3.jsonl, P1.jsonl from score_arms_on_pool1761.py")
    ap.add_argument("--out", default=str(ROOT / "outputs" / "e4_vjepa_reason" /
                                         "pool1761_arm_comparison.xlsx"))
    args = ap.parse_args()

    v10 = {r["frames_dir"]: r for r in load_jsonl(CAP / "Caption_Train4500_Mixed_1761.jsonl")}
    v12 = {r["frames_dir"]: r for r in load_jsonl(CAP / "Caption_V12_Neutral_1761_fortrain.jsonl")}
    failures = {r["frames_dir"] for r in load_jsonl(CAP / "Caption_Train4500_Failures_587.jsonl")}

    if set(v10) != set(v12):
        raise SystemExit(f"V10/V12 frames_dir mismatch: {len(set(v10)^set(v12))} rows differ")
    if len(v12) != 1761:
        raise SystemExit(f"expected 1761 windows, V12 has {len(v12)}")

    val_vids = clip_level_split([r["video_id"] for r in v12.values()])

    scores = {}  # frames_dir -> {arm: score}
    for arm in ARMS:
        p = Path(args.scores_dir) / f"{arm}.jsonl"
        if not p.exists():
            raise SystemExit(f"missing score file for arm {arm}: {p}")
        rows = load_jsonl(p)
        if len(rows) != 1761:
            raise SystemExit(f"{arm}.jsonl has {len(rows)} rows, expected 1761")
        for r in rows:
            scores.setdefault(r["frames_dir"], {})[arm] = r["score"]

    missing = [fd for fd in v12 if fd not in scores or len(scores[fd]) != len(ARMS)]
    if missing:
        raise SystemExit(f"{len(missing)} windows missing a score from at least one arm, "
                          f"e.g. {missing[:3]}")

    # ---- verification: A0 must be wrong on exactly the 587 flagged rows (how the pool was built)
    a0_wrong = {fd for fd, s in scores.items() if (s["A0"] >= 0.5) != (v12[fd]["gt_verdict"] == "YES")}
    if a0_wrong != failures:
        only_a0 = a0_wrong - failures
        only_flag = failures - a0_wrong
        print(f"[WARN] A0 re-score does not exactly match the mined-failure list: "
              f"{len(only_a0)} newly-wrong, {len(only_flag)} no-longer-wrong "
              f"(checkpoint/threshold drift from the original mining run)")
    else:
        print("[verify] A0 re-score reproduces the 587 mined failures exactly")

    # ---- assemble rows
    rows = []
    for fd, cap12 in v12.items():
        cap10 = v10[fd]
        rows.append({
            "video_id": cap12["video_id"],
            "window": cap12.get("horizon_label") or cap12.get("requested_time_to_event"),
            "split": "val" if cap12["video_id"] in val_vids else "train",
            "mined_failure": fd in failures,
            "caption_V10": f"V10: {cap10['caption']}",
            "caption_V12": f"V12: {cap12['caption']}",
            "gt": cap12["gt_verdict"],
            **{arm: scores[fd][arm] for arm in ARMS},
        })
    n_train = sum(1 for r in rows if r["split"] == "train")
    n_val = sum(1 for r in rows if r["split"] == "val")
    n_val_fail = sum(1 for r in rows if r["split"] == "val" and r["mined_failure"])
    print(f"[split] train={n_train} val={n_val} (of which {n_val_fail} are mined failures)")
    if (n_train, n_val) != (1413, 348):
        print(f"[WARN] expected 1413/348, got {n_train}/{n_val} - "
              f"check the caption file used to build the split matches training")

    wb = Workbook()

    # ================================================================ per_clip
    ws = wb.active
    ws.title = "per_clip"
    headers = (["video_id", "window", "split", "mined_failure", "caption_V10", "caption_V12",
                "gt"] + ARMS)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in rows:
        ws.append([r["video_id"], r["window"], r["split"], r["mined_failure"],
                   r["caption_V10"], r["caption_V12"], r["gt"]] + [r[a] for a in ARMS])

    n_rows = len(rows) + 1
    first_arm_col = 8  # H
    last_arm_col = first_arm_col + len(ARMS) - 1  # M
    gt_col_letter = "G"
    for i, arm in enumerate(ARMS):
        col = get_column_letter(first_arm_col + i)
        rng = f"{col}2:{col}{n_rows}"
        pass_formula = (f'OR(AND({col}2>=0.5,${gt_col_letter}2="YES"),'
                        f'AND({col}2<0.5,${gt_col_letter}2="NO"))')
        fail_formula = (f'OR(AND({col}2>=0.5,${gt_col_letter}2="NO"),'
                        f'AND({col}2<0.5,${gt_col_letter}2="YES"))')
        ws.conditional_formatting.add(rng, FormulaRule(formula=[pass_formula], fill=GREEN))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[fail_formula], fill=RED))

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_arm_col)}{n_rows}"
    widths = {"A": 10, "B": 10, "C": 8, "D": 13, "E": 55, "F": 55, "G": 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i in range(len(ARMS)):
        ws.column_dimensions[get_column_letter(first_arm_col + i)].width = 9

    # ================================================================ summary
    ws2 = wb.create_sheet("summary")
    ws2.append(["arm", "subset", "n", "fixed_FP", "fixed_FN", "broken", "still_wrong", "net"])
    for c in range(1, 9):
        ws2.cell(row=1, column=c).fill = HEADER_FILL
        ws2.cell(row=1, column=c).font = HEADER_FONT

    def is_ok(r, arm):
        return (r[arm] >= 0.5) == (r["gt"] == "YES")

    for subset_name, subset in [("train", [r for r in rows if r["split"] == "train"]),
                                ("val", [r for r in rows if r["split"] == "val"]),
                                ("all", rows)]:
        a0_fp = [r for r in subset if not is_ok(r, "A0") and r["gt"] == "NO"]
        a0_fn = [r for r in subset if not is_ok(r, "A0") and r["gt"] == "YES"]
        a0_ok = [r for r in subset if is_ok(r, "A0")]
        for arm in ARMS:
            ffp = sum(1 for r in a0_fp if is_ok(r, arm))
            ffn = sum(1 for r in a0_fn if is_ok(r, arm))
            broke = sum(1 for r in a0_ok if not is_ok(r, arm))
            still = len(a0_fp) + len(a0_fn) - ffp - ffn
            ws2.append([arm, subset_name, len(subset), ffp, ffn, broke, still, ffp + ffn - broke])
    for col, w in {"A": 8, "B": 8}.items():
        ws2.column_dimensions[col].width = w

    # ================================================================ val_only / failures_only
    for sheet_name, pred in [("val_only", lambda r: r["split"] == "val"),
                             ("failures_only", lambda r: r["mined_failure"])]:
        ws3 = wb.create_sheet(sheet_name)
        ws3.append(headers)
        for c in range(1, len(headers) + 1):
            ws3.cell(row=1, column=c).fill = HEADER_FILL
            ws3.cell(row=1, column=c).font = HEADER_FONT
        sub_rows = [r for r in rows if pred(r)]
        for r in sub_rows:
            ws3.append([r["video_id"], r["window"], r["split"], r["mined_failure"],
                        r["caption_V10"], r["caption_V12"], r["gt"]] + [r[a] for a in ARMS])
        nn = len(sub_rows) + 1
        for i, arm in enumerate(ARMS):
            col = get_column_letter(first_arm_col + i)
            rng = f"{col}2:{col}{nn}"
            pass_formula = (f'OR(AND({col}2>=0.5,${gt_col_letter}2="YES"),'
                            f'AND({col}2<0.5,${gt_col_letter}2="NO"))')
            fail_formula = (f'OR(AND({col}2>=0.5,${gt_col_letter}2="NO"),'
                            f'AND({col}2<0.5,${gt_col_letter}2="YES"))')
            ws3.conditional_formatting.add(rng, FormulaRule(formula=[pass_formula], fill=GREEN))
            ws3.conditional_formatting.add(rng, FormulaRule(formula=[fail_formula], fill=RED))
        ws3.freeze_panes = "E2"
        for col, w in widths.items():
            ws3.column_dimensions[col].width = w
        for i in range(len(ARMS)):
            ws3.column_dimensions[get_column_letter(first_arm_col + i)].width = 9

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"[wrote] {args.out}")


if __name__ == "__main__":
    main()
