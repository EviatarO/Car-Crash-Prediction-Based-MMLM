"""
select_a1fail321.py
=====================
A1-failure recovery pool: all 321 windows A1 (crash-only LoRA, the current champion,
test AP=0.900/AUC=0.904 on 677 clips) gets WRONG at threshold 0.5, out of the 1,761-pool.
See pool1761_arm_comparison.xlsx's summary_vs_A1 sheet for the origin of this count
(94 in A1's val split + 227 in A1's train split = 321).

WHY THIS SET IS RANK-INVERTED BY CONSTRUCTION (read before interpreting any AP/AUC on it):
Every row here is, by selection, one A1 scores on the WRONG side of 0.5. So every
negative A1 over-scores and every positive A1 under-scores relative to the other class -
A1's own AUC on this set is EXACTLY 0.0, not merely low. Any non-degenerate model looks
like an "improvement" here (random guessing gives 0.5). This pool cannot be the
headline metric - it exists to see whether training moves these clips at all; the real
benchmark is test AP vs A1's 0.900 on the untouched 677-clip test set (Stage 2, separate
script, deferred until these loss curves are reviewed).

Usage:
  python select_a1fail321.py \
      --a1-scores ../../outputs/e4_vjepa_reason/pool1761_scores/A1.jsonl \
      --v10-corpus ../../outputs/semantic_captions/Caption_Train4500_Mixed_1761.jsonl \
      --v12-corpus ../../outputs/semantic_captions/Caption_V12_Neutral_1761_fortrain.jsonl \
      --out-dir ../../outputs/a1fail321 \
      --val-frac 0.2 --seed 0
"""
import argparse
import json
import random
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sklearn.metrics import roc_auc_score

HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def load_jsonl(path):
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def is_ok(score, gt):
    return (score >= 0.5) == (gt == "YES")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--a1-scores", required=True)
    ap.add_argument("--v10-corpus", required=True)
    ap.add_argument("--v12-corpus", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--val-frac", type=float, default=0.2)
    ap.add_argument("--seed", type=int, default=0)
    args = ap.parse_args()

    a1_rows = load_jsonl(args.a1_scores)
    print(f"[load] A1.jsonl: {len(a1_rows)} rows")

    wrong = [r for r in a1_rows if not is_ok(r["score"], r["gt_verdict"])]
    print(f"[select] A1-wrong: {len(wrong)} rows")
    assert len(wrong) == 321, (
        f"expected 321 A1-wrong rows (per pool1761_arm_comparison.xlsx's summary_vs_A1: "
        f"94 val + 227 train), got {len(wrong)} - selection logic disagrees with the "
        f"workbook, do not proceed without reconciling this")

    y = [1 if r["gt_verdict"] == "YES" else 0 for r in wrong]
    s = [r["score"] for r in wrong]
    a1_auc = roc_auc_score(y, s)
    print(f"[sanity] A1's own AUC on its 321 failures: {a1_auc:.6f} (must be exactly 0.0 "
          f"- every row is on the wrong side of 0.5 by construction)")
    assert a1_auc == 0.0, f"A1 AUC on its own failures is {a1_auc}, expected exactly 0.0"

    unique_vids = sorted({r["video_id"] for r in wrong})
    print(f"[compose] {len(wrong)} rows / {len(unique_vids)} unique videos  "
          f"gt={dict(Counter(r['gt_verdict'] for r in wrong))}")

    # ---- caption coverage check (both corpora must cover every row - no new captioning) ----
    v10_corpus = {r["frames_dir"]: r for r in load_jsonl(args.v10_corpus)}
    v12_corpus = {r["frames_dir"]: r for r in load_jsonl(args.v12_corpus)}
    missing_v10 = [r["frames_dir"] for r in wrong if r["frames_dir"] not in v10_corpus]
    missing_v12 = [r["frames_dir"] for r in wrong if r["frames_dir"] not in v12_corpus]
    assert not missing_v10, f"{len(missing_v10)} rows missing from V10 corpus: {missing_v10[:5]}"
    assert not missing_v12, f"{len(missing_v12)} rows missing from V12 corpus: {missing_v12[:5]}"
    print(f"[verify] all {len(wrong)} rows covered by both V10 and V12 corpora - "
          f"no new captioning needed")

    # ---- split by video_id (never by row - sibling TTE/MID windows of the same video
    # must stay on the same side, or val leaks train content) ----
    rng = random.Random(args.seed)
    vids_shuffled = list(unique_vids)
    rng.shuffle(vids_shuffled)
    n_val_vids = max(1, round(len(vids_shuffled) * args.val_frac))
    val_vid_set = set(vids_shuffled[:n_val_vids])

    out_rows = []
    for r in wrong:
        out_rows.append({
            "video_id": r["video_id"], "frames_dir": r["frames_dir"],
            "gt_verdict": r["gt_verdict"],
            "horizon_label": v12_corpus[r["frames_dir"]].get("horizon_label"),
            "source": "a1_wrong",
            "a1_score": r["score"],
            "requested_time_to_event": r.get("requested_time_to_event"),
            "split": "val" if r["video_id"] in val_vid_set else "train",
        })

    n_train = sum(1 for r in out_rows if r["split"] == "train")
    n_val = sum(1 for r in out_rows if r["split"] == "val")
    print(f"[split] video-level 80/20 (seed={args.seed}): train={n_train} rows "
          f"({len(unique_vids) - n_val_vids} videos)  val={n_val} rows ({n_val_vids} videos)")
    print(f"[split] train gt: {dict(Counter(r['gt_verdict'] for r in out_rows if r['split']=='train'))}")
    print(f"[split] val   gt: {dict(Counter(r['gt_verdict'] for r in out_rows if r['split']=='val'))}")

    train_vids = {r["video_id"] for r in out_rows if r["split"] == "train"}
    assert not (train_vids & val_vid_set), "train/val video overlap - split is broken"

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(out_dir / "selection_a1fail321.jsonl", "w", encoding="utf-8") as f:
        for r in out_rows:
            f.write(json.dumps(r) + "\n")

    headers = ["video_id", "frames_dir", "gt_verdict", "horizon_label", "source",
               "a1_score", "requested_time_to_event", "split"]
    wb = Workbook()
    ws = wb.active
    ws.title = "selection_a1fail321"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in out_rows:
        ws.append([r["video_id"], r["frames_dir"], r["gt_verdict"], r["horizon_label"],
                   r["source"], round(r["a1_score"], 4), r.get("requested_time_to_event"),
                   r["split"]])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(out_rows) + 1}"
    wb.save(out_dir / "selection_a1fail321.xlsx")

    with open(out_dir / "val_vids.txt", "w", encoding="utf-8", newline="\n") as f:
        for v in sorted(val_vid_set):
            f.write(v + "\n")

    # ---- per-arm caption files: filter the 1,761 corpora down to these 321 frames_dir,
    # in the SAME row order as out_rows so downstream tooling sees a consistent pool ----
    wanted_fds = {r["frames_dir"] for r in out_rows}
    for name, corpus in [("V10", v10_corpus), ("V12", v12_corpus)]:
        sub = [corpus[r["frames_dir"]] for r in out_rows]
        out_path = out_dir / f"Caption_a1fail321_{name}.jsonl"
        with open(out_path, "w", encoding="utf-8") as f:
            for row in sub:
                f.write(json.dumps(row) + "\n")
        print(f"[wrote] {out_path} ({len(sub)} rows)")

    print(f"[wrote] {out_dir / 'selection_a1fail321.jsonl'}")
    print(f"[wrote] {out_dir / 'selection_a1fail321.xlsx'}")
    print(f"[wrote] {out_dir / 'val_vids.txt'} ({len(val_vid_set)} ids)")


if __name__ == "__main__":
    main()
