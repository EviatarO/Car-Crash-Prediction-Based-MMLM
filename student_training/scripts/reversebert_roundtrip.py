"""
ReverseBERT round-trip pilot (decoder isolation / best-case test).

Question this answers: can a single 768-d EmbeddingGemma sentence embedding be
decoded back into a FAITHFUL crash reasoning? This is the decision gate for the
"predict-embed-decode" reasoning route:

    reasoning TEXT -> EmbeddingGemma (frozen) -> 768-d ideal embedding
                   -> ReverseBERT projector (4 soft tokens) -> Qwen3-0.6B-Base (LoRA)
                   -> reconstructed TEXT

We feed each reasoning's OWN text through the encoder to get the *ideal* target
vector -- the best input the decoder could ever receive. If the decoder fails
even here, no video Predictor could rescue it -> the route is dead.

This script only does STAGE 1: generate reconstructions + BERTScore F1, and
dump an intermediate JSONL. A second script (build_roundtrip_xlsx.py) merges in
the qualitative quality judgments and builds the final workbook.

Decoder code adapted verbatim from github.com/fakerybakery/ReverseBERT (infer.py),
EmbeddingGemma variant, with CPU-safe float32 loading and greedy decoding.
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

import torch
import torch.nn as nn

REPO_ROOT = Path(__file__).resolve().parents[2]

# ---- Data sources -----------------------------------------------------------
TRAIN_JSONL = REPO_ROOT / "dataset" / "teacher_labels" / "teacher_dataset_e3a.jsonl"
VAL_JSONL = REPO_ROOT / "dataset" / "manifests" / "val_e3a.jsonl"
REVIEW_XLSX = REPO_ROOT / "outputs" / "e3a_student_90clips" / "e3a_val_epoch07_review.xlsx"

# ---- Decoder stack (EmbeddingGemma variant) ---------------------------------
ENCODER = "google/embeddinggemma-300m"
BASE_MODEL = "Qwen/Qwen3-0.6B-Base"
LORA_REPO = "mrfakename/ReverseBERT-EmbeddingGemma-300M"

# Ungated fallback (different encoder -- sanity only, NOT the exact VL-JEPA target)
GTE_ENCODER = "Alibaba-NLP/gte-base-en-v1.5"
GTE_LORA_REPO = "mrfakename/ReverseBERT-GTE-Base-EN-1.5"


class EmbeddingProjector(nn.Module):
    """Verbatim from ReverseBERT infer.py: 768 -> (4 x hidden) soft tokens."""

    def __init__(self, input_dim=768, output_dim=2048, num_tokens=4):
        super().__init__()
        self.num_tokens = num_tokens
        self.proj = nn.Sequential(
            nn.Linear(input_dim, output_dim * 2),
            nn.GELU(),
            nn.Linear(output_dim * 2, output_dim * num_tokens),
        )
        self.output_dim = output_dim

    def forward(self, embeddings):
        projected = self.proj(embeddings)
        projected = projected.view(-1, self.num_tokens, self.output_dim)
        return projected


def load_val_t_seconds() -> dict:
    """Map video_id -> t_seconds from the existing review xlsx (col A -> col C)."""
    out = {}
    if not REVIEW_XLSX.exists():
        return out
    try:
        from openpyxl import load_workbook

        wb = load_workbook(REVIEW_XLSX, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            vid = str(row[0]).strip()
            t = row[2] if len(row) > 2 else None
            out[vid] = t
        wb.close()
    except Exception as e:  # noqa: BLE001
        print(f"[warn] could not read t_seconds from review xlsx: {e}")
    return out


def _parse_reason(assistant_target) -> str:
    """teacher rows store reasoning inside assistant_target JSON {'verdict','reason'}."""
    if assistant_target is None:
        return ""
    obj = assistant_target
    if isinstance(obj, str):
        try:
            obj = json.loads(obj)
        except json.JSONDecodeError:
            return obj.strip()
    if isinstance(obj, dict):
        return str(obj.get("reason", "")).strip()
    return str(obj).strip()


def load_rows(
    split: str,
    train_jsonl: Path = TRAIN_JSONL,
    train_field: str = None,
    val_jsonl: Path = VAL_JSONL,
) -> list[dict]:
    """train_field=None -> Stage-1 behavior (teacher_dataset_e3a.jsonl, assistant_target.reason).
    train_field=<name> -> read that field directly (e.g. 'final_reasoning' from
    Teacher_Reasoning_Train_All_Clips.jsonl, matching what reversebert_finetune.py trained on)."""
    rows: list[dict] = []
    t_map = load_val_t_seconds()

    if split in ("train", "both"):
        with open(train_jsonl, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                r = json.loads(line)
                reason = (
                    str(r.get(train_field, "")).strip()
                    if train_field
                    else _parse_reason(r.get("assistant_target"))
                )
                if not reason:
                    continue
                rows.append(
                    {
                        "video_id": str(r.get("video_id", "")).strip(),
                        "dataset_split": "train",
                        "gt_verdict": r.get("gt_verdict") or r.get("final_verdict"),
                        "t_seconds": None,
                        "time_before_event_s": r.get("requested_time_to_event"),
                        "source_reasoning": reason,
                    }
                )

    if split in ("val", "both"):
        with open(val_jsonl, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                r = json.loads(line)
                reason = str(r.get("gt_reasoning_en", "")).strip()
                if not reason:
                    continue
                vid = str(r.get("video_id", "")).strip()
                rows.append(
                    {
                        "video_id": vid,
                        "dataset_split": "val",
                        "gt_verdict": r.get("gt_verdict") or r.get("final_verdict"),
                        "t_seconds": t_map.get(vid),
                        "time_before_event_s": r.get("requested_time_to_event"),
                        "source_reasoning": reason,
                    }
                )

    return rows


def load_decoder(encoder_name, base_model, lora_repo, device, projector_path=None):
    from huggingface_hub import hf_hub_download
    from sentence_transformers import SentenceTransformer
    from transformers import AutoModelForCausalLM, AutoTokenizer
    from peft import PeftModel

    print(f"[load] sentence encoder: {encoder_name}")
    encoder = SentenceTransformer(encoder_name, device=device, trust_remote_code=True)
    encoder.eval()

    print(f"[load] base LLM: {base_model} (float32/{device})")
    tokenizer = AutoTokenizer.from_pretrained(base_model)
    tokenizer.pad_token = tokenizer.eos_token
    tokenizer.padding_side = "right"
    llm = AutoModelForCausalLM.from_pretrained(base_model, torch_dtype=torch.float32)
    llm = llm.to(device)

    # lora_repo may be an HF repo id OR a local fine-tuned adapter dir
    print(f"[load] LoRA adapter: {lora_repo}")
    llm = PeftModel.from_pretrained(llm, lora_repo)
    llm.eval()

    if projector_path is None:
        proj_path = hf_hub_download(repo_id=lora_repo, filename="reverse_bert_projector.pt")
    else:
        proj_path = projector_path
    print(f"[load] projector: {proj_path}")
    projector = EmbeddingProjector(
        input_dim=768, output_dim=llm.config.hidden_size, num_tokens=4
    ).to(device)
    projector.load_state_dict(torch.load(proj_path, map_location=device, weights_only=True))
    projector.eval()

    return encoder, projector, llm, tokenizer


@torch.no_grad()
def reconstruct(text, encoder, projector, llm, tokenizer, max_new_tokens, device):
    emb = encoder.encode(text, convert_to_tensor=True).unsqueeze(0).to(device).float()
    prefix = projector(emb).float()
    out = llm.generate(
        inputs_embeds=prefix,
        max_new_tokens=max_new_tokens,
        do_sample=False,  # greedy -> reproducible
        pad_token_id=tokenizer.eos_token_id,
    )
    return tokenizer.decode(out[0], skip_special_tokens=True).strip()


def compute_bertscore(cands, refs):
    from bert_score import score

    # baseline rescaling is essential: without it, ANY two fluent English
    # sentences score ~0.8 (see Stage-1 finding). Rescaled ~0 = unrelated.
    P, R, F1 = score(cands, refs, lang="en", verbose=True, rescale_with_baseline=True)
    return [round(float(x), 4) for x in F1]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--split", choices=["train", "val", "both"], default="both")
    ap.add_argument("--limit", type=int, default=0, help="smoke: only N rows")
    ap.add_argument("--max-new-tokens", type=int, default=160)
    ap.add_argument("--use-gte", action="store_true", help="ungated GTE fallback encoder")
    ap.add_argument("--lora-path", default=None, help="local fine-tuned adapter dir (overrides emoact repo)")
    ap.add_argument("--projector-path", default=None, help="local reverse_bert_projector.pt (with --lora-path)")
    ap.add_argument("--train-jsonl", default=str(TRAIN_JSONL),
                     help="train-split source jsonl (default: teacher_dataset_e3a.jsonl, 89 rows)")
    ap.add_argument("--train-field", default=None,
                     help="read this field directly instead of assistant_target.reason "
                          "(e.g. final_reasoning, to match what reversebert_finetune.py trained on)")
    ap.add_argument("--val-jsonl", default=str(VAL_JSONL),
                     help="val-split source jsonl (default assumes repo layout; override on RunPod "
                          "where files sit flat, e.g. /workspace/decoder_ft/val_e3a.jsonl)")
    ap.add_argument(
        "--out",
        type=str,
        default=str(REPO_ROOT / "outputs" / "decoder_roundtrip" / "roundtrip_raw.jsonl"),
    )
    args = ap.parse_args()

    device = "cuda" if torch.cuda.is_available() else "cpu"
    encoder_name = GTE_ENCODER if args.use_gte else ENCODER
    lora_repo = args.lora_path or (GTE_LORA_REPO if args.use_gte else LORA_REPO)

    rows = load_rows(
        args.split,
        train_jsonl=Path(args.train_jsonl),
        train_field=args.train_field,
        val_jsonl=Path(args.val_jsonl),
    )
    if args.limit:
        # keep a mix of train+val in smoke
        rows = rows[: args.limit] if args.split != "both" else (rows[:2] + rows[-2:])[: args.limit]
    print(f"[data] {len(rows)} reasonings "
          f"(train={sum(r['dataset_split']=='train' for r in rows)}, "
          f"val={sum(r['dataset_split']=='val' for r in rows)})")

    encoder, projector, llm, tokenizer = load_decoder(
        encoder_name, BASE_MODEL, lora_repo, device, projector_path=args.projector_path
    )

    t0 = time.time()
    for i, r in enumerate(rows):
        r["reconstruction"] = reconstruct(
            r["source_reasoning"], encoder, projector, llm, tokenizer,
            args.max_new_tokens, device,
        )
        if i < 5 or (i + 1) % 10 == 0:
            print(f"\n[{i+1}/{len(rows)}] {r['video_id']} ({r['dataset_split']})")
            print(f"  SRC : {r['source_reasoning'][:200]}")
            print(f"  RECON: {r['reconstruction'][:200]}")
    print(f"\n[gen] done in {time.time()-t0:.1f}s")

    print("[score] computing BERTScore F1 ...")
    f1s = compute_bertscore(
        [r["reconstruction"] for r in rows],
        [r["source_reasoning"] for r in rows],
    )
    for r, f1 in zip(rows, f1s):
        r["bertscore_f1"] = f1

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    # quick console summary
    def _mean(xs):
        xs = [x for x in xs if x is not None]
        return sum(xs) / len(xs) if xs else float("nan")

    tr = [r["bertscore_f1"] for r in rows if r["dataset_split"] == "train"]
    va = [r["bertscore_f1"] for r in rows if r["dataset_split"] == "val"]
    print(f"\n[summary] wrote {len(rows)} rows -> {out_path}")
    print(f"  BERTScore F1  overall={_mean(f1s):.3f}  train={_mean(tr):.3f}  val={_mean(va):.3f}")


if __name__ == "__main__":
    sys.exit(main())
