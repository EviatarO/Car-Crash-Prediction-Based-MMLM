"""
zero_shot_to_xlsx.py
====================

Convert a student zero-shot inference JSONL (e.g. ``zero_shot_train_100_fixed.jsonl``
or ``zero_shot_test_fixed.jsonl``) into a review-friendly Excel workbook,
mirroring the layout of ``outputs/teacher_dataset_v11.xlsx``.

Rows where the model's ``collision_verdict`` disagrees with the ground-truth
label (or where ``parse_error`` is set) are highlighted fully in red so the
user can hand-review the bad predictions at a glance.

Usage
-----

    python student_training/scripts/zero_shot_to_xlsx.py \
        --input_jsonl outputs/zero_shot/zero_shot_train_100_fixed.jsonl \
        --output_xlsx outputs/zero_shot/zero_shot_train_100_fixed.xlsx

    python student_training/scripts/zero_shot_to_xlsx.py \
        --input_jsonl outputs/zero_shot/zero_shot_test_fixed.jsonl \
        --output_xlsx outputs/zero_shot/zero_shot_test_fixed.xlsx

If ``--output_xlsx`` is omitted, the output path defaults to the input path
with a ``.xlsx`` suffix.

Reference
---------
The Excel writer mirrors ``_write_excel()`` in
``teacher_distillation/scripts/Teacher_dataset_distill_v11.py``.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers (mirroring the teacher distill script)
# ---------------------------------------------------------------------------

_JSON_FENCE_RE = re.compile(r"^\s*```(?:json)?\s*|\s*```\s*$", re.IGNORECASE)


def _normalize_verdict(value: object) -> Optional[str]:
    """Return ``"YES"``/``"NO"`` or ``None``."""
    if value is None:
        return None
    text = str(value).strip().upper()
    return text if text in {"YES", "NO"} else None


def _dynamic_objects_to_str(value: object) -> Optional[str]:
    """Mirror of the teacher helper — serialize dynamic objects to a
    single review-friendly string."""
    if value is None:
        return None
    if isinstance(value, list):
        parts: List[str] = []
        for item in value:
            if isinstance(item, dict):
                t = item.get("type", "?")
                pos = item.get("position", "?")
                feat = item.get("feature", "")
                parts.append(f"{t} ({pos}{', ' + feat if feat else ''})")
            else:
                parts.append(str(item))
        return " | ".join(parts) if parts else None
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=True)
    return str(value)


def _parse_raw_response(raw: object) -> Dict[str, Any]:
    """Parse the ``raw_response`` string back into a dict so we can surface
    ``dynamic_objects`` / ``occlusion_check`` / ``time_to_contact`` in the
    XLSX.  Strips ```` ```json ```` fences.  Returns an empty dict on failure."""
    if not isinstance(raw, str) or not raw.strip():
        return {}
    text = raw.strip()
    # Strip a single leading/trailing code fence if present.
    text = _JSON_FENCE_RE.sub("", text).strip()
    # The regex above is anchored; if fences were present internally, try once more.
    if text.startswith("```"):
        text = text.lstrip("`")
        if text.lower().startswith("json"):
            text = text[4:]
        text = text.rstrip("`").strip()
    try:
        parsed = json.loads(text)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

COLUMN_ORDER: List[str] = [
    "video_id",
    "ground_truth",
    "gt_verdict",
    "group",
    "time_before_s",
    "collision_verdict",
    "confidence",
    "score",
    "mismatch",
    "verdict_reasoning",
    "scene_context",
    "temporal_analysis",
    "dynamic_objects",
    "occlusion_check",
    "time_to_contact",
    "latency_s",
    "parse_error",
    "review",
]


def _build_rows(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[bool]]:
    """Build the ordered row dicts and the per-row mismatch flags."""
    rows: List[Dict[str, Any]] = []
    mismatches: List[bool] = []

    for rec in records:
        gt = rec.get("ground_truth")
        try:
            gt_int = int(gt) if gt is not None else None
        except (TypeError, ValueError):
            gt_int = None

        gt_verdict = None
        if gt_int == 1:
            gt_verdict = "YES"
        elif gt_int == 0:
            gt_verdict = "NO"

        pred_verdict = _normalize_verdict(rec.get("collision_verdict"))
        parse_error = rec.get("parse_error")

        mismatch = False
        if parse_error:
            mismatch = True
        elif gt_verdict is None or pred_verdict is None:
            mismatch = True
        elif pred_verdict != gt_verdict:
            mismatch = True

        raw_parsed = _parse_raw_response(rec.get("raw_response"))

        row = {
            "video_id": rec.get("video_id"),
            "ground_truth": gt_int if gt_int is not None else gt,
            "gt_verdict": gt_verdict,
            "group": rec.get("group"),
            "time_before_s": rec.get("time_before_s"),
            "collision_verdict": rec.get("collision_verdict"),
            "confidence": rec.get("confidence"),
            "score": rec.get("score"),
            "mismatch": mismatch,
            "verdict_reasoning": rec.get("verdict_reasoning"),
            "scene_context": rec.get("scene_context"),
            "temporal_analysis": rec.get("temporal_analysis"),
            "dynamic_objects": _dynamic_objects_to_str(raw_parsed.get("dynamic_objects")),
            "occlusion_check": raw_parsed.get("occlusion_check"),
            "time_to_contact": raw_parsed.get("time_to_contact"),
            "latency_s": rec.get("latency_s"),
            "parse_error": parse_error,
            "review": "",
        }
        rows.append(row)
        mismatches.append(mismatch)

    return rows, mismatches


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _write_excel(
    rows: List[Dict[str, Any]],
    mismatches: List[bool],
    output_xlsx: Path,
) -> None:
    df = pd.DataFrame(rows, columns=COLUMN_ORDER)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_xlsx, index=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active

    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    n_cols = len(df.columns)

    # Highlight mismatch rows in red (header row is row 1).
    for row_idx, is_bad in enumerate(mismatches, start=2):
        if is_bad:
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

    # Auto-fit column widths (cap at 60).
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, len(df) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_xlsx)


# ---------------------------------------------------------------------------
# Confusion summary
# ---------------------------------------------------------------------------

def _confusion_counts(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    tp = fp = tn = fn = 0
    for row in rows:
        gt = row.get("gt_verdict")
        pred = _normalize_verdict(row.get("collision_verdict"))
        if gt == "YES" and pred == "YES":
            tp += 1
        elif gt == "YES" and pred == "NO":
            fn += 1
        elif gt == "NO" and pred == "NO":
            tn += 1
        elif gt == "NO" and pred == "YES":
            fp += 1
    return {"TP": tp, "FP": fp, "TN": tn, "FN": fn}


def _accuracy(cm: Dict[str, int]) -> float:
    total = cm["TP"] + cm["FP"] + cm["TN"] + cm["FN"]
    return (cm["TP"] + cm["TN"]) / total if total > 0 else 0.0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as fh:
        for lineno, line in enumerate(fh, start=1):
            raw = line.strip()
            if not raw:
                continue
            try:
                records.append(json.loads(raw))
            except json.JSONDecodeError as exc:
                raise RuntimeError(f"Invalid JSON at {path}:{lineno}: {exc}") from exc
    return records


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert student zero-shot JSONL to review XLSX (bad predictions highlighted red).",
    )
    parser.add_argument("--input_jsonl", required=True, type=Path, help="Path to *_fixed.jsonl")
    parser.add_argument(
        "--output_xlsx",
        type=Path,
        default=None,
        help="Path to write the XLSX (default: same stem as input, .xlsx suffix)",
    )
    args = parser.parse_args()

    input_jsonl: Path = args.input_jsonl
    if not input_jsonl.exists():
        raise SystemExit(f"Input file not found: {input_jsonl}")

    output_xlsx: Path = args.output_xlsx or input_jsonl.with_suffix(".xlsx")

    records = _read_jsonl(input_jsonl)
    rows, mismatches = _build_rows(records)
    _write_excel(rows, mismatches, output_xlsx)

    cm = _confusion_counts(rows)
    total = len(rows)
    n_mismatches = sum(1 for m in mismatches if m)
    acc = _accuracy(cm)

    print(f"Input : {input_jsonl}")
    print(f"Output: {output_xlsx}")
    print(
        f"{input_jsonl.stem}: total={total}  "
        f"TP={cm['TP']}  FN={cm['FN']}  TN={cm['TN']}  FP={cm['FP']}  "
        f"mismatches={n_mismatches}  accuracy={acc:.4f}"
    )


if __name__ == "__main__":
    main()
