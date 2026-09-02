"""
build_a1fail321_comparison.py
===============================
Per-clip comparison workbook for the A1-failure-recovery run: a1cont (crash-only
control) / v10 / v12 / v12shuf, all starting from A1's own LoRA weights
(/workspace/semsup/a1_1761/epoch_04), trained on the 321 windows A1 gets wrong.

STRUCTURALLY DIFFERENT FROM pool1761_arm_comparison.xlsx / summary_vs_A1, and the
reason this is a new script rather than a reused one: EVERY row in this pool is one
A1 already scores wrong at 0.5 - there is no A1-correct clip in this pool to break.
So "broken_FP"/"broken_FN" (present in every prior arm-comparison workbook this
project has built) are STRUCTURALLY always zero here, not a measured finding - they
are shown as "n/a (no A1-correct clips in this pool)" rather than 0, so a reader does
not mistake "nothing broke" for a real result the way a 0 would silently imply.

Only VAL rows (61 clips) get a real post-FT score - --dump-val-scores only scores the
held-out split; the 260 train rows were fit directly and were never independently
scored post-training. The per_clip sheet is val-only for this reason (not a train/val
toggle like the pool1761 workbook), and the sheet header says so explicitly.

A1's AUC on this pool's val split is EXACTLY 0.0 by construction (every val row is
also one A1 gets wrong) - the metrics sheet prints this fact inline as a caveat, not
just a number, since an AP/AUC computed on this pool means something different than
the same metric on a normal held-out set (see select_a1fail321.py's docstring).

"Selected epoch" per arm = argmax val_ap from that arm's own epoch_metrics.jsonl
(matches semsup_train.py's own --select-by convention) - printed in the metrics sheet
so it is auditable, not hidden.

Usage:
  python build_a1fail321_comparison.py \
      --results-dir ../../outputs/a1fail321/results \
      --selection ../../outputs/a1fail321/selection_a1fail321.jsonl \
      --v10-captions ../../outputs/a1fail321/Caption_a1fail321_V10.jsonl \
      --v12-captions ../../outputs/a1fail321/Caption_a1fail321_V12.jsonl \
      --out ../../outputs/a1fail321/a1fail321_arm_comparison.xlsx \
      --arms a1cont v10          # omit --arms to use all 4 once they all exist
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from sklearn.metrics import average_precision_score, roc_auc_score

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ARMS = ["a1cont", "v10", "v12", "v12shuf"]

GREEN = PatternFill(patternType="solid", bgColor="C6EFCE")
RED = PatternFill(patternType="solid", bgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BASE_FILL = PatternFill("solid", fgColor="4A5A80")
NOTE_FONT = Font(italic=True, color="666666", size=9)


def load_jsonl(path):
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def is_ok(score, gt):
    return (score >= 0.5) == (gt == "YES")


def best_epoch(results_dir, arm):
    rows = load_jsonl(results_dir / arm / "fold_01" / "epoch_metrics.jsonl")
    ranked = sorted(rows, key=lambda r: (r["val_ap"], r["epoch"]), reverse=True)
    return ranked[0]["epoch"], ranked[0]["val_ap"]


def load_val_scores(results_dir, arm, epoch):
    path = results_dir / arm / "fold_01" / f"val_scores_ep{epoch:02d}.jsonl"
    rows = load_jsonl(path)
    return {r["frames_dir"]: r["score"] for r in rows}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--results-dir", default=str(ROOT / "outputs" / "a1fail321" / "results"))
    ap.add_argument("--selection", default=str(ROOT / "outputs" / "a1fail321" / "selection_a1fail321.jsonl"))
    ap.add_argument("--v10-captions", default=str(ROOT / "outputs" / "a1fail321" / "Caption_a1fail321_V10.jsonl"))
    ap.add_argument("--v12-captions", default=str(ROOT / "outputs" / "a1fail321" / "Caption_a1fail321_V12.jsonl"))
    ap.add_argument("--out", default=str(ROOT / "outputs" / "a1fail321" / "a1fail321_arm_comparison.xlsx"))
    ap.add_argument("--arms", nargs="+", default=None,
                     help="which arms to include (default: all 4, but only those with "
                          "a completed epoch_metrics.jsonl are actually used)")
    args = ap.parse_args()

    results_dir = Path(args.results_dir)
    sel = {r["frames_dir"]: r for r in load_jsonl(args.selection)}
    v10cap = {r["frames_dir"]: r["caption"] for r in load_jsonl(args.v10_captions)}
    v12cap = {r["frames_dir"]: r["caption"] for r in load_jsonl(args.v12_captions)}

    val_fds = [fd for fd, r in sel.items() if r["split"] == "val"]
    print(f"[load] {len(sel)} total pool rows, {len(val_fds)} val rows")

    candidate_arms = args.arms or DEFAULT_ARMS
    arms, sel_epoch, sel_val_ap, scores = [], {}, {}, {}
    for arm in candidate_arms:
        metrics_path = results_dir / arm / "fold_01" / "epoch_metrics.jsonl"
        if not metrics_path.exists():
            print(f"[skip] {arm}: no epoch_metrics.jsonl yet (still running or not launched)")
            continue
        ep, ap_val = best_epoch(results_dir, arm)
        s = load_val_scores(results_dir, arm, ep)
        missing = [fd for fd in val_fds if fd not in s]
        if missing:
            print(f"[warn] {arm} ep{ep}: missing {len(missing)} val rows, e.g. {missing[:3]} - skipping arm")
            continue
        arms.append(arm)
        sel_epoch[arm] = ep
        sel_val_ap[arm] = ap_val
        scores[arm] = s
        print(f"[ready] {arm}: selected epoch {ep} (val_ap={ap_val:.4f} at training time)")

    if not arms:
        raise SystemExit("no arms have complete results yet")

    # A1's own score on this pool comes from selection (a1_score, stamped by
    # select_a1fail321.py directly from pool1761_scores/A1.jsonl)
    rows = []
    for fd in val_fds:
        r = sel[fd]
        # selection_a1fail321.jsonl leaves `horizon_label` null on all 321 rows and puts
        # the horizon in `requested_time_to_event` instead, so reading only the former
        # produced a blank column. Same fallback build_pool1761_comparison.py already uses.
        row = {"video_id": r["video_id"],
               "window": r.get("horizon_label") or r.get("requested_time_to_event"),
               "gt": r["gt_verdict"],
               "caption_V10": f"V10: {v10cap.get(fd, '(missing)')}",
               "caption_V12": f"V12: {v12cap.get(fd, '(missing)')}",
               "A1": r["a1_score"]}
        for arm in arms:
            row[arm] = scores[arm][fd]
            row[f"{arm}_gap"] = scores[arm][fd] - r["a1_score"]
        rows.append(row)

    wb = Workbook()

    # ================================================================ per_clip
    ws = wb.active
    ws.title = "per_clip"
    ws.append([f"VAL-ONLY ({len(rows)} clips) - train rows have no independent post-FT "
               "score (--dump-val-scores only scores the held-out split) and are not "
               "shown here. A1 = pre-FT score (every row here is one A1 gets WRONG at "
               "0.5, by construction - see select_a1fail321.py). Each arm's own "
               "selected epoch (argmax val_ap) is shown in the metrics sheet."])
    ws.cell(row=1, column=1).font = NOTE_FONT
    headers = (["video_id", "window", "caption_V10", "caption_V12", "gt", "A1"]
               + [c for arm in arms for c in (arm, f"{arm}_gap")])
    ws.append(headers)
    header_row = 2
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in rows:
        line = [r["video_id"], r["window"], r["caption_V10"], r["caption_V12"], r["gt"],
                round(r["A1"], 4)]
        line += [v for arm in arms for v in (round(r[arm], 4), round(r[f"{arm}_gap"], 4))]
        ws.append(line)

    n_rows = len(rows) + header_row
    gt_col_letter = "E"
    a1_col = 6
    score_cols = [a1_col] + [a1_col + 1 + 2 * i for i in range(len(arms))]
    for col_idx in score_cols:
        col = get_column_letter(col_idx)
        rng = f"{col}{header_row + 1}:{col}{n_rows}"
        pass_f = (f'OR(AND({col}{header_row+1}>=0.5,${gt_col_letter}{header_row+1}="YES"),'
                  f'AND({col}{header_row+1}<0.5,${gt_col_letter}{header_row+1}="NO"))')
        fail_f = (f'OR(AND({col}{header_row+1}>=0.5,${gt_col_letter}{header_row+1}="NO"),'
                  f'AND({col}{header_row+1}<0.5,${gt_col_letter}{header_row+1}="YES"))')
        ws.conditional_formatting.add(rng, FormulaRule(formula=[pass_f], fill=GREEN))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[fail_f], fill=RED))
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{n_rows}"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 6
    for i in range(a1_col, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 9

    # ================================================================ summary_vs_A1
    ws2 = wb.create_sheet("summary_vs_A1")
    ws2.append(["A1 baseline on this val split - EVERY row here is one A1 gets WRONG "
                "at 0.5 by construction, so A1's own AUC on this set is exactly 0.0 "
                "(not merely low) and there is NO A1-correct clip in this pool to "
                "break. broken_FP/broken_FN are therefore always n/a here, not a "
                "measured 0 - shown as n/a rather than 0 so that is not misread as "
                "'nothing broke'."])
    ws2.cell(row=1, column=1).font = NOTE_FONT
    ws2.append(["n", "A1_correct", "A1_wrong"])
    for c in range(1, 4):
        cell = ws2.cell(row=2, column=c)
        cell.fill = BASE_FILL
        cell.font = HEADER_FONT
    ws2.append([len(rows), 0, len(rows)])

    ws2.append([])
    header_row2 = ws2.max_row + 1
    headers2 = ["arm", "selected_epoch", "n", "fixed_FP", "fixed_FN", "broken_FP",
                "broken_FN", "still_wrong", "net"]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=header_row2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    a1_fp = [r for r in rows if r["gt"] == "NO"]     # A1 wrong + gt=NO -> A1 said YES (FP)
    a1_fn = [r for r in rows if r["gt"] == "YES"]    # A1 wrong + gt=YES -> A1 said NO (FN)
    for arm in arms:
        ffp = sum(1 for r in a1_fp if is_ok(r[arm], r["gt"]))
        ffn = sum(1 for r in a1_fn if is_ok(r[arm], r["gt"]))
        still = len(rows) - ffp - ffn
        ws2.append([arm, sel_epoch[arm], len(rows), ffp, ffn, "n/a", "n/a", still, ffp + ffn])
    for col, w in {"A": 9, "B": 14}.items():
        ws2.column_dimensions[col].width = w

    # ================================================================ metrics
    ws3 = wb.create_sheet("metrics")
    ws3.append([f"A1's own AUC on this val split is EXACTLY 0.0 by construction (see "
                f"note on summary_vs_A1) - read AP/AUC here as 'did training move these "
                f"clips at all', not as a normal benchmark. selected_epoch = argmax "
                f"val_ap from that arm's own training run."])
    ws3.cell(row=1, column=1).font = NOTE_FONT
    ws3.append(["arm", "selected_epoch", "n", "AP", "AUC", "acc@0.5"])
    header_row3 = ws3.max_row
    for c in range(1, 7):
        cell = ws3.cell(row=header_row3, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    y = [1 if r["gt"] == "YES" else 0 for r in rows]
    a1_scores = [r["A1"] for r in rows]
    ws3.append(["A1", "-", len(rows), round(average_precision_score(y, a1_scores), 4),
                round(roc_auc_score(y, a1_scores), 4),
                round(sum(1 for yy, ss in zip(y, a1_scores) if (ss >= 0.5) == bool(yy)) / len(y), 4)])
    for arm in arms:
        s = [r[arm] for r in rows]
        acc = sum(1 for yy, ss in zip(y, s) if (ss >= 0.5) == bool(yy)) / len(y)
        ws3.append([arm, sel_epoch[arm], len(rows), round(average_precision_score(y, s), 4),
                    round(roc_auc_score(y, s), 4), round(acc, 4)])
    for col, w in {"A": 9, "B": 14}.items():
        ws3.column_dimensions[col].width = w

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[wrote] {out_path}  (arms: {', '.join(arms)})")


if __name__ == "__main__":
    main()
