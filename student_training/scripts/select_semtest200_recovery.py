"""
select_semtest200_recovery.py
===============================
SemTest-200 clip selection, v2 - "recovery-oriented" redesign (2026-08-26).

Supersedes select_semtest200.py's pure closest-to-center fill, which left the TP side
badly ceiling-saturated (56-73/100 clips at score>=0.85, almost no headroom to rise -
see the discussion this replaces). This version builds each bucket from TWO sources
instead of one, matching the user's explicit spec:

  TP buckets (TTE_0.5/1.0/1.5), quota 33/33/34 (corrected 2026-08-26 - the first version
  of this script wrongly narrowed to the 0.3-0.5 near-boundary FN subset; the user's
  actual spec is ALL RT-eligible FN, which is close to sufficient on its own: 32/37/32
  available vs 33/33/34 needed):
    1. ALL RT-eligible FN clips (GT=YES, A0 currently wrong, response_time > TTE) -
       every one of them is a genuine recovery target, not just the ones within 0.2 of
       the boundary.
    2. Fill only the residual per-bucket shortfall (expected: 1 at TTE_0.5, 0 at
       TTE_1.0, 2 at TTE_1.5) from TP candidates (GT=YES, A0 correct, RT-eligible),
       ranked by LOWEST score first - a small patch, not the dominant source.

  TN/MID buckets (MID-4/8/10), quota 33/33/34 - now entirely FP-composed, no TN at all:
    1. ALL near-boundary (0.5<=score<0.7) FP clips.
    2. Fill the remainder from the wider FP pool (GT=NO, A0 wrong, score>=0.7 - the
       near-boundary band is already fully claimed by step 1), ranked by LOWEST score
       first (closest to the 0.7 boundary).

One window per video_id is enforced GLOBALLY within each class (a video can carry up
to 3 TTE windows - all positive-class - so the same used-video set threads across all
3 TP buckets; MID negatives never repeat a video_id at all, confirmed empirically).

Train/val split: per bucket, the combined (recovery + fill) list is sorted by score
ascending (interleaving both sources by how close each clip sits to its own boundary),
then an evenly-strided sample gives the val quota (7/7/6, same as v1) - so val draws
proportionally from both sub-populations rather than concentrating in one.

Every row is tagged with `source` in {FN_recovery, TP_fill, FP_near_boundary,
FP_fill} for provenance in downstream reporting.

Usage:
  python select_semtest200_recovery.py \
    --a0-scores ../../outputs/semtest200/A0_full4446.jsonl \
    --manifest ../../dataset/manifests/train4500_hires.jsonl \
    --train-xlsx ../../dataset/train.xlsx \
    --out-dir ../../outputs/semtest200
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

TTE_RT_MIN = {"TTE_0.5": 0.5, "TTE_1.0": 1.0, "TTE_1.5": 1.5}  # response_time > TTE
TP_BUCKETS = ["TTE_0.5", "TTE_1.0", "TTE_1.5"]
TN_BUCKETS = ["MID-4", "MID-8", "MID-10"]
QUOTA = {"TTE_0.5": 33, "TTE_1.0": 33, "TTE_1.5": 34,
         "MID-4": 33, "MID-8": 33, "MID-10": 34}
VAL_QUOTA = {"TTE_0.5": 7, "TTE_1.0": 7, "TTE_1.5": 6,
             "MID-4": 7, "MID-8": 7, "MID-10": 6}

HEADER_FILL = PatternFill("solid", fgColor="243060")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def load_a0_scores(path):
    out = {}
    for line in open(path, encoding="utf-8"):
        if not line.strip():
            continue
        r = json.loads(line)
        out[r["frames_dir"]] = {"score": float(r["score"]), "gt_verdict": r["gt_verdict"]}
    return out


def load_manifest(path):
    return [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]


def load_response_times(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["train"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        vid = f"{int(row[0]):05d}"
        out[vid] = float(row[4]) if row[4] is not None else None
    return out


def build_row(m, s, rt_val):
    return {"video_id": m["video_id"], "frames_dir": m["frames_dir"],
            "horizon_label": m["horizon_label"], "gt_verdict": s["gt_verdict"],
            "a0_score": s["score"], "response_time": rt_val}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--a0-scores", required=True)
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--train-xlsx", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--exclude-frames-dir", default=None,
                     help="path to a newline-separated file of frames_dir values to exclude "
                          "entirely from every tier - e.g. clips that failed caption QC and "
                          "need a different replacement, not the same one back.")
    ap.add_argument("--tp-fill-max", type=float, default=0.85,
                     help="upper bound (exclusive) on A0 score for tier-2 TP fill. Clips at or "
                          "above this are ceiling-saturated (no headroom to rise) and are "
                          "excluded from the pool entirely. Without this cap tier 3 can never "
                          "fire - see the classification comment. Default 0.85.")
    args = ap.parse_args()

    a0 = load_a0_scores(args.a0_scores)
    manifest = load_manifest(args.manifest)
    rt = load_response_times(args.train_xlsx)
    print(f"[load] a0_scores={len(a0)}  manifest={len(manifest)}  "
          f"response_times={len(rt)} ({sum(1 for v in rt.values() if v is not None)} non-null)")

    # ---- classify every manifest row once ----
    # Three-tier priority for TP buckets (2026-08-26, user spec):
    #   1. fn_near   - FN, RT-eligible, 0.3<=score<0.5 - MAX out this tier first.
    #   2. tp_fill    - TP, RT-eligible - fills any residual after tier 1.
    #   3. fn_wide    - FN, RT-eligible, score<0.3 - LAST resort, only if tiers 1+2
    #      together still can't fill quota.
    # Tier 2 MUST be capped at --tp-fill-max or tier 3 can NEVER fire: there are
    # 639/491/290 RT-eligible TP clips at score>=0.85 per bucket, so an uncapped tp_fill
    # absorbs every residual slot (measured 2026-08-26 - this is exactly why an earlier
    # uncapped run reproduced a 16-FN/84-TP split). The cap excludes the ceiling-saturated
    # mass: a clip already scoring 0.97 has no headroom to rise, which defeats the purpose
    # of a before/after score-movement experiment.
    fn_near = {b: [] for b in TP_BUCKETS}
    fn_wide = {b: [] for b in TP_BUCKETS}
    tp_fill = {b: [] for b in TP_BUCKETS}      # GT=YES, score>=0.5, RT-eligible
    fp_near = {b: [] for b in TN_BUCKETS}    # GT=NO, 0.5<=score<0.7
    fp_fill = {b: [] for b in TN_BUCKETS}    # GT=NO, score>=0.7

    excluded = set()
    if args.exclude_frames_dir:
        with open(args.exclude_frames_dir, encoding="utf-8") as f:
            excluded = {line.strip() for line in f if line.strip()}
        print(f"[exclude] {len(excluded)} frames_dir excluded entirely (caption-QC failures)")

    for m in manifest:
        fd = m["frames_dir"]
        if fd in excluded:
            continue
        s = a0.get(fd)
        if s is None:
            continue
        row = build_row(m, s, rt.get(m["video_id"]))
        if m["event_occurs"] == 1:
            bucket = m["horizon_label"]
            if bucket not in TP_BUCKETS:
                continue
            r_time = row["response_time"]
            rt_ok = r_time is not None and r_time > TTE_RT_MIN[bucket]
            if not rt_ok:
                continue
            if s["score"] < 0.5:
                if s["score"] >= 0.3:
                    fn_near[bucket].append(row)
                else:
                    fn_wide[bucket].append(row)
            elif s["score"] < args.tp_fill_max:
                tp_fill[bucket].append(row)
            # score >= tp_fill_max: ceiling-saturated, deliberately excluded entirely
        else:
            bucket = m["horizon_label"]
            if bucket not in TN_BUCKETS:
                continue
            if s["score"] < 0.5:
                continue
            if s["score"] < 0.7:
                fp_near[bucket].append(row)
            else:
                fp_fill[bucket].append(row)

    for b in TP_BUCKETS:
        print(f"[pool] {b}: FN_near(0.3-0.5)={len(fn_near[b])}  TP_fill_candidates={len(tp_fill[b])}  "
              f"FN_wide(<0.3)={len(fn_wide[b])}  (quota={QUOTA[b]})")
    for b in TN_BUCKETS:
        print(f"[pool] {b}: FP_near_boundary={len(fp_near[b])}  FP_fill_candidates={len(fp_fill[b])}")

    # ---- assemble TP buckets: 3-tier priority (2026-08-26 spec) ----
    #   tier 1: fn_near (0.3-0.5), MAXED OUT first
    #   tier 2: tp_fill, lowest score first, fills any residual after tier 1
    #   tier 3: fn_wide (<0.3), HIGHEST score first (closest to 0.3), only if 1+2 still short
    # Filled TIER-BY-TIER GLOBALLY, not bucket-by-bucket. A positive video_id can supply a
    # window to more than one TTE bucket, and only ONE of them may be used (global dedup),
    # so a per-bucket loop lets whichever bucket runs first consume videos a later bucket
    # needs - measured 2026-08-26: that starved TTE_1.5 by 3 despite 110 unique videos being
    # available for 100 slots. Going tier-first also matches the stated priority: every
    # bucket's tier-1 clips are claimed before any bucket falls back to tier 2, and so on.
    # Within a tier, buckets are served scarcest-supply-first for the same anti-starvation
    # reason.
    used_pos = set()
    tp_selection = {b: [] for b in TP_BUCKETS}
    shortfall = {}

    TIERS = [
        (fn_near, "FN_near", lambda r: -r["a0_score"]),   # closest to 0.5 from below first
        (tp_fill, "TP_fill", lambda r: r["a0_score"]),     # closest to 0.5 from above first
        (fn_wide, "FN_wide", lambda r: -r["a0_score"]),    # closest to 0.3 from below first
    ]
    for pool, source, rank_key in TIERS:
        # scarcest bucket first: fewest still-usable candidates relative to remaining need
        def _supply(b):
            return sum(1 for r in pool[b] if r["video_id"] not in used_pos)
        for bucket in sorted(TP_BUCKETS, key=_supply):
            for r in sorted(pool[bucket], key=rank_key):
                if len(tp_selection[bucket]) >= QUOTA[bucket]:
                    break
                if r["video_id"] in used_pos:
                    continue
                r2 = dict(r); r2["source"] = source
                tp_selection[bucket].append(r2)
                used_pos.add(r["video_id"])

    for bucket in TP_BUCKETS:
        if len(tp_selection[bucket]) < QUOTA[bucket]:
            shortfall[bucket] = QUOTA[bucket] - len(tp_selection[bucket])

    # ---- assemble MID buckets: all near-boundary FP, then lowest-score FP fill (>=0.7) ----
    used_neg = set()
    tn_selection = {}
    for bucket in TN_BUCKETS:
        chosen = []
        for r in fp_near[bucket]:                      # take ALL near-boundary FP
            if r["video_id"] in used_neg:
                continue
            r2 = dict(r); r2["source"] = "FP_near_boundary"
            chosen.append(r2); used_neg.add(r["video_id"])
        remaining = QUOTA[bucket] - len(chosen)
        fill_ranked = sorted(fp_fill[bucket], key=lambda r: r["a0_score"])  # lowest (closest to 0.7) first
        n_fill = 0
        for r in fill_ranked:
            if n_fill >= remaining:
                break
            if r["video_id"] in used_neg:
                continue
            r2 = dict(r); r2["source"] = "FP_fill"
            chosen.append(r2); used_neg.add(r["video_id"])
            n_fill += 1
        tn_selection[bucket] = chosen
        if len(chosen) < QUOTA[bucket]:
            shortfall[bucket] = QUOTA[bucket] - len(chosen)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if shortfall:
        print("\n[STOP] quota shortfall:")
        for b, n in shortfall.items():
            print(f"  {b}: short by {n}")
        with open(out_dir / "selection_recovery_shortfall.json", "w", encoding="utf-8") as f:
            json.dump({"shortfall": shortfall}, f, indent=2)
        return

    # ---- train/val split: sort combined bucket list by score ascending, evenly-strided val ----
    def assign_split(selection):
        for bucket, rows in selection.items():
            ranked = sorted(rows, key=lambda r: r["a0_score"])
            n = len(ranked)
            vq = VAL_QUOTA[bucket]
            stride = max(1, n // vq) if vq else n + 1
            val_idx = {i * stride for i in range(vq) if i * stride < n}
            for i, r in enumerate(ranked):
                r["split"] = "val" if i in val_idx else "train"

    assign_split(tp_selection)
    assign_split(tn_selection)

    all_selected = [r for b in TP_BUCKETS for r in tp_selection[b]] + \
                   [r for b in TN_BUCKETS for r in tn_selection[b]]

    n_val = sum(1 for r in all_selected if r["split"] == "val")
    n_train = len(all_selected) - n_val
    print(f"\n[selection] {len(all_selected)} windows  train={n_train}  val={n_val}")
    from collections import Counter
    print("[source breakdown]", dict(Counter(r["source"] for r in all_selected)))

    headers = ["video_id", "frames_dir", "gt_verdict", "horizon_label", "source",
               "a0_score", "response_time", "split"]

    with open(out_dir / "selection.jsonl", "w", encoding="utf-8") as f:
        for r in all_selected:
            f.write(json.dumps(r) + "\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "selection"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in all_selected:
        ws.append([r["video_id"], r["frames_dir"], r["gt_verdict"], r["horizon_label"],
                   r["source"], round(r["a0_score"], 4), r.get("response_time"), r["split"]])
    widths = {"A": 10, "B": 20, "C": 10, "D": 13, "E": 17, "F": 10, "G": 14, "H": 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(all_selected) + 1}"
    wb.save(out_dir / "selection.xlsx")

    with open(out_dir / "val_vids.txt", "w", encoding="utf-8") as f:
        for r in all_selected:
            if r["split"] == "val":
                f.write(r["video_id"] + "\n")

    print(f"[wrote] {out_dir / 'selection.jsonl'}")
    print(f"[wrote] {out_dir / 'selection.xlsx'}")
    print(f"[wrote] {out_dir / 'val_vids.txt'} ({n_val} ids)")


if __name__ == "__main__":
    main()
