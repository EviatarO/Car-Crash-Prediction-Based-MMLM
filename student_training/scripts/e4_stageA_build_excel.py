"""
e4_stageA_build_excel.py
========================
Stage A combined per-clip Excel for BADAS-Open scores. Merges the Private and
Public eval JSONLs (from e4_stageA_badas_open_eval.py) into ONE sheet.

Columns (in order; no reasoning — BADAS-Open has none):
    video_id | Public/Private | group | time_before_s | ground_truth
            | collision_verdict | score | Score Pred (>=0.5) | Correct?

Row colour follows the SCORE-based prediction (threshold 0.5), matching the
prior convention in build_test_review_excel.py: correct -> green, wrong -> red.

Usage:
    python student_training/scripts/e4_stageA_build_excel.py \
        --private outputs/e4_vjepa_reason/StageA_scorer/badas_open_private.jsonl \
        --public  outputs/e4_vjepa_reason/StageA_scorer/badas_open_public.jsonl \
        --out_xlsx outputs/e4_vjepa_reason/StageA_scorer/badas_open_StageA_scores.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
STAGE_DIR = REPO_ROOT / "outputs" / "e4_vjepa_reason" / "StageA_scorer"

THRESHOLD = 0.5
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")

COLUMNS = ["video_id", "Public/Private", "group", "time_before_s", "ground_truth",
           "collision_verdict", "score", "Score Pred (>=0.5)", "Correct?"]
WIDTHS = {"video_id": 10, "Public/Private": 14, "group": 8, "time_before_s": 14,
          "ground_truth": 13, "collision_verdict": 17, "score": 10,
          "Score Pred (>=0.5)": 16, "Correct?": 10}


def load_rows(path: Path, default_split: str) -> list[dict]:
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                r = json.loads(line)
                r.setdefault("split", default_split)
                rows.append(r)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--private", default=str(STAGE_DIR / "badas_open_private.jsonl"))
    ap.add_argument("--public",  default=str(STAGE_DIR / "badas_open_public.jsonl"))
    ap.add_argument("--out_xlsx", default=str(STAGE_DIR / "badas_open_StageA_scores.xlsx"))
    ap.add_argument("--title", default="BADAS-Open Stage A")
    args = ap.parse_args()

    rows = load_rows(Path(args.private), "Private") + load_rows(Path(args.public), "Public")
    print(f"Loaded {len(rows)} rows (Private + Public)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.title[:31]

    for c, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    n_green = n_red = 0
    for r, row in enumerate(rows, start=2):
        gt = int(row["ground_truth"])
        score = float(row["score"])
        score_pred = 1 if score >= THRESHOLD else 0
        correct = (score_pred == gt)
        n_green += correct
        n_red += (not correct)
        fill = GREEN if correct else RED

        values = {
            "video_id":          str(row.get("video_id", "")).zfill(5),
            "Public/Private":    row.get("split"),
            "group":             row.get("group"),
            "time_before_s":     row.get("time_before_s"),
            "ground_truth":      gt,
            "collision_verdict": row.get("collision_verdict"),
            "score":             round(score, 4),
            "Score Pred (>=0.5)": "YES" if score_pred else "NO",
            "Correct?":          "Y" if correct else "N",
        }
        for c, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=values[col])
            cell.fill = fill
            if col == "score":
                cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center", vertical="top")

    for c, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[col]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    out = Path(args.out_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")
    print(f"  green (score-pred==GT) = {n_green}   red (wrong) = {n_red}")


if __name__ == "__main__":
    main()
