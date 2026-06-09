"""
build_test_review_excel.py
==========================
Per-clip review sheet for the E3a Epoch-7 test results (677 clips).

Reads e3a_test_epoch07.jsonl and writes an Excel with exactly these columns,
in order:
    video_id, group, time_before_s, ground_truth, collision_verdict, score,
    verdict_reasoning

Row color follows the SCORE-based prediction (matches the published CM/F1,
threshold = 0.5), NOT the text verdict:
    correct = (score >= 0.5) == bool(ground_truth)
    correct -> soft green (C6EFCE); wrong -> soft red (FFC7CE)

Output: <metrics_dir>/e3a_test_epoch07_review.xlsx

Usage:
    python student_training/scripts/build_test_review_excel.py
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
JSONL = (REPO_ROOT / "outputs" / "e3a_student_90clips" / "e3a_test_extracted"
         / "outputs" / "trained" / "e3a_test_epoch07.jsonl")
OUT_DIR = (REPO_ROOT / "outputs" / "e3a_student_90clips" / "e3a_test_extracted"
           / "outputs" / "metrics" / "e3a_test_epoch07")
OUT_XLSX = OUT_DIR / "e3a_test_epoch07_review.xlsx"

THRESHOLD = 0.5

# Optional expected counts for the post-build sanity print (green/red).
EXPECTED_GREEN = 465
EXPECTED_RED = 212
COLUMNS = ["video_id", "group", "time_before_s", "ground_truth",
           "collision_verdict", "score", "verdict_reasoning"]

GREEN = PatternFill("solid", fgColor="C6EFCE")   # soft green   = score correct + verdict agrees
RED = PatternFill("solid", fgColor="FFC7CE")     # soft red     = score wrong + verdict wrong (both wrong)
ORANGE = PatternFill("solid", fgColor="FFE0B3")  # soft orange  = score correct BUT verdict disagrees w/ GT
STRONG_ORANGE = PatternFill("solid", fgColor="FF9933")  # strong orange = score WRONG but verdict CORRECT
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
WIDTHS = {"video_id": 10, "group": 8, "time_before_s": 14, "ground_truth": 13,
          "collision_verdict": 17, "score": 10, "verdict_reasoning": 90, "status": 22}


def load_rows(path: Path) -> list[dict]:
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def main():
    ap = argparse.ArgumentParser(description="Build per-clip review Excel from an eval JSONL")
    ap.add_argument("--jsonl", default=str(JSONL))
    ap.add_argument("--out_xlsx", default=str(OUT_XLSX))
    ap.add_argument("--title", default="E3a Test Epoch7")
    ap.add_argument("--expected_green", type=int, default=EXPECTED_GREEN)
    ap.add_argument("--expected_red", type=int, default=EXPECTED_RED)
    ap.add_argument("--color_mode", choices=["score", "tri"], default="score",
                    help="'score' = 2-colour by score>=0.5 (default); "
                         "'tri' = green/red by score + orange when score correct "
                         "but text verdict disagrees with GT (adds a 'status' column)")
    args = ap.parse_args()
    jsonl_path = Path(args.jsonl)
    out_xlsx = Path(args.out_xlsx)
    tri = args.color_mode == "tri"
    columns = COLUMNS + (["status"] if tri else [])

    rows = load_rows(jsonl_path)
    print(f"Loaded {len(rows)} rows from {jsonl_path.name}  (color_mode={args.color_mode})")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.title[:31]   # Excel sheet-name limit

    # Header
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    n_green = n_red = n_orange = n_strong = 0
    for r, row in enumerate(rows, start=2):
        vid = str(row.get("video_id", "")).zfill(5)   # keep as string, pad
        gt = int(row["ground_truth"])
        score = float(row["score"])
        score_pred = 1 if score >= THRESHOLD else 0
        score_correct = (score_pred == gt)
        verdict_pred = 1 if str(row.get("collision_verdict", "")).upper() == "YES" else 0
        verdict_correct = (verdict_pred == gt)

        if not tri:
            fill = GREEN if score_correct else RED
            status = "ok" if score_correct else "score-wrong"
            n_green += score_correct
            n_red += (not score_correct)
        elif score_correct and verdict_correct:
            fill, status = GREEN, "ok"
            n_green += 1
        elif score_correct and not verdict_correct:
            fill, status = ORANGE, "verdict-mismatch"        # score right, verdict wrong
            n_orange += 1
        elif (not score_correct) and verdict_correct:
            fill, status = STRONG_ORANGE, "score-wrong-verdict-right"  # verdict rescued
            n_strong += 1
        else:
            fill, status = RED, "both-wrong"                 # score wrong AND verdict wrong
            n_red += 1

        values = {
            "video_id": vid,
            "group": row.get("group"),
            "time_before_s": row.get("time_before_s"),
            "ground_truth": gt,
            "collision_verdict": row.get("collision_verdict"),
            "score": round(score, 4),
            "verdict_reasoning": row.get("verdict_reasoning"),
            "status": status,
        }
        for c, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=values[col])
            cell.fill = fill
            if col == "verdict_reasoning":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col == "score":
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # Column widths, freeze header, autofilter
    for c, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[col]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    print(f"Saved: {out_xlsx}")
    if tri:
        print(f"  green        (score ok, verdict ok)        = {n_green}")
        print(f"  orange       (score ok, verdict wrong)     = {n_orange}")
        print(f"  strong-orange(score wrong, verdict RIGHT)  = {n_strong}")
        print(f"  red          (both wrong)                  = {n_red}")
        print(f"  check: green+orange={n_green + n_orange} (score TP+TN={args.expected_green}); "
              f"strong+red={n_strong + n_red} (score FP+FN={args.expected_red})")
    else:
        print(f"  green (correct, score-pred==GT) = {n_green}")
        print(f"  red   (wrong)                   = {n_red}")
        print(f"  expected: green={args.expected_green} (TP+TN), red={args.expected_red} (FP+FN)")


if __name__ == "__main__":
    main()
