"""
build_epoch7_excel.py
=====================
Combines GT Excel columns (video_id, gt_verdict, t_seconds, requested_time_to_event,
verdict_reasoning_en) with Epoch 7 model reasoning, a 0-10 quality score, and
score explanation. Outputs: outputs/e3a_val_epoch07_review.xlsx
"""
import json
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]

# ── 1. GT Excel ───────────────────────────────────────────────────────────────
gt_path = REPO_ROOT / "dataset" / "teacher_dataset_GT_self_imply.xlsx"
gt_df = pd.read_excel(gt_path)
# Keep only the 5 requested columns
gt_df = gt_df[["video_id", "gt_verdict", "t_seconds",
               "requested_time_to_event", "verdict_reasoning_en"]].copy()
gt_df["video_id"] = gt_df["video_id"].astype(str).str.zfill(5)

# ── 2. Epoch-7 JSONL (hardcoded from RunPod output) ──────────────────────────
JSONL_LINES = [
    {"video_id": "00319", "ground_truth": 1, "score": 0.3007, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is approaching a crosswalk with a red fire truck stopped on the right side of the road. The fire truck is stationary and blocking the crosswalk, but the ego vehicle is not in a trajectory that would intersect with the fire truck. There is no evidence of a converging path or an obstacle in the ego vehicle's immediate path."},
    {"video_id": "00077", "ground_truth": 1, "score": 0.5703, "collision_verdict": "YES",
     "verdict_reasoning": "The ego vehicle is rapidly closing the distance to the black sedan directly ahead. The gap between the vehicles is critically small, and the closing speed is high, indicating an imminent rear-end collision if the lead vehicle were to brake suddenly."},
    {"video_id": "00687", "ground_truth": 1, "score": 0.5342, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is following a black sedan in its lane. The sedan is maintaining a stable following distance and there is no evidence of sudden braking, lane merging, or any trajectory conflict. The ego vehicle is simply continuing its normal following trajectory."},
    {"video_id": "00283", "ground_truth": 1, "score": 0.7103, "collision_verdict": "YES",
     "verdict_reasoning": "The ego vehicle is rapidly approaching a white SUV that has abruptly cut across its path from the right lane. The distance between the ego vehicle and the SUV is closing rapidly, and the trajectory indicates an imminent crossing conflict with insufficient space and time to avoid a collision."},
    {"video_id": "00147", "ground_truth": 1, "score": 0.7315, "collision_verdict": "YES",
     "verdict_reasoning": "The ego vehicle is rapidly approaching a white sedan directly ahead in its lane. In the final frames, the distance between the vehicles is critically small, indicating an imminent collision due to insufficient braking space and closing speed."},
    {"video_id": "00529", "ground_truth": 1, "score": 0.6668, "collision_verdict": "YES",
     "verdict_reasoning": "The ego vehicle is approaching a silver minivan that has stopped in its lane. The distance between the vehicles is rapidly closing, and the ego vehicle is moving directly towards the minivan. Given the proximity and the parallel trajectories, a rear-end collision is highly likely within the next few seconds."},
    {"video_id": "00493", "ground_truth": 1, "score": 0.6574, "collision_verdict": "YES",
     "verdict_reasoning": "The ego vehicle is rapidly approaching a silver sedan directly ahead in its lane. The sedan is braking, and the closing speed is high, indicating an imminent rear-end collision if the ego vehicle does not brake or decelerate in time."},
    {"video_id": "00474", "ground_truth": 1, "score": 0.7361, "collision_verdict": "YES",
     "verdict_reasoning": "The ego vehicle is rapidly approaching a yellow taxi directly ahead in its lane. The closing speed is high, and the distance between the vehicles is critically short in the final frames, indicating an imminent rear-end collision."},
    {"video_id": "00372", "ground_truth": 1, "score": 0.3777, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is stopped at a red traffic light. The lead vehicle in the adjacent lane is also stopped, and the vehicles in the same lane are maintaining their positions. There are no converging trajectories, crossing agents, or obstacles in the ego vehicle's path. The traffic signal is clear, and the scene is stable."},
    {"video_id": "01153", "ground_truth": 0, "score": 0.447, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is stopped at a red traffic light. A white sedan is crossing the intersection from left to right in the final frames. While the sedan is moving parallel to the ego vehicle, it is not on a direct collision course and has sufficient space to complete its crossing. There is no evidence of a converging trajectory or an imminent crossing conflict."},
    {"video_id": "01504", "ground_truth": 0, "score": 0.5209, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is following a red SUV in its lane. The SUV is moving forward at a steady pace, and the distance between the vehicles is stable. There are no converging trajectories, crossing vehicles, or obstacles in the ego vehicle's path. The traffic flow is orderly, and the ego vehicle is maintaining a safe following distance."},
    {"video_id": "01643", "ground_truth": 0, "score": 0.0845, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is traveling straight on a clear, empty road. There are no visible dynamic agents (pedestrians, cyclists, or other vehicles) in the scene, and the road ahead is unobstructed. The ego vehicle is maintaining a stable trajectory with no evidence of a converging path or an emerging obstacle."},
    {"video_id": "01281", "ground_truth": 0, "score": 0.4877, "collision_verdict": "YES",
     "verdict_reasoning": "The black SUV in the adjacent right lane is crossing the ego vehicle's lane. In the final frames, the SUV is directly parallel to the ego vehicle, and the distance between them is critically small, indicating an imminent lane change conflict."},
    {"video_id": "01550", "ground_truth": 0, "score": 0.293, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is maintaining a stable following distance behind the lead vehicle. There is no evidence of rapid closing speed, lane merging, or any trajectory conflict. The traffic flow is stable and predictable."},
    {"video_id": "01737", "ground_truth": 0, "score": 0.1272, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is traveling on a clear, empty road with no visible obstacles or other vehicles. There are no converging trajectories, crossing agents, or sudden obstacles to indicate an imminent collision."},
    {"video_id": "02104", "ground_truth": 0, "score": 0.3416, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is maintaining a stable trajectory in its lane. The surrounding vehicles are moving parallel or in adjacent lanes, with no evidence of converging paths, crossing trajectories, or any other trajectory conflicts. The traffic flow is orderly, and there are no visible obstacles or sudden braking events that would indicate an imminent collision."},
    {"video_id": "02117", "ground_truth": 0, "score": 0.4566, "collision_verdict": "YES",
     "verdict_reasoning": "A black SUV on the right side of the road pulls out into the ego vehicle's lane. The distance between the ego vehicle and the SUV is rapidly closing, and the SUV is crossing the ego vehicle's path. The ego vehicle is approaching the SUV at a high rate of speed, leaving insufficient time and space to stop, indicating an imminent collision."},
    {"video_id": "01552", "ground_truth": 0, "score": 0.4663, "collision_verdict": "NO",
     "verdict_reasoning": "The ego vehicle is approaching a crosswalk where a yellow school bus is parked. The bus is stationary and there are no visible pedestrians or moving agents entering the crosswalk. The ego vehicle is maintaining a safe following distance and trajectory, with no evidence of an impending collision."},
]

# ── 3. Reasoning quality scores & explanations ────────────────────────────────
# Scoring criteria: correctness of verdict (most important), scene description
# accuracy vs GT, specificity of risk factors identified (0=completely wrong, 10=perfect)
QUALITY = {
    "00319": (3,  "Wrong verdict (FN). Model identifies a fire truck but misses the actual collision agent — a car entering from the right at the intersection. The GT describes a clear cut-in risk that the model does not detect."),
    "00077": (8,  "Correct verdict (TP). Accurately identifies rapid closing gap and rear-end risk with the lead sedan. Matches GT's core observation of the sedan braking. Score well above 0.5."),
    "00687": (4,  "ScoreHead correct (TP, score=0.534≥0.5) but text generation wrong (says NO). Model describes a stable scene and misses the gray SUV drifting into the EGO lane. Score head captures some risk signal, but the reasoning fails to explain why."),
    "00283": (7,  "Correct verdict (TP). Correctly identifies a vehicle cutting across from the right with insufficient avoidance time. GT describes a pickup performing a left turn; model calls it an SUV cut-in — same risk, minor vehicle type imprecision."),
    "00147": (6,  "Correct verdict (TP), but different scenario reading. GT describes EGO deviating into a parallel vehicle's lane; model describes a forward rear-end. Right answer reached via partially wrong scene interpretation."),
    "00529": (6,  "Correct verdict (TP). Captures closing distance and collision risk. Describes vehicle as stopped minivan rather than GT's drifting silver SUV, but the closing-distance risk factor is correctly flagged."),
    "00493": (9,  "Excellent match (TP). Correctly identifies braking lead sedan and the EGO vehicle's failure to decelerate — closely mirrors the GT description. Specific and mechanistically accurate."),
    "00474": (5,  "Correct verdict (TP) but wrong scenario. GT describes a van cutting in from the right; model focuses on the yellow taxi ahead. Correct label derived from the wrong visual evidence."),
    "00372": (2,  "Wrong verdict (FN) and wrong scene description. Model describes a completely static, safe scene (stopped at red light) when the GT shows EGO approaching a sedan that stops for pedestrians — an active rear-end developing."),
    "01153": (6,  "Correct verdict (TN). Reasonably notes the crossing sedan has sufficient space and no collision course. GT confirms a smooth right turn with all vehicles in lanes. Appropriate caution applied correctly."),
    "01504": (6,  "Text verdict correct (NO/TN) but ScoreHead marginally over-predicts risk (score=0.521, just above threshold → FP by score). Reasoning correctly identifies stable following; score barely misses. Minor calibration issue near decision boundary."),
    "01643": (9,  "Excellent match (TN). Both model and GT describe an empty road with no dynamic agents. Model's very low score (0.08) reflects well-calibrated high confidence in the safe label."),
    "01281": (3,  "Wrong verdict (FP). Model sees a lane-change conflict where GT confirms only controlled closing to a braking pickup with no accident. Over-sensitive to adjacent vehicle proximity."),
    "01550": (7,  "Correct verdict (TN). Stable controlled following correctly identified. GT confirms EGO closes gap in a controlled manner — consistent with model's assessment."),
    "01737": (9,  "Excellent match (TN). Both model and GT describe an empty highway with no other vehicles. Very low score (0.13) reflects well-calibrated confidence. Clear and accurate reasoning."),
    "02104": (8,  "Correct verdict (TN). Comprehensive no-collision indicators: no converging trajectories, orderly traffic, no sudden braking. Closely matches GT's description of vehicles maintaining reasonable distances."),
    "02117": (3,  "Wrong verdict (FP). Model describes a SUV cutting into the EGO lane, but GT confirms a stationary van at a crosswalk with vehicles crossing safely. Scene description does not match GT context."),
    "01552": (5,  "Correct verdict (TN) but different scene. GT describes EGO entering a gas station with an exiting SUV; model describes a parked school bus at a crosswalk. Correct conclusion from a mismatched scene reading."),
}

# ── 4. Build merged DataFrame ─────────────────────────────────────────────────
epoch7_lookup = {r["video_id"]: r for r in JSONL_LINES}

rows = []
for _, gtr in gt_df.iterrows():
    vid = str(gtr["video_id"])
    e7 = epoch7_lookup.get(vid, {})
    q_score, q_expl = QUALITY.get(vid, (None, ""))
    score_val = e7.get("score", 0.0)
    # Use score threshold=0.5 for binary classification (matches published metrics)
    score_pred = "YES" if (isinstance(score_val, float) and score_val >= 0.5) else "NO"
    correct = "✓" if (gtr["gt_verdict"] == score_pred) else "✗"
    rows.append({
        "video_id":              vid,
        "GT Verdict":            gtr["gt_verdict"],
        "t_seconds":             gtr["t_seconds"],
        "Time Before Event (s)": gtr["requested_time_to_event"],
        "GT Reasoning (EN)":     gtr["verdict_reasoning_en"],
        "Epoch 7 Text Verdict":  e7.get("collision_verdict", ""),
        "Epoch 7 Score (0→1)":   e7.get("score", ""),
        "Score Pred (≥0.5)":     score_pred,
        "Correct?":              correct,
        "Epoch 7 Reasoning":     e7.get("verdict_reasoning", ""),
        "Quality Score (0–10)":  q_score,
        "Quality Explanation":   q_expl,
    })

merged = pd.DataFrame(rows)

# ── 5. Write Excel with formatting ────────────────────────────────────────────
out_path = REPO_ROOT / "outputs" / "e3a_val_epoch07_review.xlsx"
out_path.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "E3a Val – Epoch 7 Review"

# Styles
HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
GREEN_FILL    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL      = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
YELLOW_FILL   = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
BODY_FONT     = Font(name="Arial", size=9)
WRAP          = Alignment(wrap_text=True, vertical="top")
CENTER        = Alignment(horizontal="center", vertical="top")
THIN          = Side(style="thin", color="AAAAAA")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = list(merged.columns)

# Header row
for col_idx, col_name in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font    = HEADER_FONT
    cell.fill    = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border  = BORDER
ws.row_dimensions[1].height = 32

# Data rows
for row_idx, row in merged.iterrows():
    excel_row = row_idx + 2
    correct_flag = row["Correct?"]
    q = row["Quality Score (0–10)"]

    for col_idx, col_name in enumerate(COLS, 1):
        val = row[col_name]
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        cell.font   = BODY_FONT
        cell.border = BORDER

        # Alignment
        if col_name in ("GT Reasoning (EN)", "Epoch 7 Reasoning", "Quality Explanation"):
            cell.alignment = WRAP
        elif col_name in ("Correct?", "GT Verdict", "Epoch 7 Verdict"):
            cell.alignment = CENTER
        else:
            cell.alignment = Alignment(vertical="top")

        # Row background based on correctness
        if correct_flag == "✓":
            cell.fill = GREEN_FILL
        else:
            cell.fill = RED_FILL

        # Override quality score cell color
        if col_name == "Quality Score (0–10)":
            if isinstance(q, (int, float)):
                if q >= 7:
                    cell.fill = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
                elif q >= 5:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
            cell.alignment = CENTER
            cell.font = Font(name="Arial", size=9, bold=True)

    ws.row_dimensions[excel_row].height = 80

# Column widths
COL_WIDTHS = {
    "video_id":              12,
    "GT Verdict":            11,
    "t_seconds":             11,
    "Time Before Event (s)": 18,
    "GT Reasoning (EN)":     50,
    "Epoch 7 Text Verdict":  16,
    "Epoch 7 Score (0→1)":   16,
    "Score Pred (≥0.5)":     14,
    "Correct?":              10,
    "Epoch 7 Reasoning":     50,
    "Quality Score (0–10)":  16,
    "Quality Explanation":   55,
}
for col_idx, col_name in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)

# Freeze header
ws.freeze_panes = "A2"

wb.save(out_path)
print(f"Saved: {out_path}")

# Summary
correct_count = merged["Correct?"].value_counts().get("✓", 0)
wrong_count   = merged["Correct?"].value_counts().get("✗", 0)
avg_q         = merged["Quality Score (0–10)"].mean()
print(f"Correct verdicts: {correct_count}/18")
print(f"Wrong verdicts:   {wrong_count}/18")
print(f"Avg quality score: {avg_q:.1f}/10")
