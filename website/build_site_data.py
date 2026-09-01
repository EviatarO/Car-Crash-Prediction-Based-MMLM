"""
build_site_data.py
===================
Generates website/site_data.js (window.SITE_DATA = {...}) for the dataset-review site.

Sources:
  - dataset/train.xlsx  sheet 'train'  (all 5 columns; negatives have None times)
  - dataset/test.xlsx   sheet 'test'   (columns A-D ONLY per the user's spec: the rest of
    that workbook is chart/pivot junk, not clip metadata)
  - dataset/manifests/test_tte_curve_{public,private}_manifest.jsonl  -> t_event_s for test
    positives. Verified 2026-08-31: this covers only 284/672 test positives, and no other
    event-time source exists anywhere in either project (checked test_public_labels.csv,
    the sibling Nexar_DataSet/test*.csv - all label-only or chart junk). Uncovered
    positives play with the TN rule and get a visible "no event time" badge on the site.

Emitted per clip: id (5-digit), raw_id, label, time_of_event, time_of_alert,
response_time, group/usage (test only), video URL, thumb URL, video_missing.
URLs are ABSOLUTE-FROM-SERVER-ROOT, where the root is the Thesis/ parent directory
(two levels above MMLM_AI - the repo is nested Thesis/MMLM_For_Cars_Collision_Anticipation/
MMLM_AI, and the mp4s live in the sibling Data-Centric-... project). serve.py uses the
same root, so these URLs resolve as-is.

Output is a .js file (not .json) so index.html can load it with a plain <script> tag -
that also keeps the page functional under file:// where fetch() is blocked.

    python website/build_site_data.py
"""
import json
from pathlib import Path

import openpyxl

MMLM_AI = Path(__file__).resolve().parents[1]           # .../MMLM_AI
THESIS = MMLM_AI.parents[1]                             # .../Thesis  (server root)
NEXAR = THESIS / "Data-Centric-Crash-Prediction-Using-3LC-and-MViT" / "src" / "Nexar_DataSet"
OUT = Path(__file__).resolve().parent / "site_data.js"

# thumbnail fallback chain (first existing wins) - suffixes under dataset/train/<id5>...
TRAIN_THUMB_SUFFIXES = ["_hires_tte05", "_hires_tte10", "_hires_tte15", "_hires_mid10",
                        "_hires_neg4", "_hires_neg8", "_hires_mid0", "_hires", ""]


def url_of(p: Path) -> str:
    """Site URL for a path under the Thesis/ server root (forward slashes)."""
    return "/" + p.relative_to(THESIS).as_posix()


def find_thumb(id5: str, split: str):
    if split == "train":
        for suf in TRAIN_THUMB_SUFFIXES:
            p = MMLM_AI / "dataset" / "train" / f"{id5}{suf}" / "frame_00001.jpg"
            if p.exists():
                return url_of(p)
        return None
    for base in (MMLM_AI / "dataset" / "test_public", MMLM_AI / "dataset" / "test"):
        p = base / f"{id5}_hires" / "frame_00001.jpg"
        if p.exists():
            return url_of(p)
    return None


def build_train():
    wb = openpyxl.load_workbook(MMLM_AI / "dataset" / "train.xlsx",
                                read_only=True, data_only=True)
    ws = wb["train"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    assert [str(h).strip() for h in hdr] == \
        ["id", "time_of_event", "time_of_alert", "target", "response time"], \
        f"train.xlsx header changed: {hdr}"
    clips = []
    for r in rows:
        if r[0] is None:
            continue
        id5 = f"{int(r[0]):05d}"
        video = NEXAR / "train" / f"{id5}.mp4"
        clips.append({
            "id": id5, "raw_id": int(r[0]),
            "time_of_event": round(float(r[1]), 3) if r[1] is not None else None,
            "time_of_alert": round(float(r[2]), 3) if r[2] is not None else None,
            "target": int(r[3]),
            "response_time": round(float(r[4]), 3) if r[4] is not None else None,
            "video": url_of(video), "video_missing": not video.exists(),
            "thumb": find_thumb(id5, "train"),
        })
    clips.sort(key=lambda c: c["raw_id"])   # xlsx is sorted by response time - not useful
    return clips


def load_test_event_times():
    times = {}
    for half in ("public", "private"):
        p = MMLM_AI / "dataset" / "manifests" / f"test_tte_curve_{half}_manifest.jsonl"
        for line in open(p, encoding="utf-8"):
            r = json.loads(line)
            if r.get("t_event_s") is not None:
                times[r["video_id"]] = round(float(r["t_event_s"]), 3)
    return times


def build_test():
    times = load_test_event_times()
    wb = openpyxl.load_workbook(MMLM_AI / "dataset" / "test.xlsx",
                                read_only=True, data_only=True)
    ws = wb["test"]
    rows = ws.iter_rows(min_col=1, max_col=4, values_only=True)   # columns A-D ONLY
    hdr = next(rows)
    assert [str(h).strip() for h in hdr] == ["id", "event_occurs", "Usage", "group"], \
        f"test.xlsx A-D header changed: {hdr}"
    clips = []
    for r in rows:
        if r[0] is None:
            continue
        id5 = f"{int(r[0]):05d}"
        video = NEXAR / "test" / f"{id5}.mp4"
        clips.append({
            "id": id5, "raw_id": int(r[0]),
            "target": int(r[1]),
            "usage": str(r[2]), "group": int(r[3]) if r[3] is not None else None,
            "time_of_event": times.get(id5),      # None for all negatives + 388 positives
            "time_of_alert": None,                 # not recorded anywhere for test
            "response_time": None,
            "video": url_of(video), "video_missing": not video.exists(),
            "thumb": find_thumb(id5, "test"),
        })
    clips.sort(key=lambda c: c["raw_id"])
    return clips


def main():
    train = build_train()
    test = build_test()

    def report(name, clips):
        n_pos = sum(1 for c in clips if c["target"] == 1)
        n_vm = sum(1 for c in clips if c["video_missing"])
        n_tm = sum(1 for c in clips if c["thumb"] is None)
        n_et = sum(1 for c in clips if c["target"] == 1 and c["time_of_event"] is not None)
        print(f"[{name}] clips={len(clips)}  pos={n_pos} neg={len(clips)-n_pos}  "
              f"video_missing={n_vm}  no_thumb={n_tm}  pos_with_event_time={n_et}/{n_pos}")

    report("train", train)
    report("test", test)

    data = {"generated_from": "build_site_data.py",
            "splits": {"train": train, "test": test}}
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("window.SITE_DATA = ")
        json.dump(data, f, separators=(",", ":"))
        f.write(";\n")
    print(f"[wrote] {OUT}  ({OUT.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
